import os
import django
import openpyxl




os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

from accounts.models import Customers
from master.models import RouteMaster
from client_management.models import OutstandingAmount,CustomerOutstanding,CustomerOutstandingReport
from invoice_management.models import Invoice
from django.db import connection



import openpyxl
from django.db import connection

from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from django.db import transaction


from datetime import date
from decimal import Decimal
from django.db import transaction
from django.db.models import Sum,Count




# def update_invoice_status_for_outstanding(customer, outstanding_value):
#     cutoff = date(2025, 12, 1)
#     outstanding_value = Decimal(outstanding_value)

#     # Get invoices before cutoff (latest first)
#     invoices = Invoice.objects.filter(
#         customer=customer,
#         # created_date__date__lt=cutoff,
#         is_deleted=False
#     ).order_by("-created_date")

#     selected_unpaid = []
#     running_total = Decimal("0.00")

#     # STEP 1: Pick invoices until outstanding is covered
#     for inv in invoices:
#         amount = Decimal(inv.amout_total or 0)
#         running_total += amount
#         selected_unpaid.append(inv)

#         if running_total >= outstanding_value:
#             break

#     unpaid_ids = [inv.pk for inv in selected_unpaid]
#     paid_invoices = invoices.exclude(pk__in=unpaid_ids)

#     with transaction.atomic():

#         # STEP 2: Mark selected invoices as UNPAID with received = 0
#         for inv in selected_unpaid:
#             inv.invoice_status = "non_paid"
#             inv.amout_recieved = Decimal("0.00")
#             inv.save(update_fields=["invoice_status", "amout_recieved"])

#         # STEP 3: Mark remaining invoices as PAID
#         for inv in paid_invoices:
#             inv.invoice_status = "paid"
#             inv.amout_recieved = inv.amout_total
#             inv.save(update_fields=["invoice_status","amout_recieved"])

#         # ------------------------------------------------
#         # STEP 4: ADJUST EXTRA USING OLDEST INVOICES
#         # ------------------------------------------------

#         total_unpaid = sum(Decimal(inv.amout_total or 0) for inv in selected_unpaid)
#         extra = total_unpaid - outstanding_value

#         if extra > 0:
#             # OLDEST first
#             oldest_first = sorted(selected_unpaid, key=lambda x: x.created_date)

#             for inv in oldest_first:
#                 if extra <= 0:
#                     break

#                 invoice_total = Decimal(inv.amout_total or 0)
#                 max_receivable = invoice_total - Decimal(inv.amout_recieved or 0)

#                 if max_receivable <= 0:
#                     continue

#                 receive_now = min(extra, max_receivable)

#                 inv.amout_recieved += receive_now
#                 inv.save(update_fields=["amout_recieved"])

#                 extra -= receive_now

#     return selected_unpaid, paid_invoices



# # --------------------------------------------------
# # LOAD EXCEL AND PROCESS EACH CUSTOMER
# # --------------------------------------------------

# excel_path = "Customer_Outstanding_S13.xlsx"
# wb = openpyxl.load_workbook(excel_path)
# sheet = wb.active

# print("\nProcessing Outstanding Customers...\n")


# # Skip Excel header row (start from row 2)
# for row in sheet.iter_rows(min_row=2, values_only=True):

#     outstanding_amount = row[6]   # Column 1: Outstanding Amount
#     custom_id = str(row[1])       # Column 2: Customer ID

#     if not outstanding_amount or not custom_id:
#         continue  # skip empty rows

#     # FIND CUSTOMER
#     customer = Customers.objects.filter(custom_id=custom_id).first()

#     if not customer:
#         print(f"❌ Customer not found → ID: {custom_id}")
#         continue

#     # APPLY INVOICE LOGIC
#     unpaid, paid = update_invoice_status_for_outstanding(customer, Decimal(outstanding_amount))

#     print(f"✔ Customer {custom_id} updated:")
#     print(f"   Outstanding: {outstanding_amount}")
#     print(f"   UNPAID invoices: {[inv.invoice_no for inv in unpaid]}")
#     print(f"   PAID invoices: {[inv.invoice_no for inv in paid]}")
#     print("-----------------------------------------------------")

# print("\n🎉 FINISHED processing all customers!\n")



def mark_all_invoices_paid(customer):
    cutoff = date(2025, 12, 1)
    invoices = Invoice.objects.filter(
        customer=customer,
        # created_date__date__lt=cutoff,
        is_deleted=False
    )

    with transaction.atomic():
        for inv in invoices:
            inv.amout_recieved = inv.amout_total
            inv.invoice_status = "paid"
            inv.save(update_fields=["amout_recieved", "invoice_status"])

    print(f"✔ ALL invoices marked PAID for customer {customer.custom_id}")


# ----------------------------------------------------------------------
# Apply FIFO logic for customers IN the Excel list
# ----------------------------------------------------------------------
# def apply_outstanding_logic(customer, outstanding_value):
#     cutoff = date(2025, 12, 1)
#     outstanding_value = Decimal(outstanding_value)

#     invoices = Invoice.objects.filter(
#         customer=customer,
#         created_date__date__lt=cutoff,
#         is_deleted=False
#     ).order_by("-created_date")  # newest → oldest

#     selected_unpaid = []
#     running = Decimal("0.00")

#     for inv in invoices:
#         amt = Decimal(inv.amout_total or 0)
#         running += amt
#         selected_unpaid.append(inv)

#         if running >= outstanding_value:
#             break

#     unpaid_ids = [inv.pk for inv in selected_unpaid]
#     paid_invoices = invoices.exclude(pk__in=unpaid_ids)

#     with transaction.atomic():
#         # UNPAID invoices
#         for inv in selected_unpaid:
#             inv.invoice_status = "non_paid"
#             inv.save(update_fields=["invoice_status"])

#         # PAID invoices
#         for inv in paid_invoices:
#             inv.invoice_status = "paid"
#             inv.save(update_fields=["invoice_status"])

#     print(f"\n✔ FIFO logic applied for customer {customer.custom_id}")
#     print(f"   Outstanding: {outstanding_value}")
#     print(f"   UNPAID invoices: {[i.invoice_no for i in selected_unpaid]}")
#     print(f"   PAID invoices:   {[i.invoice_no for i in paid_invoices]}")



excel_path = "S43.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

excel_customers = {}  # dict: {"7637": 148, "9001": 200, ...}

for row in sheet.iter_rows(min_row=2, values_only=True):
    customer_id = str(row[1])   # column 2 = Customer ID
    outstanding_val = row[4] 
    print("outstanding value",outstanding_val)
       # column 5 = Outstanding Amount
    if customer_id:
        excel_customers[customer_id] = outstanding_val


# ----------------------------------------------------------------------
# PROCESS ALL ROUTE S-01 CUSTOMERS
# ----------------------------------------------------------------------
ROUTE = "S-43"

route_customers = Customers.objects.filter(
    routes__route_name=ROUTE
)

# print(f"\nProcessing route {ROUTE} — Total customers: {route_customers.count()}\n")
print(f"\Total customers in excel: {len(excel_customers)}\n")

for customer in route_customers:

    cid = str(customer.custom_id)

    if cid not in excel_customers:
        # 🚨 NOT IN EXCEL → Set ALL invoices to PAID
        mark_all_invoices_paid(customer)
        continue

print("\n🎉 Route S-01 processing complete!\n")


# def get_route_credit_invoice_summary(route_name, start_date, end_date):
#     qs = Invoice.objects.filter(
#         customer__routes__route_name=route_name,  # ✅ CORRECT
#         invoice_type='credit_invoice',
#         created_date__date__range=(start_date, end_date),
#         is_deleted=False
#     )

#     summary = qs.aggregate(
#         credit_invoice_count=Count('id'),
#         credit_invoice_total=Sum('amout_total')
#     )

#     return {
#         "credit_invoice_count": summary["credit_invoice_count"] or 0,
#         "credit_invoice_total": summary["credit_invoice_total"] or Decimal("0.00"),
#     }


# result = get_route_credit_invoice_summary(
#     route_name="S-02",
#     start_date=date(2025, 12, 1),
#     end_date=date(2025, 12, 15)
# )

# print(result)



# def reset_route_credit_invoices(route_name, start_date, end_date):
#     """
#     Set amount_received = 0 and invoice_status = non_paid
#     for credit invoices of a route within a date range
#     """

#     qs = Invoice.objects.filter(
#         customer__routes__route_name=route_name,
#         invoice_type='credit_invoice',
#         created_date__date__range=(start_date, end_date),
#         is_deleted=False
#     )

#     updated_count = qs.count()

#     with transaction.atomic():
#         qs.update(
#             amout_recieved=0,
#             invoice_status="non_paid"
#         )

#     return updated_count

# updated = reset_route_credit_invoices(
#     route_name="S-02",
#     start_date=date(2025, 12, 1),
#     end_date=date(2025, 12, 16)
# )

# print(f"Updated invoices count: {updated}")


# def get_customer_last5_credit_invoices(customer, start_date, end_date):
#     """
#     Get last 5 credit invoices for a specific customer
#     """

#     qs = Invoice.objects.filter(
#         customer=customer,
#         invoice_type='credit_invoice',
#         created_date__date__range=(start_date, end_date),
#         is_deleted=False
#     ).order_by('-created_date')[:5]

#     return [
#         {
#             "invoice_no": inv.invoice_no,
#             "date": inv.created_date.date(),
#             "total": inv.amout_total,
#             "received": inv.amout_recieved,
#             "status": inv.invoice_status
#         }
#         for inv in qs
#     ]


# customer = Customers.objects.get(custom_id="5042")

# invoices = get_customer_last5_credit_invoices(
#     customer=customer,
#     start_date=date(2025, 11, 1),
#     end_date=date(2025, 12, 15)
# )

# for inv in invoices:
#     print(inv)



# new method


CUSTOMER_IDS = [
    "15630",
    
]

print(f"\nProcessing {len(CUSTOMER_IDS)} customers...\n")

for cid in CUSTOMER_IDS:
    customer = Customers.objects.filter(custom_id=cid).first()

    if not customer:
        print(f"❌ Customer not found: {cid}")
        continue

    mark_all_invoices_paid(customer)

print("\n🎉 DONE — Selected customers processed.\n")