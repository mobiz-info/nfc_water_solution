from django.core.management.base import BaseCommand
from django.db.models import Sum
from openpyxl import Workbook
from django.utils.timezone import datetime
from client_management.models import (
    Customers,
    OutstandingAmount,
)
from sales_management.models import CollectionPayment


class Command(BaseCommand):
    help = "Compare outstanding for two dates (closing vs opening) per customer"

    def add_arguments(self, parser):
        parser.add_argument('--date1', type=str, required=True, help='First date (YYYY-MM-DD) — e.g., 2025-10-31 (closing)')
        parser.add_argument('--date2', type=str, required=True, help='Second date (YYYY-MM-DD) — e.g., 2025-11-01 (opening)')
        parser.add_argument('--route', type=str, required=False, help='Filter by route name')

    def handle(self, *args, **options):
        date1 = datetime.strptime(options['date1'], '%Y-%m-%d').date()  # Closing date
        date2 = datetime.strptime(options['date2'], '%Y-%m-%d').date()  # Opening date
        route_name = options.get('route')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Outstanding Comparison"

        headers = [
            "Sl.No", "Customer ID", "Customer Name", "Route",
            f"Outstanding as of {date1} (Closing)",
            f"Outstanding as of {date2} (Opening)",
            "Difference"
        ]
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Prepare customer queryset
        customers = Customers.objects.filter(is_guest=False)
        if route_name:
            customers = customers.filter(routes__route_name=route_name)

        total_diff = 0
        row_num = 2

        for index, customer in enumerate(customers, start=1):

            # --- Closing as of date1 (<= date1) ---
            total_outstanding_amount_1 = OutstandingAmount.objects.filter(
                customer_outstanding__customer=customer,
                customer_outstanding__product_type="amount",
                customer_outstanding__created_date__date__lte=date1
            ).aggregate(total=Sum('amount'))['total'] or 0

            total_collection_1 = CollectionPayment.objects.filter(
                customer=customer,
                created_date__date__lte=date1
            ).aggregate(total=Sum('amount_received'))['total'] or 0

            closing_outstanding = total_outstanding_amount_1 - total_collection_1

            # --- Opening as of date2 (< date2) ---
            total_outstanding_amount_2 = OutstandingAmount.objects.filter(
                customer_outstanding__customer=customer,
                customer_outstanding__product_type="amount",
                customer_outstanding__created_date__date__lt=date2
            ).aggregate(total=Sum('amount'))['total'] or 0

            total_collection_2 = CollectionPayment.objects.filter(
                customer=customer,
                created_date__date__lt=date2
            ).aggregate(total=Sum('amount_received'))['total'] or 0

            opening_outstanding = total_outstanding_amount_2 - total_collection_2

            # Difference (should ideally be 0)
            diff = round(closing_outstanding - opening_outstanding, 2)
            total_diff += diff

            if abs(diff) > 0.01 or closing_outstanding != 0 or opening_outstanding != 0:
                sheet.cell(row=row_num, column=1, value=index)
                sheet.cell(row=row_num, column=2, value=customer.custom_id)
                sheet.cell(row=row_num, column=3, value=customer.customer_name)
                sheet.cell(row=row_num, column=4, value=getattr(customer.routes, 'route_name', ''))
                sheet.cell(row=row_num, column=5, value=round(closing_outstanding, 2))
                sheet.cell(row=row_num, column=6, value=round(opening_outstanding, 2))
                sheet.cell(row=row_num, column=7, value=diff)
                row_num += 1

        # Totals row
        sheet.cell(row=row_num + 1, column=4, value="Total Difference")
        sheet.cell(row=row_num + 1, column=7, value=round(total_diff, 2))

        filename = f"Outstanding_Difference_{date1}_to_{date2}.xlsx"
        workbook.save(filename)
        self.stdout.write(self.style.SUCCESS(f"✅ Report generated: {filename}"))
