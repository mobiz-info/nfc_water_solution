# views.py
import json
import random
from decimal import Decimal
from datetime import datetime, date,timedelta

from requests import Response
from apiservices.views import find_customers
from van_management.functions import vanstock_curreptions
from van_management.models import *
from django.db import IntegrityError
from django.http import HttpResponseRedirect, JsonResponse
from django.urls import reverse, reverse_lazy
from django.views import View
from django.shortcuts import render
from django.db.models import Q, ExpressionWrapper
from django.db.models.functions import Round
from django.utils.timezone import make_aware
# from client_management.models import CustodyCustomItems
from coupon_management.models import AssignStaffCouponDetails
from master.models import RouteMaster  # Assuming you have imported RouteMaster
from accounts.models import Customers
from product.models import Product, Product_Default_Price_Level
from sales_management.models import OutstandingLog
from tax_settings.models import Tax
from .forms import *
from django.views.generic import FormView, View
from django.http import JsonResponse
from django.urls import reverse_lazy
from .forms import CashCustomerSaleForm, CreditCustomerSaleForm, CashCouponCustomerSaleForm, CreditCouponCustomerSaleForm
from accounts.models import Customers
import random
import string
from django.db.models import Sum
from django.views.generic import View
from django.shortcuts import render, redirect
from django.db.models import Sum
from .models import  OutstandingLog, SaleEntryLog, SalesExtraModel, SalesTemp, Transaction
from accounts.models import Customers
from .forms import ProductForm ,VansRouteBottleCountAddForm,VansRouteBottleCountDeductForm 
from django.db.models import F, Value
from django.db.models.functions import Coalesce
from django.views.generic import ListView
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from io import BytesIO  

from .models import *
from customer_care.models import *
from client_management.models import *
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph
import xlsxwriter
from .models import *
from openpyxl import Workbook
import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from django.utils import timezone
from van_management.models import Van_Routes,Van,VanProductStock
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from invoice_management.models import *
from van_management.models import Expense
from reportlab.lib.styles import getSampleStyleSheet
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.conf import settings
import openpyxl
from django.db import transaction
from master.functions import generate_receipt_no, log_activity,generate_invoice_no
from django.utils.timezone import now

class TransactionHistoryListView(ListView):
    model = Transaction
    template_name = 'sales_management/transaction_history.html'  # Update with your actual template path
    context_object_name = 'transaction_history_list'

class SaleEntryLogListView(ListView):
    model = SaleEntryLog
    template_name = 'sales_management/sale_entry_history.html'
    context_object_name = 'sale_entry_logs'

class OutstandingLogListView(ListView):
    model = OutstandingLog
    template_name = 'sales_management/oustanding_history.html'  # Update with your actual template path
    context_object_name = 'outstanding_logs'

# views.py

from django.http import JsonResponse

def payment_submit(request):
    if request.method == 'POST' and request.is_ajax():
        # Retrieve form data
        total_amount = request.POST.get('total_amount')
        discount = request.POST.get('discount')
        net_taxable = request.POST.get('net_taxable')
        vat = request.POST.get('vat')
        total_to_collect = request.POST.get('total_to_collect')
        amount_received = request.POST.get('amount_received')
        balance = request.POST.get('balance')

        # Print or process the form data as needed
        print("Total Amount:", total_amount)
        print("Discount:", discount)
        print("Net Taxable:", net_taxable)
        print("VAT:", vat)
        print("Total to Collect:", total_to_collect)
        print("Amount Received:", amount_received)
        print("Balance:", balance)

        
        # Return a JSON response
        return JsonResponse({'message': 'Payment data received successful'}, status=200)
    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)


class SaleEntryLogView(View):
    template_name = 'sales_management/add_sale_entry.html'

    def get(self, request, *args, **kwargs):
        # Create an instance of the form and populate it with GET data
        form = SaleEntryFilterForm(request.GET)

        # Initialize not_found to False
        not_found = False

        # Check if the form is valid
        if form.is_valid():
            # Filter the queryset based on the form data
            route_filter = form.cleaned_data.get('route_name')
            search_query = form.cleaned_data.get('search_query')

            user_li = Customers.objects.all()

            if route_filter:
                user_li = user_li.filter(routes__route_name=route_filter)

            if search_query:
                # Apply search filter on relevant fields of the Customers model
                user_li = user_li.filter(
                    Q(customer_name__icontains=search_query) |
                    Q(building_name__icontains=search_query) |
                    Q(door_house_no__icontains=search_query)
                    # Add more fields as needed
                )

            # Check if the filtered data is empty
            not_found = not user_li.exists()

        else:
            # If the form is not valid, retrieve all customers
            user_li = Customers.objects.all()

        context = {'user_li': user_li, 'form': form, 'not_found': not_found}
        return render(request, self.template_name, context)
    
class CustomerDetailsView(View):
    template_name = 'sales_management/customer_sales_detail.html'

    def get(self, request, pk, *args, **kwargs):
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)
        sales_type = user_det.sales_type
        
        # Retrieve visit schedule data from user details
        visit_schedule_data = user_det.visit_schedule

        if visit_schedule_data:
            # Define a dictionary to map week names to numbers
            week_mapping = {"week1": 1, "week2": 2, "week3": 3, "week4": 4}

            # Initialize an empty list to store the result
            result = []

            # Loop through each day and its associated weeks
            for day, weeks in visit_schedule_data.items():
                for week in weeks:
                    # Extract week number using the week_mapping dictionary
                    week_number = week_mapping.get(week)
                    # Append day, week number, and day name to the result list
                    result.append((day, week_number))

            # Sort the result by week number
            result.sort(key=lambda x: x[1])

            # Prepare data for rendering
            data_for_rendering = []
            for slno, (day, week_number) in enumerate(result, start=1):
                data_for_rendering.append({'slno': slno, 'week': week_number, 'day': day})
        else:
            # If visit_schedule_data is None, handle it appropriately
            data_for_rendering = []

        # Filter AssignStaffCouponDetails based on customer_id
        assign_staff_coupon_details = AssignStaffCouponDetails.objects.filter(
            to_customer_id=user_det.customer_id
        )

        # Join AssignStaffCouponDetails with AssignStaffCoupon and aggregate the sum of remaining_quantity
        total_remaining_quantity = assign_staff_coupon_details.aggregate(
            total_remaining_quantity=Sum('staff_coupon_assign__remaining_quantity')
        )

        # Extract the sum of remaining_quantity from the aggregation result
        sum_remaining_quantity_coupons = total_remaining_quantity.get('total_remaining_quantity', 0)

        # Fetch all data from CustodyCustomItems model related to the user
        custody_items = CustodyCustomItems.objects.filter(customer=user_det)

        # Aggregate sum of coupons, empty bottles, and cash from OutstandingLog
        outstanding_log_aggregates = OutstandingLog.objects.filter(customer=user_det).aggregate(
            total_coupons=Sum('coupons'),
            total_empty_bottles=Sum('empty_bottles'),
            total_cash=Sum('cash')
        )
        # Check if all values in outstanding_log_aggregates are None
        if all(value is None for value in outstanding_log_aggregates.values()):
            outstanding_log_aggregates = None

        # Prepare the product form
        product_form = ProductForm()

        # Remove the coupon_method field from the form if sale type is "CASH" or "CREDIT"
        if sales_type in ["CASH", "CREDIT"]:
            del product_form.fields['coupon_method']
        # Add custody_items and aggregated data to the context
        context = {
            'user_det': user_det,
            'visit_schedule_data': data_for_rendering,
            'custody_items': custody_items,
            'outstanding_log_aggregates': outstanding_log_aggregates,  # Add aggregated data to the context
            'product_form': product_form,  # Add the product form to the context
        }

        return render(request, self.template_name, context)

    def post(self, request, pk, *args, **kwargs):
        # print("dddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd")
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)

        # Process product form submission
        product_form = ProductForm(request.POST)
        if product_form.is_valid():
            # Save the product form data
            product = product_form.save(commit=False)
            product.created_by = request.user.username  # Set the created_by field
            product.save()
            
            log_activity(
                created_by=request.user.username,
                description=f"Product added for customer {user_det.customer_name}"
            )
            
            return redirect('customer_details', pk=pk)
        else:
            # If the form is not valid, re-render the page with the form errors
            context = {
                'user_det': user_det,
                'product_form': product_form,
                # Add other context data as needed
            }
            return render(request, self.template_name, context)

class GetProductsByCategoryView(View):
    def get(self, request):
        category_id = request.GET.get('category_id')
        print(category_id,'category_id')
        products = Product.objects.filter(category_id=category_id).values('product_id', 'product_name')
        print("productsproducts", products)
        return JsonResponse({'products': list(products)})
    


    

class InitiateSaleView(FormView, View):
    template_name = 'sales_management/customer_sale_form.html'
    success_url = reverse_lazy('sale_entry_log_view')

    def get_form_class(self, sales_type):
        if sales_type == 'CASH':
            return CashCustomerSaleForm
        elif sales_type == 'CREDIT':
            return CreditCustomerSaleForm
        elif sales_type == 'CASH COUPON':
            return CashCouponCustomerSaleForm
        elif sales_type == 'CREDIT COUPON':
            return CreditCouponCustomerSaleForm
        else:
            return None

    def get(self, request, *args, **kwargs):
        print("///////////////////////////////////////////////////////////////////////////")
        get_data = request.GET.dict()
        print("get_dataget_dataget_data", get_data)
        customer_id = request.GET.get('customer_id')
        product_id = request.GET.get('product_name')
        print("product_idproduct_id", product_id)
        coupon_method = request.GET.get('coupon_method')
        print("coupon_methodcoupon_method", coupon_method)
        # Fetch the customer object
        customer = Customers.objects.filter(is_guest=False, customer_id=customer_id).first()
        # Check if the customer exists
        # Filter AssignStaffCouponDetails based on customer_id
        assign_staff_coupon_details = AssignStaffCouponDetails.objects.filter(
            to_customer_id=customer_id
        )

        # Join AssignStaffCouponDetails with AssignStaffCoupon and aggregate the sum of remaining_quantity
        total_remaining_quantity = assign_staff_coupon_details.aggregate(
            total_remaining_quantity=Sum('staff_coupon_assign__remaining_quantity')
        )

        # Extract the sum of remaining_quantity from the aggregation result
        sum_remaining_quantity_coupons = total_remaining_quantity.get('total_remaining_quantity', 0)

        print("sum_remaining_quantity_couponssum_remaining_quantity_coupons", sum_remaining_quantity_coupons)
        if customer:
            if coupon_method != 'digital' or customer.sales_type not in ['CREDIT COUPON', 'CASH COUPON']:
                # Get the sales type for the customer
                sales_type = customer.sales_type
                # Instantiate the appropriate form class based on sales type
                form_class = self.get_form_class(sales_type)
                if form_class:
                    form = form_class()
                    # test 
                else:
                    return JsonResponse({'error': 'Invalid sales type'}, status=400)

                return self.render_to_response({'sum_remaining_quantity_coupons': sum_remaining_quantity_coupons, 'sales_type': sales_type, 'customer_id': customer_id, 'product_id': product_id, 'form': form, 'coupon_method': coupon_method})
                
            # Return 404 if customer does not exist
            return JsonResponse({'error': 'digital coupon found'}, status=404)
                            
        # Return 404 if customer does not exist
        return JsonResponse({'error': 'Customer not found'}, status=404)



    def generate_invoice_number(self):
        # Generate a random 10-digit invoice number
        invoice_number = ''.join(random.choices(string.digits, k=10))
        
        # Check if the generated number is unique in Transaction table
        while Transaction.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))
        
        # Check if the generated number is unique in SalesTemp table
        while SalesTemp.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))
        
        return invoice_number
    
    def delete_sales_temp_data(self,invoice_number):
        try:
            sales_temp_entry = SalesTemp.objects.get(invoice_number=invoice_number)
            sales_temp_entry.delete()
            print("Data deleted successfully from SalesTemp table for invoice number:", invoice_number)
        except SalesTemp.DoesNotExist:
            print("No data found in SalesTemp table for invoice number:", invoice_number)

    
    def post(self, request, *args, **kwargs):
        sales_type = request.POST.get('sales_type')
        customer_id = request.POST.get('customer_id')
        product_id = request.POST.get('product_id')
        customer = Customers.objects.filter(is_guest=False, customer_id=customer_id).first()
        product = Product.objects.filter(product_id=product_id).first()
        id_status = request.POST.get('status')
        # Get sales_type from form data
        print("qqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqq", request.POST)
        print("id_statusid_statusid_status", id_status)
        # Instantiate the appropriate form class based on sales type
        form_class = self.get_form_class(sales_type)
        print("form_classform_class", form_class)
        if form_class:
            print("entry_________________!")
            form = form_class(request.POST)
            print("formform", form)
            # Convert form input values to integers
            qty_needed = int(request.POST.get('qty_needed', 0))
            no_of_coupons = int(request.POST.get('no_of_coupons', 0))
            coupon_variations = int(request.POST.get('coupon_variations', 0))
            print("coupon_variationscoupon_variations", coupon_variations)
            empty_bottles = int(request.POST.get('empty_bottles', 0))
            collected_bottles = int(request.POST.get('collected_bottles', 0))
            bottle_variations = int(request.POST.get('bottle_variations', 0))
            id_status = request.POST.get('status')
            invoice_number = self.generate_invoice_number()
            cash=0
            print("id_statusid_status", id_status)
            if id_status == 'PAID':
                print("eeeeeeeeeeeeeeeeeeeeeeeee")
                invoice_number=request.POST.get('invoice_number')
                sales_temp_data = SalesTemp.objects.get(invoice_number=invoice_number).data
                customer = Customers.objects.filter(is_guest=False, customer_id=sales_temp_data['customer_id']).first()
                product = Product.objects.filter(product_id=sales_temp_data['product_id']).first()
                get_tax_percentage= Tax.objects.filter(name="vat").first()
                vat_value= get_tax_percentage/100
                print("__________________________________________________________________________________", invoice_number)
                print("sales_temp_datasales_temp_data", sales_temp_data)
                print("__________________________________________________________________________________")
                # Calculate data based on form fields
                total_amount = (qty_needed + no_of_coupons) * coupon_variations  # Dummy calculation for total_amount
                discount = total_amount * 0.1  # Dummy calculation for discount (10% of total_amount)
                net_taxable = total_amount - discount  # Dummy calculation for net_taxable
                vat = net_taxable * vat_value  # Dummy calculation for VAT (5% of net_taxable)
                received_amount = net_taxable - vat  # Dummy calculation for received_amount (net_taxable - VAT)
                balance = received_amount - total_amount  # Dummy calculation for balance (received_amount - total_amount)
                print("id_statusid_status", id_status)
                cash=total_amount
                # Generate initial invoice number
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='INCOME',
                        amount=total_amount,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=total_amount,
                        discount=discount,
                        net_taxable=net_taxable,
                        vat=vat,
                        received_amount=received_amount,
                        balance=balance,
                        product=product,
                        empty_bottles=empty_bottles,

                        # Add other fields as needed
                    )
                    # Create SalesExtraModel instance with the calculated values
                    sales_extra_model = SalesExtraModel.objects.create(
                        qty_needed=qty_needed,
                        no_of_coupons=no_of_coupons,
                        coupon_variations=coupon_variations,
                        empty_bottles=empty_bottles,
                        collected_bottles=collected_bottles,
                        bottle_variations=bottle_variations,
                        status=id_status,
                        order_number=sale_entry.order_number  # Assuming you want to link it with the same order number
                    )
                    # Create an entry in the OutstandingLog model
                    outstanding_log = OutstandingLog.objects.create(
                        customer_id=customer_id,
                        coupons=coupon_variations,
                        empty_bottles=bottle_variations,
                        cash=balance,
                        created_by=request.user  # Assuming you have a logged-in user
                    )

                    self.delete_sales_temp_data(invoice_number)

                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)

            if id_status == "FOC":
                print("FOC is working")
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='EXPENSE',
                        amount=0,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=0,
                        discount=0,
                        net_taxable=0,
                        vat=0,
                        received_amount=0,
                        balance=0,
                        product=product,
                        empty_bottles=empty_bottles,
                        # Add other fields as needed
                    )

                    # Create SalesExtraModel instance with the calculated values
                    sales_extra_model = SalesExtraModel.objects.create(
                        qty_needed=qty_needed,
                        no_of_coupons=no_of_coupons,
                        coupon_variations=coupon_variations,
                        empty_bottles=empty_bottles,
                        collected_bottles=collected_bottles,
                        bottle_variations=bottle_variations,
                        status=id_status,
                        order_number=sale_entry.order_number  # Assuming you want to link it with the same order number
                    )
                    outstanding_log = OutstandingLog.objects.create(
                        customer_id=customer_id,
                        coupons=coupon_variations,
                        empty_bottles=bottle_variations,
                        cash=0,
                        created_by=request.user  # Assuming you have a logged-in user
                    )

                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)
                
            if id_status == "PENDING":
                print("PENDING is working")
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='INCOME',
                        amount=0,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=0,
                        discount=0,
                        net_taxable=0,
                        vat=0,
                        received_amount=0,
                        balance=0,
                        product=product,
                        empty_bottles=empty_bottles,
                        # Add other fields as needed
                    )
                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)
                
                
            if id_status == "CUSTODY":
                print("CUSTODY is working")
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='INCOME',
                        amount=0,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=0,
                        discount=0,
                        net_taxable=0,
                        vat=0,
                        received_amount=0,
                        balance=0,
                        product=product,
                        empty_bottles=empty_bottles,
                        # Add other fields as needed
                    )
                    outstanding_log = OutstandingLog.objects.create(
                        customer_id=customer_id,
                        coupons=coupon_variations,
                        empty_bottles=bottle_variations,
                        cash=0,
                        created_by=request.user  # Assuming you have a logged-in user
                    )
                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)
            # else:
            #     # Print form name
            #     form_name = form.__class__.__name__  # Get the class name of the form
            #     print(f"{form_name} is not valid. Errors:")
            #     for field, errors in form.errors.items():
            #         print(f"- Field: {field}, Errors: {', '.join(errors)}")
            #     # Return a JsonResponse with the error message and appropriate status code
            #     return JsonResponse({'error': f'{form_name} is not valid'}, status=400)


        # Ensure that you return a response in all code paths
        return JsonResponse({'success': 'Sale entry created successfully'}, status=200)



from django.http import JsonResponse

class PaymentForm(View):
    template_name = 'sales_management/customer_sales_detail.html'

    def generate_invoice_number(self):
        # Generate a random 10-digit invoice number
        invoice_number = ''.join(random.choices(string.digits, k=10))
        # Check if the generated number is unique in Transaction table
        while Transaction.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))
        # Check if the generated number is unique in SalesTemp table
        while SalesTemp.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))        
        return invoice_number

    def post(self, request, *args, **kwargs):
        print("gggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggg", request.POST)
        # Retrieve data from the POST request
        customer_id = request.POST.get('customer_id')
        product_id = request.POST.get('product_id')
        sales_type = request.POST.get('sales_type')
        qty_needed = request.POST.get('qty_needed')
        empty_bottles = request.POST.get('empty_bottles')
        collected_bottles = request.POST.get('collected_bottles')
        status = request.POST.get('status')
        discount_percentage = float(request.POST.get('discount_percentage'))/100
        print('statusstatusstatusstatusstatus', status)
        invoice_number = self.generate_invoice_number()
        if status == "PAID":
            # Create a dictionary to store the data
            data = {
                'customer_id': customer_id,
                'product_id': product_id,
                'sales_type': sales_type,
                'qty_needed': qty_needed,
                'empty_bottles': empty_bottles,
                'collected_bottles': collected_bottles,
                'status': status,
                'discount_percentage': discount_percentage,
            }
            # Create an instance of SalesTemp model and save it to the database
            sales_temp = SalesTemp.objects.create(invoice_number=invoice_number, data=data)

        # Fetch the rate from the Customers table based on the customer_id
        customer = Customers.objects.filter(is_guest=False, customer_id=customer_id).first()
        if customer:
            customer_type = customer.customer_type
            get_customer_default_price = Product_Default_Price_Level.objects.filter(customer_type=customer_type, product_id=product_id).first()
            if get_customer_default_price:
                rate = float(get_customer_default_price.rate)
            else:
                # Handle the case where the rate is not available
                return JsonResponse({'error': 'Rate not available for the customer and product'}, status=400)
            # Fetch tax options from the Tax model
            tax_options = Tax.objects.all().values_list('id', 'name')
            # Calculate the total_amount by multiplying rate with qty_needed
            # Add None at the beginning of the tax options list
            tax_options_list = list(tax_options)
            tax_options_list.insert(0, (None, 'None'))
            
            total_amount = float(qty_needed) * rate

            print("7//////////////////////////////////////////////////////")
            print("qty_needed", qty_needed, "rateraterate", rate)
            print("total_amounttotal_amount", total_amount, type(total_amount))
            print("discount_percentage", discount_percentage, type(discount_percentage))


            discount_amount = total_amount * discount_percentage
            net_taxable = total_amount - discount_amount
            # Calculate VAT (5% of total_amount
            # Query to retrieve product tax percentage
            product_tax_percentage = Product.objects.filter(product_id=product_id).annotate(
                tax_percentage=Coalesce(F('tax__percentage'), Value(0))
            ).values_list('tax_percentage', flat=True).first()

            vat_percentage=product_tax_percentage/100
            print("vat_percentage", vat_percentage)
            vat = total_amount * vat_percentage

            total_to_collect = total_amount + vat
            # Calculate total_to_collect (net_taxable plus VAT)
        else:
            # Handle the case where customer is not found
            return JsonResponse({'error': 'Customer not found'}, status=400)

        # Prepare the response data
        response_data = {
            'total_amount': total_amount,
            'discount_amount': discount_amount,
            'net_taxable':net_taxable,
            'total_to_collect': total_to_collect,
            'vat':vat,
            'invoice_number':invoice_number,
            'sales_type': sales_type,
            'status': status,
        }
        print("response_dataresponse_data", response_data)

        # Return the response
        return JsonResponse(response_data)


class CalculateTotaltoCollect(View):
    def get(self, request):
        print("colllectAmount__calc")
        vat_value = request.GET.get('vat_value')
        # products = Product.objects.filter(vat_value=vat_value).values('product_id', 'product_name')
        # print("productsproducts", products)
        return JsonResponse({'products': list(vat_value)})



# class CouponSaleView(View):
#     template_name = 'sales_management/coupon_sale.html'

#     def get(self, request, *args, **kwargs):
#         # Create an instance of the form and populate it with GET data
#         form = SaleEntryFilterForm(request.GET)

#         # Initialize not_found to False
#         not_found = False

#         # Check if the form is valid
#         if form.is_valid():
#             # Filter the queryset based on the form data
#             route_filter = form.cleaned_data.get('route_name')
#             search_query = form.cleaned_data.get('search_query')

#             user_li = Customers.objects.all()

#             if route_filter:
#                 user_li = user_li.filter(routes__route_name=route_filter)

#             if search_query:
#                 # Apply search filter on relevant fields of the Customers model
#                 user_li = user_li.filter(
#                     Q(customer_name__icontains=search_query) |
#                     Q(building_name__icontains=search_query) |
#                     Q(door_house_no__icontains=search_query)
#                     # Add more fields as needed
#                 )

#             not_found = not user_li.exists()

#         else:
#             # If the form is not valid, retrieve all customers
#             user_li = Customers.objects.all()

#         context = {'user_li': user_li, 'form': form, 'not_found': not_found}
#         return render(request, self.template_name, context)



#     def post(self, request, *args, **kwargs):
        
    
class CouponSaleView(View):
    template_name = 'sales_management/coupon_sale.html'

    def get(self, request, *args, **kwargs):
        # Create an instance of the form and populate it with GET data
        form = SaleEntryFilterForm(request.GET)

        # Initialize not_found to False
        not_found = False

        # Check if the form is valid
        if form.is_valid():
            # Filter the queryset based on the form data
            route_filter = form.cleaned_data.get('route_name')
            search_query = form.cleaned_data.get('search_query')

            user_li = Customers.objects.all()

            if route_filter:
                user_li = user_li.filter(routes__route_name=route_filter)

            if search_query:
                user_li = user_li.filter(
                    Q(customer_name__icontains=search_query) |
                    Q(building_name__icontains=search_query) |
                    Q(door_house_no__icontains=search_query)
                )

            not_found = not user_li.exists()

        else:
            user_li = CustomerCoupons.objects.filter()

        balance_coupons = user_li
        print(balance_coupons,'balance_coupons')

        # Get the manual book type last purchased
        # manual_book_type_last_purchased = user_li.latest('created_date').deposit_type

        context = {'user_li': user_li, 'form': form, 'not_found': not_found}
        return render(request, self.template_name, context)
    

class DetailsView(View):
    template_name = 'sales_management/detail.html'

    def get(self, request, pk, *args, **kwargs):
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)
        sales_type = user_det.sales_type
        
        # Retrieve visit schedule data from user details
        visit_schedule_data = user_det.visit_schedule

        if visit_schedule_data:
            # Define a dictionary to map week names to numbers
            week_mapping = {"week1": 1, "week2": 2, "week3": 3, "week4": 4}

            # Initialize an empty list to store the result
            result = []

            # Loop through each day and its associated weeks
            for day, weeks in visit_schedule_data.items():
                for week in weeks:
                    # Extract week number using the week_mapping dictionary
                    week_number = week_mapping.get(week)
                    # Append day, week number, and day name to the result list
                    result.append((day, week_number))

            # Sort the result by week number
            # result.sort(key=lambda x: x[1])

            # Prepare data for rendering
            data_for_rendering = []
            for slno, (day, week_number) in enumerate(result, start=1):
                data_for_rendering.append({'slno': slno, 'week': week_number, 'day': day})
        else:
            # If visit_schedule_data is None, handle it appropriately
            data_for_rendering = []

        # Filter AssignStaffCouponDetails based on customer_id
        assign_staff_coupon_details = AssignStaffCouponDetails.objects.filter(
            to_customer_id=user_det.customer_id
        )

        # Join AssignStaffCouponDetails with AssignStaffCoupon and aggregate the sum of remaining_quantity
        total_remaining_quantity = assign_staff_coupon_details.aggregate(
            total_remaining_quantity=Sum('staff_coupon_assign__remaining_quantity')
        )

        # Extract the sum of remaining_quantity from the aggregation result
        sum_remaining_quantity_coupons = total_remaining_quantity.get('total_remaining_quantity', 0)

        # Fetch all data from CustodyCustomItems model related to the user
        custody_items = CustodyCustomItems.objects.filter(customer=user_det)

        # Aggregate sum of coupons, empty bottles, and cash from OutstandingLog
        outstanding_log_aggregates = OutstandingLog.objects.filter(customer=user_det).aggregate(
            total_coupons=Sum('coupons'),
            total_empty_bottles=Sum('empty_bottles'),
            total_cash=Sum('cash')
        )
        # Check if all values in outstanding_log_aggregates are None
        if all(value is None for value in outstanding_log_aggregates.values()):
            outstanding_log_aggregates = None

        # Prepare the product form
        product_form = ProductForm()

        # Remove the coupon_method field from the form if sale type is "CASH" or "CREDIT"
        if sales_type in ["CASH", "CREDIT"]:
            del product_form.fields['coupon_method']
        # Add custody_items and aggregated data to the context
        context = {
            'user_det': user_det,
            'visit_schedule_data': data_for_rendering,
            'custody_items': custody_items,
            'outstanding_log_aggregates': outstanding_log_aggregates,  # Add aggregated data to the context
            'product_form': product_form,  # Add the product form to the context
        }

        return render(request, self.template_name, context)

    def post(self, request, pk, *args, **kwargs):
        print("dddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd")
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)

        # Process product form submission
        product_form = ProductForm(request.POST)
        if product_form.is_valid():
            # Save the product form data
            product = product_form.save(commit=False)
            product.created_by = request.user.username  # Set the created_by field
            product.save()
            return redirect('customer_details', pk=pk)
        else:
            # If the form is not valid, re-render the page with the form errors
            context = {
                'user_det': user_det,
                'product_form': product_form,
                # Add other context data as needed
            }
            return render(request, self.template_name, context)
        
#------------------SALES REPORT-------------------------        

        
def salesreport(request):
    instances = CustomUser.objects.filter(user_type='Salesman')
    print("instances",instances)
    return render(request, 'sales_management/sales_report.html', {'instances': instances})

def salesreportview(request, salesman):
    salesman = get_object_or_404(CustomUser, pk=salesman)
    customer_supplies = CustomerSupply.objects.filter(salesman=salesman)
    customer_supply_items = CustomerSupplyItems.objects.filter(customer_supply__in=customer_supplies)

    # Calculate counts of products for each sales type
    cash_coupon_counts = customer_supplies.filter(customer__sales_type='CASH COUPON').aggregate(total_products=Count('customersupplyitems'))

    # print("cash_coupon_counts",cash_coupon_counts)
    credit_coupon_counts = customer_supplies.filter(customer__sales_type='CREDIT COUPON').aggregate(total_products=Count('customersupplyitems'))
    # print("credit_coupon_counts",credit_coupon_counts)
    cash_counts = customer_supplies.filter(customer__sales_type='CASH').aggregate(total_products=Count('customersupplyitems'))
    # print("cash_counts",cash_counts)
    credit_counts = customer_supplies.filter(customer__sales_type='CREDIT').aggregate(total_products=Count('customersupplyitems'))
    # print("credit_counts",credit_counts)
    
    context = {
        'salesman': salesman,
        'customer_supplies':customer_supplies,
        'customer_supply_items':customer_supply_items,
        'cash_coupon_counts': cash_coupon_counts['total_products'],
        'credit_coupon_counts': credit_coupon_counts['total_products'],
        'cash_counts': cash_counts['total_products'],
        'credit_counts': credit_counts['total_products'],
    }

    return render(request, 'sales_management/salesreportview.html', context)

def download_salesreport_pdf(request):
    # Retrieve sales report data
    customer_supplies = CustomerSupply.objects.all()
    # Check if customer_supplies is not empty
    if customer_supplies:
        # Create the HTTP response with PDF content type and attachment filename
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

        # Create a PDF document
        pdf_buffer = BytesIO()
        pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        data = []

        # Add headers
        headers = ['Customer Name', 'Customer Address', 'Customer Type', 'Customer Sales Type', 
                   'Cash Coupon Quantity','Credit Coupon Quantity','Cash Quantity','Credit Quantity']
        data.append(headers)


        # Add data to the PDF document
        sl_no = 1
        for supply in customer_supplies:
            try:
                cash_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH COUPON').count()
                credit_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT COUPON').count()
                cash_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH').count()
                credit_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT').count()
            except Exception as e:
                print(e)
                cash_coupon_counts = 0
                credit_coupon_counts = 0
                cash_counts = 0
                credit_counts = 0
            
            # Append data for each supply
            data.append([
                sl_no,
                supply.customer.customer_name,
                f"{supply.customer.building_name} {supply.customer.door_house_no}",
                supply.customer.customer_type,
                supply.customer.sales_type,
                cash_coupon_counts,
                credit_coupon_counts,
                cash_counts,
                credit_counts
            ])
            sl_no += 1

        table = Table(data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])
        table.setStyle(style)

        # Add the table to the PDF document
        elements = [table]
        pdf.build(elements)

        # Get the value of the BytesIO buffer and write it to the response
        pdf_value = pdf_buffer.getvalue()
        pdf_buffer.close()
        response.write(pdf_value)

        return response
    else:
        return HttpResponse("No data available for sales report.")
    

def download_salesreport_excel(request):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ['Sl No', 'Customer Name', 'Customer Address', 'Customer Type', 
               'Customer Sales Type', 'Cash Coupon Quantity', 'Credit Coupon Quantity', 
               'Cash Quantity', 'Credit Quantity']
    worksheet.append(headers)

    # Retrieve sales report data
    customer_supplies = CustomerSupply.objects.all()

    # Add data to the worksheet
    sl_no = 1
    for supply in customer_supplies:
        try:
            
            cash_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH COUPON').count()
            print("cash_coupon_counts",cash_coupon_counts)
            credit_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT COUPON').count()
            print("credit_coupon_counts",credit_coupon_counts)
            cash_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH').count()
            print("cash_counts",cash_counts)
            credit_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT').count()
            print("credit_counts",credit_counts)

        except Exception as e:
            print(e)
            cash_coupon_counts = 0
            credit_coupon_counts = 0
            cash_counts = 0
            credit_counts = 0
        
        # Write data for each supply to the worksheet
        row_data = [
            sl_no,
            supply.customer.customer_name,
            f"{supply.customer.building_name} {supply.customer.door_house_no}",
            supply.customer.customer_type,
            supply.customer.sales_type,
            cash_coupon_counts,
            credit_coupon_counts,
            cash_counts,
            credit_counts
        ]
        worksheet.append(row_data)

        sl_no += 1

    # Create HTTP response with Excel content type and attachment filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

    # Write the workbook data into the response
    workbook.save(response)

    # Write headers
    headers = ['Customer Name', 'Customer Address', 'Product', 'Quantity', 'Amount']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write data to the Excel workbook
    row = 1  # Start from row 1 since row 0 is for headers
    for supply in customer_supplies:
        for item in supply.customersupplyitems_set.all():
            customer_address = ", ".join(filter(None, [supply.customer.building_name, supply.customer.door_house_no, supply.customer.floor_no]))
            product_name = "Product Not Found"
            if item.product:  # Check if product is not None
                product_name = item.product.product_name
            worksheet.write(row, 0, supply.customer.customer_name)
            worksheet.write(row, 1, customer_address)
            worksheet.write(row, 2, product_name)
            worksheet.write(row, 3, item.quantity)
            worksheet.write(row, 4, item.amount)
            row += 1

    workbook.close()

    return response

#------------------Collection Report-------------------------                
# def collectionreport(request):
#     start_date = None
#     end_date = None
#     selected_date = None
#     selected_route_id = None
#     selected_route = None
#     template = 'sales_management/collection_report.html'
#     colectionpayment = CollectionPayment.objects.all()
    
#     routes = RouteMaster.objects.all()
#     route_counts = {}
#     today = datetime.datetime.today()
    
#     if request.method == 'POST':
#         start_date = request.POST.get('start_date')
#         end_date = request.POST.get('end_date')
#         selected_date = request.POST.get('date')
#         selected_route_id = request.POST.get('selected_route_id')
#         if start_date and end_date:
#             colectionpayment = colectionpayment.filter(customer_supply__created_date__range=[start_date, end_date])
#         elif selected_date:
#             colectionpayment = colectionpayment.filter(customer_supply__created_date=selected_date)
        
#         if selected_route_id:
#             selected_route = RouteMaster.objects.get(id=selected_route_id)
#             colectionpayment = colectionpayment.filter(customer__routes__route_name=selected_route)
    
#     # /
    
#     context = {
#         'colectionpayment': colectionpayment, 
#         'routes': routes, 
#         'route_counts': route_counts, 
#         'today': today,
#         'start_date': start_date, 
#         'end_date': end_date, 
#         'selected_date': selected_date, 
#         'selected_route_id': selected_route_id, 
#         'selected_route': selected_route,
        
#     }
#     return render(request, template, context)



# def collection_report_excel(request):
#     instances = CollectionPayment.objects.all()
#     route_filter = request.GET.get('route_name')
#     start_date_str = request.GET.get('start_date')
#     end_date_str = request.GET.get('end_date')
    
#     if start_date_str and end_date_str:
#         start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         instances = instances.filter(customer__customer_supply__created_date__range=[start_date, end_date])
    
#     print('route_filter :', route_filter)
#     if route_filter and route_filter != '' and route_filter != 'None':
#         instances = instances.filter(routes__route_name=route_filter)

#     route_li = RouteMaster.objects.all()
#     serial_number = 1
#     for customer in instances:
#         customer.serial_number = serial_number
#         serial_number += 1
#     data = {
#         'Serial Number': [customer.serial_number for customer in instances],
#         'Date': [customer.customer_supply.created_date.date() for customer in instances],
#         'Customer name': [customer.customer.customer_name for customer in instances],
#         'Mobile No': [customer.customer.mobile_no for customer in instances],
#         'Route': [customer.customer.routes.route_name if customer.customer.routes else '' for customer in instances],
#         'Building Name': [customer.customer.building_name for customer in instances],
#         'House No': [customer.customer.door_house_no if customer.customer.door_house_no else 'Nil' for customer in instances],
#         'Receipt No/Reference No': [customer.customer_supply.reference_number for customer in instances],
#         'Amount': [customer.amount for customer in instances],
#         'Mode of Payment': [customer.payment_method for customer in instances],

#     }
#     df = pd.DataFrame(data)

#     buffer = BytesIO()
#     with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
#         df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)
#         workbook = writer.book
#         worksheet = writer.sheets['Sheet1']
#         table_border_format = workbook.add_format({'border':1})
#         worksheet.conditional_format(4, 0, len(df.index)+4, len(df.columns) - 1, {'type':'cell', 'criteria': '>', 'value':0, 'format':table_border_format})
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'border': 1})
#         worksheet.merge_range('A1:J2', f'National Water', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A3:J3', f'    Collection Report   ', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A4:J4', '', merge_format)
    
#     filename = f"Collection Report.xlsx"
#     response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'inline; filename = "{filename}"'
#     return response


# def dailycollectionreport(request):
#     instances = CollectionPayment.objects.all()
#     route_filter = request.GET.get('route_name')
#     start_date_str = request.GET.get('start_date')

#     if start_date_str :
#         start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         instances = instances.filter(created_date__range=start_date)
#     if route_filter:
#             instances = instances.filter(routes__route_name=route_filter)
#     route_li = RouteMaster.objects.all()
    
#     context = {'instances': instances,'route_li':route_li}
#     return render(request, 'sales_management/daily_collection_report.html', context)


# def daily_collection_report_excel(request):
#     instances = CollectionPayment.objects.all()
#     route_filter = request.GET.get('route_name')
#     start_date_str = request.GET.get('start_date')
    
#     if start_date_str :
#         start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         instances = instances.filter(customer__customer_supply__created_date__range=start_date)
    
#     print('route_filter :', route_filter)
#     if route_filter and route_filter != '' and route_filter != 'None':
#         instances = instances.filter(routes__route_name=route_filter)

#     route_li = RouteMaster.objects.all()
#     serial_number = 1
#     for customer in instances:
#         customer.serial_number = serial_number
#         serial_number += 1
#     data = {
#         'Serial Number': [customer.serial_number for customer in instances],
#         'Customer name': [customer.customer.customer_name for customer in instances],
#         'Mobile No': [customer.customer.mobile_no for customer in instances],
#         'Route': [customer.customer.routes.route_name if customer.customer.routes else '' for customer in instances],
#         'Building Name': [customer.customer.building_name for customer in instances],
#         'House No': [customer.customer.door_house_no if customer.customer.door_house_no else 'Nil' for customer in instances],
#         'Receipt No/Reference No': [customer.customer_supply.reference_number for customer in instances],
#         'Amount': [customer.amount for customer in instances],
#         'Mode of Payment': [customer.payment_method for customer in instances],
#         'Invoice': [customer.invoice.invoice_no for customer in instances],
#         'Invoice Reference No': [customer.invoice.reference_no  for customer in instances],


#     }
#     df = pd.DataFrame(data)

#     buffer = BytesIO()
#     with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
#         df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)
#         workbook = writer.book
#         worksheet = writer.sheets['Sheet1']
#         table_border_format = workbook.add_format({'border':1})
#         worksheet.conditional_format(4, 0, len(df.index)+4, len(df.columns) - 1, {'type':'cell', 'criteria': '>', 'value':0, 'format':table_border_format})
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'border': 1})
#         worksheet.merge_range('A1:J2'SanaSana Water', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A3:J3', f'    Daily Collection Report   ', merge_format)
#         # worksheet.merge_range('E3:H3', f'Date: {def_date}', merge_format)
#         # worksheet.merge_range('I3:M3', f'Total bottle: {total_bottle}', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A4:J4', '', merge_format)
    
#     filename = f"Daily Collection Report.xlsx" 
#     response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'inline; filename = "{filename}"'
#     return response

#------------------Product-Route wise sales report
from itertools import chain
def product_route_salesreport(request):
    template = 'sales_management/product_route_salesreport.html'
    filter_data = {}

    selected_product_id = request.GET.get('selected_product_id')
    selected_product = None
    if selected_product_id:
        selected_product = get_object_or_404(ProdutItemMaster, id=selected_product_id)
        filter_data["product_id"] = selected_product.pk

    route_filter = request.GET.get('route_name')

    # Start with all customers
    user_li = Customers.objects.all()
    query = request.GET.get("q")
    if query:
        user_li = user_li.filter(
            Q(custom_id__icontains=query) |
            Q(customer_name__icontains=query) |
            Q(sales_type__icontains=query) |
            Q(mobile_no__icontains=query) |
            Q(routes__route_name__icontains=query) |
            Q(location__location_name__icontains=query) |
            Q(building_name__icontains=query)
        )

    if route_filter:
        user_li = user_li.filter(routes__route_name=route_filter)
        filter_data["route_name"] = route_filter

    route_li = RouteMaster.objects.all()
    customersupplyitems = CustomerSupplyItems.objects.all()
    coupons_collected = CustomerSupplyCoupon.objects.all()
    products = ProdutItemMaster.objects.all()

    # Date filters
    if request.GET.get('start_date'):
        start_date = datetime.datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()

    if request.GET.get('end_date'):
        end_date = datetime.datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    else:
        end_date = datetime.datetime.today().date()

    customersupplyitems = customersupplyitems.filter(
        customer_supply__created_date__date__gte=start_date,
        customer_supply__created_date__date__lte=end_date
    )
    coupons_collected = coupons_collected.filter(
        customer_supply__created_date__date__gte=start_date,
        customer_supply__created_date__date__lte=end_date,
        customer_supply__customer__sales_type='CASH COUPON'
    )

    # Format for HTML input
    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    filter_data['end_date'] = end_date.strftime('%Y-%m-%d')

    combined_coupons = []
    totals = {}

    # ✅ If coupons selected
    if selected_product and selected_product.category and selected_product.category.category_name.lower() == "coupons":
        customersupplyitems = CustomerSupplyItems.objects.none()  # no direct rows for coupons

        # manual/free coupons
        coupons_collected = CustomerSupplyCoupon.objects.filter(
            customer_supply__created_date__date__gte=start_date,
            customer_supply__created_date__date__lte=end_date,
            customer_supply__customer__sales_type="CASH COUPON",
            leaf__coupon__coupon_type__coupon_type_name=selected_product
        )
        if route_filter:
            coupons_collected = coupons_collected.filter(
                customer_supply__customer__routes__route_name=route_filter
            )

        # digital coupons
        digital_coupons = CustomerSupplyDigitalCoupon.objects.filter(
            customer_supply__created_date__date__gte=start_date,
            customer_supply__created_date__date__lte=end_date
        )
        if route_filter:
            digital_coupons = digital_coupons.filter(
                customer_supply__customer__routes__route_name=route_filter
            )

        # Merge both into one list for template
        combined_coupons = list(chain(coupons_collected, digital_coupons))

        # Totals
        total_manual_coupon = coupons_collected.aggregate(total=Count('leaf'))['total'] or 0
        total_free_coupon = coupons_collected.aggregate(total=Count('free_leaf'))['total'] or 0
        total_digital_coupon = digital_coupons.aggregate(total=Sum('count'))['total'] or 0

        totals = {
            'total_quantity': 0,
            'total_empty_bottle': coupons_collected.aggregate(
                total=Sum('customer_supply__collected_empty_bottle')
            )['total'] or 0,
            'total_amount_collected': 0,
            'total_coupon_collected': total_manual_coupon + total_free_coupon + total_digital_coupon,
            'manual_coupon_total': total_manual_coupon,
            'free_coupon_total': total_free_coupon,
            'digital_coupon_total': total_digital_coupon,
        }

    else:
        # ✅ Normal product handling
        if selected_product:
            customersupplyitems = customersupplyitems.filter(product=selected_product)
            if route_filter:
                customersupplyitems = customersupplyitems.filter(
                    customer_supply__customer__routes__route_name=route_filter
                )

        totals = customersupplyitems.aggregate(
            total_quantity=Sum('quantity'),
            total_empty_bottle=Sum('customer_supply__collected_empty_bottle'),
            total_amount_collected=Sum('amount'),
        )

        total_coupon_collected = CustomerSupplyCoupon.objects.filter(
            customer_supply__in=customersupplyitems.values('customer_supply')
        ).annotate(
            leaf_count=Count('leaf')
        ).aggregate(
            total=Sum('leaf_count')
        )['total'] or 0

        totals['total_coupon_collected'] = total_coupon_collected

    # Log activity
    log_activity(
        created_by=request.user,
        description=f"Generated product route sales report with filters: {filter_data}"
    )

    context = {
        'customersupplyitems': customersupplyitems.order_by("-customer_supply__created_date"),
        'products': products,
        'filter_data': filter_data,
        'combined_coupons': combined_coupons,
        'selected_product': selected_product,
        'route_li': route_li,
        'totals': totals,
    }
    return render(request, template, context)



def download_product_sales_excel(request):
    selected_product_id = request.GET.get('selected_product_id')
    route_filter = request.GET.get('route_name')

    start_date = datetime.datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    end_date = datetime.datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()

    selected_product = None
    if selected_product_id:
        selected_product = get_object_or_404(ProdutItemMaster, id=selected_product_id)

    customersupplyitems = CustomerSupplyItems.objects.filter(
        customer_supply__created_date__date__gte=start_date, customer_supply__created_date__date__lte=end_date
    )

    if selected_product:
        customersupplyitems = customersupplyitems.filter(product=selected_product)

    if route_filter:
        customersupplyitems = customersupplyitems.filter(
            customer_supply__customer__routes__route_name=route_filter
        )

    coupons_collected = CustomerSupplyCoupon.objects.filter(
        customer_supply__in=customersupplyitems.values('customer_supply')
    ).annotate(leaf_count=Count('leaf'))

    data = []

    total_quantity = 0
    total_empty_bottle = 0
    total_amount_collected = 0
    total_coupon_collected = 0

    for item in customersupplyitems:
        coupon_count = coupons_collected.filter(customer_supply=item.customer_supply).count()

        row = {
            'Time of Supply': item.customer_supply.created_date.strftime('%d/%m/%Y'),
            'Invoice No': item.customer_supply.reference_number,
            'Route Name': item.customer_supply.customer.routes.route_name,
            'Customer Name': item.customer_supply.customer.customer_name,
            'Customer ID': item.customer_supply.customer.custom_id,
            'Mode of Supply': item.customer_supply.customer.sales_type,
            'Quantity': item.quantity,
            'Empty Bottle Collected': item.customer_supply.collected_empty_bottle,
            'Amount Collected': item.amount,
            'Coupon Collected': coupon_count
        }
        data.append(row)

        # Add to totals
        total_quantity += item.quantity
        total_empty_bottle += item.customer_supply.collected_empty_bottle
        total_amount_collected += item.amount
        total_coupon_collected += coupon_count

    # Create DataFrame
    df = pd.DataFrame(data)

    # Add footer row with totals
    footer = {
        'Time of Supply': 'Total',
        'Invoice No': '',
        'Route Name': '',
        'Customer Name': '',
        'Customer ID': '',
        'Mode of Supply': '',
        'Quantity': total_quantity,
        'Empty Bottle Collected': total_empty_bottle,
        'Amount Collected': total_amount_collected,
        'Coupon Collected': total_coupon_collected
    }

    # Convert footer to DataFrame and concatenate with original DataFrame
    footer_df = pd.DataFrame([footer])
    df = pd.concat([df, footer_df], ignore_index=True)

    # Write DataFrame to Excel buffer
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Product Sales', index=False)

        # Accessing the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Product Sales']

        # Format the footer row
        
        # Apply the format to the footer row (last row)
        worksheet.set_row(len(df), None)

    # Prepare HTTP response with Excel file
    filename = f"Product_Sales_Report_{start_date}_{end_date}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Log activity
    log_activity(
        created_by=request.user,
        description=f"Generated product route sales Excel report"
    )
    
    return response


def download_product_sales_print(request):
    template = 'sales_management/product_route_sales_print.html'
    filter_data = {}

    selected_product_id = request.GET.get('selected_product_id')

    selected_product = None
    if selected_product_id:
        selected_product = get_object_or_404(ProdutItemMaster, id=selected_product_id)
        filter_data["product_id"] = selected_product.pk

    route_filter = request.GET.get('route_name')

    # Start with all customers
    user_li = Customers.objects.all()

    query = request.GET.get("q")
    if query:
        user_li = user_li.filter(
            Q(custom_id__icontains=query) |
            Q(customer_name__icontains=query) |
            Q(sales_type__icontains=query) |
            Q(mobile_no__icontains=query) |
            Q(routes__route_name__icontains=query) |
            Q(location__location_name__icontains=query) |
            Q(building_name__icontains=query)
        )

    if route_filter:
        user_li = user_li.filter(routes__route_name=route_filter)
        filter_data["route_name"] = route_filter

    route_li = RouteMaster.objects.all()
    customersupplyitems = CustomerSupplyItems.objects.all()
    coupons_collected = CustomerSupplyCoupon.objects.all()
    products = ProdutItemMaster.objects.all()
    # today = datetime.datetime.today().date()
    
    if request.GET.get('start_date'):
        start_date = datetime.datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()
    if request.GET.get('end_date'):
        end_date = datetime.datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    else:
        end_date = datetime.datetime.today().date()

    customersupplyitems = customersupplyitems.filter(
        customer_supply__created_date__date__gte=start_date, customer_supply__created_date__date__lte=end_date
    )
    coupons_collected = coupons_collected.filter(
        customer_supply__created_date__date__gte=start_date, customer_supply__created_date__date__lte=end_date,
        customer_supply__customer__sales_type='CASH COUPON'
    )

    # Convert to the required format for the HTML date input
    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    filter_data['end_date'] = end_date.strftime('%Y-%m-%d')

    # Apply product filter if selected_product is provided
    if selected_product:
        customersupplyitems = customersupplyitems.filter(product=selected_product)

    # Apply route filter if route_filter is provided
    if route_filter:
        customersupplyitems = customersupplyitems.filter(
            customer_supply__customer__routes__route_name=route_filter
        )
        coupons_collected = coupons_collected.filter(
            customer_supply__customer__routes__route_name=route_filter,
            customer_supply__customer__sales_type='CASH COUPON'
        )
        filter_data["route_name"] = route_filter
        
    totals = customersupplyitems.aggregate(
        total_quantity=Sum('quantity'),
        total_empty_bottle=Sum('customer_supply__collected_empty_bottle'),
        total_amount_collected=Sum('amount'),
    )
    
    total_coupon_collected = CustomerSupplyCoupon.objects.filter(
        customer_supply__in=customersupplyitems.values('customer_supply')
    ).annotate(leaf_count=Count('leaf')).aggregate(total=Sum('leaf_count'))['total'] or 0

    totals['total_coupon_collected'] = total_coupon_collected
    
    log_activity(
        created_by=request.user,
        description=f"Generated product route sales print report with filters: {filter_data}"
    )
    context = {
        'customersupplyitems': customersupplyitems.order_by("-customer_supply__created_date"),
        'products': products,
        'filter_data': filter_data,
        'coupons_collected': coupons_collected,
        'route_li': route_li,
        'totals':totals,
    }
    return render(request, template, context)


# def yearmonthsalesreport(request):
#     user_li = Customers.objects.all()
#     # user_li = user_li.filter(routes__route_name=route_filter)
#     route_li = RouteMaster.objects.all()

#     context = {
#         'user_li': user_li, 
#         'route_li': route_li,
        
#             }
#     # route_li = RouteMaster.objects.all()
#     # print(route_li,'route_li')
#     # context = {'route_li': route_li}

#     return render(request,'sales_management/yearmonthsalesreport.html',context)
def yearmonthsalesreport(request):
    user_li = Customers.objects.all()
    route_li = RouteMaster.objects.all()

    route_sales = []
    current_year = datetime.datetime.now().year
    current_month = datetime.datetime.now().month


    for route in route_li:
        ytd_sales = CustomerSupply.objects.filter(customer__routes=route, created_date__year=current_year).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0
        mtd_sales = CustomerSupply.objects.filter(customer__routes=route, created_date__year=current_year, created_date__month=current_month).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0

        route_sales.append({
            'route': route,
            'ytd_sales': ytd_sales,
            'mtd_sales': mtd_sales,
            'year': current_year,
        })

    context = {
        'user_li': user_li, 
        'route_li' : route_li,
        'route_sales': route_sales,
    }

    return render(request, 'sales_management/yearmonthsalesreport.html', context)

# def yearmonthsalesreportview(request,route_id):
#     route = RouteMaster.objects.get(route_id = route_id)
#     print(route,'route')
#     route_customer = Customers.objects.filter(is_guest=False, routes__route_name=route)
#     print(route_customer,'route_customer')

#     context = {
#         'route_customer': route_customer
#         }
    
#     return render(request, 'sales_management/yearmonthsalesreportview.html',context)
# def yearmonthsalesreportview(request, route_id):
#     route = RouteMaster.objects.get(route_id=route_id)
#     route_customers = Customers.objects.filter(is_guest=False, routes__route_name=route)

#     # Calculate YTD and MTD sales for each customer
#     today = date.today()
#     ytd_sales = CustomerSupply.objects.filter(customer__in=route_customers, created_date__year=today.year).aggregate(total_ytd_sales=Sum('grand_total'))['total_ytd_sales'] or 0
#     mtd_sales = CustomerSupply.objects.filter(customer__in=route_customers, created_date__year=today.year, created_date__month=today.month).aggregate(total_mtd_sales=Sum('grand_total'))['total_mtd_sales'] or 0

#     # Associate YTD and MTD sales with each customer
#     customers_with_sales = []
#     for customer in route_customers:
#         ytd_customer_sales = CustomerSupply.objects.filter(customer=customer, created_date__year=today.year).aggregate(total_ytd_sales=Sum('grand_total'))['total_ytd_sales'] or 0
#         mtd_customer_sales = CustomerSupply.objects.filter(customer=customer, created_date__year=today.year, created_date__month=today.month).aggregate(total_mtd_sales=Sum('grand_total'))['total_mtd_sales'] or 0

#         customers_with_sales.append({
#             'customer': customer,
#             'ytd_sales': ytd_customer_sales,
#             'mtd_sales': mtd_customer_sales
#         })

#     context = {
#         'route': route,
#         'customers_with_sales': customers_with_sales,
#         'ytd_sales': ytd_sales,
#         'mtd_sales': mtd_sales
#     }

#     return render(request, 'sales_management/yearmonthsalesreportview.html', context)

def yearmonthsalesreportview(request, route_id):
    route = RouteMaster.objects.get(route_id=route_id)
    current_year = datetime.datetime.now().year

    yearly_sales = CustomerSupply.objects.filter(customer__routes=route, created_date__year=current_year).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0

    customers = Customers.objects.filter(is_guest=False, routes__route_name=route)
    monthly_sales = []
    for customer in customers:
        monthly_sales_data = []
        for month in range(1, 13):
            month_date = datetime(current_year, month, 1)

            month_name = month_date.strftime("%B")  # Get month name
            monthly_sales_amount = CustomerSupply.objects.filter(customer=customer, created_date__year=current_year, created_date__month=month).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0
            monthly_sales_data.append({month_name: monthly_sales_amount})  # Append month name and sales amount
        monthly_sales.append({'customer': customer, 'monthly_sales': monthly_sales_data})

    context = {
        'route_id': route_id,
        'yearly_sales': yearly_sales,
        'monthly_sales': monthly_sales,
    }
    return render(request, 'sales_management/yearmonthsalesreportview.html', context)

#---------------------New Sales Report-----------------------------


def customerSales_report(request):
    filter_data = {}

    total_amount = 0
    total_discount = 0
    total_net_payable = 0
    total_vat = 0
    total_grand_total = 0
    total_amount_recieved = 0

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sales_type_filter = request.GET.get('sales_type')
    route_filter = request.GET.get('route')

    if start_date_str and end_date_str:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()
        end_date = datetime.datetime.today().date()

    filter_data = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'sales_type': sales_type_filter,
        'route': route_filter,
    }

    # Filter CustomerSupply data
    sales = CustomerSupply.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).exclude(customer__sales_type__in=["CASH COUPON", "CREDIT COUPON"])
    
    if sales_type_filter:
        sales = sales.filter(customer__sales_type=sales_type_filter)
        
    if route_filter:
        sales = sales.filter(customer__routes__pk=route_filter)
    
    sales = sales.order_by("-created_date")
    

    # Filter CustomerCoupon data
    coupons = CustomerCoupon.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    )
    
    if sales_type_filter:
        coupons = coupons.filter(payment_type=sales_type_filter)
    
    if route_filter:
        coupons = coupons.filter(customer__routes__pk=route_filter)

    coupons = coupons.order_by("-created_date")

    # Organize the data for rendering in the template
    sales_report_data = []

    # Process CustomerSupply data
    for sale in sales:
        sales_report_data.append({
            'date': sale.created_date.date(),
            'ref_invoice_no': sale.reference_number,
            'invoice_number': sale.invoice_no,
            'customer_name': sale.customer.customer_name,
            'custom_id': sale.customer.custom_id,
            'building_name': sale.customer.building_name,
            'sales_type': sale.customer.sales_type,
            'route_name': sale.customer.routes.route_name,
            'salesman': sale.customer.sales_staff.get_fullname(),
            'amount': sale.grand_total,
            'discount': sale.discount,
            'net_taxable': sale.subtotal,
            'vat_amount': sale.vat,
            'grand_total': sale.grand_total,
            'amount_collected': sale.amount_recieved,
        })
        
        total_amount += sale.grand_total
        total_discount += sale.discount
        total_net_payable += sale.net_payable
        total_vat += sale.vat
        total_grand_total += sale.grand_total
        total_amount_recieved += sale.amount_recieved

    # Process CustomerCoupon data
    for coupon in coupons:
        sales_report_data.append({
            'date': coupon.created_date.date(),
            'ref_invoice_no': coupon.reference_number,
            'invoice_number': coupon.invoice_no,
            'customer_name': coupon.customer.customer_name,
            'custom_id': coupon.customer.custom_id,
            'building_name': coupon.customer.building_name,
            'sales_type': coupon.customer.sales_type,
            'route_name': coupon.customer.routes.route_name,
            'salesman': coupon.customer.sales_staff.get_fullname(),
            'amount': coupon.grand_total,
            'discount': coupon.discount,
            'net_taxable': coupon.net_amount,
            'vat_amount': Tax.objects.get(name="VAT").percentage,
            'grand_total': coupon.grand_total,
            'amount_collected': coupon.amount_recieved,
        })
        
        total_amount += coupon.grand_total
        total_discount += coupon.discount
        total_net_payable += coupon.net_amount
        total_vat += Tax.objects.get(name="VAT").percentage
        total_grand_total += coupon.grand_total
        total_amount_recieved += coupon.amount_recieved

    context = {
        'customersales': sales_report_data,
        'total_amount': total_amount,
        'total_discount': total_discount,
        'total_net_payable': total_net_payable,
        'total_vat': total_vat,
        'total_grand_total': total_grand_total,
        'total_amount_recieved': total_amount_recieved,
        'filter_data': filter_data,
        'sales_types': ['CASH', 'CREDIT'],  # Available filter options
        'route_li': RouteMaster.objects.all(),
    }
    return render(request, 'sales_management/customerSales_report.html', context)
    # filter_data = {}

    # total_amount = 0
    # total_discount = 0
    # total_net_payable = 0
    # total_vat = 0
    # total_grand_total = 0
    # total_amount_recieved = 0

    # start_date_str = request.GET.get('start_date')
    # end_date_str = request.GET.get('end_date')
    # sales_type_filter = request.GET.get('sales_type')

    # if start_date_str and end_date_str:
    #     start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
    #     end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    # else:
    #     start_date = datetime.datetime.today().date()
    #     end_date = datetime.datetime.today().date()

    # filter_data = {
    #     'start_date': start_date.strftime('%Y-%m-%d'),
    #     'end_date': end_date.strftime('%Y-%m-%d'),
    #     'sales_type': sales_type_filter,
    # }

    # # Filter CustomerSupply data
    # sales = CustomerSupply.objects.select_related('customer', 'salesman').filter(
    #     created_date__date__gte=start_date,
    #     created_date__date__lte=end_date
    # ).exclude(customer__sales_type__in=["CASH COUPON", "CREDIT COUPON"])
    
    # if sales_type_filter:
    #     sales = sales.filter(customer__sales_type=sales_type_filter)
    
    # sales = sales.order_by("-created_date")

    # # Filter CustomerCoupon data
    # coupons = CustomerCoupon.objects.select_related('customer', 'salesman').filter(
    #     created_date__date__gte=start_date,
    #     created_date__date__lte=end_date
    # )
    
    # if sales_type_filter:
    #     coupons = coupons.filter(payment_type=sales_type_filter)

    # coupons = coupons.order_by("-created_date")

    # # Filter CustodyCustom data
    # custody_customs = CustodyCustom.objects.select_related('customer').filter(
    #     created_date__date__gte=start_date,
    #     created_date__date__lte=end_date
    # )

    # if sales_type_filter:
    #     custody_customs = custody_customs.filter(customer__sales_type=sales_type_filter)

    # custody_customs = custody_customs.order_by("-created_date")

    # # Organize the data for rendering in the template
    # sales_report_data = []

    # # Process CustomerSupply data
    # for sale in sales:
    #     sales_report_data.append({
    #         'date': sale.created_date.date(),
    #         'ref_invoice_no': sale.reference_number,
    #         'invoice_number': sale.invoice_no,
    #         'customer_name': sale.customer.customer_name,
    #         'building_name': sale.customer.building_name,
    #         'sales_type': sale.customer.sales_type,
    #         'route_name': sale.customer.routes.route_name,
    #         'salesman': sale.customer.sales_staff.get_fullname(),
    #         'amount': sale.grand_total,
    #         'discount': sale.discount,
    #         'net_taxable': sale.subtotal,
    #         'vat_amount': sale.vat,
    #         'grand_total': sale.grand_total,
    #         'amount_collected': sale.amount_recieved,
    #     })
        
    #     total_amount += sale.grand_total
    #     total_discount += sale.discount
    #     total_net_payable += sale.net_payable
    #     total_vat += sale.vat
    #     total_grand_total += sale.grand_total
    #     total_amount_recieved += sale.amount_recieved

    # # Process CustomerCoupon data
    # for coupon in coupons:
    #     sales_report_data.append({
    #         'date': coupon.created_date.date(),
    #         'ref_invoice_no': coupon.reference_number,
    #         'invoice_number': coupon.invoice_no,
    #         'customer_name': coupon.customer.customer_name,
    #         'building_name': coupon.customer.building_name,
    #         'sales_type': coupon.customer.sales_type,
    #         'route_name': coupon.customer.routes.route_name,
    #         'salesman': coupon.customer.sales_staff.get_fullname(),
    #         'amount': coupon.grand_total,
    #         'discount': coupon.discount,
    #         'net_taxable': coupon.net_amount,
    #         'vat_amount': Tax.objects.get(name="VAT").percentage,
    #         'grand_total': coupon.grand_total,
    #         'amount_collected': coupon.amount_recieved,
    #     })
        
    #     total_amount += coupon.grand_total
    #     total_discount += coupon.discount
    #     total_net_payable += coupon.net_amount
    #     total_vat += Tax.objects.get(name="VAT").percentage
    #     total_grand_total += coupon.grand_total
    #     total_amount_recieved += coupon.amount_recieved

    # # Process CustodyCustom data
    # for custody in custody_customs:
    #     # Fetch related items
    #     custody_items = CustodyCustomItems.objects.filter(custody_custom=custody)
        
    #     for item in custody_items:
    #         sales_report_data.append({
    #             'date': custody.created_date.date(),
    #             'ref_invoice_no': custody.reference_no,
    #             'invoice_number': custody.agreement_no,
    #             'customer_name': custody.customer.customer_name,
    #             'building_name': custody.customer.building_name,
    #             'sales_type': custody.customer.sales_type,
    #             'route_name': custody.customer.routes.route_name,
    #             'salesman': custody.created_by,  # Assuming created_by is the sales staff
    #             'amount': item.amount,
    #             'discount': 0,  # Adjust as needed
    #             'net_taxable': item.amount - (item.can_deposite_chrge or 0),  # Adjust as needed
    #             'vat_amount': Tax.objects.get(name="VAT").percentage,
    #             'grand_total': item.amount,  # Adjust as needed
    #             'amount_collected': custody.amount_collected,
    #         })
            
    #         total_amount += item.amount
    #         # Adjust the totals based on your requirements
    #         total_grand_total += item.amount
    #         total_amount_recieved += custody.amount_collected

    # context = {
    #     'customersales': sales_report_data,
    #     'total_amount': total_amount,
    #     'total_discount': total_discount,
    #     'total_net_payable': total_net_payable,
    #     'total_vat': total_vat,
    #     'total_grand_total': total_grand_total,
    #     'total_amount_recieved': total_amount_recieved,
    #     'filter_data': filter_data,
    #     'sales_types': ['CASH', 'CREDIT'],  # Available filter options
    # }
    # return render(request, 'sales_management/customerSales_report.html', context)


def customerSales_Excel_report(request):
    filter_data = {}

    total_amount = 0
    total_discount = 0
    total_net_payable = 0
    total_vat = 0
    total_grand_total = 0
    total_amount_recieved = 0

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sales_type_filter = request.GET.get('sales_type')

    if start_date_str and end_date_str:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()
        end_date = datetime.datetime.today().date()

    filter_data = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'sales_type': sales_type_filter,
    }

    # Filter CustomerSupply data
    sales = CustomerSupply.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).exclude(customer__sales_type__in=["CASH COUPON", "CREDIT COUPON"])

    if sales_type_filter:
        sales = sales.filter(customer__sales_type=sales_type_filter)

    sales = sales.order_by("-created_date")

    # Filter CustomerCoupon data
    coupons = CustomerCoupon.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    )

    if sales_type_filter:
        coupons = coupons.filter(payment_type=sales_type_filter)

    coupons = coupons.order_by("-created_date")

    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Customer Sales Report"

    # Define column headers
    headers = [
        'Date', 'Reference Invoice No', 'Invoice Number', 'Customer Name','Customer Id',
        'Building Name', 'Sales Type', 'Route Name', 'Salesman',
        'Amount', 'Discount', 'Net Taxable', 'VAT Amount', 'Grand Total', 'Amount Collected'
    ]
    sheet.append(headers)

    # Process CustomerSupply data
    for sale in sales:
        row = [
            sale.created_date.date(),
            sale.reference_number,
            sale.invoice_no,
            sale.customer.customer_name,
            sale.customer.custom_id,
            sale.customer.building_name,
            sale.customer.sales_type,
            sale.customer.routes.route_name,
            sale.customer.sales_staff.get_fullname(),
            sale.grand_total,
            sale.discount,
            sale.subtotal,
            sale.vat,
            sale.grand_total,
            sale.amount_recieved,
        ]
        sheet.append(row)

        total_amount += sale.grand_total
        total_discount += sale.discount
        total_net_payable += sale.net_payable
        total_vat += sale.vat
        total_grand_total += sale.grand_total
        total_amount_recieved += sale.amount_recieved

    # Process CustomerCoupon data
    for coupon in coupons:
        row = [
            coupon.created_date.date(),
            coupon.reference_number,
            coupon.invoice_no,
            coupon.customer.customer_name,
            coupon.customer.custom_id,
            coupon.customer.building_name,
            coupon.customer.sales_type,
            coupon.customer.routes.route_name,
            coupon.customer.sales_staff.get_fullname(),
            coupon.grand_total,
            coupon.discount,
            coupon.net_amount,
            Tax.objects.get(name="VAT").percentage,
            coupon.grand_total,
            coupon.amount_recieved,
        ]
        sheet.append(row)

        total_amount += coupon.grand_total
        total_discount += coupon.discount
        total_net_payable += coupon.net_amount
        total_vat += Tax.objects.get(name="VAT").percentage
        total_grand_total += coupon.grand_total
        total_amount_recieved += coupon.amount_recieved

    # Add totals row
    totals_row = [
        '', '', '', '', '', '', '', '','',  # Empty cells for non-applicable columns
        total_amount, total_discount, total_net_payable, total_vat, total_grand_total, total_amount_recieved
    ]
    sheet.append(totals_row)

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=CustomerSalesReport_{start_date_str}_to_{end_date_str}.xlsx'
    
    # Save workbook to the response
    workbook.save(response)

    return response


def customerSales_Print_report(request):
    filter_data = {}

    total_amount = 0
    total_discount = 0
    total_net_payable = 0
    total_vat = 0
    total_grand_total = 0
    total_amount_recieved = 0

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    sales_type_filter = request.GET.get('sales_type')

    if start_date_str and end_date_str:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()
        end_date = datetime.datetime.today().date()

    filter_data = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'sales_type': sales_type_filter,
    }

    # Filter CustomerSupply data
    sales = CustomerSupply.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).exclude(customer__sales_type__in=["CASH COUPON", "CREDIT COUPON"])

    if sales_type_filter:
        sales = sales.filter(customer__sales_type=sales_type_filter)

    sales = sales.order_by("-created_date")

    # Filter CustomerCoupon data
    coupons = CustomerCoupon.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    )

    if sales_type_filter:
        coupons = coupons.filter(payment_type=sales_type_filter)

    coupons = coupons.order_by("-created_date")

    # Organize the data for rendering in the template
    sales_report_data = []

    # Process CustomerSupply data
    for sale in sales:
        sales_report_data.append({
            'date': sale.created_date.date(),
            'ref_invoice_no': sale.reference_number,
            'invoice_number': sale.invoice_no,
            'customer_name': sale.customer.customer_name,
            'custom_id': sale.customer.custom_id,
            'building_name': sale.customer.building_name,
            'sales_type': sale.customer.sales_type,
            'route_name': sale.customer.routes.route_name,
            'salesman': sale.customer.sales_staff.get_fullname(),
            'amount': sale.grand_total,
            'discount': sale.discount,
            'net_taxable': sale.subtotal,
            'vat_amount': sale.vat,
            'grand_total': sale.grand_total,
            'amount_collected': sale.amount_recieved,
        })

        total_amount += sale.grand_total
        total_discount += sale.discount
        total_net_payable += sale.net_payable
        total_vat += sale.vat
        total_grand_total += sale.grand_total
        total_amount_recieved += sale.amount_recieved

    # Process CustomerCoupon data
    for coupon in coupons:
        sales_report_data.append({
            'date': coupon.created_date.date(),
            'ref_invoice_no': coupon.reference_number,
            'invoice_number': coupon.invoice_no,
            'customer_name': coupon.customer.customer_name,
            'custom_id': coupon.customer.custom_id,
            'building_name': coupon.customer.building_name,
            'sales_type': coupon.customer.sales_type,
            'route_name': coupon.customer.routes.route_name,
            'salesman': coupon.customer.sales_staff.get_fullname(),
            'amount': coupon.grand_total,
            'discount': coupon.discount,
            'net_taxable': coupon.net_amount,
            'vat_amount': Tax.objects.get(name="VAT").percentage,
            'grand_total': coupon.grand_total,
            'amount_collected': coupon.amount_recieved,
        })

        total_amount += coupon.grand_total
        total_discount += coupon.discount
        total_net_payable += coupon.net_amount
        total_vat += Tax.objects.get(name="VAT").percentage
        total_grand_total += coupon.grand_total
        total_amount_recieved += coupon.amount_recieved

    context = {
        'customersales': sales_report_data,
        'total_amount': total_amount,
        'total_discount': total_discount,
        'total_net_payable': total_net_payable,
        'total_vat': total_vat,
        'total_grand_total': total_grand_total,
        'total_amount_recieved': total_amount_recieved,
        'filter_data': filter_data,
        'sales_types': ['CASH', 'CREDIT'],  # Available filter options
    }
    return render(request, 'sales_management/customerSales_report_pdf.html', context)

#------------------Collection Report-------------------------                



def collection_report(request):
    filter_data = {}
    selected_route_id = request.GET.get('route_name')
    template = 'sales_management/collection_report.html'

    # Fetch routes and set default date range
    routes = RouteMaster.objects.all()
    today = datetime.datetime.today()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    start_date = today.date()
    end_date = today.date() + timedelta(days=1)
    
    # Parse date filters if provided
    if start_date_str and end_date_str:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    filter_data['end_date'] = end_date.strftime('%Y-%m-%d')

    # Fetch collection payments within the date range
    collection_payments = CollectionPayment.objects.filter(
        created_date__date__range=[start_date, end_date]
    ).values(
        'customer__custom_id', 
        'customer__customer_name',
        'customer__mobile_no',
        'customer__routes__route_name',
        'customer__building_name',
        'customer__door_house_no',
        'created_date__date',
        'payment_method',
        'customer__sales_type'
    ).annotate(
        total_amount=Sum('collectionitems__amount'),
        total_discount=Sum('collectionitems__invoice__discount'),
        total_net_taxable=Sum('collectionitems__invoice__net_taxable'),
        total_vat=Sum('collectionitems__invoice__vat'),
        collected_amount=Sum('collectionitems__amount_received')
    ).order_by('-created_date')

    # Filter by route if selected
    if selected_route_id:
        selected_route = RouteMaster.objects.get(route_name=selected_route_id)
        collection_payments = collection_payments.filter(customer__routes__route_name=selected_route.route_name)
        filter_data['selected_route'] = selected_route_id

    # Compute Grand Totals
    grand_total_amount = sum(item['total_amount'] for item in collection_payments if item['total_amount'] is not None)
    grand_total_collected = sum(item['collected_amount'] for item in collection_payments if item['collected_amount'] is not None)
    grand_total_balance = grand_total_amount - grand_total_collected  # Balance = Amount - Collected

    # Prepare the context for the template
    context = {
        'collection_payments': collection_payments, 
        'routes': routes, 
        'today': today,
        'filter_data': filter_data,
        'grand_total_amount': grand_total_amount,
        'grand_total_collected': grand_total_collected,
        'grand_total_balance': grand_total_balance,
    }

    return render(request, template, context)
    # filter_data = {}
    # selected_route_id = request.GET.get('route_name')
    # template = 'sales_management/collection_report.html'
    
    # routes = RouteMaster.objects.all()
    # today = datetime.datetime.today()
    
    # start_date_str = request.GET.get('start_date')
    # end_date_str = request.GET.get('end_date')
    
    # start_date = datetime.datetime.today().date()
    # end_date = datetime.datetime.today().date() + timedelta(days=1)

    # if start_date_str and end_date_str:
    #     start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
    #     end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    # filter_data['end_date'] = end_date.strftime('%Y-%m-%d')
    
    # # Filter and count collection payments
    # collection_payments = CollectionItems.objects.filter(
    #     collection_payment__created_date__date__range=[start_date, end_date]
    # ).values(
    #     'collection_payment__customer__custom_id', 
    #     'collection_payment__customer__customer_name',
    #     'collection_payment__customer__mobile_no',
    #     'collection_payment__customer__routes__route_name',
    #     'collection_payment__customer__building_name',
    #     'collection_payment__customer__door_house_no',
    #     'collection_payment__created_date__date',
    #     'collection_payment__payment_method',
    #     'collection_payment__customer__sales_type',
        
        
    # ).annotate(
    #     count_amount=Sum('amount'),
    #     count_balance=Sum('balance'),
    #     count_amount_received=Sum('amount_received')
    # ).order_by('-collection_payment__created_date__date')
    
    # if selected_route_id:
    #     selected_route = RouteMaster.objects.get(route_name=selected_route_id)
    #     collection_payments = collection_payments.filter(collection_payment__customer__routes__route_name=selected_route.route_name)
    #     filter_data['selected_route'] = selected_route_id
    
    # context = {
    #     'collection_payments': collection_payments, 
    #     'routes': routes, 
    #     'today': today,
    #     'filter_data': filter_data,
    # }
    # return render(request, template, context)


def collection_report_excel(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    selected_route_id = request.GET.get('route_name')

    # Set default date range
    start_date = datetime.datetime.today().date()
    end_date = datetime.datetime.today().date() + timedelta(days=1)

    # Parse provided dates if available
    if start_date_str and end_date_str:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Query the filtered data
    collection_payments = CollectionItems.objects.filter(
        collection_payment__created_date__date__range=[start_date, end_date]
    ).values(
        'collection_payment__customer__custom_id', 
        'collection_payment__customer__customer_name',
        'collection_payment__customer__mobile_no',
        'collection_payment__customer__routes__route_name',
        'collection_payment__customer__building_name',
        'collection_payment__customer__door_house_no',
        'collection_payment__created_date__date',
        'collection_payment__payment_method',
        'collection_payment__customer__sales_type'
    ).annotate(
        count_amount=Sum('amount'),
        count_balance=Sum('balance'),
        count_amount_received=Sum('amount_received')
    ).order_by('-collection_payment__created_date__date')

    if selected_route_id:
        selected_route = get_object_or_404(RouteMaster, route_name=selected_route_id)
        collection_payments = collection_payments.filter(collection_payment__customer__routes__route_name=selected_route.route_name)

    # Create an HttpResponse object with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=collection_report_filtered.xlsx'

    # Create an Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(response, {'in_memory': True})
    worksheet = workbook.add_worksheet()

    # Define the header row
    headers = [
        'Sl No', 'Date & Time', 'Customer/Mobile No', 'Route',
        'Building Name/No', 'Amount', 'Amount Received', 'Balance', 'Mode of Payment','Sales Type'
    ]

    # Write the header row
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)

    # Write the data rows
    for row_num, data in enumerate(collection_payments, start=1):
        worksheet.write(row_num, 0, row_num)  # Sl No
        worksheet.write(row_num, 1, data['collection_payment__created_date__date'].strftime('%Y-%m-%d'))
        worksheet.write(row_num, 2, f"{data['collection_payment__customer__customer_name']} ")
        worksheet.write(row_num, 3, data['collection_payment__customer__routes__route_name'])
        worksheet.write(row_num, 4, f"{data['collection_payment__customer__building_name']} / {data['collection_payment__customer__door_house_no']}")
        worksheet.write(row_num, 5, data['count_amount'])
        worksheet.write(row_num, 6, data['count_amount_received'])
        worksheet.write(row_num, 7, data['count_balance'])
        worksheet.write(row_num, 8, data['collection_payment__payment_method'])
        worksheet.write(row_num, 8, data['collection_payment__customer__sales_type'])

    # Close the workbook and write the data to the response
    workbook.close()
    return response

def print_collection_report(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    selected_route_id = request.GET.get('route_name')

    # Set default date range
    start_date = datetime.datetime.today().date()
    end_date = datetime.datetime.today().date() + timedelta(days=1)

    # Parse provided dates if available
    if start_date_str and end_date_str:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Query the filtered data
    collection_payments = CollectionItems.objects.filter(
        collection_payment__created_date__date__range=[start_date, end_date]
    ).values(
        'collection_payment__customer__custom_id', 
        'collection_payment__customer__customer_name',
        'collection_payment__customer__mobile_no',
        'collection_payment__customer__routes__route_name',
        'collection_payment__customer__building_name',
        'collection_payment__customer__door_house_no',
        'collection_payment__created_date__date',
        'collection_payment__payment_method',
        'collection_payment__customer__sales_type'
    ).annotate(
        count_amount=Sum('amount'),
        count_balance=Sum('balance'),
        count_amount_received=Sum('amount_received')
    ).order_by('-collection_payment__created_date__date')

    if selected_route_id:
        selected_route = get_object_or_404(RouteMaster, route_name=selected_route_id)
        collection_payments = collection_payments.filter(collection_payment__customer__routes__route_name=selected_route.route_name)

    context = {
        'collection_payments': collection_payments,
        'start_date': start_date.strftime('%d/%m/%Y'),
        'end_date': end_date.strftime('%d/%m/%Y'),
        'selected_route': selected_route_id,
    }

    return render(request, 'sales_management/collection_report_print.html', context)

#-----------------Suspense Report--------------------------
from .forms import SuspenseCollectionForm

def suspense_report(request):
    start_date = request.GET.get('start_date')
    print("start_date", start_date)
    filter_data = {}

    if start_date:
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    
    van_instances = Van.objects.all()
    if request.user.branch_id :
        van_instances = van_instances.filter(branch_id=request.user.branch_id)


    context = {
        'van_instances': van_instances,
        'filter_data': filter_data,
    }

    return render(request, 'sales_management/suspense_report.html', context)

# def create_suspense_collection(request,id,date):
#     van_instance = Van.objects.get(pk=id)
#     salesman = van_instance.salesman
    
#     expenses_instanses = Expense.objects.filter(date_created=date,van__pk=id)
#     today_expense = expenses_instanses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
    
#     # cash sales amount collected
#     supply_amount_collected = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,customer__sales_type="CASH").aggregate(total_amount=Sum('amount_recieved'))['total_amount'] or 0
#     coupon_amount_collected = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,customer__sales_type="CASH").aggregate(total_amount=Sum('amount_recieved'))['total_amount'] or 0
#     cash_sales_amount_collected = supply_amount_collected + coupon_amount_collected
    
#     # collection details
#     dialy_collections = CollectionPayment.objects.filter(created_date__date=date,salesman_id=salesman,amount_received__gt=0)
    
#     credit_sales_amount_collected = dialy_collections.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
#     total_sales_amount_collected = cash_sales_amount_collected + credit_sales_amount_collected
#     net_payble = total_sales_amount_collected - today_expense
    
#     amount_paid = SuspenseCollection.objects.filter(date=date,salesman=salesman).aggregate(total_amount=Sum('amount_paid'))['total_amount'] or 0
#     amount_payeble = net_payble - amount_paid
    
#     if request.method == 'POST':
#         form = SuspenseCollectionForm(request.POST,initial={'payable_amount': amount_payeble})
#         if Decimal(request.POST.get("amount_paid")) <= amount_payeble :
#             if form.is_valid():
#                 suspense_collection = form.save(commit=False)
#                 suspense_collection.date = date  # Set the created_date
#                 suspense_collection.created_date = datetime.datetime.today().date()
#                 suspense_collection.salesman = salesman
#                 suspense_collection.route = Van_Routes.objects.filter(van=van_instance).first().routes
#                 suspense_collection.cash_sale_amount = cash_sales_amount_collected
#                 suspense_collection.credit_sale_amount = credit_sales_amount_collected
#                 suspense_collection.expense = today_expense
#                 suspense_collection.net_payeble_amount = amount_payeble  # Set the net_payeble_amount field
#                 # Calculate amount_balance
#                 amount_paid = form.cleaned_data['amount_paid']
#                 amount_balance = amount_payeble - amount_paid
#                 suspense_collection.amount_balance = amount_balance
#                 suspense_collection.save()
                
                
#                 action = request.POST.get('action')
#                 date = request.POST.get('date')
#                 route_name = request.POST.get('route_name')
                
#                 if action == 'print':
#                     return redirect(f'/print-dsr/?date={date}&route_name={route_name}')
#                 else:
#                     return redirect('suspense_report')
#             else:
#                 print("Form errors:", form.errors)
#         else:
#             print("graeterthan")
#     else:
#         form = SuspenseCollectionForm(initial={'payable_amount': amount_payeble})
#         print("form",form)
    
#     return render(request, 'sales_management/create_suspense_collection.html', {'form': form})
def create_suspense_collection(request, id, date):
    van_instance = Van.objects.get(pk=id)
    salesman = van_instance.salesman
    
    expenses_instanses = Expense.objects.filter(date_created=date, van__pk=id)
    today_expense = expenses_instanses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
    
    supply_amount_collected = CustomerSupply.objects.filter(created_date__date=date, salesman=salesman, amount_recieved__gt=0).aggregate(total_amount=Sum('amount_recieved'))['total_amount'] or 0
    coupon_amount_collected = CustomerCoupon.objects.filter(created_date__date=date, salesman=salesman, amount_recieved__gt=0).aggregate(total_amount=Sum('amount_recieved'))['total_amount'] or 0
    cash_sales_amount_collected = supply_amount_collected + coupon_amount_collected
    
    dialy_collections = CollectionPayment.objects.filter(created_date__date=date, salesman_id=salesman, amount_received__gt=0)
    credit_sales_amount_collected = dialy_collections.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
    total_sales_amount_collected = cash_sales_amount_collected + credit_sales_amount_collected
    net_payble = total_sales_amount_collected - today_expense
    
    amount_paid = SuspenseCollection.objects.filter(date=date, salesman=salesman).aggregate(total_amount=Sum('amount_paid'))['total_amount'] or 0
    amount_payeble = net_payble - amount_paid
    
    if request.method == 'POST':
        form = SuspenseCollectionForm(request.POST, initial={'payable_amount': amount_payeble})
        if Decimal(request.POST.get("amount_paid")) <= amount_payeble:
            if form.is_valid():
                suspense_collection = form.save(commit=False)
                suspense_collection.date = date
                suspense_collection.created_date = datetime.datetime.today().date()
                suspense_collection.salesman = salesman
                suspense_collection.route = Van_Routes.objects.filter(van=van_instance).first().routes
                suspense_collection.cash_sale_amount = cash_sales_amount_collected
                suspense_collection.credit_sale_amount = credit_sales_amount_collected
                suspense_collection.expense = today_expense
                suspense_collection.net_payeble_amount = amount_payeble
                amount_paid = form.cleaned_data['amount_paid']
                amount_balance = amount_payeble - amount_paid
                suspense_collection.amount_balance = amount_balance
                suspense_collection.save()
                
                if request.POST.get('action') == 'submit_and_print':
                    return JsonResponse({
                        'success': True,
                        'date': date,
                        'route_name': Van_Routes.objects.filter(van=van_instance).first().routes.route_name
                    })
                else:
                    return redirect('suspense_report')
            else:
                return JsonResponse({'success': False, 'errors': form.errors.as_json()})
        else:
            return JsonResponse({'success': False, 'errors': 'Amount paid cannot be greater than payable amount'})
    else:
        form = SuspenseCollectionForm(initial={'payable_amount': amount_payeble})
    
    return render(request, 'sales_management/create_suspense_collection.html', {'form': form})


from django.utils import timezone

def suspense_report_excel(request):
    start_date = request.GET.get('start_date')
    filter_data = {}

    if start_date:
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')

    invoices = InvoiceDailyCollection.objects.filter(created_date__gte=start_date)\
                                              .select_related('customer__routes')\
                                              .select_related('customer__sales_staff')

    suspense_collections = SuspenseCollection.objects.filter(created_date__gte=start_date)

    # Create a BytesIO object to save workbook data
    buffer = BytesIO()

    # Create a new Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    worksheet.title = "Suspense Report"

    # Write header row
    headers = ['Sl No', 'Created Date', 'Invoice Type', 'Route', 'Salesman', 'Opening Suspense', 'Paid', 'Closing Suspense']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write data rows
    row = 1
    for invoice in invoices:
        created_date = invoice.created_date.replace(tzinfo=None)
        worksheet.write(row, 0, row)  # Sl No
        worksheet.write(row, 1, created_date)  # Created Date
        worksheet.write(row, 2, invoice.invoice.invoice_type)  # Invoice Type
        worksheet.write(row, 4, invoice.customer.sales_staff.username)  # Salesman
        worksheet.write(row, 5, invoice.amount)  # Opening Suspense
        # Fetch the corresponding SuspenseCollection for this invoice
        related_suspense_collection = suspense_collections.filter(salesman=invoice.customer.sales_staff).first()
        if related_suspense_collection:
            worksheet.write(row, 6, related_suspense_collection.amount_paid)  # Paid
            worksheet.write(row, 7, related_suspense_collection.amount_balance)  # Closing Suspense
        else:
            worksheet.write(row, 6, '')  # Paid
            worksheet.write(row, 7, '')  # Closing Suspense

        # Extract and write route information
        if invoice.customer.routes:
            route_info = f"{invoice.customer.routes.route_name}"
            worksheet.write(row, 3, route_info)  # Route
        else:
            worksheet.write(row, 3, '')  # Route

        row += 1

    # Close the workbook
    workbook.close()

    # Create HttpResponse object to return the Excel file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Suspense_Report.xlsx'
    return response


def suspense_report_print(request):
    start_date = request.GET.get('start_date')
    filter_data = {}

    if start_date:
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')

    invoices = InvoiceDailyCollection.objects.filter(created_date__gte=start_date)\
                                              .select_related('customer__routes')\
                                              .select_related('customer__sales_staff')

    suspense_collections = SuspenseCollection.objects.filter(created_date__gte=start_date)

    # Prepare data for the PDF report
    data = [['Sl No', 'Created Date', 'Invoice Type', 'Route', 'Salesman', 'Opening Suspense', 'Paid', 'Closing Suspense']]

    for invoice in invoices:
        created_date = invoice.created_date.replace(tzinfo=None)
        salesman_name = invoice.customer.sales_staff.username if invoice.customer.sales_staff else ''
        route_info = f"{invoice.customer.routes.route_name}" if invoice.customer.routes else ''

        related_suspense_collection = suspense_collections.filter(salesman=invoice.customer.sales_staff).first()
        paid = related_suspense_collection.amount_paid if related_suspense_collection else ''
        balance = related_suspense_collection.amount_balance if related_suspense_collection else ''

        data.append([invoice.id, created_date, invoice.invoice.invoice_type, route_info, salesman_name, invoice.amount, paid, balance])

    # Create a BytesIO object to save PDF data
    buffer = BytesIO()

    # Create a new PDF document
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Create a table with the data
    table = Table(data)
    
    # Add style to the table
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), 'grey'),
                        ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), 'lightgrey'),
                        ('GRID', (0, 0), (-1, -1), 1, 'black')])

    table.setStyle(style)
    
    # Add table to the document
    doc.build([table])

    # Create HttpResponse object to return the PDF file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Suspense_Report.pdf'
    return response




#-----------------DSR cash sales Report--------------------------

def cashsales_report(request):
    filter_data = {}
    data_filter = False
    cash_total_net_taxable = 0
    cash_total_vat = 0
    cash_total_subtotal = 0
    cash_total_amount_recieved = 0
    cash_sale_recharge_count = 0
    cash_total_quantity = 0
    cash_total_qty = 0
    stock_report_total = 0

    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    salesman_id = ""
    cash_sales = CustomerSupply.objects.none
    recharge_cash_sales = CustomerCoupon.objects.none
    products = ProdutItemMaster.objects.none
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    route_name = request.GET.get('route_name')

    if date_from and date_to:
        date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
        filter_data['filter_date_from'] = date_from.strftime('%Y-%m-%d')
        filter_data['filter_date_to'] = date_to.strftime('%Y-%m-%d')
        data_filter = True
    else:
        date_from = datetime.datetime.today().date()
        date_to = date_from
        filter_data['filter_date_from'] = date_from.strftime('%Y-%m-%d')
        filter_data['filter_date_to'] = date_to.strftime('%Y-%m-%d')

    if route_name:
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        if van_route:
            salesman = van_route.van.salesman
            salesman_id = salesman.pk
            filter_data['route_name'] = route_name

            ##### stock report ####
            products = ProdutItemMaster.objects.filter()
            van_instances = Van.objects.filter(salesman=salesman)
            van_product_stock = VanProductStock.objects.filter(
                created_date__range=[date_from, date_to],
                van__in=van_instances,
                product__product_name="5 Gallon"
            )
            stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0

            ### cash sales ####
            cash_sales = CustomerSupply.objects.filter(
                created_date__date__range=[date_from, date_to],
                salesman=salesman,
                amount_recieved__gt=0
            ).exclude(customer__sales_type="CASH COUPON")
            cash_total_net_taxable = cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
            cash_total_vat = cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
            cash_total_subtotal = cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
            cash_total_received = cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
            cash_total_quantity = cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

            recharge_cash_sales = CustomerCoupon.objects.filter(
                created_date__date__range=[date_from, date_to],
                salesman=salesman,
                amount_recieved__gt=0
            )
            cash_sale_recharge_net_payeble = recharge_cash_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
            cash_sale_recharge_vat_total = 0
            cash_sale_recharge_grand_total = recharge_cash_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
            cash_sale_recharge_amount_recieved = recharge_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
            cash_total_net_taxable = cash_total_net_taxable + cash_sale_recharge_net_payeble
            cash_total_vat = cash_total_vat + cash_sale_recharge_vat_total
            cash_total_subtotal = cash_total_subtotal + cash_sale_recharge_grand_total
            cash_total_amount_recieved = cash_total_received + cash_sale_recharge_amount_recieved

            total_cash_sales_count = cash_sales.count() + recharge_cash_sales.count()

            cash_sale_recharge_count = recharge_cash_sales.count()
            cash_total_qty = cash_total_quantity + cash_sale_recharge_count

    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        'cash_sales': cash_sales,
        'recharge_cash_sales': recharge_cash_sales,
        'cash_total_net_taxable': cash_total_net_taxable,
        'cash_total_vat': cash_total_vat,
        'cash_total_subtotal': cash_total_subtotal,
        'cash_total_amount_recieved': cash_total_amount_recieved,
        'cash_total_qty': cash_total_qty,
        'filter_data': filter_data,
    }

    return render(request, 'sales_management/dsr_cash_sales_report.html', context)


def cashsales_report_print(request):
    filter_data = {}
    data_filter = False
    cash_total_net_taxable = 0
    cash_total_vat = 0
    cash_total_subtotal = 0
    cash_total_amount_recieved = 0
    cash_sale_recharge_count=0
    cash_total_quantity = 0
    cash_total_qty = 0
    stock_report_total = 0
   
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    cash_sales = CustomerSupply.objects.none
    recharge_cash_sales = CustomerCoupon.objects.none
    products = ProdutItemMaster.objects.none
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        
        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=salesman)
        van_product_stock = VanProductStock.objects.filter(created_date=date,van=van_instances,product__product_name="5 Gallon")
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
     
        ### cash sales ####
        cash_sales = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__gt=0).exclude(customer__sales_type="CASH COUPON")
        cash_total_net_taxable = cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        cash_total_vat = cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        cash_total_subtotal = cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        cash_total_received = cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        cash_total_quantity = cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        recharge_cash_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__gt=0)
        cash_sale_recharge_net_payeble = recharge_cash_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        cash_sale_recharge_vat_total = 0
        cash_sale_recharge_grand_total = recharge_cash_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        cash_sale_recharge_amount_recieved = recharge_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        cash_total_net_taxable = cash_total_net_taxable + cash_sale_recharge_net_payeble 
        cash_total_vat = cash_total_vat + cash_sale_recharge_vat_total 
        cash_total_subtotal = cash_total_subtotal + cash_sale_recharge_grand_total 
        cash_total_amount_recieved = cash_total_received + cash_sale_recharge_amount_recieved 
        
        total_cash_sales_count = cash_sales.count() + recharge_cash_sales.count()
        
        cash_sale_recharge_count = recharge_cash_sales.count()
        cash_total_qty = cash_total_quantity + cash_sale_recharge_count
        
       
      
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        
        'routes_instances': routes_instances,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        
        #cash sales
        'cash_sales': cash_sales,
        'recharge_cash_sales': recharge_cash_sales,
        'cash_total_net_taxable':cash_total_net_taxable,
        'cash_total_vat':cash_total_vat,
        'cash_total_subtotal': cash_total_subtotal,
        'cash_total_amount_recieved': cash_total_amount_recieved,
        'cash_total_qty': cash_total_qty,
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),
    }
    
    return render(request, 'sales_management/dsr_cash_sales_print.html', context)



#-----------------DSR credit sales Report--------------------------

def creditsales_report(request):
    filter_data = {}
    data_filter = False
    credit_total_net_taxable = 0
    credit_total_vat = 0
    credit_total_subtotal = 0
    credit_total_amount_recieved = 0
    credit_total_qty = 0
    credit_total_quantity = 0
    credit_sale_recharge_count = 0
    stock_report_total = 0
   
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    credit_sales = CustomerSupply.objects.none
    recharge_credit_sales = CustomerCoupon.objects.none
    products = ProdutItemMaster.objects.none
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')

    if from_date:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
    else:
        from_date = datetime.datetime.today().date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
    if to_date:
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        to_date = datetime.datetime.today().date()
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
     
        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=salesman)
       
        ### credit sales ####
        credit_sales = CustomerSupply.objects.filter(created_date__date__range=[from_date, to_date],salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])

        credit_total_net_taxable = credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        credit_total_vat = credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        credit_total_subtotal = credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        credit_total_received = credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        credit_total_quantity = credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        recharge_credit_sales = CustomerCoupon.objects.filter(created_date__date__range=[from_date, to_date],salesman=salesman,amount_recieved__lte=0)
        credit_sale_recharge_net_payeble = recharge_credit_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        credit_sale_recharge_vat_total = 0
        
        credit_sale_recharge_grand_total = recharge_credit_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        credit_sale_recharge_amount_recieved = recharge_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        credit_total_net_taxable = credit_total_net_taxable + credit_sale_recharge_net_payeble 
        credit_total_vat = credit_total_vat + credit_sale_recharge_vat_total 
        credit_total_subtotal = credit_total_subtotal + credit_sale_recharge_grand_total
        credit_total_amount_recieved = credit_total_received + credit_sale_recharge_amount_recieved
        
        total_credit_sales_count = credit_sales.count() + recharge_credit_sales.count()
        credit_sale_recharge_count = recharge_credit_sales.count()
        credit_total_qty = credit_total_quantity + credit_sale_recharge_count
        
      
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        
        'routes_instances': routes_instances,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        
        
        # credit sales
        'credit_sales': credit_sales,
        'recharge_credit_sales': recharge_credit_sales,
        'credit_total_net_taxable':credit_total_net_taxable,
        'credit_total_vat':credit_total_vat,
        'credit_total_subtotal':credit_total_subtotal,
        'credit_total_amount_recieved': credit_total_amount_recieved,
        'credit_total_qty': credit_total_qty,
        'filter_data': filter_data,
    }
    
    return render(request, 'sales_management/dsr_credit_sales_report.html', context)


def creditsales_report_print(request):
    filter_data = {}
    data_filter = False
    credit_total_net_taxable = 0
    credit_total_vat = 0
    credit_total_subtotal = 0
    credit_total_amount_recieved = 0
    credit_total_qty = 0
    credit_total_quantity = 0
    credit_sale_recharge_count = 0
    stock_report_total = 0
   
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    credit_sales = CustomerSupply.objects.none
    recharge_credit_sales = CustomerCoupon.objects.none
    products = ProdutItemMaster.objects.none
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
     
        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=salesman)
       
        ### credit sales ####
        credit_sales = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])

        credit_total_net_taxable = credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        credit_total_vat = credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        credit_total_subtotal = credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        credit_total_received = credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        credit_total_quantity = credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        recharge_credit_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__lte=0)
        credit_sale_recharge_net_payeble = recharge_credit_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        credit_sale_recharge_vat_total = 0
        
        credit_sale_recharge_grand_total = recharge_credit_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        credit_sale_recharge_amount_recieved = recharge_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        credit_total_net_taxable = credit_total_net_taxable + credit_sale_recharge_net_payeble 
        credit_total_vat = credit_total_vat + credit_sale_recharge_vat_total 
        credit_total_subtotal = credit_total_subtotal + credit_sale_recharge_grand_total
        credit_total_amount_recieved = credit_total_received + credit_sale_recharge_amount_recieved
        
        total_credit_sales_count = credit_sales.count() + recharge_credit_sales.count()
        credit_sale_recharge_count = recharge_credit_sales.count()
        credit_total_qty = credit_total_quantity + credit_sale_recharge_count
        
      
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        
        'routes_instances': routes_instances,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        
        
        # credit sales
        'credit_sales': credit_sales,
        'recharge_credit_sales': recharge_credit_sales,
        'credit_total_net_taxable':credit_total_net_taxable,
        'credit_total_vat':credit_total_vat,
        'credit_total_subtotal':credit_total_subtotal,
        'credit_total_amount_recieved': credit_total_amount_recieved,
        'credit_total_qty': credit_total_qty,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),
    }
    
    return render(request, 'sales_management/dsr_credit_sales_print.html', context)
#----------DSR Coupon Sales--------------

def dsr_coupon_sales(request):
    filter_data = {}
    data_filter = False
    total_cash_sales_count = 0
    total_credit_sales_count = 0
    manual_coupon_total = 0
    digital_coupon_total = 0
    total_coupon_sales_count = 0
    coupon_total_qty = 0
    
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    coupon_sales = CustomerSupply.objects.none
    routes_instances = RouteMaster.objects.all()
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')
    
    if from_date:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
    else:
        from_date = datetime.datetime.today().date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
    if to_date:
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        to_date = datetime.datetime.today().date()
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        
        # Coupon sales
        coupon_sales = CustomerSupply.objects.filter(created_date__date__range=(from_date, to_date),salesman=salesman,customer__sales_type="CASH COUPON")
        manual_coupon_total = CustomerSupplyCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(Count('leaf'))['leaf__count']
        digital_coupon_total = CustomerSupplyDigitalCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(total_count=Sum('count'))['total_count'] or 0
        
        total_coupon_sales_count = coupon_sales.count()
        
        total_sales_count = total_cash_sales_count + total_credit_sales_count + total_coupon_sales_count
        coupon_total_qty = coupon_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
       
        
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
     
        # coupon sales
        'coupon_sales': coupon_sales,
        'manual_coupon_total':manual_coupon_total,
        'digital_coupon_total':digital_coupon_total,
        'total_coupon_sales_count': total_coupon_sales_count,
        'coupon_total_qty':coupon_total_qty,
      
        'filter_data': filter_data,
    }
    
    return render(request, 'sales_management/dsr_coupons_sales.html', context)

def dsr_coupons_sales_print(request):
    filter_data = {}
    data_filter = False
    total_cash_sales_count = 0
    total_credit_sales_count = 0
    manual_coupon_total = 0
    digital_coupon_total = 0
    total_coupon_sales_count = 0
    coupon_total_qty = 0
   
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    coupon_sales = CustomerSupply.objects.none
    routes_instances = RouteMaster.objects.all()
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        
        # Coupon sales
        coupon_sales = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,customer__sales_type="CASH COUPON")
        manual_coupon_total = CustomerSupplyCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(Count('leaf'))['leaf__count']
        digital_coupon_total = CustomerSupplyDigitalCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(total_count=Sum('count'))['total_count'] or 0
        
        total_coupon_sales_count = coupon_sales.count()
        
        total_sales_count = total_cash_sales_count + total_credit_sales_count + total_coupon_sales_count
        coupon_total_qty = coupon_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
       
        
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
     
        # coupon sales
        'coupon_sales': coupon_sales,
        'manual_coupon_total':manual_coupon_total,
        'digital_coupon_total':digital_coupon_total,
        'total_coupon_sales_count': total_coupon_sales_count,
        'coupon_total_qty':coupon_total_qty,
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),
    }
    
    return render(request, 'sales_management/dsr_coupon_sales_print.html', context)

#----------DSR Coupon Book Sales--------------

def dsr_coupon_book_sales(request):
    # Initialize counts and variables
    filter_data = {}
    data_filter = False
    salesman_id = ""
    van_route = None

    # Retrieve all route instances
    routes_instances = RouteMaster.objects.all()

    from_date= request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')

    if from_date:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
    else:
        from_date = datetime.datetime.today().date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
    
    if to_date:
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        to_date = datetime.datetime.today().date()
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')

    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        route_id = van_route.routes.pk if van_route else None
        filter_data['route_name'] = route_name
    else:
        route_id = None

    # Ensure van_route is defined
    if van_route:
        # Retrieve salesman
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        customer_coupons = CustomerCoupon.objects.filter(salesman=salesman, created_date__date__range=[from_date, to_date])
    else:
        customer_coupons = []

    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'customer_coupons': customer_coupons,
        'routes_instances': routes_instances,
        'filter_data': filter_data,
    }

    return render(request, 'sales_management/dsr_coupon_book_sales.html', context)



def dsr_coupon_book_sales_print(request):
    filter_data = {}
    data_filter = False
    salesman_id = ""
    van_route = None

    routes_instances = RouteMaster.objects.all()
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')

    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')

    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        route_id = van_route.routes.pk if van_route else None
        filter_data['route_name'] = route_name
    else:
        route_id = None

    if van_route:
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        customer_coupons = CustomerCoupon.objects.filter(salesman=salesman, created_date__date=date)
    else:
        customer_coupons = []

    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'customer_coupons': customer_coupons,
        'routes_instances': routes_instances,
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),
    }

    return render(request, 'sales_management/dsr_coupon_book_sales_print.html', context)
#-----------------DSR FOC Customers--------------------------

def dsr_foc_customers(request):
    # Initialize counts and variables
    filter_data = {}
    data_filter = False
    salesman_id = ""
    van_route = None
    foc_customers = CustomerSupply.objects.none


    # Retrieve all route instances
    routes_instances = RouteMaster.objects.all()

    # Get filter parameters from request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')

    # Set date range and filter data
    if from_date and to_date:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        # Default to today's date if no range is provided
        from_date = to_date = datetime.datetime.today().date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')

    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        route_id = van_route.routes.pk if van_route else None
        filter_data['route_name'] = route_name
    else:
        route_id = None

    # Ensure van_route is defined
    if van_route:
        # Retrieve salesman
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        foc_customers = CustomerSupply.objects.filter(created_date__date__range=[from_date, to_date], customer__sales_type='FOC', salesman=salesman) or CustomerSupply.objects.filter(created_date__date__range=[from_date, to_date], allocate_bottle_to_free__gt=0, salesman=salesman)
    else:
        foc_customers = []

    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'foc_customers': foc_customers,
        'routes_instances': routes_instances,
        'filter_data': filter_data,
    }

    return render(request, 'sales_management/dsr_foc_customers.html', context)

def dsr_foc_customers_print(request):
        # Initialize counts and variables
    filter_data = {}
    data_filter = False
    salesman_id = ""
    van_route = None
    foc_customers = CustomerSupply.objects.none


    # Retrieve all route instances
    routes_instances = RouteMaster.objects.all()

    # Get filter parameters from request
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')

    # Set date and filter data
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')

    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        route_id = van_route.routes.pk if van_route else None
        filter_data['route_name'] = route_name
    else:
        route_id = None

    # Ensure van_route is defined
    if van_route:
        # Retrieve salesman
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        foc_customers = CustomerSupply.objects.filter(created_date__date=date, customer__sales_type='FOC', salesman=salesman) or CustomerSupply.objects.filter(created_date__date=date, allocate_bottle_to_free__gt=0, salesman=salesman)
    else:
        foc_customers = []

    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'foc_customers': foc_customers,
        'routes_instances': routes_instances,
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),

    }

    return render(request, 'sales_management/dsr_foc_customers_print.html', context)



#-----------------DSR Stock Report--------------------------

def dsr_stock_report(request):
    filter_data = {}
    data_filter = False
    stock_report_total = 0

    # Get date filters and route name from the request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')

    # If no filters are applied, set default values
    if not from_date or not to_date:
        from_date = to_date = datetime.datetime.today().date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
        data_filter = True

    # Get all routes initially
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none()

    # If a specific route is selected, filter based on the route
    if route_name:
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        if van_route:
            filter_data['route_name'] = route_name
            van_instances = Van.objects.filter(salesman=van_route.van.salesman)
            van_product_stock = VanProductStock.objects.filter(
                created_date__range=(from_date, to_date),
                van__in=van_instances,
                product__product_name="5 Gallon"
            )
            stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
    else:
        # If no specific route is selected, fetch data for all routes
        van_product_stock = VanProductStock.objects.filter(
            created_date__range=(from_date, to_date),
            product__product_name="5 Gallon"
        )
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0

    context = {
        'data_filter': data_filter,
        'van_route': van_route if route_name else None,
        'routes_instances': routes_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        'filter_data': filter_data,
    }

    return render(request, 'sales_management/dsr_stock_report.html', context)

def dsr_stock_report_print(request):
    filter_data = {}
    data_filter = False
    stock_report_total = 0

    # Get date filters and route name from the request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')

    # If no filters are applied, set default values
    if not from_date or not to_date:
        from_date = to_date = datetime.datetime.today().date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['to_date'] = to_date.strftime('%Y-%m-%d')
        data_filter = True

    # Get all routes initially
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none()

    # If a specific route is selected, filter based on the route
    if route_name:
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        if van_route:
            filter_data['route_name'] = route_name
            van_instances = Van.objects.filter(salesman=van_route.van.salesman)
            van_product_stock = VanProductStock.objects.filter(
                created_date__range=(from_date, to_date),
                van__in=van_instances,
                product__product_name="5 Gallon"
            )
            stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
    else:
        # If no specific route is selected, fetch data for all routes
        van_product_stock = VanProductStock.objects.filter(
            created_date__range=(from_date, to_date),
            product__product_name="5 Gallon"
        )
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0

   
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    products = ProdutItemMaster.objects.none
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none
    
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
      
        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=salesman)
        van_product_stock = VanProductStock.objects.filter(created_date__range=(from_date, to_date),van=van_instances,product__product_name="5 Gallon")
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
       
        'routes_instances': routes_instances,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        'filter_data': filter_data,
    }
    
    return render(request, 'sales_management/dsr_stock_report_print.html', context)

def dsr_expense(request):
    filter_data = {}
    data_filter = False
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    expenses_instanses = Expense.objects.none
    routes_instances = RouteMaster.objects.all()
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        
        ### expenses ####
        expenses_instanses = Expense.objects.filter(expense_date=date,van=van_route.van)
        
       
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
        # expenses
        'expenses_instanses': expenses_instanses,
        
        'filter_data': filter_data,
    }
    return render(request, 'sales_management/dsr_expense.html', context)

def dsr_expense_print(request):
    filter_data = {}
    data_filter = False
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    expenses_instanses = Expense.objects.none
    routes_instances = RouteMaster.objects.all()
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        
        ### expenses ####
        expenses_instanses = Expense.objects.filter(expense_date=date,van=van_route.van)
        
       
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
        # expenses
        'expenses_instanses': expenses_instanses,
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),

    }
    return render(request, 'sales_management/dsr_expense_print.html', context)
def dsr_five_gallon_rates(request):
    filter_data = {}
    data_filter = False
   
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    routes_instances = RouteMaster.objects.all()
    unique_amounts = CustomerCouponItems.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        # 5 gallon rate based
        unique_amounts = set(CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon").values_list('customer_supply__customer__rate', flat=True))
      
        
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
        # 5 gallon rate based
        'five_gallon_rates': unique_amounts,
        'filter_data': filter_data,
    }
    
    return render(request, 'sales_management/dsr_five_gallon_rates.html', context)
def dsr_five_gallon_rates_print(request):
    filter_data = {}
    data_filter = False
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    routes_instances = RouteMaster.objects.all()
    unique_amounts = CustomerCouponItems.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        # 5 gallon rate based
        unique_amounts = set(CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon").values_list('customer_supply__customer__rate', flat=True))
      
        
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
        # 5 gallon rate based
        'five_gallon_rates': unique_amounts,
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),
    }
    
    return render(request, 'sales_management/dsr_five_gallon_rates_print.html', context)

def dsr_credit_outstanding(request):
    filter_data = {}
    data_filter = False
    outstanding_credit_notes_total_amount = 0
    outstanding_credit_notes_received_amount = 0
    outstanding_credit_notes_balance = 0
    
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    routes_instances = RouteMaster.objects.all()
    dialy_collections = InvoiceDailyCollection.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        # collection details
        dialy_collections = CollectionPayment.objects.filter(salesman_id=salesman,amount_received__gt=0) 
        # credit outstanding
        # outstanding_credit_notes = Invoice.objects.filter(invoice_type="credit_invoice",customer__sales_staff=salesman).exclude(created_date__date__gt=date)
        outstanding_credit_notes_total_amount = OutstandingAmount.objects.filter(customer_outstanding__created_date__date__lte=date,customer_outstanding__product_type="amount",customer_outstanding__customer__routes=van_route.routes).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        total_amount_received_upto_today = dialy_collections.filter(created_date__date__lte=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_total_amount = outstanding_credit_notes_total_amount - total_amount_received_upto_today

        # outstanding_credit_notes_total_amount = outstanding_credit_notes_total_amount - dialy_collections.filter(created_date__date__lte=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_received_amount = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_balance = outstanding_credit_notes_total_amount - outstanding_credit_notes_received_amount

       
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
        'outstanding_credit_notes_total_amount' : outstanding_credit_notes_total_amount,
        'outstanding_credit_notes_received_amount' : outstanding_credit_notes_received_amount,
        'outstanding_credit_notes_balance': outstanding_credit_notes_balance,
        'filter_data': filter_data,
    }
    
    return render(request, 'sales_management/dsr_credit_outstanding.html', context)

def dsr_credit_outstanding_print(request):
    filter_data = {}
    data_filter = False
    outstanding_credit_notes_total_amount = 0
    outstanding_credit_notes_received_amount = 0
    outstanding_credit_notes_balance = 0
    
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    routes_instances = RouteMaster.objects.all()
    dialy_collections = InvoiceDailyCollection.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        # collection details
        dialy_collections = CollectionPayment.objects.filter(salesman_id=salesman,amount_received__gt=0) 
        # credit outstanding
        # outstanding_credit_notes = Invoice.objects.filter(invoice_type="credit_invoice",customer__sales_staff=salesman).exclude(created_date__date__gt=date)
        outstanding_credit_notes_total_amount = OutstandingAmount.objects.filter(customer_outstanding__created_date__date__lte=date,customer_outstanding__product_type="amount",customer_outstanding__customer__routes=van_route.routes).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        total_amount_received_upto_today = dialy_collections.filter(created_date__date__lte=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_total_amount = outstanding_credit_notes_total_amount - total_amount_received_upto_today

        # outstanding_credit_notes_total_amount = outstanding_credit_notes_total_amount - dialy_collections.filter(created_date__date__lte=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_received_amount = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_balance = outstanding_credit_notes_total_amount - outstanding_credit_notes_received_amount

       
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'routes_instances': routes_instances,
        'outstanding_credit_notes_total_amount' : outstanding_credit_notes_total_amount,
        'outstanding_credit_notes_received_amount' : outstanding_credit_notes_received_amount,
        'outstanding_credit_notes_balance': outstanding_credit_notes_balance,
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),
    }
    
    return render(request, 'sales_management/dsr_credit_outstanding_print.html', context)

def visitstatistics_report(request):
    filter_data = {}
    data_filter = False

    # Initialize overall counts and variables
    routes_statistics = []
    routes_instances = RouteMaster.objects.all()

    # Get filter parameters from request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')

    # Set date range and filter data
    if from_date and to_date:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['filter_from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['filter_to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        today = datetime.datetime.today().date()
        from_date = today
        to_date = today
        filter_data['filter_from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['filter_to_date'] = to_date.strftime('%Y-%m-%d')

    # If a specific route is selected, filter by it
    if route_name:
        data_filter = True
        routes_instances = routes_instances.filter(route_name=route_name)
        filter_data['route_name'] = route_name

    # Loop through each route to calculate statistics
    for route in routes_instances:
        van_route = Van_Routes.objects.filter(routes=route).first()

        if van_route:
            salesman = van_route.van.salesman
            driver = van_route.van.driver  # Ensure driver is a valid related field

            # Get names
            salesman_name = salesman.get_fullname() if salesman else "N/A"
            driver_name = driver.get_fullname() if driver else "N/A"

            # Calculate statistics for the current route
            new_customers_count = Customers.objects.filter(is_guest=False, 
                created_date__date__range=[from_date, to_date],
                sales_staff_id=salesman
            ).count()
            emergency_supply_count = DiffBottlesModel.objects.filter(
                created_date__date__range=[from_date, to_date],
                assign_this_to_id=salesman
            ).count()
            visited_customers_count = CustomerSupply.objects.filter(
                salesman_id=salesman,
                created_date__date__range=[from_date, to_date]
            ).distinct().count()

            # Find today's customers
            todays_customers = find_customers(request, str(from_date), route.pk)
            
            # Ensure todays_customers is not None
            if todays_customers is None:
                planned_visit_count = 0
            else:
                planned_visit_count = len(todays_customers)

            # Calculate non-visited customers as an absolute value
            non_visited_count = abs(planned_visit_count - visited_customers_count)

            # Add route statistics to the list
            routes_statistics.append({
                'salesman_id': salesman.pk,
                'route': route.route_name,
                'new_customers_count': new_customers_count,
                'planned_visit_count': planned_visit_count,
                'visited_customers_count': visited_customers_count,
                'non_visited_count': non_visited_count,
                'emergency_supply_count': emergency_supply_count,
                'salesman_name': salesman_name,
                'driver_name': driver_name,
            })

    context = {
        'routes_statistics': routes_statistics,
        'filter_data': filter_data,
        'data_filter': data_filter,
        'routes_instances': routes_instances,
    }

    return render(request, 'sales_management/dsr_visit_statistics_report.html', context)


def visitstatistics_report_print(request):
    filter_data = {}
    data_filter = False
    routes_statistics = []
    
    # Get filter parameters from request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    route_name = request.GET.get('route_name')
    
    # Set date range and filter data
    if from_date and to_date:
        from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        filter_data['filter_from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['filter_to_date'] = to_date.strftime('%Y-%m-%d')
    else:
        today = datetime.datetime.today().date()
        from_date = today
        to_date = today
        filter_data['filter_from_date'] = from_date.strftime('%Y-%m-%d')
        filter_data['filter_to_date'] = to_date.strftime('%Y-%m-%d')

    # Filter routes if a specific route is selected
    routes_instances = RouteMaster.objects.all()
    if route_name:
        data_filter = True
        routes_instances = routes_instances.filter(route_name=route_name)
        filter_data['route_name'] = route_name

    # Loop through each route to calculate statistics
    for route in routes_instances:
        van_route = Van_Routes.objects.filter(routes=route).first()

        if van_route:
            salesman = van_route.van.salesman
            driver = van_route.van.driver

            # Get names
            salesman_name = salesman.get_fullname() if salesman else "N/A"
            driver_name = driver.get_fullname() if driver else "N/A"

            # Calculate statistics for the current route
            new_customers_count = Customers.objects.filter(is_guest=False, 
                created_date__date__range=[from_date, to_date],
                sales_staff_id=salesman
            ).count()
            emergency_supply_count = DiffBottlesModel.objects.filter(
                created_date__date__range=[from_date, to_date],
                assign_this_to_id=salesman
            ).count()
            visited_customers_count = CustomerSupply.objects.filter(
                salesman_id=salesman,
                created_date__date__range=[from_date, to_date]
            ).distinct().count()

            # Find today's customers
            todays_customers = find_customers(request, str(from_date), route.pk)
            planned_visit_count = len(todays_customers) if todays_customers else 0

            # Calculate non-visited customers as an absolute value
            non_visited_count = abs(planned_visit_count - visited_customers_count)

            # Add route statistics to the list
            routes_statistics.append({
                'salesman_id': salesman.pk,
                'route': route.route_name,
                'new_customers_count': new_customers_count,
                'planned_visit_count': planned_visit_count,
                'visited_customers_count': visited_customers_count,
                'non_visited_count': non_visited_count,
                'emergency_supply_count': emergency_supply_count,
                'salesman_name': salesman_name,
                'driver_name': driver_name,
            })

    context = {
        'routes_statistics': routes_statistics,
        'filter_data': filter_data,
        'data_filter': data_filter,
    }

    return render(request, 'sales_management/dsr_visit_statistics_report_print.html', context)



def fivegallonrelated_report(request):
    user_id = request.user.id
    start_date = request.GET.get('start_date')
    filter_data = {}

    if start_date:
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=start_date, collected_empty_bottle__gt=0, salesman_id=user_id).count()

    # empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=start_date, collected_empty_bottle__gt=0, **filter_data).count()
    empty_bottle_pending = CustomerSupply.objects.filter(created_date__date=start_date, allocate_bottle_to_pending__gt=0, salesman_id=user_id).count()
    print(empty_bottle_pending,'empty_bottle_pending')
    coupons_collected = CustomerSupplyCoupon.objects.filter(customer_supply__created_date__date=start_date, customer_supply__salesman_id=user_id).aggregate(total_coupons=Count('leaf'))['total_coupons']
    total_supplied_quantity = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=start_date, customer_supply__salesman_id=user_id).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
    total_collected_amount = CustomerSupply.objects.filter(created_date__date=start_date, salesman_id=user_id).aggregate(total_collected_amount=Sum('net_payable'))['total_collected_amount'] or 0
    total_pending_amount = CustomerSupply.objects.filter(created_date__date=start_date, salesman_id=user_id).aggregate(total_pending_amount=Sum('grand_total') - Sum('net_payable'))['total_pending_amount'] or 0
    mode_of_supply = CustomerSupply.objects.filter(created_date__date=start_date, salesman_id=user_id).values('customer__sales_type').annotate(total=Count('customer__sales_type'))

    context = {
        'empty_bottles_collected': empty_bottles_collected,
        'empty_bottle_pending': empty_bottle_pending,
        'coupons_collected': coupons_collected,
        'total_supplied_quantity': total_supplied_quantity,
        'total_collected_amount': total_collected_amount,
        'total_pending_amount': total_pending_amount,
        'mode_of_supply' :mode_of_supply
    }

    return render(request, 'sales_management/dsr_fivegallonrelated_report.html', context)

    
def bottlecount_report(request):
    user_id = request.user.id
    start_date = request.GET.get('start_date')
    filter_data = {}
    if start_date:
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    
    total_empty_bottles = CustomerCustodyStock.objects.filter(customer__sales_staff_id=user_id).aggregate(total_quantity=Sum('quantity'))['total_quantity']
    if total_empty_bottles is None:
        total_empty_bottles = 0

    total_supplied_bottles = CustomerSupply.objects.filter(created_date__date=start_date).aggregate(total_bottles=Sum('collected_empty_bottle'))['total_bottles']
    # closing_stock_count = VanStock.objects.filter(created_date=start_date,stock_type="closing").aggregate(total_count=Sum('count'))['total_count'] or 0,
    closing_stock_count = VanStock.objects.filter(created_date=start_date, stock_type='closing').count() or 0
    damage_bottle_count = VanProductItems.objects.filter(van_stock__created_date=start_date, van_stock__stock_type='damage').aggregate(total_damage=Sum('count'))['total_damage'] or 0
    pending_bottle_count = CustomerSupply.objects.filter(created_date__date=start_date,salesman_id=user_id).aggregate(total_pending=Sum('allocate_bottle_to_pending'))['total_pending'] or 0
    if pending_bottle_count is None:
        pending_bottle_count = 0

    total_count = total_empty_bottles + total_supplied_bottles + closing_stock_count + damage_bottle_count + pending_bottle_count

    context = {
        'total_empty_bottles': total_empty_bottles,
        'total_supplied_bottles':total_supplied_bottles,
        'closing_stock_count': closing_stock_count,
        'damage_bottle_count': damage_bottle_count,
        'pending_bottle_count': pending_bottle_count,
        'total_count': total_count,
    }

    return render(request, 'sales_management/dsr_bottlecount_report.html', context)


def dsr_summary(request):
    
    filter_data = {}
    data_filter = False
    total_count = 0
    non_visited_count = 0
    new_customers_count = 0
    planned_visit_count = 0
    total_empty_bottles = 0
    closing_stock_count = 0
    damage_bottle_count = 0
    pending_bottle_count = 0
    total_supplied_bottles = 0
    emergency_supply_count = 0
    visited_customers_count = 0
    cash_total_net_taxable = 0
    cash_total_vat = 0
    cash_total_subtotal = 0
    cash_total_amount_recieved = 0
    cash_sale_recharge_count=0
    cash_total_quantity = 0
    cash_total_qty = 0
    credit_total_net_taxable = 0
    credit_total_vat = 0
    credit_total_subtotal = 0
    credit_total_amount_recieved = 0
    credit_total_qty = 0
    credit_total_quantity = 0
    credit_sale_recharge_count = 0
    in_hand_amount = 0
    today_expense = 0
    today_payable = 0
    suspense_paid_amount = 0
    suspense_balance_amount = 0
    todays_opening_outstanding_amount = 0
    credit_outstanding_collected_amount = 0
    outstanding_closing_balance = 0
    outstanding_total_amount_collected = 0
    cash_sales_amount_collected = 0
    credit_sales_amount_collected = 0
    total_sales_amount_collected = 0
    total_cash_sales_count = 0
    total_credit_sales_count = 0
    total_sales_count = 0
    no_of_collected_cheque = 0
    collected_cheque_amount = 0
    balance_in_hand = 0
    net_payble = 0
    stock_report_total = 0
    manual_coupon_total = 0
    digital_coupon_total = 0
    total_coupon_sales_count = 0
    coupon_total_qty = 0
    total_outstandaing_manual_coupons = 0
    total_outstandaing_digital_coupons =0
    total_debit_amount_count = 0
    total_credit_amount_count = 0
    total_coupon_amount_count = 0
    
    five_gallon_cash_total_net_taxable = 0
    five_gallon_cash_total_vat = 0
    five_gallon_cash_total_subtotal = 0
    five_gallon_cash_total_received = 0
    five_gallon_cash_total_quantity = 0
    other_cash_total_net_taxable = 0
    other_cash_total_vat = 0
    other_cash_total_subtotal = 0
    other_cash_total_received = 0
    other_cash_total_quantity = 0
    cash_sale_recharge_net_payeble = 0
    cash_sale_recharge_vat_total = 0
    cash_sale_recharge_grand_total = 0
    cash_sale_recharge_amount_recieved = 0
    
    five_gallon_credit_total_net_taxable = 0
    five_gallon_credit_total_vat = 0
    five_gallon_credit_total_subtotal = 0
    five_gallon_credit_total_received = 0
    five_gallon_credit_total_quantity = 0
    other_credit_total_net_taxable = 0
    other_credit_total_vat = 0
    other_credit_total_subtotal = 0
    other_credit_total_received = 0
    other_credit_total_quantity = 0
    credit_sale_recharge_net_payeble = 0
    credit_sale_recharge_vat_total = 0
    credit_sale_recharge_grand_total = 0
    credit_sale_recharge_amount_recieved = 0
    foc_total_quantity = 0
    
    customer_custody_instances = CustodyCustom.objects.none
    collected_amount_from_custody_issue = 0
    customer_custody_items_fgallon_instances = CustodyCustomItems.objects.none
    customer_custody_items_fgallon_count = 0
    customer_custody_items_nonfgallon_instances = CustodyCustomItems.objects.none
    customer_custody_items_nonfgallon_count = 0
   
    salesman_id =  None
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    products = ProdutItemMaster.objects.none
    cash_sales = CustomerSupply.objects.none
    expenses_instanses = Expense.objects.none
    credit_sales = CustomerSupply.objects.none
    coupon_sales = CustomerSupply.objects.none
    foc_customers = CustomerSupply.objects.none
    routes_instances = RouteMaster.objects.all()
    customer_coupons = CustomerCoupon.objects.none
    other_cash_sales = CustomerSupply.objects.none
    other_credit_sales = CustomerSupply.objects.none
    van_product_stock = VanProductStock.objects.none
    unique_amounts = CustomerCouponItems.objects.none
    recharge_cash_sales = CustomerCoupon.objects.none
    recharge_credit_sales = CustomerCoupon.objects.none
    five_gallon_cash_sales = CustomerSupply.objects.none
    damage_bottle_instances = VanSaleDamage.objects.none
    five_gallon_credit_sales = CustomerSupply.objects.none
    dialy_collections = InvoiceDailyCollection.objects.none
    pending_bottle_customer_instances = CustomerSupply.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    salesman_id = request.GET.get('salesman_id') or None
    salesman = None
    if salesman_id: 
        salesman = CustomUser.objects.filter(
        pk=salesman_id,
        user_type="Salesman"
        ).first()
        salesman_id = salesman.pk   

    print("salesmanId:",salesman_id)
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        
    yesterday_date = date - datetime.timedelta(days=1)
    salesman_list = CustomUser.objects.filter(user_type="Salesman").order_by('first_name')
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        
        filter_data['route_name'] = route_name
        filter_data['salesman_id'] = str(salesman_id)
        #new customers created
        new_customers_qs = Customers.objects.filter(
            is_guest=False,
            created_date__date=date,
            routes__route_name=route_name
        )

        if salesman_id:
            new_customers_qs = new_customers_qs.filter(
                sales_staff_id=salesman_id
            )

        new_customers_count = new_customers_qs.count()
        #emergency supply
        emergency_supply_qs = DiffBottlesModel.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            emergency_supply_qs = emergency_supply_qs.filter(
                assign_this_to_id=salesman_id
            )

        emergency_supply_count = emergency_supply_qs.count()
        #actual visit
        visited_customers_qs = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            visited_customers_qs = visited_customers_qs.filter(
                salesman_id=salesman_id
            )

        visited_customers_count = visited_customers_qs.distinct().count()
        todays_customers = find_customers(request, str(date), van_route.routes.pk)
        if todays_customers :
            planned_visit_count = len(todays_customers)
        else:
            planned_visit_count = 0
        non_visited_count = planned_visit_count - visited_customers_count
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        van_salesman = van_route.van.salesman
        

        ##### stock report #### 
        products =  ProdutItemMaster.objects.filter(product_name="5 Gallon")
        print(products)
        van_instances = Van.objects.get(salesman=van_salesman)
        van_product_stock = VanProductStock.objects.filter(created_date=date,van=van_instances,product__product_name="5 Gallon")
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        # vanstock_curreptions(van_instances.pk,date)
        
        #### Bottle Count ####
        total_empty_bottles = van_product_stock.aggregate(totalempty_bottle=Sum('empty_can_count'))['totalempty_bottle'] or 0
        total_supplied_bottles =  van_product_stock.aggregate(total_sold=Sum('sold_count'))['total_sold'] or 0
        pending_bottle_count = van_product_stock.aggregate(total_pending=Sum('pending_count'))['total_pending'] or 0
        damage_bottle_count = van_product_stock.aggregate(total_damage=Sum('damage_count'))['total_damage'] or 0
        closing_stock_count = van_product_stock.aggregate(total_closing=Sum('closing_count'))['total_closing'] or 0
        total_count = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        #### coupon sales count ####
        customer_coupons=CustomerCoupon.objects.filter(created_date__date=date,customer__routes__route_name=route_name)
        if salesman_id:
            customer_coupons = customer_coupons.filter(
                salesman_id=salesman_id
            )
        
        five_gallon_supply = CustomerSupplyItems.objects.filter(
            customer_supply__supply_date__date=date,
            product__product_name="5 Gallon",
            customer_supply__customer__routes__route_name=route_name
        )

      

        if salesman_id:
            five_gallon_supply = five_gallon_supply.filter(
                customer_supply__salesman_id=salesman_id
            )

        five_gallon_supply = five_gallon_supply.values_list(
            'customer_supply__pk', flat=True
        )

        other_supply = CustomerSupplyItems.objects.filter(
        customer_supply__customer__routes__route_name=route_name
        ).exclude(
            product__product_name="5 Gallon"
        )

        if salesman_id:
            other_supply = other_supply.filter(
                customer_supply__salesman_id=salesman_id
            )

        other_supply = other_supply.filter(
            customer_supply__created_date__date=date
        ).values_list('customer_supply__pk', flat=True)
        ### Cash Sales Start ###
        five_gallon_cash_sales = CustomerSupply.objects.filter(
            pk__in=five_gallon_supply,
            supply_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__gt=0
        ).exclude(customer__sales_type="CASH COUPON")

        if salesman_id:
            five_gallon_cash_sales = five_gallon_cash_sales.filter(
                salesman_id=salesman_id
            )

        other_cash_sales = CustomerSupply.objects.filter(
            pk__in=other_supply,
            supply_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__gt=0
        ).exclude(customer__sales_type="CASH COUPON")

        if salesman_id:
            other_cash_sales = other_cash_sales.filter(
                salesman_id=salesman_id
            )
        # Aggregating for five_gallon_cash_sales
        five_gallon_cash_total_net_taxable = five_gallon_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_cash_total_vat = five_gallon_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_cash_total_subtotal = five_gallon_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_cash_total_received = five_gallon_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_cash_total_quantity = five_gallon_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_cash_sales (corrected to use other_cash_sales)
        other_cash_total_net_taxable = other_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_cash_total_vat = other_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_cash_total_subtotal = other_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_cash_total_received = other_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_cash_total_quantity = other_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Combining both cash sales querysets
        cash_sales = five_gallon_cash_sales.union(other_cash_sales)
        cash_total_net_taxable = five_gallon_cash_total_net_taxable + other_cash_total_net_taxable
        cash_total_vat = five_gallon_cash_total_vat + other_cash_total_vat
        cash_total_subtotal = five_gallon_cash_total_subtotal + other_cash_total_subtotal
        cash_total_received = five_gallon_cash_total_received + other_cash_total_received
        cash_total_quantity = five_gallon_cash_total_quantity + other_cash_total_quantity
        
        recharge_cash_sales = CustomerCoupon.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__gt=0
        )

       
        if salesman_id:
            recharge_cash_sales = recharge_cash_sales.filter(
                salesman_id=salesman_id
            )
        cash_sale_recharge_net_payeble = recharge_cash_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        cash_sale_recharge_vat_total = 0
        cash_sale_recharge_grand_total = recharge_cash_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        cash_sale_recharge_amount_recieved = recharge_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        cash_total_net_taxable = cash_total_net_taxable + cash_sale_recharge_net_payeble 
        cash_total_vat = cash_total_vat + cash_sale_recharge_vat_total 
        cash_total_subtotal = cash_total_subtotal + cash_sale_recharge_grand_total 
        cash_total_amount_recieved = cash_total_received + cash_sale_recharge_amount_recieved 
        
        total_cash_sales_count = cash_sales.count() + recharge_cash_sales.count()
        
        cash_sale_recharge_count = recharge_cash_sales.count()
        cash_total_qty = cash_total_quantity + cash_sale_recharge_count
        ### Cash Sales End ###
        
        ### credit sales Start ####
        five_gallon_credit_sales = CustomerSupply.objects.filter(
            pk__in=five_gallon_supply,
            supply_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__lte=0
            ).exclude(
                customer__sales_type__in=["FOC", "CASH COUPON"]
        )

        if salesman_id:
            five_gallon_credit_sales = five_gallon_credit_sales.filter(
                salesman_id=salesman_id
            )
        other_credit_sales = CustomerSupply.objects.filter(
        pk__in=other_supply,
        supply_date__date=date,
        customer__routes__route_name=route_name,
        amount_recieved__lte=0
        ).exclude(
            customer__sales_type__in=["FOC", "CASH COUPON"]
        )

        if salesman_id:
            other_credit_sales = other_credit_sales.filter(
                salesman_id=salesman_id
            )
        
        # Aggregating for five_gallon_credit_sales
        five_gallon_credit_total_net_taxable = five_gallon_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_credit_total_vat = five_gallon_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_credit_total_subtotal = five_gallon_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_credit_total_received = five_gallon_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_credit_total_quantity = five_gallon_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_credit_sales (corrected to use other_credit_sales)
        other_credit_total_net_taxable = other_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_credit_total_vat = other_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_credit_total_subtotal = other_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_credit_total_received = other_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_credit_total_quantity = other_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Combining both cash sales querysets
        credit_sales = five_gallon_credit_sales.union(other_credit_sales)
        credit_total_net_taxable = five_gallon_credit_total_net_taxable + other_credit_total_net_taxable
        credit_total_vat = five_gallon_credit_total_vat + other_credit_total_vat
        credit_total_subtotal = five_gallon_credit_total_subtotal + other_credit_total_subtotal
        credit_total_received = five_gallon_credit_total_received + other_credit_total_received
        credit_total_quantity = five_gallon_credit_total_quantity + other_credit_total_quantity
        
        recharge_credit_sales = CustomerCoupon.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__lte=0
        )

        if salesman_id:
            recharge_credit_sales = recharge_credit_sales.filter(
                salesman_id=salesman_id
            )
        credit_sale_recharge_net_payeble = recharge_credit_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        credit_sale_recharge_vat_total = 0
        credit_sale_recharge_grand_total = recharge_credit_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        credit_sale_recharge_amount_recieved = recharge_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        credit_total_net_taxable = credit_total_net_taxable + credit_sale_recharge_net_payeble 
        credit_total_vat = credit_total_vat + credit_sale_recharge_vat_total 
        credit_total_subtotal = credit_total_subtotal + credit_sale_recharge_grand_total
        credit_total_amount_recieved = credit_total_received + credit_sale_recharge_amount_recieved
        
        total_credit_sales_count = credit_sales.count() + recharge_credit_sales.count()
        credit_sale_recharge_count = recharge_credit_sales.count()
        credit_total_qty = credit_total_quantity + credit_sale_recharge_count
        ### credit sales End ####
        
        # Coupon sales
        coupon_sales = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name,
            customer__sales_type="CASH COUPON"
        )

        if salesman_id:
            coupon_sales = coupon_sales.filter(
                salesman_id=salesman_id
            )
        for sale in coupon_sales:
            sale.outstanding_manual_coupons = sale.outstanding_manual_coupons()
            sale.outstanding_digital_coupons = sale.outstanding_digital_coupons()
        manual_coupon_total = CustomerSupplyCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(Count('leaf'))['leaf__count']
        digital_coupon_total = CustomerSupplyDigitalCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(total_count=Sum('count'))['total_count'] or 0
        
        total_coupon_sales_count = coupon_sales.count()
        
        total_sales_count = total_cash_sales_count + total_credit_sales_count + total_coupon_sales_count
        coupon_total_qty = coupon_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Calculate total outstanding digital and manual coupons
        total_outstandaing_manual_coupons = sum(sale.outstanding_manual_coupons for sale in coupon_sales)
        total_outstandaing_digital_coupons = sum(sale.outstanding_digital_coupons for sale in coupon_sales)
        
        ### expenses ####
        expenses_instanses = Expense.objects.filter(expense_date=date,route__route_name=route_name)
        today_expense = expenses_instanses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
        
        ### suspense ###
        suspense_collections = SuspenseCollection.objects.filter(
            created_date__date=date,
            route__route_name=route_name
        )
        if salesman_id:
            suspense_collections = suspense_collections.filter(
                salesman_id=salesman_id
            )
        cash_sales_amount = suspense_collections.aggregate(total_cash_sale=Sum('cash_sale_amount'))['total_cash_sale'] or 0
        credit_sales_amount = suspense_collections.aggregate(total_credit_sale=Sum('credit_sale_amount'))['total_credit_sale'] or 0
        
        in_hand_amount = cash_sales_amount + credit_sales_amount
        today_payable = in_hand_amount - today_expense
        suspense_paid_amount = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
        suspense_balance_amount = today_payable - suspense_paid_amount
        
        # collection details
        dialy_collections = CollectionPayment.objects.filter(
            customer__is_guest=False,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            dialy_collections = dialy_collections.filter(
                salesman_id=salesman_id
            )
        # credit outstanding
        qs = OutstandingAmount.objects.filter(
            customer_outstanding__product_type="amount",
            customer_outstanding__outstanding_date__date__lte=yesterday_date,
            customer_outstanding__customer__is_guest=False,
            customer_outstanding__customer__routes__route_name=route_name
        )

        if salesman_id:
            qs = qs.filter(
                customer_outstanding__created_by=salesman_id
            )

        outstanding_amount_upto_yesterday = (
            qs.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        )
        dialy_colection_upto_yesterday = dialy_collections.filter(created_date__date__lte=yesterday_date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        todays_opening_outstanding_amount = outstanding_amount_upto_yesterday - dialy_colection_upto_yesterday
        credit_outstanding_collected_amount = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_closing_balance = todays_opening_outstanding_amount + credit_total_subtotal - credit_outstanding_collected_amount
        
        print(outstanding_amount_upto_yesterday)
        print(dialy_colection_upto_yesterday)
        print(credit_outstanding_collected_amount)
        # pending customers
        pending_bottle_customer_instances = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name,
            allocate_bottle_to_pending__gt=0,
        )

       

        if salesman_id:
            pending_bottle_customer_instances = pending_bottle_customer_instances.filter(
                salesman_id=salesman_id
            )
        # 5 gallon rate based
        customer_supplies = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name
        )
        if salesman_id:
            customer_supplies = customer_supplies.filter(
                salesman_id=salesman_id
            )
        invoice_nos = customer_supplies.values_list('invoice_no', flat=True)
        invoice_items = InvoiceItems.objects.filter(invoice__invoice_no__in=invoice_nos)

        rate_expr = ExpressionWrapper(
            F('rate') / F('qty'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )

        # Annotate with calculated_rate
        invoice_items = invoice_items.annotate(calculated_rate=Round(rate_expr, 2))

        # ✅ Exclude zero or null values
        invoice_items = invoice_items.exclude(calculated_rate__isnull=True).exclude(calculated_rate=0)

        # Get unique non-zero rates

        five_gallon_rates = CustomerSupplyItems.objects.filter(
            customer_supply__supply_date__date=date,
            customer_supply__customer__routes__route_name=route_name,
            product__product_name="5 Gallon",
            quantity__gt=0,
            amount__gt=0
        )

        # ✅ Apply salesman filter ONLY if provided
        if salesman:
            five_gallon_rates = five_gallon_rates.filter(
                customer_supply__salesman_id=salesman
            )

        five_gallon_rates = (
            five_gallon_rates
            .annotate(
                calculated_rate=ExpressionWrapper(
                    F("amount") / F("quantity"),
                    output_field=DecimalField(max_digits=10, decimal_places=2)
                )
            )
            .values_list("calculated_rate", flat=True)
            .exclude(calculated_rate=0)
            .distinct()
            .order_by("calculated_rate")
        )

        
        unique_amounts = five_gallon_rates
        # unique_amounts = set(CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon").values_list('customer_supply__customer__rate', flat=True))
        five_gallon_rate_wise_instances = CustomerSupplyItems.objects.filter(
            customer_supply__supply_date__date=date,
            customer_supply__customer__routes__route_name=route_name,
            product__product_name="5 Gallon"
        )

        if salesman_id:
            five_gallon_rate_wise_instances = five_gallon_rate_wise_instances.filter(
                customer_supply__salesman_id=salesman_id
            )
        total_debit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved__gt=0).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_credit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved=0).exclude(customer_supply__customer__sales_type__in=["FOC","CASH COUPON"]).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_coupon_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__customer__sales_type="CASH COUPON").aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        # Custody Issue 
        customer_custody_instances = CustodyCustom.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            customer_custody_instances = customer_custody_instances.filter(
                created_by=salesman_id
            )
        collected_amount_from_custody_issue = customer_custody_instances.aggregate(total_amount=Sum('amount_collected'))['total_amount'] or 0
        
        customer_custody_items_fgallon_instances = CustodyCustomItems.objects.filter(custody_custom__pk__in=customer_custody_instances.values_list('pk'),product__product_name="5 Gallon")
        customer_custody_items_fgallon_count = customer_custody_items_fgallon_instances.aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
        # custody other products
        customer_custody_items_nonfgallon_instances = CustodyCustomItems.objects.filter(custody_custom__pk__in=customer_custody_instances.values_list('pk')).exclude(product__product_name="5 Gallon")
        customer_custody_items_nonfgallon_count = customer_custody_items_nonfgallon_instances.aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
        
        # cash sales amount collected
        supply_amount_collected = cash_total_amount_recieved
        cash_sales_amount_collected = supply_amount_collected
        
        dialy_collections = dialy_collections.filter(created_date__date=date)
        credit_sales_amount_collected = dialy_collections.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        total_sales_amount_collected = cash_sales_amount_collected + credit_sales_amount_collected
        
        cheque_collection = CollectionPayment.objects.filter(
            payment_method="CHEQUE",
            created_date__date=date,
            customer__routes__route_name=route_name
        )

       
        if salesman_id:
            cheque_collection = cheque_collection.filter(
                salesman_id=salesman_id
            )
        no_of_collected_cheque = cheque_collection.count()
        collected_cheque_amount = cheque_collection.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        
        balance_in_hand = total_sales_amount_collected + collected_amount_from_custody_issue - collected_cheque_amount - today_expense
        net_payble = total_sales_amount_collected + collected_amount_from_custody_issue - today_expense
        foc_qs_1 = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__sales_type='FOC',
            customer__routes__route_name=route_name
        )

        foc_qs_2 = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name,
            allocate_bottle_to_free__gt=0,
        )
        if salesman_id:
            foc_qs_1 = foc_qs_1.filter(
                salesman_id=salesman_id
            )
            foc_qs_2 = foc_qs_2.filter(
                salesman_id=salesman_id
            )
        foc_customers = (
             foc_qs_1 | foc_qs_2
            
        )
        for foc in foc_customers:
            if foc.allocate_bottle_to_free != 0:
                foc_total_quantity =foc_total_quantity + foc.allocate_bottle_to_free
            else:
                foc_total_quantity =foc_total_quantity + foc.get_total_supply_qty()
                
        damage_bottle_instances = VanSaleDamage.objects.filter(van=van_instances,created_date__date=date)
        
    context = {
        'van_route': van_route,
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        # visit statistics
        'routes_instances': routes_instances,
        'non_visited_count': non_visited_count,
        'planned_visit_count': planned_visit_count,
        'new_customers_count': new_customers_count,
        'emergency_supply_count': emergency_supply_count,
        'visited_customers_count': visited_customers_count,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        # pending customers
        'pending_bottle_customer_instances': pending_bottle_customer_instances,
        # Bottle Count
        'total_count': total_count,
        'total_empty_bottles': total_empty_bottles,
        'closing_stock_count': closing_stock_count,
        'damage_bottle_count': damage_bottle_count,
        'pending_bottle_count': pending_bottle_count,
        'total_supplied_bottles':total_supplied_bottles,
        #coupon book sale
        'customer_coupons':customer_coupons,
        # five gallon cash sales
        'five_gallon_cash_sales': five_gallon_cash_sales or [],
        'five_gallon_cash_total_vat': five_gallon_cash_total_vat,
        'five_gallon_cash_total_subtotal': five_gallon_cash_total_subtotal,
        'five_gallon_cash_total_received': five_gallon_cash_total_received,
        'five_gallon_cash_total_quantity': five_gallon_cash_total_quantity,
        'five_gallon_cash_total_net_taxable': five_gallon_cash_total_net_taxable,
        # other cash sales
        'other_cash_sales': other_cash_sales,
        'other_cash_total_vat': other_cash_total_vat,
        'other_cash_total_subtotal': other_cash_total_subtotal,
        'other_cash_total_received': other_cash_total_received,
        'other_cash_total_quantity': other_cash_total_quantity,
        'other_cash_total_net_taxable': other_cash_total_net_taxable,
        #cash sales
        'recharge_cash_sales': recharge_cash_sales,
        'cash_sale_recharge_vat_total': cash_sale_recharge_vat_total,
        'cash_sale_recharge_grand_total': cash_sale_recharge_grand_total,
        'cash_sale_recharge_net_payeble': cash_sale_recharge_net_payeble,
        'cash_sale_recharge_amount_recieved': cash_sale_recharge_amount_recieved,
        
        'cash_total_vat':cash_total_vat,
        'cash_total_qty': cash_total_qty,
        'cash_total_subtotal': cash_total_subtotal,
        'cash_total_net_taxable':cash_total_net_taxable,
        'cash_total_amount_recieved': cash_total_amount_recieved,
        # five gallon credit sales
        'five_gallon_credit_sales': five_gallon_credit_sales,
        'five_gallon_credit_total_vat': five_gallon_credit_total_vat,
        'five_gallon_credit_total_subtotal': five_gallon_credit_total_subtotal,
        'five_gallon_credit_total_received': five_gallon_credit_total_received,
        'five_gallon_credit_total_quantity': five_gallon_credit_total_quantity,
        'five_gallon_credit_total_net_taxable': five_gallon_credit_total_net_taxable,
        # credit sales
        'credit_sales': credit_sales,
        'recharge_credit_sales': recharge_credit_sales,
        'credit_sale_recharge_vat_total': credit_sale_recharge_vat_total,
        'credit_sale_recharge_net_payeble': credit_sale_recharge_net_payeble,
        'credit_sale_recharge_grand_total': credit_sale_recharge_grand_total,
        'credit_sale_recharge_amount_recieved': credit_sale_recharge_amount_recieved,
        
        'credit_total_vat':credit_total_vat,
        'credit_total_qty': credit_total_qty,
        'credit_total_subtotal':credit_total_subtotal,
        'credit_total_net_taxable':credit_total_net_taxable,
        'credit_total_amount_recieved': credit_total_amount_recieved,
        # coupon sales
        'coupon_sales': coupon_sales,
        'coupon_total_qty':coupon_total_qty,
        'manual_coupon_total':manual_coupon_total,
        'digital_coupon_total':digital_coupon_total,
        'total_coupon_sales_count': total_coupon_sales_count,
        'total_outstandaing_manual_coupons':total_outstandaing_manual_coupons,
        'total_outstandaing_digital_coupons':total_outstandaing_digital_coupons,
        
        # expenses
        'expenses_instanses': expenses_instanses,
        # suspense
        'today_expense': today_expense,
        'today_payable': today_payable,
        'in_hand_amount': in_hand_amount,
        'suspense_paid_amount': suspense_paid_amount,
        'suspense_balance_amount': suspense_balance_amount,
        'outstanding_closing_balance': outstanding_closing_balance,
        'outstanding_total_amount_collected':outstanding_total_amount_collected,
        'todays_opening_outstanding_amount' : todays_opening_outstanding_amount,
        'credit_outstanding_collected_amount' : credit_outstanding_collected_amount,
        # 5 gallon rate based
        'five_gallon_rates': unique_amounts,
        'total_debit_amount_count': total_debit_amount_count,
        'total_credit_amount_count': total_credit_amount_count,
        'total_coupon_amount_count': total_coupon_amount_count,
        # dialy collections
        'dialy_collections': dialy_collections,
        # sales amount collected
        'total_sales_count': total_sales_count,
        'no_of_collected_cheque': no_of_collected_cheque,
        'total_cash_sales_count': total_cash_sales_count,
        'collected_cheque_amount': collected_cheque_amount,
        'total_credit_sales_count': total_credit_sales_count,
        'cash_sales_amount_collected': cash_sales_amount_collected,
        'total_sales_amount_collected': total_sales_amount_collected,
        'credit_sales_amount_collected': credit_sales_amount_collected,
        # custody details
        'customer_custody_instances': customer_custody_instances,
        'collected_amount_from_custody_issue': collected_amount_from_custody_issue,
        'customer_custody_items_fgallon_instances': customer_custody_items_fgallon_instances,
        'customer_custody_items_fgallon_count': customer_custody_items_fgallon_count,
        'customer_custody_items_nonfgallon_instances': customer_custody_items_nonfgallon_instances,
        'customer_custody_items_nonfgallon_count': customer_custody_items_nonfgallon_count,
        
        'damage_bottle_instances': damage_bottle_instances,
        
        'net_payble': net_payble,
        'balance_in_hand': balance_in_hand,
        
        'filter_data': filter_data,
        # FOC customer
        'foc_customers':foc_customers,
        'foc_total_quantity':foc_total_quantity,
        'salesman_list':salesman_list,
        'selected_salesman': salesman,
        
        
    }

    print("fetch completed")
    
    return render(request, 'sales_management/dsr_summary.html', context)



def dsr_summary1(request):
    
    filter_data = {}
    data_filter = False
    total_count = 0
    non_visited_count = 0
    new_customers_count = 0
    planned_visit_count = 0
    total_empty_bottles = 0
    closing_stock_count = 0
    damage_bottle_count = 0
    pending_bottle_count = 0
    total_supplied_bottles = 0
    emergency_supply_count = 0
    visited_customers_count = 0
    cash_total_net_taxable = 0
    cash_total_vat = 0
    cash_total_subtotal = 0
    cash_total_amount_recieved = 0
    cash_sale_recharge_count=0
    cash_total_quantity = 0
    cash_total_qty = 0
    credit_total_net_taxable = 0
    credit_total_vat = 0
    credit_total_subtotal = 0
    credit_total_amount_recieved = 0
    credit_total_qty = 0
    credit_total_quantity = 0
    credit_sale_recharge_count = 0
    in_hand_amount = 0
    today_expense = 0
    today_payable = 0
    suspense_paid_amount = 0
    suspense_balance_amount = 0
    outstanding_credit_notes_total_amount = 0
    outstanding_credit_notes_received_amount = 0
    outstanding_credit_notes_balance = 0
    outstanding_total_amount_collected = 0
    cash_sales_amount_collected = 0
    credit_sales_amount_collected = 0
    total_sales_amount_collected = 0
    total_cash_sales_count = 0
    total_credit_sales_count = 0
    total_sales_count = 0
    no_of_collected_cheque = 0
    collected_cheque_amount = 0
    balance_in_hand = 0
    net_payble = 0
    stock_report_total = 0
    manual_coupon_total = 0
    digital_coupon_total = 0
    total_coupon_sales_count = 0
    coupon_total_qty = 0
    total_outstandaing_manual_coupons = 0
    total_outstandaing_digital_coupons =0
    total_debit_amount_count = 0
    total_credit_amount_count = 0
    total_coupon_amount_count = 0
    
    five_gallon_cash_total_net_taxable = 0
    five_gallon_cash_total_vat = 0
    five_gallon_cash_total_subtotal = 0
    five_gallon_cash_total_received = 0
    five_gallon_cash_total_quantity = 0
    other_cash_total_net_taxable = 0
    other_cash_total_vat = 0
    other_cash_total_subtotal = 0
    other_cash_total_received = 0
    other_cash_total_quantity = 0
    cash_sale_recharge_net_payeble = 0
    cash_sale_recharge_vat_total = 0
    cash_sale_recharge_grand_total = 0
    cash_sale_recharge_amount_recieved = 0
    
    five_gallon_credit_total_net_taxable = 0
    five_gallon_credit_total_vat = 0
    five_gallon_credit_total_subtotal = 0
    five_gallon_credit_total_received = 0
    five_gallon_credit_total_quantity = 0
    other_credit_total_net_taxable = 0
    other_credit_total_vat = 0
    other_credit_total_subtotal = 0
    other_credit_total_received = 0
    other_credit_total_quantity = 0
    credit_sale_recharge_net_payeble = 0
    credit_sale_recharge_vat_total = 0
    credit_sale_recharge_grand_total = 0
    credit_sale_recharge_amount_recieved = 0
    foc_total_quantity = 0 
   
    salesman_id =  ""
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    products = ProdutItemMaster.objects.none
    cash_sales = CustomerSupply.objects.none
    expenses_instanses = Expense.objects.none
    credit_sales = CustomerSupply.objects.none
    coupon_sales = CustomerSupply.objects.none
    foc_customers = CustomerSupply.objects.none
    routes_instances = RouteMaster.objects.all()
    customer_coupons = CustomerCoupon.objects.none
    other_cash_sales = CustomerSupply.objects.none
    other_credit_sales = CustomerSupply.objects.none
    van_product_stock = VanProductStock.objects.none
    unique_amounts = CustomerCouponItems.objects.none
    recharge_cash_sales = CustomerCoupon.objects.none
    recharge_credit_sales = CustomerCoupon.objects.none
    five_gallon_cash_sales = CustomerSupply.objects.none
    damage_bottle_instances = VanSaleDamage.objects.none
    five_gallon_credit_sales = CustomerSupply.objects.none
    dialy_collections = InvoiceDailyCollection.objects.none
    pending_bottle_customer_instances = CustomerSupply.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    salesman_id = request.GET.get("salesman_id")
    
    if salesman_id:
        filter_data["salesman_id"] = salesman_id
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        
        if van_route and van_route.van:
            salesman = van_route.van.salesman
            # salesman_id = salesman.pk
            filter_data["route_name"] = route_name
    
    
    
    
        #new customers created
        new_customers_count = Customers.objects.filter(is_guest=False, created_date__date=date,sales_staff_id=salesman).count()
        #emergency supply
        emergency_supply_count = DiffBottlesModel.objects.filter(created_date__date=date, assign_this_to_id=salesman).count()
        #actual visit
        visited_customers_count = CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).distinct().count()
        todays_customers = find_customers(request, str(date), van_route.routes.pk)
        if todays_customers :
            planned_visit_count = len(todays_customers)
        else:
            planned_visit_count = 0
        non_visited_count = planned_visit_count - visited_customers_count
        
        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=salesman)
        van_product_stock = VanProductStock.objects.filter(created_date=date,van=van_instances,product__product_name="5 Gallon")
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        # vanstock_curreptions(van_instances.pk,date)
        
        #### Bottle Count ####
        total_empty_bottles = van_product_stock.aggregate(totalempty_bottle=Sum('empty_can_count'))['totalempty_bottle'] or 0
        total_supplied_bottles =  van_product_stock.aggregate(total_sold=Sum('sold_count'))['total_sold'] or 0
        pending_bottle_count = van_product_stock.aggregate(total_pending=Sum('pending_count'))['total_pending'] or 0
        damage_bottle_count = van_product_stock.aggregate(total_damage=Sum('damage_count'))['total_damage'] or 0
        closing_stock_count = van_product_stock.aggregate(total_closing=Sum('closing_count'))['total_closing'] or 0
        total_count = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        #### coupon sales count ####
        customer_coupons=CustomerCoupon.objects.filter(salesman=salesman,created_date__date=date)
        
        five_gallon_supply = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").values_list('customer_supply__pk', flat=True)
        other_supply = CustomerSupplyItems.objects.exclude(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").values_list('customer_supply__pk', flat=True)
        ### Cash Sales Start ###
        five_gallon_cash_sales = CustomerSupply.objects.filter(pk__in=five_gallon_supply,created_date__date=date,salesman=salesman,amount_recieved__gt=0).exclude(customer__sales_type="CASH COUPON")
        other_cash_sales = CustomerSupply.objects.filter(pk__in=other_supply,created_date__date=date,salesman=salesman,amount_recieved__gt=0).exclude(customer__sales_type="CASH COUPON")
        # Aggregating for five_gallon_cash_sales
        five_gallon_cash_total_net_taxable = five_gallon_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_cash_total_vat = five_gallon_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_cash_total_subtotal = five_gallon_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_cash_total_received = five_gallon_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_cash_total_quantity = five_gallon_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_cash_sales (corrected to use other_cash_sales)
        other_cash_total_net_taxable = other_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_cash_total_vat = other_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_cash_total_subtotal = other_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_cash_total_received = other_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_cash_total_quantity = other_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Combining both cash sales querysets
        cash_sales = five_gallon_cash_sales.union(other_cash_sales)
        cash_total_net_taxable = five_gallon_cash_total_net_taxable + other_cash_total_net_taxable
        cash_total_vat = five_gallon_cash_total_vat + other_cash_total_vat
        cash_total_subtotal = five_gallon_cash_total_subtotal + other_cash_total_subtotal
        cash_total_received = five_gallon_cash_total_received + other_cash_total_received
        cash_total_quantity = five_gallon_cash_total_quantity + other_cash_total_quantity
        
        recharge_cash_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__gt=0)
        cash_sale_recharge_net_payeble = recharge_cash_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        cash_sale_recharge_vat_total = 0
        cash_sale_recharge_grand_total = recharge_cash_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        cash_sale_recharge_amount_recieved = recharge_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        cash_total_net_taxable = cash_total_net_taxable + cash_sale_recharge_net_payeble 
        cash_total_vat = cash_total_vat + cash_sale_recharge_vat_total 
        cash_total_subtotal = cash_total_subtotal + cash_sale_recharge_grand_total 
        cash_total_amount_recieved = cash_total_received + cash_sale_recharge_amount_recieved 
        
        total_cash_sales_count = cash_sales.count() + recharge_cash_sales.count()
        
        cash_sale_recharge_count = recharge_cash_sales.count()
        cash_total_qty = cash_total_quantity + cash_sale_recharge_count
        ### Cash Sales End ###
        
        ### credit sales Start ####
        five_gallon_credit_sales = CustomerSupply.objects.filter(pk__in=five_gallon_supply,created_date__date=date,salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])
        other_credit_sales = CustomerSupply.objects.filter(pk__in=other_supply,created_date__date=date,salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])
        
        # Aggregating for five_gallon_credit_sales
        five_gallon_credit_total_net_taxable = five_gallon_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_credit_total_vat = five_gallon_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_credit_total_subtotal = five_gallon_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_credit_total_received = five_gallon_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_credit_total_quantity = five_gallon_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_credit_sales (corrected to use other_credit_sales)
        other_credit_total_net_taxable = other_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_credit_total_vat = other_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_credit_total_subtotal = other_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_credit_total_received = other_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_credit_total_quantity = other_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Combining both cash sales querysets
        credit_sales = five_gallon_credit_sales.union(other_credit_sales)
        credit_total_net_taxable = five_gallon_credit_total_net_taxable + other_credit_total_net_taxable
        credit_total_vat = five_gallon_credit_total_vat + other_credit_total_vat
        credit_total_subtotal = five_gallon_credit_total_subtotal + other_credit_total_subtotal
        credit_total_received = five_gallon_credit_total_received + other_credit_total_received
        credit_total_quantity = five_gallon_credit_total_quantity + other_credit_total_quantity
        
        recharge_credit_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__lte=0)
        credit_sale_recharge_net_payeble = recharge_credit_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        credit_sale_recharge_vat_total = 0
        credit_sale_recharge_grand_total = recharge_credit_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        credit_sale_recharge_amount_recieved = recharge_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        credit_total_net_taxable = credit_total_net_taxable + credit_sale_recharge_net_payeble 
        credit_total_vat = credit_total_vat + credit_sale_recharge_vat_total 
        credit_total_subtotal = credit_total_subtotal + credit_sale_recharge_grand_total
        credit_total_amount_recieved = credit_total_received + credit_sale_recharge_amount_recieved
        
        total_credit_sales_count = credit_sales.count() + recharge_credit_sales.count()
        credit_sale_recharge_count = recharge_credit_sales.count()
        credit_total_qty = credit_total_quantity + credit_sale_recharge_count
        ### credit sales End ####
        
        # Coupon sales
        coupon_sales = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,customer__sales_type="CASH COUPON")
        for sale in coupon_sales:
            sale.outstanding_manual_coupons = sale.outstanding_manual_coupons()
            sale.outstanding_digital_coupons = sale.outstanding_digital_coupons()
        manual_coupon_total = CustomerSupplyCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(Count('leaf'))['leaf__count']
        digital_coupon_total = CustomerSupplyDigitalCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(total_count=Sum('count'))['total_count'] or 0
        
        total_coupon_sales_count = coupon_sales.count()
        
        total_sales_count = total_cash_sales_count + total_credit_sales_count + total_coupon_sales_count
        coupon_total_qty = coupon_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Calculate total outstanding digital and manual coupons
        total_outstandaing_manual_coupons = sum(sale.outstanding_manual_coupons for sale in coupon_sales)
        total_outstandaing_digital_coupons = sum(sale.outstanding_digital_coupons for sale in coupon_sales)
        
        ### expenses ####
        expenses_instanses = Expense.objects.filter(expense_date=date,van=van_route.van)
        today_expense = expenses_instanses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
        
        ### suspense ###
        suspense_collections = SuspenseCollection.objects.filter(created_date__date=date,salesman=salesman)
        cash_sales_amount = suspense_collections.aggregate(total_cash_sale=Sum('cash_sale_amount'))['total_cash_sale'] or 0
        credit_sales_amount = suspense_collections.aggregate(total_credit_sale=Sum('credit_sale_amount'))['total_credit_sale'] or 0
        
        in_hand_amount = cash_sales_amount + credit_sales_amount
        today_payable = in_hand_amount - today_expense
        suspense_paid_amount = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
        suspense_balance_amount = today_payable - suspense_paid_amount
        
        # collection details
        dialy_collections = CollectionPayment.objects.filter(customer__routes__route_name=route_name,customer__is_guest=False)
        # credit outstanding
        # outstanding_credit_notes = Invoice.objects.filter(invoice_type="credit_invoice",customer__sales_staff=salesman).exclude(created_date__date__gt=date)
        outstanding_credit_notes_total_amount = OutstandingAmount.objects.filter(customer_outstanding__created_date__date__lte=date,customer_outstanding__product_type="amount",customer_outstanding__customer__routes=van_route.routes).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        dialy_colection_upto__yesterday = dialy_collections.filter(created_date__date__lt=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_total_amount = outstanding_credit_notes_total_amount - dialy_colection_upto__yesterday
        outstanding_credit_notes_received_amount = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_balance = outstanding_credit_notes_total_amount - outstanding_credit_notes_received_amount
        outstanding_total_amount_collected = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0

        
        # pending customers
        pending_bottle_customer_instances = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,allocate_bottle_to_pending__gt=0)
        # 5 gallon rate based
        customer_supplies = CustomerSupply.objects.filter(created_date__date=date,salesman_id=salesman_id)
        invoice_nos = customer_supplies.values_list('invoice_no', flat=True)
        invoice_items = InvoiceItems.objects.filter(invoice__invoice_no__in=invoice_nos)

        rate_expr = ExpressionWrapper(
            F('rate') / F('qty'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )

        # Annotate with calculated_rate
        invoice_items = invoice_items.annotate(calculated_rate=Round(rate_expr, 2))

        # ✅ Exclude zero or null values
        invoice_items = invoice_items.exclude(calculated_rate__isnull=True).exclude(calculated_rate=0)

        # Get unique non-zero rates
        unique_amounts = set(invoice_items.values_list('calculated_rate', flat=True))
        # unique_amounts = set(CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon").values_list('customer_supply__customer__rate', flat=True))
        five_gallon_rate_wise_instances = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon")
        total_debit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved__gt=0).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_credit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved=0).exclude(customer_supply__customer__sales_type__in=["FOC","CASH COUPON"]).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_coupon_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__customer__sales_type="CASH COUPON").aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        # cash sales amount collected
        supply_amount_collected = cash_total_amount_recieved
        cash_sales_amount_collected = supply_amount_collected
        
        dialy_collections = dialy_collections.filter(created_date__date=date)
        credit_sales_amount_collected = dialy_collections.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        total_sales_amount_collected = cash_sales_amount_collected + credit_sales_amount_collected
        
        cheque_collection = CollectionPayment.objects.filter(payment_method="CHEQUE",created_date__date=date,salesman=salesman)
        no_of_collected_cheque = cheque_collection.count()
        collected_cheque_amount = cheque_collection.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        
        balance_in_hand = total_sales_amount_collected - collected_cheque_amount - today_expense
        net_payble = total_sales_amount_collected - today_expense
        
        foc_customers = (
            CustomerSupply.objects.filter(created_date__date=date, customer__sales_type='FOC', salesman=salesman) |
            CustomerSupply.objects.filter(created_date__date=date, salesman=salesman, allocate_bottle_to_free__gt=0)
        )
        for foc in foc_customers:
            if foc.allocate_bottle_to_free != 0:
                foc_total_quantity =foc_total_quantity + foc.allocate_bottle_to_free
            else:
                foc_total_quantity =foc_total_quantity + foc.get_total_supply_qty()
                
        damage_bottle_instances = VanSaleDamage.objects.filter(van=van_instances,created_date__date=date)
        
    context = {
        'van_route': van_route,
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        # 'salesman_instances':salesman_instances,
        # visit statistics
        'routes_instances': routes_instances,
        'non_visited_count': non_visited_count,
        'planned_visit_count': planned_visit_count,
        'new_customers_count': new_customers_count,
        'emergency_supply_count': emergency_supply_count,
        'visited_customers_count': visited_customers_count,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        # pending customers
        'pending_bottle_customer_instances': pending_bottle_customer_instances,
        # Bottle Count
        'total_count': total_count,
        'total_empty_bottles': total_empty_bottles,
        'closing_stock_count': closing_stock_count,
        'damage_bottle_count': damage_bottle_count,
        'pending_bottle_count': pending_bottle_count,
        'total_supplied_bottles':total_supplied_bottles,
        #coupon book sale
        'customer_coupons':customer_coupons,
        # five gallon cash sales
        'five_gallon_cash_sales': five_gallon_cash_sales or [],
        'five_gallon_cash_total_vat': five_gallon_cash_total_vat,
        'five_gallon_cash_total_subtotal': five_gallon_cash_total_subtotal,
        'five_gallon_cash_total_received': five_gallon_cash_total_received,
        'five_gallon_cash_total_quantity': five_gallon_cash_total_quantity,
        'five_gallon_cash_total_net_taxable': five_gallon_cash_total_net_taxable,
        # other cash sales
        'other_cash_sales': other_cash_sales,
        'other_cash_total_vat': other_cash_total_vat,
        'other_cash_total_subtotal': other_cash_total_subtotal,
        'other_cash_total_received': other_cash_total_received,
        'other_cash_total_quantity': other_cash_total_quantity,
        'other_cash_total_net_taxable': other_cash_total_net_taxable,
        #cash sales
        'recharge_cash_sales': recharge_cash_sales,
        'cash_sale_recharge_vat_total': cash_sale_recharge_vat_total,
        'cash_sale_recharge_grand_total': cash_sale_recharge_grand_total,
        'cash_sale_recharge_net_payeble': cash_sale_recharge_net_payeble,
        'cash_sale_recharge_amount_recieved': cash_sale_recharge_amount_recieved,
        
        'cash_total_vat':cash_total_vat,
        'cash_total_qty': cash_total_qty,
        'cash_total_subtotal': cash_total_subtotal,
        'cash_total_net_taxable':cash_total_net_taxable,
        'cash_total_amount_recieved': cash_total_amount_recieved,
        # five gallon credit sales
        'five_gallon_credit_sales': five_gallon_credit_sales,
        'five_gallon_credit_total_vat': five_gallon_credit_total_vat,
        'five_gallon_credit_total_subtotal': five_gallon_credit_total_subtotal,
        'five_gallon_credit_total_received': five_gallon_credit_total_received,
        'five_gallon_credit_total_quantity': five_gallon_credit_total_quantity,
        'five_gallon_credit_total_net_taxable': five_gallon_credit_total_net_taxable,
        # credit sales
        'credit_sales': credit_sales,
        'recharge_credit_sales': recharge_credit_sales,
        'credit_sale_recharge_vat_total': credit_sale_recharge_vat_total,
        'credit_sale_recharge_net_payeble': credit_sale_recharge_net_payeble,
        'credit_sale_recharge_grand_total': credit_sale_recharge_grand_total,
        'credit_sale_recharge_amount_recieved': credit_sale_recharge_amount_recieved,
        
        'credit_total_vat':credit_total_vat,
        'credit_total_qty': credit_total_qty,
        'credit_total_subtotal':credit_total_subtotal,
        'credit_total_net_taxable':credit_total_net_taxable,
        'credit_total_amount_recieved': credit_total_amount_recieved,
        # coupon sales
        'coupon_sales': coupon_sales,
        'coupon_total_qty':coupon_total_qty,
        'manual_coupon_total':manual_coupon_total,
        'digital_coupon_total':digital_coupon_total,
        'total_coupon_sales_count': total_coupon_sales_count,
        'total_outstandaing_manual_coupons':total_outstandaing_manual_coupons,
        'total_outstandaing_digital_coupons':total_outstandaing_digital_coupons,
        
        # expenses
        'expenses_instanses': expenses_instanses,
        # suspense
        'today_expense': today_expense,
        'today_payable': today_payable,
        'in_hand_amount': in_hand_amount,
        'suspense_paid_amount': suspense_paid_amount,
        'suspense_balance_amount': suspense_balance_amount,
        'outstanding_credit_notes_balance': outstanding_credit_notes_balance,
        'outstanding_total_amount_collected':outstanding_total_amount_collected,
        'outstanding_credit_notes_total_amount' : outstanding_credit_notes_total_amount,
        'outstanding_credit_notes_received_amount' : outstanding_credit_notes_received_amount,
        # 5 gallon rate based
        'five_gallon_rates': unique_amounts,
        'total_debit_amount_count': total_debit_amount_count,
        'total_credit_amount_count': total_credit_amount_count,
        'total_coupon_amount_count': total_coupon_amount_count,
        # dialy collections
        'dialy_collections': dialy_collections,
        # sales amount collected
        'total_sales_count': total_sales_count,
        'no_of_collected_cheque': no_of_collected_cheque,
        'total_cash_sales_count': total_cash_sales_count,
        'collected_cheque_amount': collected_cheque_amount,
        'total_credit_sales_count': total_credit_sales_count,
        'cash_sales_amount_collected': cash_sales_amount_collected,
        'total_sales_amount_collected': total_sales_amount_collected,
        'credit_sales_amount_collected': credit_sales_amount_collected,
        
        'damage_bottle_instances': damage_bottle_instances,
        
        'net_payble': net_payble,
        'balance_in_hand': balance_in_hand,
        
        'filter_data': filter_data,
        # FOC customer
        'foc_customers':foc_customers,
        'foc_total_quantity':foc_total_quantity,
        'selected_salesman': salesman,
        
    }
    
    return render(request, 'sales_management/dsr_summary1.html', context)


def get_salesmen_by_route(request):
    route_name = request.GET.get('route_name')
    if route_name:
        van_routes = Van_Routes.objects.filter(routes__route_name=route_name)
        salesmen = [
            {"id": route.van.salesman.pk, "name": route.van.salesman.username}
            for route in van_routes if route.van and route.van.salesman
        ]
        return JsonResponse({"salesmen": salesmen})
    return JsonResponse({"salesmen": []})


def print_dsr_summary(request):
    
    filter_data = {}
    data_filter = False
    total_count = 0
    non_visited_count = 0
    new_customers_count = 0
    planned_visit_count = 0
    total_empty_bottles = 0
    closing_stock_count = 0
    damage_bottle_count = 0
    pending_bottle_count = 0
    total_supplied_bottles = 0
    emergency_supply_count = 0
    visited_customers_count = 0
    cash_total_net_taxable = 0
    cash_total_vat = 0
    cash_total_subtotal = 0
    cash_total_amount_recieved = 0
    cash_sale_recharge_count=0
    cash_total_quantity = 0
    cash_total_qty = 0
    credit_total_net_taxable = 0
    credit_total_vat = 0
    credit_total_subtotal = 0
    credit_total_amount_recieved = 0
    credit_total_qty = 0
    credit_total_quantity = 0
    credit_sale_recharge_count = 0
    in_hand_amount = 0
    today_expense = 0
    today_payable = 0
    suspense_paid_amount = 0
    suspense_balance_amount = 0
    todays_opening_outstanding_amount = 0
    credit_outstanding_collected_amount = 0
    outstanding_closing_balance = 0
    outstanding_total_amount_collected = 0
    cash_sales_amount_collected = 0
    credit_sales_amount_collected = 0
    total_sales_amount_collected = 0
    total_cash_sales_count = 0
    total_credit_sales_count = 0
    total_sales_count = 0
    no_of_collected_cheque = 0
    collected_cheque_amount = 0
    balance_in_hand = 0
    net_payble = 0
    stock_report_total = 0
    manual_coupon_total = 0
    digital_coupon_total = 0
    total_coupon_sales_count = 0
    coupon_total_qty = 0
    total_outstandaing_manual_coupons = 0
    total_outstandaing_digital_coupons =0
    total_debit_amount_count = 0
    total_credit_amount_count = 0
    total_coupon_amount_count = 0
    
    five_gallon_cash_total_net_taxable = 0
    five_gallon_cash_total_vat = 0
    five_gallon_cash_total_subtotal = 0
    five_gallon_cash_total_received = 0
    five_gallon_cash_total_quantity = 0
    other_cash_total_net_taxable = 0
    other_cash_total_vat = 0
    other_cash_total_subtotal = 0
    other_cash_total_received = 0
    other_cash_total_quantity = 0
    cash_sale_recharge_net_payeble = 0
    cash_sale_recharge_vat_total = 0
    cash_sale_recharge_grand_total = 0
    cash_sale_recharge_amount_recieved = 0
    
    five_gallon_credit_total_net_taxable = 0
    five_gallon_credit_total_vat = 0
    five_gallon_credit_total_subtotal = 0
    five_gallon_credit_total_received = 0
    five_gallon_credit_total_quantity = 0
    other_credit_total_net_taxable = 0
    other_credit_total_vat = 0
    other_credit_total_subtotal = 0
    other_credit_total_received = 0
    other_credit_total_quantity = 0
    credit_sale_recharge_net_payeble = 0
    credit_sale_recharge_vat_total = 0
    credit_sale_recharge_grand_total = 0
    credit_sale_recharge_amount_recieved = 0
    foc_total_quantity = 0
    
    customer_custody_instances = CustodyCustom.objects.none
    collected_amount_from_custody_issue = 0
    customer_custody_items_fgallon_instances = CustodyCustomItems.objects.none
    customer_custody_items_fgallon_count = 0
    customer_custody_items_nonfgallon_instances = CustodyCustomItems.objects.none
    customer_custody_items_nonfgallon_count = 0
   
    salesman_id =  None
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    products = ProdutItemMaster.objects.none
    cash_sales = CustomerSupply.objects.none
    expenses_instanses = Expense.objects.none
    credit_sales = CustomerSupply.objects.none
    coupon_sales = CustomerSupply.objects.none
    foc_customers = CustomerSupply.objects.none
    routes_instances = RouteMaster.objects.all()
    customer_coupons = CustomerCoupon.objects.none
    other_cash_sales = CustomerSupply.objects.none
    other_credit_sales = CustomerSupply.objects.none
    van_product_stock = VanProductStock.objects.none
    unique_amounts = CustomerCouponItems.objects.none
    recharge_cash_sales = CustomerCoupon.objects.none
    recharge_credit_sales = CustomerCoupon.objects.none
    five_gallon_cash_sales = CustomerSupply.objects.none
    damage_bottle_instances = VanSaleDamage.objects.none
    five_gallon_credit_sales = CustomerSupply.objects.none
    dialy_collections = InvoiceDailyCollection.objects.none
    pending_bottle_customer_instances = CustomerSupply.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    raw_salesman_id = request.GET.get('salesman_id')

# normalize
    salesman_id = raw_salesman_id if raw_salesman_id not in ("", None, "None") else None
    salesman = None
    if salesman_id: 
        salesman = CustomUser.objects.filter(
        pk=salesman_id,
        user_type="Salesman"
        ).first()

        salesman_id = salesman.pk if salesman else None
         

    print("salesmanId:",salesman_id)
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        
    yesterday_date = date - datetime.timedelta(days=1)
    salesman_list = CustomUser.objects.filter(user_type="Salesman")
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        
        filter_data['route_name'] = route_name
        filter_data['salesman_id'] = str(salesman_id)
        #new customers created
        new_customers_qs = Customers.objects.filter(
            is_guest=False,
            created_date__date=date,
            routes__route_name=route_name
        )

        if salesman_id:
            new_customers_qs = new_customers_qs.filter(
                sales_staff_id=salesman_id
            )

        new_customers_count = new_customers_qs.count()
        #emergency supply
        emergency_supply_qs = DiffBottlesModel.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            emergency_supply_qs = emergency_supply_qs.filter(
                assign_this_to_id=salesman_id
            )

        emergency_supply_count = emergency_supply_qs.count()
        #actual visit
        visited_customers_qs = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            visited_customers_qs = visited_customers_qs.filter(
                salesman_id=salesman_id
            )

        visited_customers_count = visited_customers_qs.distinct().count()
        todays_customers = find_customers(request, str(date), van_route.routes.pk)
        if todays_customers :
            planned_visit_count = len(todays_customers)
        else:
            planned_visit_count = 0
        non_visited_count = planned_visit_count - visited_customers_count
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        van_salesman = van_route.van.salesman
        

        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=van_salesman)
        van_product_stock = VanProductStock.objects.filter(created_date=date,van=van_instances,product__product_name="5 Gallon")
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        # vanstock_curreptions(van_instances.pk,date)
        
        #### Bottle Count ####
        total_empty_bottles = van_product_stock.aggregate(totalempty_bottle=Sum('empty_can_count'))['totalempty_bottle'] or 0
        total_supplied_bottles =  van_product_stock.aggregate(total_sold=Sum('sold_count'))['total_sold'] or 0
        pending_bottle_count = van_product_stock.aggregate(total_pending=Sum('pending_count'))['total_pending'] or 0
        damage_bottle_count = van_product_stock.aggregate(total_damage=Sum('damage_count'))['total_damage'] or 0
        closing_stock_count = van_product_stock.aggregate(total_closing=Sum('closing_count'))['total_closing'] or 0
        total_count = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        #### coupon sales count ####
        customer_coupons=CustomerCoupon.objects.filter(created_date__date=date,customer__routes__route_name=route_name)
        if salesman_id:
            customer_coupons = customer_coupons.filter(
                salesman_id=salesman_id
            )
        
        five_gallon_supply = CustomerSupplyItems.objects.filter(
            customer_supply__supply_date__date=date,
            product__product_name="5 Gallon",
            customer_supply__customer__routes__route_name=route_name
        )

      

        if salesman_id:
            five_gallon_supply = five_gallon_supply.filter(
                customer_supply__salesman_id=salesman_id
            )

        five_gallon_supply = five_gallon_supply.values_list(
            'customer_supply__pk', flat=True
        )

        other_supply = CustomerSupplyItems.objects.filter(
        customer_supply__customer__routes__route_name=route_name
        ).exclude(
            product__product_name="5 Gallon"
        )

        if salesman_id:
            other_supply = other_supply.filter(
                customer_supply__salesman_id=salesman_id
            )

        other_supply = other_supply.filter(
            customer_supply__supply_date__date=date
        ).values_list('customer_supply__pk', flat=True)
        ### Cash Sales Start ###
        five_gallon_cash_sales = CustomerSupply.objects.filter(
            pk__in=five_gallon_supply,
            supply_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__gt=0
        ).exclude(customer__sales_type="CASH COUPON")

        if salesman_id:
            five_gallon_cash_sales = five_gallon_cash_sales.filter(
                salesman_id=salesman_id
            )

        other_cash_sales = CustomerSupply.objects.filter(
            pk__in=other_supply,
            supply_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__gt=0
        ).exclude(customer__sales_type="CASH COUPON")

        if salesman_id:
            other_cash_sales = other_cash_sales.filter(
                salesman_id=salesman_id
            )
        # Aggregating for five_gallon_cash_sales
        five_gallon_cash_total_net_taxable = five_gallon_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_cash_total_vat = five_gallon_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_cash_total_subtotal = five_gallon_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_cash_total_received = five_gallon_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_cash_total_quantity = five_gallon_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_cash_sales (corrected to use other_cash_sales)
        other_cash_total_net_taxable = other_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_cash_total_vat = other_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_cash_total_subtotal = other_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_cash_total_received = other_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_cash_total_quantity = other_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Combining both cash sales querysets
        cash_sales = five_gallon_cash_sales.union(other_cash_sales)
        cash_total_net_taxable = five_gallon_cash_total_net_taxable + other_cash_total_net_taxable
        cash_total_vat = five_gallon_cash_total_vat + other_cash_total_vat
        cash_total_subtotal = five_gallon_cash_total_subtotal + other_cash_total_subtotal
        cash_total_received = five_gallon_cash_total_received + other_cash_total_received
        cash_total_quantity = five_gallon_cash_total_quantity + other_cash_total_quantity
        
        recharge_cash_sales = CustomerCoupon.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__gt=0
        )

       
        if salesman_id:
            recharge_cash_sales = recharge_cash_sales.filter(
                salesman_id=salesman_id
            )
        cash_sale_recharge_net_payeble = recharge_cash_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        cash_sale_recharge_vat_total = 0
        cash_sale_recharge_grand_total = recharge_cash_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        cash_sale_recharge_amount_recieved = recharge_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        cash_total_net_taxable = cash_total_net_taxable + cash_sale_recharge_net_payeble 
        cash_total_vat = cash_total_vat + cash_sale_recharge_vat_total 
        cash_total_subtotal = cash_total_subtotal + cash_sale_recharge_grand_total 
        cash_total_amount_recieved = cash_total_received + cash_sale_recharge_amount_recieved 
        
        total_cash_sales_count = cash_sales.count() + recharge_cash_sales.count()
        
        cash_sale_recharge_count = recharge_cash_sales.count()
        cash_total_qty = cash_total_quantity + cash_sale_recharge_count
        ### Cash Sales End ###
        
        ### credit sales Start ####
        five_gallon_credit_sales = CustomerSupply.objects.filter(
            pk__in=five_gallon_supply,
            supply_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__lte=0
            ).exclude(
                customer__sales_type__in=["FOC", "CASH COUPON"]
        )

        if salesman_id:
            five_gallon_credit_sales = five_gallon_credit_sales.filter(
                salesman_id=salesman_id
            )
        other_credit_sales = CustomerSupply.objects.filter(
        pk__in=other_supply,
        supply_date__date=date,
        customer__routes__route_name=route_name,
        amount_recieved__lte=0
        ).exclude(
            customer__sales_type__in=["FOC", "CASH COUPON"]
        )

        if salesman_id:
            other_credit_sales = other_credit_sales.filter(
                salesman_id=salesman_id
            )
        
        # Aggregating for five_gallon_credit_sales
        five_gallon_credit_total_net_taxable = five_gallon_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_credit_total_vat = five_gallon_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_credit_total_subtotal = five_gallon_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_credit_total_received = five_gallon_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_credit_total_quantity = five_gallon_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_credit_sales (corrected to use other_credit_sales)
        other_credit_total_net_taxable = other_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_credit_total_vat = other_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_credit_total_subtotal = other_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_credit_total_received = other_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_credit_total_quantity = other_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Combining both cash sales querysets
        credit_sales = five_gallon_credit_sales.union(other_credit_sales)
        credit_total_net_taxable = five_gallon_credit_total_net_taxable + other_credit_total_net_taxable
        credit_total_vat = five_gallon_credit_total_vat + other_credit_total_vat
        credit_total_subtotal = five_gallon_credit_total_subtotal + other_credit_total_subtotal
        credit_total_received = five_gallon_credit_total_received + other_credit_total_received
        credit_total_quantity = five_gallon_credit_total_quantity + other_credit_total_quantity
        
        recharge_credit_sales = CustomerCoupon.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name,
            amount_recieved__lte=0
        )

        if salesman_id:
            recharge_credit_sales = recharge_credit_sales.filter(
                salesman_id=salesman_id
            )
        credit_sale_recharge_net_payeble = recharge_credit_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        credit_sale_recharge_vat_total = 0
        credit_sale_recharge_grand_total = recharge_credit_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        credit_sale_recharge_amount_recieved = recharge_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        credit_total_net_taxable = credit_total_net_taxable + credit_sale_recharge_net_payeble 
        credit_total_vat = credit_total_vat + credit_sale_recharge_vat_total 
        credit_total_subtotal = credit_total_subtotal + credit_sale_recharge_grand_total
        credit_total_amount_recieved = credit_total_received + credit_sale_recharge_amount_recieved
        
        total_credit_sales_count = credit_sales.count() + recharge_credit_sales.count()
        credit_sale_recharge_count = recharge_credit_sales.count()
        credit_total_qty = credit_total_quantity + credit_sale_recharge_count
        ### credit sales End ####
        
        # Coupon sales
        coupon_sales = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name,
            customer__sales_type="CASH COUPON"
        )

        if salesman_id:
            coupon_sales = coupon_sales.filter(
                salesman_id=salesman_id
            )
        for sale in coupon_sales:
            sale.outstanding_manual_coupons = sale.outstanding_manual_coupons()
            sale.outstanding_digital_coupons = sale.outstanding_digital_coupons()
        manual_coupon_total = CustomerSupplyCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(Count('leaf'))['leaf__count']
        digital_coupon_total = CustomerSupplyDigitalCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(total_count=Sum('count'))['total_count'] or 0
        
        total_coupon_sales_count = coupon_sales.count()
        
        total_sales_count = total_cash_sales_count + total_credit_sales_count + total_coupon_sales_count
        coupon_total_qty = coupon_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Calculate total outstanding digital and manual coupons
        total_outstandaing_manual_coupons = sum(sale.outstanding_manual_coupons for sale in coupon_sales)
        total_outstandaing_digital_coupons = sum(sale.outstanding_digital_coupons for sale in coupon_sales)
        
        ### expenses ####
        expenses_instanses = Expense.objects.filter(expense_date=date,route__route_name=route_name)
        today_expense = expenses_instanses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
        
        ### suspense ###
        suspense_collections = SuspenseCollection.objects.filter(
            created_date__date=date,
            route__route_name=route_name
        )
        if salesman_id:
            suspense_collections = suspense_collections.filter(
                salesman_id=salesman_id
            )
        cash_sales_amount = suspense_collections.aggregate(total_cash_sale=Sum('cash_sale_amount'))['total_cash_sale'] or 0
        credit_sales_amount = suspense_collections.aggregate(total_credit_sale=Sum('credit_sale_amount'))['total_credit_sale'] or 0
        
        in_hand_amount = cash_sales_amount + credit_sales_amount
        today_payable = in_hand_amount - today_expense
        suspense_paid_amount = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
        suspense_balance_amount = today_payable - suspense_paid_amount
        
        # collection details
        dialy_collections = CollectionPayment.objects.filter(
            customer__is_guest=False,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            dialy_collections = dialy_collections.filter(
                salesman_id=salesman_id
            )
        # credit outstanding
        qs = OutstandingAmount.objects.filter(
            customer_outstanding__product_type="amount",
            customer_outstanding__created_date__date__lte=yesterday_date,
            customer_outstanding__customer__is_guest=False,
            customer_outstanding__customer__routes__route_name=route_name
        )

        if salesman_id:
            qs = qs.filter(
                customer_outstanding__created_by=salesman_id
            )

        outstanding_amount_upto_yesterday = (
            qs.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        )
        dialy_colection_upto_yesterday = dialy_collections.filter(created_date__date__lte=yesterday_date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        todays_opening_outstanding_amount = outstanding_amount_upto_yesterday - dialy_colection_upto_yesterday
        credit_outstanding_collected_amount = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_closing_balance = todays_opening_outstanding_amount + credit_total_subtotal - credit_outstanding_collected_amount
        
        print(outstanding_amount_upto_yesterday)
        print(dialy_colection_upto_yesterday)
        print(credit_outstanding_collected_amount)
        # pending customers
        pending_bottle_customer_instances = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name,
            allocate_bottle_to_pending__gt=0,
        )

       

        if salesman_id:
            pending_bottle_customer_instances = pending_bottle_customer_instances.filter(
                salesman_id=salesman_id
            )
        # 5 gallon rate based
        customer_supplies = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name
        )
        if salesman_id:
            customer_supplies = customer_supplies.filter(
                salesman_id=salesman_id
            )
        invoice_nos = customer_supplies.values_list('invoice_no', flat=True)
        invoice_items = InvoiceItems.objects.filter(invoice__invoice_no__in=invoice_nos)

        rate_expr = ExpressionWrapper(
            F('rate') / F('qty'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )

        # Annotate with calculated_rate
        invoice_items = invoice_items.annotate(calculated_rate=Round(rate_expr, 2))

        # ✅ Exclude zero or null values
        invoice_items = invoice_items.exclude(calculated_rate__isnull=True).exclude(calculated_rate=0)

        five_gallon_rates = CustomerSupplyItems.objects.filter(
            customer_supply__supply_date__date=date,
            customer_supply__customer__routes__route_name=route_name,
            product__product_name="5 Gallon",
            quantity__gt=0,
            amount__gt=0
        )

        # ✅ Apply salesman filter ONLY if provided
        if salesman:
            five_gallon_rates = five_gallon_rates.filter(
                customer_supply__salesman_id=salesman
            )

        five_gallon_rates = (
            five_gallon_rates
            .annotate(
                calculated_rate=ExpressionWrapper(
                    F("amount") / F("quantity"),
                    output_field=DecimalField(max_digits=10, decimal_places=2)
                )
            )
            .values_list("calculated_rate", flat=True)
            .exclude(calculated_rate=0)
            .distinct()
            .order_by("calculated_rate")
        )

        # Get unique non-zero rates
        unique_amounts = five_gallon_rates
        # unique_amounts = set(CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon").values_list('customer_supply__customer__rate', flat=True))
        five_gallon_rate_wise_instances = CustomerSupplyItems.objects.filter(
            customer_supply__supply_date__date=date,
            customer_supply__customer__routes__route_name=route_name,
            product__product_name="5 Gallon"
        )

        if salesman_id:
            five_gallon_rate_wise_instances = five_gallon_rate_wise_instances.filter(
                customer_supply__salesman_id=salesman_id
            )
        total_debit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved__gt=0).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_credit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved=0).exclude(customer_supply__customer__sales_type__in=["FOC","CASH COUPON"]).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_coupon_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__customer__sales_type="CASH COUPON").aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        # Custody Issue 
        customer_custody_instances = CustodyCustom.objects.filter(
            created_date__date=date,
            customer__routes__route_name=route_name
        )

        if salesman_id:
            customer_custody_instances = customer_custody_instances.filter(
                created_by=salesman_id
            )
        collected_amount_from_custody_issue = customer_custody_instances.aggregate(total_amount=Sum('amount_collected'))['total_amount'] or 0
        
        customer_custody_items_fgallon_instances = CustodyCustomItems.objects.filter(custody_custom__pk__in=customer_custody_instances.values_list('pk'),product__product_name="5 Gallon")
        customer_custody_items_fgallon_count = customer_custody_items_fgallon_instances.aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
        # custody other products
        customer_custody_items_nonfgallon_instances = CustodyCustomItems.objects.filter(custody_custom__pk__in=customer_custody_instances.values_list('pk')).exclude(product__product_name="5 Gallon")
        customer_custody_items_nonfgallon_count = customer_custody_items_nonfgallon_instances.aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
        
        # cash sales amount collected
        supply_amount_collected = cash_total_amount_recieved
        cash_sales_amount_collected = supply_amount_collected
        
        dialy_collections = dialy_collections.filter(created_date__date=date)
        credit_sales_amount_collected = dialy_collections.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        total_sales_amount_collected = cash_sales_amount_collected + credit_sales_amount_collected
        
        cheque_collection = CollectionPayment.objects.filter(
            payment_method="CHEQUE",
            created_date__date=date,
            customer__routes__route_name=route_name
        )

       
        if salesman_id:
            cheque_collection = cheque_collection.filter(
                salesman_id=salesman_id
            )
        no_of_collected_cheque = cheque_collection.count()
        collected_cheque_amount = cheque_collection.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        
        balance_in_hand = total_sales_amount_collected + collected_amount_from_custody_issue - collected_cheque_amount - today_expense
        net_payble = total_sales_amount_collected + collected_amount_from_custody_issue - today_expense
        foc_qs_1 = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__sales_type='FOC',
            customer__routes__route_name=route_name
        )

        foc_qs_2 = CustomerSupply.objects.filter(
            supply_date__date=date,
            customer__routes__route_name=route_name,
            allocate_bottle_to_free__gt=0,
        )
        if salesman_id:
            foc_qs_1 = foc_qs_1.filter(
                salesman_id=salesman_id
            )
            foc_qs_2 = foc_qs_2.filter(
                salesman_id=salesman_id
            )
        foc_customers = (
             foc_qs_1 | foc_qs_2
            
        )
        for foc in foc_customers:
            if foc.allocate_bottle_to_free != 0:
                foc_total_quantity =foc_total_quantity + foc.allocate_bottle_to_free
            else:
                foc_total_quantity =foc_total_quantity + foc.get_total_supply_qty()
                
        damage_bottle_instances = VanSaleDamage.objects.filter(van=van_instances,created_date__date=date)
        
    context = {
        'van_route': van_route,
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        # visit statistics
        'routes_instances': routes_instances,
        'non_visited_count': non_visited_count,
        'planned_visit_count': planned_visit_count,
        'new_customers_count': new_customers_count,
        'emergency_supply_count': emergency_supply_count,
        'visited_customers_count': visited_customers_count,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        # pending customers
        'pending_bottle_customer_instances': pending_bottle_customer_instances,
        # Bottle Count
        'total_count': total_count,
        'total_empty_bottles': total_empty_bottles,
        'closing_stock_count': closing_stock_count,
        'damage_bottle_count': damage_bottle_count,
        'pending_bottle_count': pending_bottle_count,
        'total_supplied_bottles':total_supplied_bottles,
        #coupon book sale
        'customer_coupons':customer_coupons,
        # five gallon cash sales
        'five_gallon_cash_sales': five_gallon_cash_sales or [],
        'five_gallon_cash_total_vat': five_gallon_cash_total_vat,
        'five_gallon_cash_total_subtotal': five_gallon_cash_total_subtotal,
        'five_gallon_cash_total_received': five_gallon_cash_total_received,
        'five_gallon_cash_total_quantity': five_gallon_cash_total_quantity,
        'five_gallon_cash_total_net_taxable': five_gallon_cash_total_net_taxable,
        # other cash sales
        'other_cash_sales': other_cash_sales,
        'other_cash_total_vat': other_cash_total_vat,
        'other_cash_total_subtotal': other_cash_total_subtotal,
        'other_cash_total_received': other_cash_total_received,
        'other_cash_total_quantity': other_cash_total_quantity,
        'other_cash_total_net_taxable': other_cash_total_net_taxable,
        #cash sales
        'recharge_cash_sales': recharge_cash_sales,
        'cash_sale_recharge_vat_total': cash_sale_recharge_vat_total,
        'cash_sale_recharge_grand_total': cash_sale_recharge_grand_total,
        'cash_sale_recharge_net_payeble': cash_sale_recharge_net_payeble,
        'cash_sale_recharge_amount_recieved': cash_sale_recharge_amount_recieved,
        
        'cash_total_vat':cash_total_vat,
        'cash_total_qty': cash_total_qty,
        'cash_total_subtotal': cash_total_subtotal,
        'cash_total_net_taxable':cash_total_net_taxable,
        'cash_total_amount_recieved': cash_total_amount_recieved,
        # five gallon credit sales
        'five_gallon_credit_sales': five_gallon_credit_sales,
        'five_gallon_credit_total_vat': five_gallon_credit_total_vat,
        'five_gallon_credit_total_subtotal': five_gallon_credit_total_subtotal,
        'five_gallon_credit_total_received': five_gallon_credit_total_received,
        'five_gallon_credit_total_quantity': five_gallon_credit_total_quantity,
        'five_gallon_credit_total_net_taxable': five_gallon_credit_total_net_taxable,
        # credit sales
        'credit_sales': credit_sales,
        'recharge_credit_sales': recharge_credit_sales,
        'credit_sale_recharge_vat_total': credit_sale_recharge_vat_total,
        'credit_sale_recharge_net_payeble': credit_sale_recharge_net_payeble,
        'credit_sale_recharge_grand_total': credit_sale_recharge_grand_total,
        'credit_sale_recharge_amount_recieved': credit_sale_recharge_amount_recieved,
        
        'credit_total_vat':credit_total_vat,
        'credit_total_qty': credit_total_qty,
        'credit_total_subtotal':credit_total_subtotal,
        'credit_total_net_taxable':credit_total_net_taxable,
        'credit_total_amount_recieved': credit_total_amount_recieved,
        # coupon sales
        'coupon_sales': coupon_sales,
        'coupon_total_qty':coupon_total_qty,
        'manual_coupon_total':manual_coupon_total,
        'digital_coupon_total':digital_coupon_total,
        'total_coupon_sales_count': total_coupon_sales_count,
        'total_outstandaing_manual_coupons':total_outstandaing_manual_coupons,
        'total_outstandaing_digital_coupons':total_outstandaing_digital_coupons,
        
        # expenses
        'expenses_instanses': expenses_instanses,
        # suspense
        'today_expense': today_expense,
        'today_payable': today_payable,
        'in_hand_amount': in_hand_amount,
        'suspense_paid_amount': suspense_paid_amount,
        'suspense_balance_amount': suspense_balance_amount,
        'outstanding_closing_balance': outstanding_closing_balance,
        'outstanding_total_amount_collected':outstanding_total_amount_collected,
        'todays_opening_outstanding_amount' : todays_opening_outstanding_amount,
        'credit_outstanding_collected_amount' : credit_outstanding_collected_amount,
        # 5 gallon rate based
        'five_gallon_rates': unique_amounts,
        'total_debit_amount_count': total_debit_amount_count,
        'total_credit_amount_count': total_credit_amount_count,
        'total_coupon_amount_count': total_coupon_amount_count,
        # dialy collections
        'dialy_collections': dialy_collections,
        # sales amount collected
        'total_sales_count': total_sales_count,
        'no_of_collected_cheque': no_of_collected_cheque,
        'total_cash_sales_count': total_cash_sales_count,
        'collected_cheque_amount': collected_cheque_amount,
        'total_credit_sales_count': total_credit_sales_count,
        'cash_sales_amount_collected': cash_sales_amount_collected,
        'total_sales_amount_collected': total_sales_amount_collected,
        'credit_sales_amount_collected': credit_sales_amount_collected,
        # custody details
        'customer_custody_instances': customer_custody_instances,
        'collected_amount_from_custody_issue': collected_amount_from_custody_issue,
        'customer_custody_items_fgallon_instances': customer_custody_items_fgallon_instances,
        'customer_custody_items_fgallon_count': customer_custody_items_fgallon_count,
        'customer_custody_items_nonfgallon_instances': customer_custody_items_nonfgallon_instances,
        'customer_custody_items_nonfgallon_count': customer_custody_items_nonfgallon_count,
        
        'damage_bottle_instances': damage_bottle_instances,
        
        'net_payble': net_payble,
        'balance_in_hand': balance_in_hand,
        
        'filter_data': filter_data,
        # FOC customer
        'foc_customers':foc_customers,
        'foc_total_quantity':foc_total_quantity,
        'salesman_list':salesman_list,
        'selected_salesman': salesman,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),

        
    }
    print("selected_salesman:",salesman)
    
    return render(request, 'sales_management/dsr_summary_print.html', context)

# import openpyxl
# from django.http import HttpResponse
# from django.template.loader import render_to_string
# from datetime import datetime
# import io

# def export_dsr_summary_data_to_excel(request):
#     filter_data = {}
#     new_customers_count = 0
#     emergency_supply_count = 0
#     visited_customers_count = 0
#     non_visited_count = 0
#     planned_visit_count = 0
#     empty_bottles_collected = 0
#     empty_bottle_pending = 0
#     coupons_collected = 0
#     total_supplied_quantity = 0
#     total_collected_amount = 0
#     total_pending_amount = 0
#     mode_of_supply = 0
#     total_empty_bottles = 0
#     total_supplied_bottles = 0
#     closing_stock_count = 0
#     damage_bottle_count = 0
#     pending_bottle_count = 0
#     total_count = 0
#     cash_total_net_taxable = 0
#     cash_total_vat = 0
#     cash_total_amout_total = 0
#     credit_total_net_taxable = 0
#     credit_total_vat = 0
#     credit_total_amout_total = 0
#     suspense_opening_amount = 0
#     suspense_paid_amount = 0
#     suspense_balance_amount = 0
    
#     van_instances = Van.objects.none
#     van_route = Van_Routes.objects.none
#     cash_invoices = Invoice.objects.none
#     credit_invoices = Invoice.objects.none
#     products = ProdutItemMaster.objects.none
#     expenses_instanses = Expense.objects.none
#     routes_instances = RouteMaster.objects.all()
#     van_product_stock = VanProductStock.objects.none
#     customer_coupon_items = CustomerCouponItems.objects.none
    
#     date = request.GET.get('date')
#     route_name = request.GET.get('route_name')
    
#     if date:
#         date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
#     else:
#         date = datetime.datetime.today().date()
#     filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
#     if route_name:
#         van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
#         salesman = van_route.van.salesman
#         filter_data['route_name'] = route_name
#         #new customers created
#         new_customers_count = Customers.objects.filter(is_guest=False, created_date__date=date,sales_staff_id=salesman).count()
#         #emergency supply
#         emergency_supply_count = DiffBottlesModel.objects.filter(created_date__date=date, assign_this_to_id=salesman).count()
#         #actual visit
#         visited_customers_count = CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).distinct().count()
#         todays_customers = find_customers(request, str(date), van_route.routes.pk)
#         planned_visit_count = len(todays_customers)
#         non_visited_count = planned_visit_count - visited_customers_count
        
#         ##### stock report #### 
#         products = ProdutItemMaster.objects.filter()
#         van_instances = Van.objects.filter(salesman=salesman)
#         van_product_stock = VanProductStock.objects.filter()
        
#         #### 5 Gallon Related ###
#         empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=date, collected_empty_bottle__gt=0, salesman_id=salesman).count()
#         # empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=date, collected_empty_bottle__gt=0, **filter_data).count()
#         empty_bottle_pending = CustomerSupply.objects.filter(created_date__date=date, allocate_bottle_to_pending__gt=0, salesman_id=salesman).count()
#         coupons_collected = CustomerSupplyCoupon.objects.filter(customer_supply__created_date__date=date, customer_supply__salesman_id=salesman).aggregate(total_coupons=Count('leaf'))['total_coupons']
#         total_supplied_quantity = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date, customer_supply__salesman_id=salesman).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
#         total_collected_amount = CustomerSupply.objects.filter(created_date__date=date, salesman_id=salesman).aggregate(total_collected_amount=Sum('net_payable'))['total_collected_amount'] or 0
#         total_pending_amount = CustomerSupply.objects.filter(created_date__date=date, salesman_id=salesman).aggregate(total_pending_amount=Sum('grand_total') - Sum('net_payable'))['total_pending_amount'] or 0
#         mode_of_supply = CustomerSupply.objects.filter(created_date__date=date, salesman_id=salesman).values('customer__sales_type').annotate(total=Count('customer__sales_type')) or 0
        
#         #### Bottle Count ####
#         total_empty_bottles = CustodyCustomItems.objects.filter(custody_custom__customer__sales_staff_id=salesman,custody_custom__created_date__date=date).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0

#         total_supplied_bottles = CustomerSupply.objects.filter(created_date__date=date).aggregate(total_bottles=Sum('collected_empty_bottle'))['total_bottles'] or 0
#         # closing_stock_count = VanStock.objects.filter(created_date=date,stock_type="closing").aggregate(total_count=Sum('count'))['total_count'] or 0,
#         closing_stock_count = VanStock.objects.filter(created_date__date=date, stock_type='closing').count() or 0
#         damage_bottle_count = VanProductItems.objects.filter(van_stock__created_date__date=date, van_stock__stock_type='damage').aggregate(total_damage=Sum('count'))['total_damage'] or 0
#         pending_bottle_count = CustomerSupply.objects.filter(created_date__date=date,salesman_id=salesman).aggregate(total_pending=Sum('allocate_bottle_to_pending'))['total_pending'] or 0

#         total_count = total_empty_bottles + total_supplied_bottles + closing_stock_count + damage_bottle_count + pending_bottle_count
        
#         #### coupon sales count ####
#         customer_coupon_items=CustomerCouponItems.objects.filter(customer_coupon__salesman=salesman,customer_coupon__created_date__date=date).order_by("-customer_coupon__created_date")
        
#         ### cash sales ####
#         cash_invoices = Invoice.objects.filter(created_date__date=date, invoice_type="cash_invoice")
#         cash_total_net_taxable = cash_invoices.aggregate(total_net_taxable=Sum('net_taxable'))['total_net_taxable'] or 0
#         cash_total_vat = cash_invoices.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
#         cash_total_amout_total = cash_invoices.aggregate(total_amout_total=Sum('amout_total'))['total_amout_total'] or 0
        
#         ### credit sales ####
#         credit_invoices = Invoice.objects.filter(created_date__date=date, invoice_type="credit_invoice")
#         credit_total_net_taxable = credit_invoices.aggregate(total_net_taxable=Sum('net_taxable'))['total_net_taxable'] or 0
#         credit_total_vat = credit_invoices.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
#         credit_total_amout_total = credit_invoices.aggregate(total_amout_total=Sum('amout_total'))['total_amout_total'] or 0
        
#         ### expenses ####
#         expenses_instanses = Expense.objects.filter(date_created=date,van__salesman=salesman)
        
#         ### suspense ###
#         suspense_collections = SuspenseCollection.objects.filter(created_date__date=date,salesman=salesman)
#         cash_sales_amount = suspense_collections.aggregate(total_cash_sale=Sum('cash_sale_amount'))['total_cash_sale'] or 0
#         credit_sales_amount = suspense_collections.aggregate(total_credit_sale=Sum('credit_sale_amount'))['total_credit_sale'] or 0
        
#         suspense_opening_amount = cash_sales_amount + credit_sales_amount
#         suspense_paid_amount = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
#         suspense_balance_amount = suspense_opening_amount - suspense_paid_amount
        

#     context = {
#         'van_route': van_route,
#         # visit statistics
#         'new_customers_count': new_customers_count,
#         'emergency_supply_count': emergency_supply_count,
#         'visited_customers_count': visited_customers_count,
#         'non_visited_count': non_visited_count,
#         'planned_visit_count': planned_visit_count,
#         'routes_instances': routes_instances,
#         # stock report
#         'products': products,
#         'van_instances': van_instances,
#         'van_product_stock': van_product_stock,
#         # 5 Gallon Related
#         'empty_bottles_collected': empty_bottles_collected,
#         'empty_bottle_pending': empty_bottle_pending,
#         'coupons_collected': coupons_collected,
#         'total_supplied_quantity': total_supplied_quantity,
#         'total_collected_amount': total_collected_amount,
#         'total_pending_amount': total_pending_amount,
#         'mode_of_supply' :mode_of_supply,
#         # Bottle Count
#         'total_empty_bottles': total_empty_bottles,
#         'total_supplied_bottles':total_supplied_bottles,
#         'closing_stock_count': closing_stock_count,
#         'damage_bottle_count': damage_bottle_count,
#         'pending_bottle_count': pending_bottle_count,
#         'total_count': total_count,
#         #coupon book sale
#         'customer_coupon_items':customer_coupon_items,
#         #cash sales
#         'cash_invoices': cash_invoices,
#         'cash_total_net_taxable':cash_total_net_taxable,
#         'cash_total_vat':cash_total_vat,
#         'cash_total_amout_total':cash_total_amout_total,
#         # credit sales
#         'credit_invoices': credit_invoices,
#         'credit_total_net_taxable':credit_total_net_taxable,
#         'credit_total_vat':credit_total_vat,
#         'credit_total_amout_total':credit_total_amout_total,
#         # expenses
#         'expenses_instanses': expenses_instanses,
#         # suspense
#         'suspense_opening_amount': suspense_opening_amount,
#         'suspense_paid_amount': suspense_paid_amount,
#         'suspense_balance_amount': suspense_balance_amount,
        
#         'filter_data': filter_data
#     }
#     html_content = render_to_string('sales_management/dsr_summary.html', context)

#     # Create a new Excel workbook
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     worksheet.title = "DSR Summary"

#     # Convert HTML content to plain text and write it to the Excel file
#     for row_idx, row in enumerate(html_content.split('\n'), start=1):
#         for col_idx, value in enumerate(row.split(','), start=1):
#             worksheet.cell(row=row_idx, column=col_idx, value=value)

#     # Save the workbook to an in-memory buffer
#     output = io.BytesIO()
#     workbook.save(output)
#     output.seek(0)  # Move the buffer's cursor to the beginning

#     # Prepare response to send the Excel file
#     response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="dialy-summery-report{date}.xlsx"'

#     return response

# def export_daily_summary_report(request):
#     filter_data = {}
#     new_customers_count = 0
#     emergency_supply_count = 0
#     visited_customers_count = 0
#     non_visited_count = 0
#     planned_visit_count = 0
#     empty_bottles_collected = 0
#     empty_bottle_pending = 0
#     coupons_collected = 0
#     total_supplied_quantity = 0
#     total_collected_amount = 0
#     total_pending_amount = 0
#     mode_of_supply = 0
#     total_empty_bottles = 0
#     total_supplied_bottles = 0
#     closing_stock_count = 0
#     damage_bottle_count = 0
#     pending_bottle_count = 0
#     total_count = 0
#     cash_total_net_taxable = 0
#     cash_total_vat = 0
#     cash_total_amout_total = 0
#     credit_total_net_taxable = 0
#     credit_total_vat = 0
#     credit_total_amout_total = 0
#     suspense_opening_amount = 0
#     suspense_paid_amount = 0
#     suspense_balance_amount = 0
    
#     van_instances = Van.objects.none
#     van_route = Van_Routes.objects.none
#     cash_invoices = Invoice.objects.none
#     credit_invoices = Invoice.objects.none
#     products = ProdutItemMaster.objects.none
#     expenses_instanses = Expense.objects.none
#     routes_instances = RouteMaster.objects.all()
#     van_product_stock = VanProductStock.objects.none
#     customer_coupon_items = CustomerCouponItems.objects.none
    
#     date = request.GET.get('date')
#     route_name = request.GET.get('route_name')
    
#     if date:
#         date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
#     else:
#         date = datetime.datetime.today().date()
#     filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
#     if route_name:
#         van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
#         salesman = van_route.van.salesman
#         filter_data['route_name'] = route_name
#         #new customers created
#         new_customers_count = Customers.objects.filter(is_guest=False, created_date__date=date,sales_staff_id=salesman).count()
#         #emergency supply
#         emergency_supply_count = DiffBottlesModel.objects.filter(created_date__date=date, assign_this_to_id=salesman).count()
#         #actual visit
#         visited_customers_count = CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).distinct().count()
#         todays_customers = find_customers(request, str(date), van_route.routes.pk)
#         planned_visit_count = len(todays_customers)
#         non_visited_count = planned_visit_count - visited_customers_count
        
#         ##### stock report #### 
#         products = ProdutItemMaster.objects.filter()
#         van_instances = Van.objects.filter(salesman=salesman)
#         van_product_stock = VanProductStock.objects.filter()
        
#         #### 5 Gallon Related ###
#         empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=date, collected_empty_bottle__gt=0, salesman_id=salesman).count()
#         # empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=date, collected_empty_bottle__gt=0, **filter_data).count()
#         empty_bottle_pending = CustomerSupply.objects.filter(created_date__date=date, allocate_bottle_to_pending__gt=0, salesman_id=salesman).count()
#         coupons_collected = CustomerSupplyCoupon.objects.filter(customer_supply__created_date__date=date, customer_supply__salesman_id=salesman).aggregate(total_coupons=Count('leaf'))['total_coupons']
#         total_supplied_quantity = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date, customer_supply__salesman_id=salesman).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
#         total_collected_amount = CustomerSupply.objects.filter(created_date__date=date, salesman_id=salesman).aggregate(total_collected_amount=Sum('net_payable'))['total_collected_amount'] or 0
#         total_pending_amount = CustomerSupply.objects.filter(created_date__date=date, salesman_id=salesman).aggregate(total_pending_amount=Sum('grand_total') - Sum('net_payable'))['total_pending_amount'] or 0
#         mode_of_supply = CustomerSupply.objects.filter(created_date__date=date, salesman_id=salesman).values('customer__sales_type').annotate(total=Count('customer__sales_type')) or 0
        
#         #### Bottle Count ####
#         total_empty_bottles = CustodyCustomItems.objects.filter(custody_custom__customer__sales_staff_id=salesman,custody_custom__created_date__date=date).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0

#         total_supplied_bottles = CustomerSupply.objects.filter(created_date__date=date).aggregate(total_bottles=Sum('collected_empty_bottle'))['total_bottles'] or 0
#         # closing_stock_count = VanStock.objects.filter(created_date=date,stock_type="closing").aggregate(total_count=Sum('count'))['total_count'] or 0,
#         closing_stock_count = VanStock.objects.filter(created_date__date=date, stock_type='closing').count() or 0
#         damage_bottle_count = VanProductItems.objects.filter(van_stock__created_date__date=date, van_stock__stock_type='damage').aggregate(total_damage=Sum('count'))['total_damage'] or 0
#         pending_bottle_count = CustomerSupply.objects.filter(created_date__date=date,salesman_id=salesman).aggregate(total_pending=Sum('allocate_bottle_to_pending'))['total_pending'] or 0

#         total_count = total_empty_bottles + total_supplied_bottles + closing_stock_count + damage_bottle_count + pending_bottle_count
        
#         #### coupon sales count ####
#         customer_coupon_items=CustomerCouponItems.objects.filter(customer_coupon__salesman=salesman,customer_coupon__created_date__date=date).order_by("-customer_coupon__created_date")
        
#         ### cash sales ####
#         cash_invoices = Invoice.objects.filter(created_date__date=date, invoice_type="cash_invoice")
#         cash_total_net_taxable = cash_invoices.aggregate(total_net_taxable=Sum('net_taxable'))['total_net_taxable'] or 0
#         cash_total_vat = cash_invoices.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
#         cash_total_amout_total = cash_invoices.aggregate(total_amout_total=Sum('amout_total'))['total_amout_total'] or 0
        
#         ### credit sales ####
#         credit_invoices = Invoice.objects.filter(created_date__date=date, invoice_type="credit_invoice")
#         credit_total_net_taxable = credit_invoices.aggregate(total_net_taxable=Sum('net_taxable'))['total_net_taxable'] or 0
#         credit_total_vat = credit_invoices.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
#         credit_total_amout_total = credit_invoices.aggregate(total_amout_total=Sum('amout_total'))['total_amout_total'] or 0
        
#         ### expenses ####
#         expenses_instanses = Expense.objects.filter(date_created=date,van__salesman=salesman)
        
#         ### suspense ###
#         suspense_collections = SuspenseCollection.objects.filter(created_date__date=date,salesman=salesman)
#         cash_sales_amount = suspense_collections.aggregate(total_cash_sale=Sum('cash_sale_amount'))['total_cash_sale'] or 0
#         credit_sales_amount = suspense_collections.aggregate(total_credit_sale=Sum('credit_sale_amount'))['total_credit_sale'] or 0
        
#         suspense_opening_amount = cash_sales_amount + credit_sales_amount
#         suspense_paid_amount = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
#         suspense_balance_amount = suspense_opening_amount - suspense_paid_amount

#     # Create DataFrames for each table
#     visit_statistics_df = pd.DataFrame({
#         'New Customer Created': [new_customers_count],
#         'Planned Visit': [planned_visit_count],
#         'Emergency Supply': [emergency_supply_count],
#         'Non Visited': [non_visited_count]
#     })

#     # Define the Excel writer
#     excel_writer = pd.ExcelWriter('daily_summary_report.xlsx', engine='openpyxl')

#     # Write each DataFrame to a specific sheet in the Excel file
#     visit_statistics_df.to_excel(excel_writer, index=False, sheet_name='Visit Statistics', startrow=1, startcol=1)

#     # Get the workbook and the worksheet
#     workbook = excel_writer.book
#     worksheet = excel_writer.sheets['Visit Statistics']

#     # Set column widths
#     column_widths = [20, 20, 20, 20]  # Adjust the widths as needed
#     for i, width in enumerate(column_widths):
#         worksheet.column_dimensions[worksheet.cell(row=1, column=i+1).column_letter].width = width

#     # Add headers
#     for idx, value in enumerate(visit_statistics_df.columns):
#         worksheet.cell(row=1, column=idx+1, value=value)

#     # Save the Excel file
#     excel_writer.save()

#     # Define the file path
#     file_path = 'daily_summary_report.xlsx'

#     # Prepare response
#     with open(file_path, 'rb') as file:
#         response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
#         response['Content-Disposition'] = 'attachment; filename="daily_summary_report.xlsx"'
#     return response

def bottle_transactions(request):
    template = 'sales_management/bottle_transactions.html'
    filter_data = {}

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    query = request.GET.get("q")
    route_filter = request.GET.get('route_name')

    # Start with all customers
    user_li = Customers.objects.all()

    # Apply filters if they exist
    if query:
        user_li = user_li.filter(
            Q(custom_id__icontains=query) |
            Q(customer_name__icontains=query) |
            Q(sales_type__icontains=query) |
            Q(mobile_no__icontains=query) |
            Q(routes__route_name__icontains=query) |
            Q(location__location_name__icontains=query) |
            Q(building_name__icontains=query)
        )

    if route_filter:
        user_li = user_li.filter(routes__route_name=route_filter)
        filter_data["route_name"] = route_filter

    # Get all route names for the dropdown
    route_li = RouteMaster.objects.all()

    customersupplyitems = CustomerSupplyItems.objects.all()
    today = datetime.datetime.today().date()
    
    customersupplyitems = customersupplyitems.filter(
                    product__product_name='5 Gallon',
    )

    if start_date and end_date:
        customersupplyitems = customersupplyitems.filter(
            customer_supply__created_date__range=[start_date, end_date],
        )
        filter_data['start_date'] = start_date
        filter_data['end_date'] = end_date

    if route_filter:
        customersupplyitems = customersupplyitems.filter(
            customer_supply__customer__routes__route_name=route_filter
        )

    # Aggregate data for the selected route
    aggregated_data = customersupplyitems.aggregate(
        total_collected_empty_bottle=Sum('customer_supply__collected_empty_bottle'),
        total_fivegallon_quantity=Sum('quantity')
    )

    context = {
        'customersupplyitems': customersupplyitems.order_by("-customer_supply__created_date"),
        'today': today,
        'filter_data': filter_data,
        'route_li': route_li,
        'aggregated_data': aggregated_data,
    }

    return render(request, template, context)

#------------------------------Bottle Count-------------------------------------

def van_route_bottle_count(request):
    selected_route = request.GET.get('route_name', '')
    date_str = request.GET.get("filter_date")
    
    if date_str:
        try:
            date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date = datetime.datetime.today().date()
    else:
        date = datetime.datetime.today().date()

    if selected_route:
        vans = Van.objects.filter(van_master__routes__route_name=selected_route)
    else:
        vans = Van.objects.all()

    routes = RouteMaster.objects.all()
    
    instances = BottleCount.objects.filter(created_date__date=date,van__in=vans)

    context = {
        'instances': instances,
        'routes': routes,
        'filter_data': {
            'selected_route': selected_route,
            'selected_date': date_str if date_str else datetime.datetime.today().date(),
        }
    }

    return render(request, 'sales_management/van_bottle_count.html', context)

def VansRouteBottleCountAdd(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)

    if request.method == 'POST':
        form = VansRouteBottleCountAddForm(request.POST)
        if form.is_valid():
            bottle_count = form.save(commit=False)
            bottle_count.van = van
            bottle_count.created_by = request.user.username
            bottle_count.save()

            return redirect('van_route_bottle_count')
    else:
        form = VansRouteBottleCountAddForm()

    context = {
        'form': form,
        'van': van,
    }
    return render(request, 'sales_management/van_route_bottle_count_add.html', context)

def VansRouteBottleCountDeduct(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)

    if request.method == 'POST':
        form = VansRouteBottleCountDeductForm(request.POST)
        if form.is_valid():
            bottle_count = form.save(commit=False)
            bottle_count.van = van
            bottle_count.created_by = request.user.username
            bottle_count.save()

            return redirect('van_route_bottle_count')
    else:
        form = VansRouteBottleCountDeductForm()

    context = {
        'form': form,
        'van': van,
    }
    return render(request, 'sales_management/van_route_bottle_count_deduct.html', context)

#-----------------DSR Outstanding Amount Collected Customers--------------------------

def outstanding_amount_collected(request):
# Initialize counts and variables
    filter_data = {}
    data_filter = False
    salesman_id = ""
    van_route = None
    dialy_collections = CollectionPayment.objects.none()
    outstanding_total_amount_collected = 0
    outstanding_credit_notes_balance = 0

    # Retrieve all route instances
    routes_instances = RouteMaster.objects.all()

    # Get filter parameters from request
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')

    # Set date and filter data
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')

    if route_name:
        data_filter = True
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        route_id = van_route.routes.pk if van_route else None
        filter_data['route_name'] = route_name
    else:
        route_id = None

    # Ensure van_route is defined
    if van_route:
        # Retrieve salesman
        salesman = van_route.van.salesman
        salesman_id = salesman.pk

        # Collection details
        dialy_collections = CollectionPayment.objects.filter(salesman_id=salesman, amount_received__gt=0)

        # Credit outstanding
        outstanding_credit_notes_total_amount = OutstandingAmount.objects.filter(
            customer_outstanding__created_date__date__lte=date,
            customer_outstanding__product_type="amount",
            customer_outstanding__customer__routes=van_route.routes
        ).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        
        outstanding_credit_notes_total_amount -= dialy_collections.filter(
            created_date__date__lte=date
        ).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        
        outstanding_credit_notes_received_amount = dialy_collections.filter(
            created_date__date=date
        ).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        
        outstanding_credit_notes_balance = outstanding_credit_notes_total_amount - outstanding_credit_notes_received_amount
        outstanding_total_amount_collected = dialy_collections.filter(
            created_date__date=date
        ).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
    else:
        dialy_collections = []

    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        'dialy_collections': dialy_collections,
        'routes_instances': routes_instances,
        'filter_data': filter_data,
        'outstanding_total_amount_collected': outstanding_total_amount_collected,
        'outstanding_credit_notes_balance': outstanding_credit_notes_balance,
    }
    
    return render(request, 'sales_management/dsr_outstanding_amount_collected.html', context)


def dsr(request):
    
    filter_data = {}
    data_filter = False
    new_customers_count = 0
    emergency_supply_count = 0
    visited_customers_count = 0
    non_visited_count = 0
    planned_visit_count = 0
    total_empty_bottles = 0
    total_supplied_bottles = 0
    closing_stock_count = 0
    damage_bottle_count = 0
    pending_bottle_count = 0
    total_count = 0
    cash_total_net_taxable = 0
    cash_total_vat = 0
    cash_total_subtotal = 0
    cash_total_amount_recieved = 0
    cash_sale_recharge_count=0
    cash_total_quantity = 0
    cash_total_qty = 0
    credit_total_net_taxable = 0
    credit_total_vat = 0
    credit_total_subtotal = 0
    credit_total_amount_recieved = 0
    credit_total_qty = 0
    credit_total_quantity = 0
    credit_sale_recharge_count = 0
    in_hand_amount = 0
    today_expense = 0
    today_payable = 0
    suspense_paid_amount = 0
    suspense_balance_amount = 0
    outstanding_credit_notes_total_amount = 0
    outstanding_credit_notes_received_amount = 0
    outstanding_credit_notes_balance = 0
    outstanding_total_amount_collected = 0
    cash_sales_amount_collected = 0
    credit_sales_amount_collected = 0
    total_sales_amount_collected = 0
    total_cash_sales_count = 0
    total_credit_sales_count = 0
    total_sales_count = 0
    no_of_collected_cheque = 0
    collected_cheque_amount = 0
    balance_in_hand = 0
    net_payble = 0
    stock_report_total = 0
    manual_coupon_total = 0
    digital_coupon_total = 0
    total_coupon_sales_count = 0
    coupon_total_qty = 0
    total_debit_amount_count = 0
    total_credit_amount_count = 0
    total_coupon_amount_count = 0
    
    five_gallon_cash_total_net_taxable = 0
    five_gallon_cash_total_vat = 0
    five_gallon_cash_total_subtotal = 0
    five_gallon_cash_total_received = 0
    five_gallon_cash_total_quantity = 0
    other_cash_total_net_taxable = 0
    other_cash_total_vat = 0
    other_cash_total_subtotal = 0
    other_cash_total_received = 0
    other_cash_total_quantity = 0
    cash_sale_recharge_net_payeble = 0
    cash_sale_recharge_vat_total = 0
    cash_sale_recharge_grand_total = 0
    cash_sale_recharge_amount_recieved = 0
    
    five_gallon_credit_total_net_taxable = 0
    five_gallon_credit_total_vat = 0
    five_gallon_credit_total_subtotal = 0
    five_gallon_credit_total_received = 0
    five_gallon_credit_total_quantity = 0
    other_credit_total_net_taxable = 0
    other_credit_total_vat = 0
    other_credit_total_subtotal = 0
    other_credit_total_received = 0
    other_credit_total_quantity = 0
    credit_sale_recharge_net_payeble = 0
    credit_sale_recharge_vat_total = 0
    credit_sale_recharge_grand_total = 0
    credit_sale_recharge_amount_recieved = 0
   
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    cash_sales = CustomerSupply.objects.none
    credit_sales = CustomerSupply.objects.none
    five_gallon_cash_sales = CustomerSupply.objects.none
    five_gallon_credit_sales = CustomerSupply.objects.none
    other_cash_sales = CustomerSupply.objects.none
    other_credit_sales = CustomerSupply.objects.none
    coupon_sales = CustomerSupply.objects.none
    recharge_cash_sales = CustomerCoupon.objects.none
    recharge_credit_sales = CustomerCoupon.objects.none
    products = ProdutItemMaster.objects.none
    expenses_instanses = Expense.objects.none
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none
    customer_coupons = CustomerCoupon.objects.none
    unique_amounts = CustomerCouponItems.objects.none
    dialy_collections = InvoiceDailyCollection.objects.none
    pending_bottle_customer_instances = CustomerSupply.objects.none
    foc_customers = CustomerSupply.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        #new customers created
        new_customers_count = Customers.objects.filter(is_guest=False, created_date__date=date,sales_staff_id=salesman).count()
        #emergency supply
        emergency_supply_count = DiffBottlesModel.objects.filter(created_date__date=date, assign_this_to_id=salesman).count()
        #actual visit
        visited_customers_count = CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).distinct().count()
        todays_customers = find_customers(request, str(date), van_route.routes.pk)
        if todays_customers :
            planned_visit_count = len(todays_customers)
        else:
            planned_visit_count = 0
        non_visited_count = planned_visit_count - visited_customers_count
        
        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=salesman)
        van_product_stock = VanProductStock.objects.filter(created_date=date,van=van_instances,product__product_name="5 Gallon")
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        #### Bottle Count ####
        total_empty_bottles = van_product_stock.aggregate(totalempty_bottle=Sum('empty_can_count'))['totalempty_bottle'] or 0
        total_supplied_bottles =  van_product_stock.aggregate(total_sold=Sum('sold_count'))['total_sold'] or 0
        pending_bottle_count = van_product_stock.aggregate(total_pending=Sum('pending_count'))['total_pending'] or 0
        damage_bottle_count = van_product_stock.aggregate(total_damage=Sum('damage_count'))['total_damage'] or 0
        closing_stock_count = van_product_stock.aggregate(total_closing=Sum('closing_count'))['total_closing'] or 0
        total_count = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        #### coupon sales count ####
        customer_coupons=CustomerCoupon.objects.filter(salesman=salesman,created_date__date=date)
        
        five_gallon_supply = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").values_list('customer_supply__pk', flat=True)
        other_supply = CustomerSupplyItems.objects.exclude(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").values_list('customer_supply__pk', flat=True)
        ### Cash Sales Start ###
        five_gallon_cash_sales = CustomerSupply.objects.filter(pk__in=five_gallon_supply,created_date__date=date,salesman=salesman,amount_recieved__gt=0).exclude(customer__sales_type="CASH COUPON")
        other_cash_sales = CustomerSupply.objects.filter(pk__in=other_supply,created_date__date=date,salesman=salesman,amount_recieved__gt=0).exclude(customer__sales_type="CASH COUPON")
        # Aggregating for five_gallon_cash_sales
        five_gallon_cash_total_net_taxable = five_gallon_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_cash_total_vat = five_gallon_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_cash_total_subtotal = five_gallon_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_cash_total_received = five_gallon_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_cash_total_quantity = five_gallon_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_cash_sales (corrected to use other_cash_sales)
        other_cash_total_net_taxable = other_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_cash_total_vat = other_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_cash_total_subtotal = other_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_cash_total_received = other_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_cash_total_quantity = other_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Combining both cash sales querysets
        cash_sales = five_gallon_cash_sales.union(other_cash_sales)
        cash_total_net_taxable = five_gallon_cash_total_net_taxable + other_cash_total_net_taxable
        cash_total_vat = five_gallon_cash_total_vat + other_cash_total_vat
        cash_total_subtotal = five_gallon_cash_total_subtotal + other_cash_total_subtotal
        cash_total_received = five_gallon_cash_total_received + other_cash_total_received
        cash_total_quantity = five_gallon_cash_total_quantity + other_cash_total_quantity
        
        recharge_cash_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__gt=0)
        cash_sale_recharge_net_payeble = recharge_cash_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        cash_sale_recharge_vat_total = 0
        cash_sale_recharge_grand_total = recharge_cash_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        cash_sale_recharge_amount_recieved = recharge_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        cash_total_net_taxable = cash_total_net_taxable + cash_sale_recharge_net_payeble 
        cash_total_vat = cash_total_vat + cash_sale_recharge_vat_total 
        cash_total_subtotal = cash_total_subtotal + cash_sale_recharge_grand_total 
        cash_total_amount_recieved = cash_total_received + cash_sale_recharge_amount_recieved 
        
        total_cash_sales_count = cash_sales.count() + recharge_cash_sales.count()
        
        cash_sale_recharge_count = recharge_cash_sales.count()
        cash_total_qty = cash_total_quantity + cash_sale_recharge_count
        ### Cash Sales End ###
        
        ### credit sales Start ####
        five_gallon_credit_sales = CustomerSupply.objects.filter(pk__in=five_gallon_supply,created_date__date=date,salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])
        other_credit_sales = CustomerSupply.objects.filter(pk__in=other_supply,created_date__date=date,salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])
        
        # Aggregating for five_gallon_credit_sales
        five_gallon_credit_total_net_taxable = five_gallon_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_credit_total_vat = five_gallon_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_credit_total_subtotal = five_gallon_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_credit_total_received = five_gallon_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_credit_total_quantity = five_gallon_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_credit_sales (corrected to use other_credit_sales)
        other_credit_total_net_taxable = other_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_credit_total_vat = other_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_credit_total_subtotal = other_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_credit_total_received = other_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_credit_total_quantity = other_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Combining both cash sales querysets
        credit_sales = five_gallon_credit_sales.union(other_credit_sales)
        credit_total_net_taxable = five_gallon_credit_total_net_taxable + other_credit_total_net_taxable
        credit_total_vat = five_gallon_credit_total_vat + other_credit_total_vat
        credit_total_subtotal = five_gallon_credit_total_subtotal + other_credit_total_subtotal
        credit_total_received = five_gallon_credit_total_received + other_credit_total_received
        credit_total_quantity = five_gallon_credit_total_quantity + other_credit_total_quantity
        
        recharge_credit_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__lte=0)
        credit_sale_recharge_net_payeble = recharge_credit_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        credit_sale_recharge_vat_total = 0
        credit_sale_recharge_grand_total = recharge_credit_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        credit_sale_recharge_amount_recieved = recharge_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        credit_total_net_taxable = credit_total_net_taxable + credit_sale_recharge_net_payeble 
        credit_total_vat = credit_total_vat + credit_sale_recharge_vat_total 
        credit_total_subtotal = credit_total_subtotal + credit_sale_recharge_grand_total
        credit_total_amount_recieved = credit_total_received + credit_sale_recharge_amount_recieved
        
        total_credit_sales_count = credit_sales.count() + recharge_credit_sales.count()
        credit_sale_recharge_count = recharge_credit_sales.count()
        credit_total_qty = credit_total_quantity + credit_sale_recharge_count
        ### credit sales End ####
        
        # Coupon sales
        coupon_sales = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,customer__sales_type="CASH COUPON")
        manual_coupon_total = CustomerSupplyCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(Count('leaf'))['leaf__count']
        digital_coupon_total = CustomerSupplyDigitalCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(total_count=Sum('count'))['total_count'] or 0
        
        total_coupon_sales_count = coupon_sales.count()
        
        total_sales_count = total_cash_sales_count + total_credit_sales_count + total_coupon_sales_count
        coupon_total_qty = coupon_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        ### expenses ####
        expenses_instanses = Expense.objects.filter(expense_date=date,van=van_route.van)
        today_expense = expenses_instanses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
        
        ### suspense ###
        suspense_collections = SuspenseCollection.objects.filter(created_date__date=date,salesman=salesman)
        cash_sales_amount = suspense_collections.aggregate(total_cash_sale=Sum('cash_sale_amount'))['total_cash_sale'] or 0
        credit_sales_amount = suspense_collections.aggregate(total_credit_sale=Sum('credit_sale_amount'))['total_credit_sale'] or 0
        
        in_hand_amount = cash_sales_amount + credit_sales_amount
        today_payable = in_hand_amount - today_expense
        suspense_paid_amount = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
        suspense_balance_amount = today_payable - suspense_paid_amount
        
        # collection details
        dialy_collections = CollectionPayment.objects.filter(salesman_id=salesman)
        # credit outstanding
        # outstanding_credit_notes = Invoice.objects.filter(invoice_type="credit_invoice",customer__sales_staff=salesman).exclude(created_date__date__gt=date)
        outstanding_credit_notes_total_amount = OutstandingAmount.objects.filter(customer_outstanding__created_date__date__lte=date,customer_outstanding__product_type="amount",customer_outstanding__customer__routes=van_route.routes).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        dialy_colection_upto__yesterday = dialy_collections.filter(created_date__date__lt=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_total_amount = outstanding_credit_notes_total_amount - dialy_colection_upto__yesterday
        outstanding_credit_notes_received_amount = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_balance = outstanding_credit_notes_total_amount - outstanding_credit_notes_received_amount
        outstanding_total_amount_collected = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0

        
        # pending customers
        pending_bottle_customer_instances = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,allocate_bottle_to_pending__gt=0)
        # 5 gallon rate based
        unique_amounts = set(CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon").values_list('customer_supply__customer__rate', flat=True))
        five_gallon_rate_wise_instances = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon")
        total_debit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved__gt=0).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_credit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved=0).exclude(customer_supply__customer__sales_type__in=["FOC","CASH COUPON"]).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_coupon_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__customer__sales_type="CASH COUPON").aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        # cash sales amount collected
        supply_amount_collected = cash_total_amount_recieved
        cash_sales_amount_collected = supply_amount_collected
        
        dialy_collections = dialy_collections.filter(created_date__date=date)
        credit_sales_amount_collected = dialy_collections.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        total_sales_amount_collected = cash_sales_amount_collected + credit_sales_amount_collected
        
        cheque_collection = CollectionPayment.objects.filter(payment_method="CHEQUE",created_date__date=date,salesman=salesman)
        no_of_collected_cheque = cheque_collection.count()
        collected_cheque_amount = cheque_collection.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        
        balance_in_hand = total_sales_amount_collected - collected_cheque_amount - today_expense
        net_payble = total_sales_amount_collected - today_expense
        
        
        foc_customers = CustomerSupply.objects.filter(created_date__date=date, customer__sales_type='FOC', salesman=salesman) or CustomerSupply.objects.filter(created_date__date=date, salesman=salesman, allocate_bottle_to_free__gt=0)
        
        # if van_product_stock.empty_can_count != CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).aggregate(collected_empty_bottle=Sum('collected_empty_bottle'))['collected_empty_bottle'] or 0 :
        #     van_product_stock.empty_can_count = CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).aggregate(collected_empty_bottle=Sum('collected_empty_bottle'))['collected_empty_bottle'] or 0
            
        # if van_product_stock.pending_count :
        #     van_product_stock.pending_count = CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).aggregate(allocate_bottle_to_pending=Sum('allocate_bottle_to_pending'))['allocate_bottle_to_pending'] or 0
        
        # if van_product_stock.sold_count :
        #     van_product_stock.sold_count = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").exclude(customer_supply__customer__sales_type__in=["FOC"]).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
            
        # if van_product_stock.stock :
        #     van_product_stock.stock = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").exclude(customer_supply__customer__sales_type__in=["FOC"]).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
            
        # if van_product_stock.foc :
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        # visit statistics
        'new_customers_count': new_customers_count,
        'emergency_supply_count': emergency_supply_count,
        'visited_customers_count': visited_customers_count,
        'non_visited_count': non_visited_count,
        'planned_visit_count': planned_visit_count,
        'routes_instances': routes_instances,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        # pending customers
        'pending_bottle_customer_instances': pending_bottle_customer_instances,
        # Bottle Count
        'total_empty_bottles': total_empty_bottles,
        'total_supplied_bottles':total_supplied_bottles,
        'closing_stock_count': closing_stock_count,
        'damage_bottle_count': damage_bottle_count,
        'pending_bottle_count': pending_bottle_count,
        'total_count': total_count,
        #coupon book sale
        'customer_coupons':customer_coupons,
        # five gallon cash sales
        'five_gallon_cash_sales': five_gallon_cash_sales or [],
        'five_gallon_cash_total_net_taxable': five_gallon_cash_total_net_taxable,
        'five_gallon_cash_total_vat': five_gallon_cash_total_vat,
        'five_gallon_cash_total_subtotal': five_gallon_cash_total_subtotal,
        'five_gallon_cash_total_received': five_gallon_cash_total_received,
        'five_gallon_cash_total_quantity': five_gallon_cash_total_quantity,
        # other cash sales
        'other_cash_sales': other_cash_sales,
        'other_cash_total_net_taxable': other_cash_total_net_taxable,
        'other_cash_total_vat': other_cash_total_vat,
        'other_cash_total_subtotal': other_cash_total_subtotal,
        'other_cash_total_received': other_cash_total_received,
        'other_cash_total_quantity': other_cash_total_quantity,
        #cash sales
        'recharge_cash_sales': recharge_cash_sales,
        'cash_sale_recharge_net_payeble': cash_sale_recharge_net_payeble,
        'cash_sale_recharge_vat_total': cash_sale_recharge_vat_total,
        'cash_sale_recharge_grand_total': cash_sale_recharge_grand_total,
        'cash_sale_recharge_amount_recieved': cash_sale_recharge_amount_recieved,
        
        'cash_total_net_taxable':cash_total_net_taxable,
        'cash_total_vat':cash_total_vat,
        'cash_total_subtotal': cash_total_subtotal,
        'cash_total_amount_recieved': cash_total_amount_recieved,
        'cash_total_qty': cash_total_qty,
        # five gallon credit sales
        'five_gallon_credit_sales': five_gallon_credit_sales,
        'five_gallon_credit_total_net_taxable': five_gallon_credit_total_net_taxable,
        'five_gallon_credit_total_vat': five_gallon_credit_total_vat,
        'five_gallon_credit_total_subtotal': five_gallon_credit_total_subtotal,
        'five_gallon_credit_total_received': five_gallon_credit_total_received,
        'five_gallon_credit_total_quantity': five_gallon_credit_total_quantity,
        # credit sales
        'credit_sales': credit_sales,
        'recharge_credit_sales': recharge_credit_sales,
        'credit_sale_recharge_net_payeble': credit_sale_recharge_net_payeble,
        'credit_sale_recharge_vat_total': credit_sale_recharge_vat_total,
        'credit_sale_recharge_grand_total': credit_sale_recharge_grand_total,
        'credit_sale_recharge_amount_recieved': credit_sale_recharge_amount_recieved,
        
        'credit_total_net_taxable':credit_total_net_taxable,
        'credit_total_vat':credit_total_vat,
        'credit_total_subtotal':credit_total_subtotal,
        'credit_total_amount_recieved': credit_total_amount_recieved,
        'credit_total_qty': credit_total_qty,
        # coupon sales
        'coupon_sales': coupon_sales,
        'manual_coupon_total':manual_coupon_total,
        'digital_coupon_total':digital_coupon_total,
        'total_coupon_sales_count': total_coupon_sales_count,
        'coupon_total_qty':coupon_total_qty,
        # expenses
        'expenses_instanses': expenses_instanses,
        # suspense
        'in_hand_amount': in_hand_amount,
        'today_expense': today_expense, 
        'today_payable': today_payable,
        'suspense_paid_amount': suspense_paid_amount,
        'suspense_balance_amount': suspense_balance_amount,
        'outstanding_credit_notes_total_amount' : outstanding_credit_notes_total_amount,
        'outstanding_credit_notes_received_amount' : outstanding_credit_notes_received_amount,
        'outstanding_credit_notes_balance': outstanding_credit_notes_balance,
        'outstanding_total_amount_collected':outstanding_total_amount_collected,
        # 5 gallon rate based
        'five_gallon_rates': unique_amounts,
        'total_debit_amount_count': total_debit_amount_count,
        'total_credit_amount_count': total_credit_amount_count,
        'total_coupon_amount_count': total_coupon_amount_count,
        # dialy collections
        'dialy_collections': dialy_collections,
        # sales amount collected
        'cash_sales_amount_collected': cash_sales_amount_collected,
        'credit_sales_amount_collected': credit_sales_amount_collected,
        'total_sales_amount_collected': total_sales_amount_collected,
        'total_cash_sales_count': total_cash_sales_count,
        'total_credit_sales_count': total_credit_sales_count,
        'total_sales_count': total_sales_count,
        'no_of_collected_cheque': no_of_collected_cheque,
        'collected_cheque_amount': collected_cheque_amount,
        
        'balance_in_hand': balance_in_hand,
        'net_payble': net_payble,
        
        'filter_data': filter_data,
        # FOC customer
        'foc_customers':foc_customers,
    }
    
    return render(request, 'sales_management/new_dsr_summary.html', context)

def print_dsr(request):
    
    filter_data = {}
    data_filter = False
    new_customers_count = 0
    emergency_supply_count = 0
    visited_customers_count = 0
    non_visited_count = 0
    planned_visit_count = 0
    total_empty_bottles = 0
    total_supplied_bottles = 0
    closing_stock_count = 0
    damage_bottle_count = 0
    pending_bottle_count = 0
    total_count = 0
    cash_total_net_taxable = 0
    cash_total_vat = 0
    cash_total_subtotal = 0
    cash_total_amount_recieved = 0
    cash_sale_recharge_count=0
    cash_total_quantity = 0
    cash_total_qty = 0
    credit_total_net_taxable = 0
    credit_total_vat = 0
    credit_total_subtotal = 0
    credit_total_amount_recieved = 0
    credit_total_qty = 0
    credit_total_quantity = 0
    credit_sale_recharge_count = 0
    in_hand_amount = 0
    today_expense = 0
    today_payable = 0
    suspense_paid_amount = 0
    suspense_balance_amount = 0
    outstanding_credit_notes_total_amount = 0
    outstanding_credit_notes_received_amount = 0
    outstanding_credit_notes_balance = 0
    outstanding_total_amount_collected = 0
    cash_sales_amount_collected = 0
    credit_sales_amount_collected = 0
    total_sales_amount_collected = 0
    total_cash_sales_count = 0
    total_credit_sales_count = 0
    total_sales_count = 0
    no_of_collected_cheque = 0
    collected_cheque_amount = 0
    balance_in_hand = 0
    net_payble = 0
    stock_report_total = 0
    manual_coupon_total = 0
    digital_coupon_total = 0
    total_coupon_sales_count = 0
    coupon_total_qty = 0
    total_debit_amount_count = 0
    total_credit_amount_count = 0
    total_coupon_amount_count = 0
    
    five_gallon_cash_total_net_taxable = 0
    five_gallon_cash_total_vat = 0
    five_gallon_cash_total_subtotal = 0
    five_gallon_cash_total_received = 0
    five_gallon_cash_total_quantity = 0
    other_cash_total_net_taxable = 0
    other_cash_total_vat = 0
    other_cash_total_subtotal = 0
    other_cash_total_received = 0
    other_cash_total_quantity = 0
    cash_sale_recharge_net_payeble = 0
    cash_sale_recharge_vat_total = 0
    cash_sale_recharge_grand_total = 0
    cash_sale_recharge_amount_recieved = 0
    
    five_gallon_credit_total_net_taxable = 0
    five_gallon_credit_total_vat = 0
    five_gallon_credit_total_subtotal = 0
    five_gallon_credit_total_received = 0
    five_gallon_credit_total_quantity = 0
    other_credit_total_net_taxable = 0
    other_credit_total_vat = 0
    other_credit_total_subtotal = 0
    other_credit_total_received = 0
    other_credit_total_quantity = 0
    credit_sale_recharge_net_payeble = 0
    credit_sale_recharge_vat_total = 0
    credit_sale_recharge_grand_total = 0
    credit_sale_recharge_amount_recieved = 0
   
    van_instances = Van.objects.none
    van_route = Van_Routes.objects.none
    salesman_id =  ""
    cash_sales = CustomerSupply.objects.none
    credit_sales = CustomerSupply.objects.none
    five_gallon_cash_sales = CustomerSupply.objects.none
    five_gallon_credit_sales = CustomerSupply.objects.none
    other_cash_sales = CustomerSupply.objects.none
    other_credit_sales = CustomerSupply.objects.none
    coupon_sales = CustomerSupply.objects.none
    recharge_cash_sales = CustomerCoupon.objects.none
    recharge_credit_sales = CustomerCoupon.objects.none
    products = ProdutItemMaster.objects.none
    expenses_instanses = Expense.objects.none
    routes_instances = RouteMaster.objects.all()
    van_product_stock = VanProductStock.objects.none
    customer_coupons = CustomerCoupon.objects.none
    unique_amounts = CustomerCouponItems.objects.none
    dialy_collections = InvoiceDailyCollection.objects.none
    pending_bottle_customer_instances = CustomerSupply.objects.none
    foc_customers = CustomerSupply.objects.none
    
    date = request.GET.get('date')
    route_name = request.GET.get('route_name')
    
    if date:
        date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    
    if route_name:
        data_filter = True
        
        van_route = Van_Routes.objects.filter(routes__route_name=route_name).first()
        salesman = van_route.van.salesman
        salesman_id = salesman.pk
        filter_data['route_name'] = route_name
        #new customers created
        new_customers_count = Customers.objects.filter(is_guest=False, created_date__date=date,sales_staff_id=salesman).count()
        #emergency supply
        emergency_supply_count = DiffBottlesModel.objects.filter(created_date__date=date, assign_this_to_id=salesman).count()
        #actual visit
        visited_customers_count = CustomerSupply.objects.filter(salesman_id=salesman, created_date__date=date).distinct().count()
        todays_customers = find_customers(request, str(date), van_route.routes.pk)
        if todays_customers :
            planned_visit_count = len(todays_customers)
        else:
            planned_visit_count = 0
        non_visited_count = planned_visit_count - visited_customers_count
        
        ##### stock report #### 
        products = ProdutItemMaster.objects.filter()
        van_instances = Van.objects.get(salesman=salesman)
        van_product_stock = VanProductStock.objects.filter(created_date=date,van=van_instances,product__product_name="5 Gallon")
        stock_report_total = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        #### Bottle Count ####
        total_empty_bottles = van_product_stock.aggregate(totalempty_bottle=Sum('empty_can_count'))['totalempty_bottle'] or 0
        total_supplied_bottles =  van_product_stock.aggregate(total_sold=Sum('sold_count'))['total_sold'] or 0
        pending_bottle_count = van_product_stock.aggregate(total_pending=Sum('pending_count'))['total_pending'] or 0
        damage_bottle_count = van_product_stock.aggregate(total_damage=Sum('damage_count'))['total_damage'] or 0
        closing_stock_count = van_product_stock.aggregate(total_closing=Sum('closing_count'))['total_closing'] or 0
        total_count = van_product_stock.aggregate(total_stock=Sum('stock'))['total_stock'] or 0
        
        #### coupon sales count ####
        customer_coupons=CustomerCoupon.objects.filter(salesman=salesman,created_date__date=date)
        
        five_gallon_supply = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").values_list('customer_supply__pk', flat=True)
        other_supply = CustomerSupplyItems.objects.exclude(customer_supply__created_date__date=date,customer_supply__salesman=salesman,product__product_name="5 Gallon").values_list('customer_supply__pk', flat=True)
        ### Cash Sales Start ###
        five_gallon_cash_sales = CustomerSupply.objects.filter(pk__in=five_gallon_supply,created_date__date=date,salesman=salesman,amount_recieved__gt=0).exclude(customer__sales_type="CASH COUPON")
        other_cash_sales = CustomerSupply.objects.filter(pk__in=other_supply,created_date__date=date,salesman=salesman,amount_recieved__gt=0).exclude(customer__sales_type="CASH COUPON")

        # Aggregating for five_gallon_cash_sales
        five_gallon_cash_total_net_taxable = five_gallon_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_cash_total_vat = five_gallon_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_cash_total_subtotal = five_gallon_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_cash_total_received = five_gallon_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_cash_total_quantity = five_gallon_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_cash_sales (corrected to use other_cash_sales)
        other_cash_total_net_taxable = other_cash_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_cash_total_vat = other_cash_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_cash_total_subtotal = other_cash_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_cash_total_received = other_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_cash_total_quantity = other_cash_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Combining both cash sales querysets
        cash_sales = five_gallon_cash_sales.union(other_cash_sales)
        cash_total_net_taxable = five_gallon_cash_total_net_taxable + other_cash_total_net_taxable
        cash_total_vat = five_gallon_cash_total_vat + other_cash_total_vat
        cash_total_subtotal = five_gallon_cash_total_subtotal + other_cash_total_subtotal
        cash_total_received = five_gallon_cash_total_received + other_cash_total_received
        cash_total_quantity = five_gallon_cash_total_quantity + other_cash_total_quantity
        
        recharge_cash_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__gt=0)
        cash_sale_recharge_net_payeble = recharge_cash_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        cash_sale_recharge_vat_total = 0
        cash_sale_recharge_grand_total = recharge_cash_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        cash_sale_recharge_amount_recieved = recharge_cash_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        cash_total_net_taxable = cash_total_net_taxable + cash_sale_recharge_net_payeble 
        cash_total_vat = cash_total_vat + cash_sale_recharge_vat_total 
        cash_total_subtotal = cash_total_subtotal + cash_sale_recharge_grand_total 
        cash_total_amount_recieved = cash_total_received + cash_sale_recharge_amount_recieved 
        
        total_cash_sales_count = cash_sales.count() + recharge_cash_sales.count()
        
        cash_sale_recharge_count = recharge_cash_sales.count()
        cash_total_qty = cash_total_quantity + cash_sale_recharge_count
        ### Cash Sales End ###
        
        ### credit sales Start ####
        five_gallon_credit_sales = CustomerSupply.objects.filter(pk__in=five_gallon_supply,created_date__date=date,salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])
        other_credit_sales = CustomerSupply.objects.filter(pk__in=other_supply,created_date__date=date,salesman=salesman,amount_recieved__lte=0).exclude(customer__sales_type__in=["FOC","CASH COUPON"])
        
        # Aggregating for five_gallon_credit_sales
        five_gallon_credit_total_net_taxable = five_gallon_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        five_gallon_credit_total_vat = five_gallon_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        five_gallon_credit_total_subtotal = five_gallon_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        five_gallon_credit_total_received = five_gallon_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        five_gallon_credit_total_quantity = five_gallon_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0

        # Aggregating for other_credit_sales (corrected to use other_credit_sales)
        other_credit_total_net_taxable = other_credit_sales.aggregate(total_net_taxable=Sum('net_payable'))['total_net_taxable'] or 0
        other_credit_total_vat = other_credit_sales.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
        other_credit_total_subtotal = other_credit_sales.aggregate(total_subtotal=Sum('subtotal'))['total_subtotal'] or 0
        other_credit_total_received = other_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        other_credit_total_quantity = other_credit_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        # Combining both cash sales querysets
        credit_sales = five_gallon_credit_sales.union(other_credit_sales)
        credit_total_net_taxable = five_gallon_credit_total_net_taxable + other_credit_total_net_taxable
        credit_total_vat = five_gallon_credit_total_vat + other_credit_total_vat
        credit_total_subtotal = five_gallon_credit_total_subtotal + other_credit_total_subtotal
        credit_total_received = five_gallon_credit_total_received + other_credit_total_received
        credit_total_quantity = five_gallon_credit_total_quantity + other_credit_total_quantity
        
        recharge_credit_sales = CustomerCoupon.objects.filter(created_date__date=date,salesman=salesman,amount_recieved__lte=0)
        credit_sale_recharge_net_payeble = recharge_credit_sales.aggregate(total_net_amount=Sum('net_amount'))['total_net_amount'] or 0
        credit_sale_recharge_vat_total = 0
        credit_sale_recharge_grand_total = recharge_credit_sales.aggregate(total_grand_total=Sum('grand_total'))['total_grand_total'] or 0
        credit_sale_recharge_amount_recieved = recharge_credit_sales.aggregate(total_amount_recieved=Sum('amount_recieved'))['total_amount_recieved'] or 0
        
        credit_total_net_taxable = credit_total_net_taxable + credit_sale_recharge_net_payeble 
        credit_total_vat = credit_total_vat + credit_sale_recharge_vat_total 
        credit_total_subtotal = credit_total_subtotal + credit_sale_recharge_grand_total
        credit_total_amount_recieved = credit_total_received + credit_sale_recharge_amount_recieved
        
        total_credit_sales_count = credit_sales.count() + recharge_credit_sales.count()
        credit_sale_recharge_count = recharge_credit_sales.count()
        credit_total_qty = credit_total_quantity + credit_sale_recharge_count
        ### credit sales End ####
        
        # Coupon sales
        coupon_sales = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,customer__sales_type="CASH COUPON")
        manual_coupon_total = CustomerSupplyCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(Count('leaf'))['leaf__count']
        digital_coupon_total = CustomerSupplyDigitalCoupon.objects.filter(customer_supply__in=coupon_sales).aggregate(total_count=Sum('count'))['total_count'] or 0
        
        total_coupon_sales_count = coupon_sales.count()
        
        total_sales_count = total_cash_sales_count + total_credit_sales_count + total_coupon_sales_count
        coupon_total_qty = coupon_sales.aggregate(total_quantity=Sum('customersupplyitems__quantity'))['total_quantity'] or 0
        
        ### expenses ####
        expenses_instanses = Expense.objects.filter(expense_date=date,van=van_route.van)
        today_expense = expenses_instanses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
        
        ### suspense ###
        suspense_collections = SuspenseCollection.objects.filter(created_date__date=date,salesman=salesman)
        cash_sales_amount = suspense_collections.aggregate(total_cash_sale=Sum('cash_sale_amount'))['total_cash_sale'] or 0
        credit_sales_amount = suspense_collections.aggregate(total_credit_sale=Sum('credit_sale_amount'))['total_credit_sale'] or 0
        
        in_hand_amount = cash_sales_amount + credit_sales_amount
        today_payable = in_hand_amount - today_expense
        suspense_paid_amount = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
        suspense_balance_amount = today_payable - suspense_paid_amount
        
        # collection details
        dialy_collections = CollectionPayment.objects.filter(salesman_id=salesman)
        # credit outstanding
        # outstanding_credit_notes = Invoice.objects.filter(invoice_type="credit_invoice",customer__sales_staff=salesman).exclude(created_date__date__gt=date)
        outstanding_credit_notes_total_amount = OutstandingAmount.objects.filter(customer_outstanding__created_date__date__lte=date,customer_outstanding__product_type="amount",customer_outstanding__customer__routes=van_route.routes).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        dialy_colection_upto__yesterday = dialy_collections.filter(created_date__date__lt=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_total_amount = outstanding_credit_notes_total_amount - dialy_colection_upto__yesterday
        outstanding_credit_notes_received_amount = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        outstanding_credit_notes_balance = outstanding_credit_notes_total_amount - outstanding_credit_notes_received_amount
        outstanding_total_amount_collected = dialy_collections.filter(created_date__date=date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0

        
        # pending customers
        pending_bottle_customer_instances = CustomerSupply.objects.filter(created_date__date=date,salesman=salesman,allocate_bottle_to_pending__gt=0)
        five_gallon_rates = (
            CustomerSupplyItems.objects.filter(
                customer_supply__created_date__date=date,
                customer_supply__salesman_id=salesman,
                product__product_name="5 Gallon"
            )
            .values_list("rate", flat=True)  # ✅ SUPPLY RATE
            .distinct()
        )
        # 5 gallon rate based
        unique_amounts = set(five_gallon_rates)
        five_gallon_rate_wise_instances = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=date,customer_supply__salesman_id=salesman,product__product_name="5 Gallon")
        total_debit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved__gt=0).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_credit_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__amount_recieved=0).exclude(customer_supply__customer__sales_type__in=["FOC","CASH COUPON"]).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_coupon_amount_count = five_gallon_rate_wise_instances.filter(customer_supply__customer__sales_type="CASH COUPON").aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        # cash sales amount collected
        supply_amount_collected = cash_total_amount_recieved
        cash_sales_amount_collected = supply_amount_collected
        
        dialy_collections = dialy_collections.filter(created_date__date=date)
        credit_sales_amount_collected = dialy_collections.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        total_sales_amount_collected = cash_sales_amount_collected + credit_sales_amount_collected
        
        cheque_collection = CollectionPayment.objects.filter(payment_method="CHEQUE",created_date__date=date,salesman=salesman)
        no_of_collected_cheque = cheque_collection.count()
        collected_cheque_amount = cheque_collection.aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        
        balance_in_hand = total_sales_amount_collected - collected_cheque_amount - today_expense
        net_payble = total_sales_amount_collected - today_expense
        
        
        foc_customers = CustomerSupply.objects.filter(created_date__date=date, customer__sales_type='FOC', salesman=salesman) or CustomerSupply.objects.filter(created_date__date=date, salesman=salesman, allocate_bottle_to_free__gt=0)

        
        
    context = {
        'data_filter': data_filter,
        'salesman_id': salesman_id,
        'van_route': van_route,
        # visit statistics
        'new_customers_count': new_customers_count,
        'emergency_supply_count': emergency_supply_count,
        'visited_customers_count': visited_customers_count,
        'non_visited_count': non_visited_count,
        'planned_visit_count': planned_visit_count,
        'routes_instances': routes_instances,
        # stock report
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        'stock_report_total': stock_report_total,
        # pending customers
        'pending_bottle_customer_instances': pending_bottle_customer_instances,
        # Bottle Count
        'total_empty_bottles': total_empty_bottles,
        'total_supplied_bottles':total_supplied_bottles,
        'closing_stock_count': closing_stock_count,
        'damage_bottle_count': damage_bottle_count,
        'pending_bottle_count': pending_bottle_count,
        'total_count': total_count,
        #coupon book sale
        'customer_coupons':customer_coupons,
        # five gallon cash sales
        'five_gallon_cash_sales': five_gallon_cash_sales or [],
        'five_gallon_cash_total_net_taxable': five_gallon_cash_total_net_taxable,
        'five_gallon_cash_total_vat': five_gallon_cash_total_vat,
        'five_gallon_cash_total_subtotal': five_gallon_cash_total_subtotal,
        'five_gallon_cash_total_received': five_gallon_cash_total_received,
        'five_gallon_cash_total_quantity': five_gallon_cash_total_quantity,
        # other cash sales
        'other_cash_sales': other_cash_sales,
        'other_cash_total_net_taxable': other_cash_total_net_taxable,
        'other_cash_total_vat': other_cash_total_vat,
        'other_cash_total_subtotal': other_cash_total_subtotal,
        'other_cash_total_received': other_cash_total_received,
        'other_cash_total_quantity': other_cash_total_quantity,
        #cash sales
        'recharge_cash_sales': recharge_cash_sales,
        'cash_sale_recharge_net_payeble': cash_sale_recharge_net_payeble,
        'cash_sale_recharge_vat_total': cash_sale_recharge_vat_total,
        'cash_sale_recharge_grand_total': cash_sale_recharge_grand_total,
        'cash_sale_recharge_amount_recieved': cash_sale_recharge_amount_recieved,
        
        'cash_total_net_taxable':cash_total_net_taxable,
        'cash_total_vat':cash_total_vat,
        'cash_total_subtotal': cash_total_subtotal,
        'cash_total_amount_recieved': cash_total_amount_recieved,
        'cash_total_qty': cash_total_qty,
        # five gallon credit sales
        'five_gallon_credit_sales': five_gallon_credit_sales,
        'five_gallon_credit_total_net_taxable': five_gallon_credit_total_net_taxable,
        'five_gallon_credit_total_vat': five_gallon_credit_total_vat,
        'five_gallon_credit_total_subtotal': five_gallon_credit_total_subtotal,
        'five_gallon_credit_total_received': five_gallon_credit_total_received,
        'five_gallon_credit_total_quantity': five_gallon_credit_total_quantity,
        # credit sales
        'credit_sales': credit_sales,
        'recharge_credit_sales': recharge_credit_sales,
        'credit_sale_recharge_net_payeble': credit_sale_recharge_net_payeble,
        'credit_sale_recharge_vat_total': credit_sale_recharge_vat_total,
        'credit_sale_recharge_grand_total': credit_sale_recharge_grand_total,
        'credit_sale_recharge_amount_recieved': credit_sale_recharge_amount_recieved,
        
        'credit_total_net_taxable':credit_total_net_taxable,
        'credit_total_vat':credit_total_vat,
        'credit_total_subtotal':credit_total_subtotal,
        'credit_total_amount_recieved': credit_total_amount_recieved,
        'credit_total_qty': credit_total_qty,
        # coupon sales
        'coupon_sales': coupon_sales,
        'manual_coupon_total':manual_coupon_total,
        'digital_coupon_total':digital_coupon_total,
        'total_coupon_sales_count': total_coupon_sales_count,
        'coupon_total_qty':coupon_total_qty,
        # expenses
        'expenses_instanses': expenses_instanses,
        # suspense
        'in_hand_amount': in_hand_amount,
        'today_expense': today_expense, 
        'today_payable': today_payable,
        'suspense_paid_amount': suspense_paid_amount,
        'suspense_balance_amount': suspense_balance_amount,
        'outstanding_credit_notes_total_amount' : outstanding_credit_notes_total_amount,
        'outstanding_credit_notes_received_amount' : outstanding_credit_notes_received_amount,
        'outstanding_credit_notes_balance': outstanding_credit_notes_balance,
        'outstanding_total_amount_collected':outstanding_total_amount_collected,
        # 5 gallon rate based
        'five_gallon_rates': unique_amounts,
        'total_debit_amount_count': total_debit_amount_count,
        'total_credit_amount_count': total_credit_amount_count,
        'total_coupon_amount_count': total_coupon_amount_count,
        # dialy collections
        'dialy_collections': dialy_collections,
        # sales amount collected
        'cash_sales_amount_collected': cash_sales_amount_collected,
        'credit_sales_amount_collected': credit_sales_amount_collected,
        'total_sales_amount_collected': total_sales_amount_collected,
        'total_cash_sales_count': total_cash_sales_count,
        'total_credit_sales_count': total_credit_sales_count,
        'total_sales_count': total_sales_count,
        'no_of_collected_cheque': no_of_collected_cheque,
        'collected_cheque_amount': collected_cheque_amount,
        
        'balance_in_hand': balance_in_hand,
        'net_payble': net_payble,
        
        'filter_data': filter_data,
        # FOC customer
        'foc_customers':foc_customers,
        
        'filter_data': filter_data,
        'filter_date_formatted': date.strftime('%d-%m-%Y'),
    }
    
    return render(request, 'sales_management/new_dsr_summary_print.html', context)

from django.core.paginator import Paginator

def collection_list_view(request):
    
    filter_data = {}

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    query = request.GET.get("q")
    route_filter = request.GET.get('route_name')

    # Initialize queryset
    instances = CollectionPayment.objects.select_related('customer', 'salesman').all().order_by("-created_date")

    if route_filter:
        instances = instances.filter(customer__routes__route_name=route_filter)

    if query:
        instances = instances.filter(
            Q(customer__customer_name__icontains=query) |
            Q(receipt_number__icontains=query) |
            Q(payment_method__icontains=query)
        )
        
    
    today = datetime.datetime.today().date()
    if start_date and end_date:
        start = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        instances = instances.filter(created_date__gte=start, created_date__lt=end)
        filter_data['start_date'] = start_date
        filter_data['end_date'] = end_date
    

   
    # Fetch all routes for the dropdown
    route_li = RouteMaster.objects.all()

    # # Implement pagination (optional if using el_pagination_tags)
    # paginator = Paginator(instances, 20)  # Show 20 records per page
    # page_number = request.GET.get('page')
    # page_obj = paginator.get_page(page_number)

    context = {
        'instances': instances, 
        'today': today,
        'filter_data': filter_data,
        'route_li': route_li,
    }

    return render(request, 'sales_management/collection_list.html', context)

def delete_collection_payment(request, pk):
    
    # customer = get_object_or_404(Customers, customer_id=customer_id)
    
    # Retrieve all matching collection payments
    collection_payments = CollectionPayment.objects.filter(pk=pk)

    if not collection_payments.exists():
        return JsonResponse({"error": "No matching collection payment found."}, status=404)

    for collection_payment in collection_payments:
        # Retrieve related collection items for each collection payment
        collection_items = CollectionItems.objects.filter(collection_payment=collection_payment)
        
        for item in collection_items:
            payment_amount = item.amount_received
            invoice = item.invoice
            log_activity(request.user, f"Invoice {invoice} adjusted: Amount received reduced by {payment_amount}")
            # Adjust the invoice amount_received
            invoice.amout_recieved -= payment_amount
            if invoice.amout_recieved < 0:
                invoice.amout_recieved = 0
            invoice.invoice_status = 'non_paid'
            invoice.save()

            # Adjust the outstanding report if it exists
            if CustomerOutstandingReport.objects.filter(customer=invoice.customer, product_type="amount").exists():
                outstanding_instance = CustomerOutstandingReport.objects.get(customer=invoice.customer, product_type="amount")
                outstanding_instance.value += payment_amount
                outstanding_instance.save()

            # Delete each collection item
            item.delete()
        log_activity(request.user.id, f"Deleted collection payment: {collection_payment.receipt_number}")
        # Delete the collection payment after its items are handled
        collection_payment.delete()

    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Receipt and associated data successfully deleted and reversed.",
        "redirect": "true",
        "redirect_url": reverse('collection_list'),
    }
    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


def coupon_sales_report_view(request):
    # Retrieve filter parameters from GET request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sales_type = request.GET.get('sales_type')
    route_name = request.GET.get('route_name')

    # Set date range filter or use today's date if not provided
    if not (start_date and end_date):
        start_datetime = datetime.datetime.today().date()
        end_datetime = datetime.datetime.today().date()
    else:
        start_datetime = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end_datetime = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

    # Query coupon sales within the date range
    coupon_sales = CustomerCouponItems.objects.filter(
        customer_coupon__created_date__date__range=[start_datetime, end_datetime]
    )
    
    print("After date filter:", coupon_sales)
    if sales_type :
        coupon_sales = coupon_sales.filter(customer_coupon__customer__sales_type=sales_type)
    print("After sales_type filter:", coupon_sales)
    if route_name:
        coupon_sales = coupon_sales.filter(customer_coupon__customer__routes__route_name=route_name)
    print("After route_name filter:", coupon_sales)

    # Calculate total sums
    total_rate = coupon_sales.aggregate(total=Sum('rate'))['total'] or 0
    total_amount_collected = coupon_sales.aggregate(total=Sum('customer_coupon__amount_recieved'))['total'] or 0
    total_balance = coupon_sales.aggregate(total=Sum('customer_coupon__balance'))['total'] or 0
    total_per_leaf_rate = sum(coupon.get_per_leaf_rate() for coupon in coupon_sales if coupon.get_per_leaf_rate() is not None)



    # Fetch routes for dropdown
    routes_instances = RouteMaster.objects.all()

    # Prepare filter data for the template
    filter_data = {
        'start_date': start_date,
        'end_date': end_date,
        'sales_type': sales_type,
        'route_name': route_name
    }

    context = {
        'coupon_sales': coupon_sales,  # Pass the actual queryset
        'total_rate': total_rate,
        'total_amount_collected': total_amount_collected,
        'total_balance': total_balance,
        'total_per_leaf_rate':total_per_leaf_rate,
        'filter_data': filter_data,
        'routes_instances': routes_instances,
        'data_filter': any([start_date, end_date, sales_type, route_name])  # To show reset filter button
    }
    print("context",context)

    return render(request, 'sales_management/coupon_sales_report.html', context)

def coupon_sales_excel_view(request):
    # Retrieve filter parameters from GET request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sales_type = request.GET.get('sales_type')
    route_name = request.GET.get('route_name')

    # Set date range filter or use today's date if not provided
    if not (start_date and end_date):
        start_datetime = datetime.date.today()
        end_datetime = datetime.date.today()
    else:
        start_datetime = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end_datetime = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

    # Query coupon sales within the date range
    coupon_sales = CustomerCouponItems.objects.filter(
        customer_coupon__created_date__date__range=[start_datetime, end_datetime]
    )
    
    if sales_type:
        coupon_sales = coupon_sales.filter(customer_coupon__customer__sales_type=sales_type)
    
    if route_name:
        coupon_sales = coupon_sales.filter(customer_coupon__customer__routes__route_name=route_name)

    # Prepare data for Excel
    data = []
    for sale in coupon_sales:
        data.append({
            'Coupon Method': sale.customer_coupon.coupon_method,
            'Book Number': sale.coupon.book_num,
            'Customer Name': sale.customer_coupon.customer.customer_name,
            'Customer ID': sale.customer_coupon.customer.custom_id,
            'Sales Type': sale.customer_coupon.customer.sales_type,
            'Route Name': sale.customer_coupon.customer.routes.route_name,
            'No of Leaflets': sale.coupon.no_of_leaflets,
            'Used Leaflets': sale.get_used_leaflets(),
            'Balance Coupons': sale.get_unused_leaflets(),
            'Rate': sale.rate,
            'Per Leaf Rate': sale.get_per_leaf_rate(),
            'Amount Collected': sale.customer_coupon.amount_recieved,
            'Balance': sale.customer_coupon.balance
        })

    # Create DataFrame
    df = pd.DataFrame(data)

    # Calculate footer (sum for numeric columns)
    footer = {
        'Coupon Method': 'Total',
        'Book Number': '',
        'Customer Name': '',
        'Customer ID': '',
        'Sales Type': '',
        'Route Name': '',
        'No of Leaflets': '',
        'Used Leaflets': df['Used Leaflets'].sum(),
        'Balance Coupons': df['Balance Coupons'].sum(),
        'Rate': df['Rate'].sum(),
        'Per Leaf Rate': df['Per Leaf Rate'].sum(),  # Assuming you want to sum per leaf rate
        'Amount Collected': df['Amount Collected'].sum(),
        'Balance': df['Balance'].sum()
    }

    # Convert the footer to a DataFrame and concatenate with the original DataFrame
    footer_df = pd.DataFrame([footer])
    df = pd.concat([df, footer_df], ignore_index=True)

    # Write DataFrame to Excel buffer
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Coupon Sales', index=False)

    # Prepare HTTP response with Excel file
    filename = f"Coupon_Sales_Report_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def coupon_sales_print_view(request):
    # Retrieve filter parameters from GET request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sales_type = request.GET.get('sales_type')
    route_name = request.GET.get('route_name')

    # Set date range filter or use today's date if not provided
    if not (start_date and end_date):
        start_datetime = datetime.datetime.today().date()
        end_datetime = datetime.datetime.today().date()
    else:
        start_datetime = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end_datetime = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

    # Query coupon sales within the date range
    coupon_sales = CustomerCouponItems.objects.filter(
        customer_coupon__created_date__date__range=[start_datetime, end_datetime]
    )
    
    if sales_type:
        coupon_sales = coupon_sales.filter(customer_coupon__customer__sales_type=sales_type)
    
    if route_name:
        coupon_sales = coupon_sales.filter(customer_coupon__customer__routes__route_name=route_name)

    # Calculate total sums
    total_rate = coupon_sales.aggregate(total=Sum('rate'))['total'] or 0
    total_amount_collected = coupon_sales.aggregate(total=Sum('customer_coupon__amount_recieved'))['total'] or 0
    total_balance = coupon_sales.aggregate(total=Sum('customer_coupon__balance'))['total'] or 0
    total_per_leaf_rate = sum(coupon.get_per_leaf_rate() for coupon in coupon_sales if coupon.get_per_leaf_rate() is not None)

    # Prepare filter data for the template
    filter_data = {
        'start_date': start_date,
        'end_date': end_date,
        'sales_type': sales_type,
        'route_name': route_name
    }

    context = {
        'coupon_sales': coupon_sales,  # Pass the actual queryset
        'total_rate': total_rate,
        'total_amount_collected': total_amount_collected,
        'total_balance': total_balance,
        'total_per_leaf_rate': total_per_leaf_rate,
        'filter_data': filter_data,
    }

    return render(request, 'sales_management/coupon_sales_print_report.html', context)

def receipt_list_view(request):
    
    filter_data = {}

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    query = request.GET.get("q")
    route_filter = request.GET.get('route_name')

    receipts = Receipt.objects.all().order_by('-created_date')
    if route_filter:
        receipts = receipts.filter(customer__routes__route_name=route_filter)
        filter_data['route_name'] = route_filter


    if query:
        receipts = receipts.filter(
            Q(customer__customer_name__icontains=query) |
            Q(receipt_number__icontains=query) |
            Q(customer__custom_id__icontains=query)|
            Q(invoice_number__icontains=query)
        )
        
    
    today = datetime.datetime.today().date()
    if start_date and end_date:
        receipts = receipts.filter(created_date__range=[start_date, end_date])
        filter_data['start_date'] = start_date
        filter_data['end_date'] = end_date
    

   
    # Fetch all routes for the dropdown
    route_li = RouteMaster.objects.all()

    context = {
        'receipts': receipts, 
        'today': today,
        'filter_data': filter_data,
        'route_li': route_li,
    }

    return render(request, 'sales_management/receipt_list.html', context)

def receipt_list_print(request):
    filter_data = {}

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    query = request.GET.get("q")
    route_filter = request.GET.get('route_name')

    receipts = Receipt.objects.all().order_by('-created_date')
    if route_filter:
        receipts = receipts.filter(customer__routes__route_name=route_filter)
        filter_data['route_name'] = route_filter


    if query:
        receipts = receipts.filter(
            Q(customer__customer_name__icontains=query) |
            Q(receipt_number__icontains=query) |
            Q(customer__custom_id__icontains=query)|
            Q(invoice_number__icontains=query)
        )
        
    
    today = datetime.datetime.today().date()
    if start_date and end_date:
        receipts = receipts.filter(created_date__range=[start_date, end_date])
        filter_data['start_date'] = start_date
        filter_data['end_date'] = end_date
    

   
    # Fetch all routes for the dropdown
    route_li = RouteMaster.objects.all()

    context = {
        'receipts': receipts, 
        'today': today,
        'filter_data': filter_data,
        'route_li': route_li,
    }

    return render(request, 'sales_management/receipt_list_print.html', context)


def receipt_list_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")
    route_filter = request.GET.get('route_name')

    receipts = Receipt.objects.all().order_by('-created_date')
    
    if route_filter:
        receipts = receipts.filter(customer__routes__route_name=route_filter)

    if query:
        receipts = receipts.filter(
            Q(customer__customer_name__icontains=query) |
            Q(receipt_number__icontains=query) |
            Q(customer__custom_id__icontains=query) |
            Q(invoice_number__icontains=query)
        )

    if start_date and end_date:
        receipts = receipts.filter(created_date__range=[start_date, end_date])

    # Prepare data for the Excel file
    data = []
    for receipt in receipts:
        data.append({
            'Date Time': receipt.created_date.astimezone(timezone.utc).replace(tzinfo=None).strftime('%d-%m-%Y'),
            'Customer Name': receipt.customer.customer_name,
            'Customer Code': receipt.customer.custom_id,
            'Building Name': receipt.customer.building_name,
            'Room No': receipt.customer.door_house_no,
            'Route': receipt.customer.routes.route_name,  # Adjust if needed
            'Receipt Number': receipt.receipt_number,
            'Amount': receipt.amount_received,
            'Against Invoice': receipt.invoice_number,
        })

    # Create a Pandas DataFrame
    df = pd.DataFrame(data)

    # Create an Excel file in memory
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Receipts', index=False)

    # Prepare response
    buffer.seek(0)
    filename = "Receipt_List.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def delete_receipt(request, receipt_number, customer_id):
    
    response_data = {
        "status": "false",
        "title": "Error",
        "message": "An error occurred during deletion.",
    }

    receipt = get_object_or_404(Receipt, receipt_number=receipt_number, customer=customer_id)
    transaction_type = receipt.transaction_type
    invoice_number = receipt.invoice_number
    

    
    if transaction_type =="supply":
        try:
            
            customer_outstanding = CustomerOutstanding.objects.filter(
                customer=receipt.customer,
                product_type="amount"
            ).first()
            if not customer_outstanding:
                customer_outstanding = CustomerOutstanding.objects.create(
                    customer=receipt.customer,
                    product_type="amount",
                    invoice_no=invoice_number,
                    created_by=request.user,
                    created_date=timezone.now()
                )
            
            # Proceed with outstanding_amount logic
            outstanding_amount, created = OutstandingAmount.objects.get_or_create(
                customer_outstanding=customer_outstanding,
                customer_outstanding__customer=receipt.customer,
                customer_outstanding__invoice_no=receipt.invoice_number,
                defaults={'amount': 0} 
            )

            if created or outstanding_amount.amount == 0:
                outstanding_amount.amount += receipt.amount_received
                outstanding_amount.save()

            outstanding_instance, created = CustomerOutstandingReport.objects.get_or_create(
                customer=receipt.customer,
                product_type="amount",
                defaults={'value': 0}
            )

            outstanding_instance.value += receipt.amount_received
            outstanding_instance.save()
            
            invoice = Invoice.objects.get(invoice_no=invoice_number, customer=receipt.customer)
            invoice.amout_recieved = 0  
            invoice.amout_total = receipt.amount_received  
            invoice.invoice_status = "non_paid"
            invoice.save()

        except CustomerOutstandingReport.DoesNotExist:
            response_data["message"] = "Customer outstanding report not found."
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')


            
    elif transaction_type == "collection":
        try:
            customer_outstanding = CustomerOutstanding.objects.filter(
                customer=receipt.customer,
                product_type="amount"
            ).first()
            if not customer_outstanding:
                customer_outstanding = CustomerOutstanding.objects.create(
                    customer=receipt.customer,
                    product_type="amount",
                    invoice_no=invoice_number,
                    created_by=request.user,
                    created_date=timezone.now()
                )
            
            # Proceed with outstanding_amount logic
            outstanding_amount, created = OutstandingAmount.objects.get_or_create(
                customer_outstanding=customer_outstanding,
                customer_outstanding__customer=receipt.customer,
                customer_outstanding__invoice_no=receipt.invoice_number,
                defaults={'amount': 0} 
            )

            if created or outstanding_amount.amount == 0:
                outstanding_amount.amount += receipt.amount_received
                outstanding_amount.save()

            outstanding_instance, created = CustomerOutstandingReport.objects.get_or_create(
                customer=receipt.customer,
                product_type="amount",
                defaults={'value': 0}
            )

            outstanding_instance.value += receipt.amount_received
            outstanding_instance.save()

        except CustomerOutstandingReport.DoesNotExist:
            response_data["message"] = "Customer outstanding report not found."
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')

        try:
            collection_payment = CollectionPayment.objects.get(receipt_number=receipt_number, customer=customer_id)
            collection_items = CollectionItems.objects.filter(collection_payment=collection_payment)
            
            for item in collection_items:
                invoices = item.invoice
                
                if invoices:
                    
                    invoice = Invoice.objects.get(invoice_no=invoices.invoice_no, customer=receipt.customer)
                        
                    invoice.amout_recieved = 0 
                    invoice.amout_total = item.amount_received 
                    invoice.invoice_status = "non_paid" 
                    invoice.save()
            
        except CollectionPayment.DoesNotExist:
            response_data["message"] = "Collection payment not found."
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
        
    elif transaction_type=="coupon_rechange":
        try:
            customer_outstanding = CustomerOutstanding.objects.filter(
                customer=receipt.customer,
                product_type="amount"
            ).first()
            if not customer_outstanding:
                customer_outstanding = CustomerOutstanding.objects.create(
                    customer=receipt.customer,
                    product_type="amount",
                    invoice_no=invoice_number,
                    created_by=request.user,
                    created_date=timezone.now()
                )
            
            # Proceed with outstanding_amount logic
            outstanding_amount, created = OutstandingAmount.objects.get_or_create(
                customer_outstanding=customer_outstanding,
                customer_outstanding__customer=receipt.customer,
                customer_outstanding__invoice_no=receipt.invoice_number,
                defaults={'amount': 0} 
            )

            if created or outstanding_amount.amount == 0:
                outstanding_amount.amount += receipt.amount_received
                outstanding_amount.save()

            outstanding_instance, created = CustomerOutstandingReport.objects.get_or_create(
                customer=receipt.customer,
                product_type="amount",
                defaults={'value': 0}
            )

            outstanding_instance.value += receipt.amount_received
            outstanding_instance.save()

        except CustomerOutstandingReport.DoesNotExist:
            response_data["message"] = "Customer outstanding report not found."
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
 
        try:
            invoice = Invoice.objects.get(invoice_no=invoice_number, customer_id=customer_id)
            invoice.amout_recieved = 0  
            invoice.amout_total = receipt.amount_received  
            invoice.invoice_status = "non_paid"
            invoice.save()
            
            
        except Invoice.DoesNotExist:
            response_data["message"] = "Invoice not found."
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    

    
    receipt.delete()
    
    log_activity(request.user, f"Successfully deleted receipt {receipt_number} and reversed associated transactions.")
        
    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Receipt and associated data successfully deleted and reversed.",
        "redirect": "true",
        "redirect_url": reverse('receipt_list'),
    }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

# from django.db.models.functions import ExtractMonth

# def monthly_sales_report(request):
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')

#     if start_date:
#         start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
#     if end_date:
#         end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

#     months = range(1, 13)
    
#     report_data = {}
    
#     customer_sales = (
#         CustomerSupply.objects
#         .values('customer__customer_name', 'created_date__month')
#         .annotate(total_sales=Sum('grand_total'))
#         .order_by('customer__customer_name', 'created_date__month')
#     )
    
#     if start_date and end_date:
#         customer_sales = customer_sales.filter(created_date__range=(start_date, end_date))

#     for entry in customer_sales:
#         customer_name = entry['customer__customer_name']
#         month = entry['created_date__month']
#         total_sales = entry['total_sales'] or 0  
        
#         if customer_name not in report_data:
#             report_data[customer_name] = {month: 0 for month in months}
        
#         report_data[customer_name][month] += total_sales  

#     context = {
#         'report_data': report_data,
#         'months': months,
#         'filter_data': {
#             'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
#             'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
#         },
#     }
    
#     return render(request, 'sales_management/monthly_sales_report.html', context)

def monthly_sales_report(request):
    month = int(request.GET.get('month', datetime.datetime.now().month))
    year = datetime.datetime.now().year  
    monthly_sales = (
        CustomerSupply.objects
        .filter(created_date__year=year, created_date__month=month)
        .values('salesman__id', 'salesman__username')
        .annotate(
            total_grand_total=Sum('grand_total'),
            total_discount=Sum('discount'),
            total_net_payable=Sum('net_payable'),
            total_vat=Sum('vat'),
            total_subtotal=Sum('subtotal'),
            total_amount_received=Sum('amount_recieved'),
        )
        .order_by('salesman__username')
    )

    sales_data = []
    for sale in monthly_sales:
        salesman_id = sale['salesman__id']

        total_supply_qty = CustomerSupplyItems.objects.filter(
            customer_supply__salesman_id=salesman_id,
            customer_supply__created_date__year=year,
            customer_supply__created_date__month=month
        ).aggregate(total_qty=Sum('quantity'))['total_qty'] or 0

        manual_coupon_count = CustomerSupplyCoupon.objects.filter(
            customer_supply__salesman_id=salesman_id,
            customer_supply__created_date__year=year,
            customer_supply__created_date__month=month
        ).aggregate(total_leaf=Count('leaf') + Count('free_leaf'))['total_leaf'] or 0

        digital_coupon_count = CustomerSupplyDigitalCoupon.objects.filter(
            customer_supply__salesman_id=salesman_id,
            customer_supply__created_date__year=year,
            customer_supply__created_date__month=month
        ).aggregate(total_count=Sum('count'))['total_count'] or 0

        sale['total_supply_qty'] = total_supply_qty
        sale['total_coupon_received'] = {
            "manual_coupon": manual_coupon_count,
            "digital_coupon": digital_coupon_count
        }

        sales_data.append(sale)

    month_choices = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]

    context = {
        'month': month,
        'sales_data': sales_data,
        'month_choices': month_choices,  
    }
    return render(request, 'sales_management/salesman_monthly_sales.html', context)

import calendar
def monthly_sales_report_print(request):
    month = int(request.GET.get('month', datetime.datetime.now().month))
    year = datetime.datetime.now().year  

    month_name = calendar.month_name[month]

    monthly_sales = (
        CustomerSupply.objects
        .filter(created_date__year=year, created_date__month=month)
        .values('salesman__id', 'salesman__username')
        .annotate(
            total_grand_total=Sum('grand_total'),
            total_discount=Sum('discount'),
            total_net_payable=Sum('net_payable'),
            total_vat=Sum('vat'),
            total_subtotal=Sum('subtotal'),
            total_amount_received=Sum('amount_recieved'),
        )
        .order_by('salesman__username')
    )

    sales_data = []
    for sale in monthly_sales:
        salesman_id = sale['salesman__id']

        total_supply_qty = CustomerSupplyItems.objects.filter(
            customer_supply__salesman_id=salesman_id,
            customer_supply__created_date__year=year,
            customer_supply__created_date__month=month
        ).aggregate(total_qty=Sum('quantity'))['total_qty'] or 0

        manual_coupon_count = CustomerSupplyCoupon.objects.filter(
            customer_supply__salesman_id=salesman_id,
            customer_supply__created_date__year=year,
            customer_supply__created_date__month=month
        ).aggregate(total_leaf=Count('leaf') + Count('free_leaf'))['total_leaf'] or 0

        digital_coupon_count = CustomerSupplyDigitalCoupon.objects.filter(
            customer_supply__salesman_id=salesman_id,
            customer_supply__created_date__year=year,
            customer_supply__created_date__month=month
        ).aggregate(total_count=Sum('count'))['total_count'] or 0

        sale['total_supply_qty'] = total_supply_qty
        sale['total_coupon_received'] = {
            "manual_coupon": manual_coupon_count,
            "digital_coupon": digital_coupon_count
        }

        sales_data.append(sale)

    context = {
        'month': month_name,  
        'sales_data': sales_data,
    }
    return render(request, 'sales_management/salesman_monthly_sales_print.html', context)



def detailed_sales_report(request):

    current_date = datetime.datetime.now()
    default_from_date = current_date.replace(day=1)
    default_to_date = (default_from_date + timedelta(days=1))

    # Parse from_date and to_date from query params
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    route_name = request.GET.get('route_name') or None

    try:
        start_date = make_aware(datetime.datetime.strptime(from_date_str, "%Y-%m-%d")) if from_date_str else make_aware(default_from_date)
        end_date = make_aware(datetime.datetime.strptime(to_date_str, "%Y-%m-%d")) if to_date_str else make_aware(default_to_date)
    except ValueError:
        start_date, end_date = make_aware(default_from_date), make_aware(default_to_date)

    # year_choices = [(str(year), str(year)) for year in range(current_date.year - 5, current_date.year + 1)]
    # month_choices = [(f"{month:02}", calendar.month_name[month]) for month in range(1, 13)]
    
    all_routes = RouteMaster.objects.all()

    report_routes = all_routes
    if route_name:
        report_routes = report_routes.filter(route_name=route_name)

    return render(request, 'sales_management/route_sales_report.html', {
        'routes': report_routes,
        'all_routes': all_routes,
        # 'month_choices': month_choices,
        # 'year_choices': year_choices,
        'selected_year': start_date.year,
        'selected_month': f"{start_date.month:02}",
        'start_date': start_date,
        'end_date': end_date,
        'from_date': start_date.date().isoformat(),
        'to_date': end_date.date().isoformat(),
    })
   
def print_sales_report(request):
    current_year = datetime.datetime.now().year
    current_month = datetime.datetime.now().month

    # Choices for dropdowns
    year_choices = [(str(year), str(year)) for year in range(current_year - 5, current_year + 1)]
    month_choices = [(f"{month:02}", calendar.month_name[month]) for month in range(1, 13)]

    # GET Params
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    try:
        if from_date_str and to_date_str:
            start_date = make_aware(datetime.datetime.strptime(from_date_str, "%Y-%m-%d"))
            end_date = make_aware(datetime.datetime.strptime(to_date_str, "%Y-%m-%d"))
            selected_year = start_date.year
            selected_month = f"{start_date.month:02}"
        else:
            selected_year = selected_year or str(current_year)
            selected_month = selected_month or str(current_month).zfill(2)
            start_date = datetime.datetime.strptime(f"{selected_year}-{selected_month}-01", "%Y-%m-%d")
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
            start_date = make_aware(start_date)
            end_date = make_aware(end_date)
    except ValueError as e:
        start_date, end_date = None, None
        print(f"Error parsing dates: {e}")

    routes = RouteMaster.objects.all()

    return render(request, 'sales_management/print_sales_report.html', {
        'routes': routes,
        'month_choices': month_choices,
        'year_choices': year_choices,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'from_date': start_date.date().isoformat() if start_date else '',
        'to_date': end_date.date().isoformat() if end_date else '',
        'start_date': start_date,
        'end_date': end_date,
    })
    
def routewise_sales_report(request, route_id):
    today = datetime.datetime.now()
    current_year = today.year
    current_month = today.month

    year_choices = [(str(year), str(year)) for year in range(current_year - 5, current_year + 1)]
    month_choices = [(f"{month:02}", calendar.month_name[month]) for month in range(1, 13)]

    selected_year = request.GET.get('year', str(current_year))
    selected_month = request.GET.get('month') or f"{current_month:02}"  

    try:
        start_date = datetime.datetime.strptime(f"{selected_year}-{selected_month}-01", "%Y-%m-%d")
        next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_date = next_month - timedelta(days=1)

        start_date, end_date = make_aware(start_date), make_aware(end_date)
    except ValueError as e:
        start_date, end_date = None, None

    route = RouteMaster.objects.get(route_id=route_id)
    
    van_route = Van_Routes.objects.filter(routes=route_id).first()
    
    coupon_items_instances = CouponType.objects.all()
    
    if start_date and end_date:
        date_list = []
        current = start_date
        while current <= end_date:
            date_list.append(current)
            current += timedelta(days=1)
    else:
        date_list = []

    return render(request, 'sales_management/routewise_sales_report.html', {
        'route': route,
        'van_route': van_route,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'start_date': start_date,
        'end_date': end_date,
        'year_choices': year_choices,
        'month_choices': month_choices,
        'current_year': current_year,  
        'current_month': f"{current_month:02}",
        'coupon_items_instances': coupon_items_instances,
        'coupon_items_instances_length': coupon_items_instances.count(),
        'date_list': date_list,
    })
    
def print_routewise_sales_report(request, route_id):
    today = datetime.datetime.now()
    current_year = today.year
    current_month = today.month

    year_choices = [(str(year), str(year)) for year in range(current_year - 5, current_year + 1)]
    month_choices = [(f"{month:02}", calendar.month_name[month]) for month in range(1, 13)]

    selected_year = request.GET.get('year', str(current_year))
    selected_month = request.GET.get('month') or f"{current_month:02}"  

    try:
        start_date = datetime.datetime.strptime(f"{selected_year}-{selected_month}-01", "%Y-%m-%d")
        next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_date = next_month - timedelta(days=1)

        start_date, end_date = make_aware(start_date), make_aware(end_date)
    except ValueError as e:
        start_date, end_date = None, None
    
    route = RouteMaster.objects.get(route_id=route_id)
    
    van_route = Van_Routes.objects.filter(routes=route_id).first()

    return render(request, 'sales_management/print_routewise_sales_report.html', {
        'route': route,
        'van_route': van_route,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'start_date': start_date,
        'end_date': end_date,
        'year_choices': year_choices,
        'month_choices': month_choices,
        'current_year': current_year, 
        'current_month': f"{current_month:02}", 
    })
    
def offload_list(request):
    # Get filter values from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    route_name = request.GET.get('route_name')
    query = request.GET.get('q')

    # Parse dates and handle time zone aware filtering
    try:
        if start_date:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        if end_date:
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        start_date = None
        end_date = None

    # Start with full querysets
    offloads = Offload.objects.select_related('salesman', 'van', 'product').order_by('-created_date')
    offload_coupons = OffloadCoupon.objects.select_related('salesman', 'van', 'coupon').order_by('-created_date')

    # Apply date filtering if both dates are provided
    if start_date and end_date:
        offloads = offloads.filter(created_date__date__range=(start_date.date(), end_date.date()))
        offload_coupons = offload_coupons.filter(created_date__date__range=(start_date.date(), end_date.date()))

    # Apply route name filtering
    if route_name:
        offloads = offloads.filter(van__van_master__routes__route_name=route_name)
        offload_coupons = offload_coupons.filter(van__van_master__routes__route_name=route_name)

    # Apply search query filtering
    if query:
        offloads = offloads.filter(
            Q(van__van_make__icontains=query) |
            Q(salesman__first_name__icontains=query) |
            Q(salesman__last_name__icontains=query) |
            Q(product__product_name__icontains=query) |
            Q(stock_type__icontains=query)
        )
        offload_coupons = offload_coupons.filter(
            Q(van__van_make__icontains=query) |
            Q(salesman__first_name__icontains=query) |
            Q(salesman__last_name__icontains=query) |
            Q(coupon__book_num__icontains=query) |
            Q(stock_type__icontains=query)
        )

    # Prepare combined data for the template
    offloaded_data = []

    for offload in offloads:
        offloaded_data.append({
            'created_date': offload.created_date,
            'salesman': offload.salesman.get_fullname,
            'van': offload.van.van_make,
            'product_name': offload.product.product_name,
            'quantity': offload.quantity,
            'stock_type': offload.stock_type,
            'entry_type': 'product',
            'route_name': offload.van.get_vans_routes(),
        })

    for coupon in offload_coupons:
        offloaded_data.append({
            'created_date': coupon.created_date,
            'salesman': coupon.salesman.get_fullname,
            'van': coupon.van.van_make,
            'product_name': coupon.coupon.book_num,
            'quantity': coupon.quantity,
            'stock_type': coupon.stock_type,
            'entry_type': 'coupon',
            'route_name': coupon.van.get_vans_routes(),
        })

    # Fetch distinct van routes for the dropdown
    van_routes = Van_Routes.objects.select_related('routes').distinct()

    # Pass data to the template
    return render(request, 'sales_management/offload_list.html', {
        'offloaded_data': offloaded_data,
        'filter_data': {
            'filter_date_from': start_date.strftime('%Y-%m-%d') if start_date else '',
            'filter_date_to': end_date.strftime('%Y-%m-%d') if end_date else '',
            'route_name': route_name or '',
            'q': query or '',
        },
        'van_routes': van_routes,
    })

def offload_list_print(request):
        # Get filter values from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    route_name = request.GET.get('route_name')
    query = request.GET.get('q')

    # Parse dates and handle time zone aware filtering
    try:
        if start_date:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        if end_date:
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        start_date = None
        end_date = None

    # Start with full querysets
    offloads = Offload.objects.select_related('salesman', 'van', 'product').order_by('-created_date')
    offload_coupons = OffloadCoupon.objects.select_related('salesman', 'van', 'coupon').order_by('-created_date')

    # Apply date filtering if both dates are provided
    if start_date and end_date:
        offloads = offloads.filter(created_date__date__range=(start_date.date(), end_date.date()))
        offload_coupons = offload_coupons.filter(created_date__date__range=(start_date.date(), end_date.date()))

    # Apply route name filtering
    if route_name:
        offloads = offloads.filter(van__van_master__routes__route_name=route_name)
        offload_coupons = offload_coupons.filter(van__van_master__routes__route_name=route_name)

    # Apply search query filtering
    if query:
        offloads = offloads.filter(
            Q(van__van_make__icontains=query) |
            Q(salesman__first_name__icontains=query) |
            Q(salesman__last_name__icontains=query) |
            Q(product__product_name__icontains=query) |
            Q(stock_type__icontains=query)
        )
        offload_coupons = offload_coupons.filter(
            Q(van__van_make__icontains=query) |
            Q(salesman__first_name__icontains=query) |
            Q(salesman__last_name__icontains=query) |
            Q(coupon__book_num__icontains=query) |
            Q(stock_type__icontains=query)
        )

    # Prepare combined data for the template
    offloaded_data = []

    for offload in offloads:
        offloaded_data.append({
            'created_date': offload.created_date,
            'salesman': offload.salesman.get_fullname,
            'van': offload.van.van_make,
            'product_name': offload.product.product_name,
            'quantity': offload.quantity,
            'stock_type': offload.stock_type,
            'entry_type': 'product',
            'route_name': offload.van.get_vans_routes(),
        })

    for coupon in offload_coupons:
        offloaded_data.append({
            'created_date': coupon.created_date,
            'salesman': coupon.salesman.get_fullname,
            'van': coupon.van.van_make,
            'product_name': coupon.coupon.book_num,
            'quantity': coupon.quantity,
            'stock_type': coupon.stock_type,
            'entry_type': 'coupon',
            'route_name': coupon.van.get_vans_routes(),
        })

    # Fetch distinct van routes for the dropdown
    van_routes = Van_Routes.objects.select_related('routes').distinct()

    # Pass data to the template
    return render(request, 'sales_management/offload_list_print.html', {
        'offloaded_data': offloaded_data,
        'filter_data': {
            'filter_date_from': start_date.strftime('%Y-%m-%d') if start_date else '',
            'filter_date_to': end_date.strftime('%Y-%m-%d') if end_date else '',
            'route_name': route_name or '',
            'q': query or '',
        },
        'van_routes': van_routes,
    })
    
def download_offload_excel(request):
    # Get filter values from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    route_name = request.GET.get('route_name')
    query = request.GET.get('q')

    # Parse dates
    try:
        if start_date:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        if end_date:
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        start_date = None
        end_date = None

    # Query data
    offloads = Offload.objects.select_related('salesman', 'van', 'product').order_by('-created_date')
    offload_coupons = OffloadCoupon.objects.select_related('salesman', 'van', 'coupon').order_by('-created_date')

    if start_date and end_date:
        offloads = offloads.filter(created_date__date__range=(start_date.date(), end_date.date()))
        offload_coupons = offload_coupons.filter(created_date__date__range=(start_date.date(), end_date.date()))

    if route_name:
        offloads = offloads.filter(van__van_master__routes__route_name=route_name)
        offload_coupons = offload_coupons.filter(van__van_master__routes__route_name=route_name)

    if query:
        offloads = offloads.filter(
            Q(van__van_make__icontains=query) |
            Q(salesman__first_name__icontains=query) |
            Q(salesman__last_name__icontains=query) |
            Q(product__product_name__icontains=query) |
            Q(stock_type__icontains=query)
        )
        offload_coupons = offload_coupons.filter(
            Q(van__van_make__icontains=query) |
            Q(salesman__first_name__icontains=query) |
            Q(salesman__last_name__icontains=query) |
            Q(coupon__book_num__icontains=query) |
            Q(stock_type__icontains=query)
        )

    # Create the Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Offloaded Data"

    # Add headers
    sheet.append([
        'Created Date', 'Salesman', 'Van','Route Name', 'Product Name',
        'Quantity', 'Stock Type'
    ])

    # Add data to the Excel sheet
    for offload in offloads:
        # Remove timezone information and format the datetime
        created_date = offload.created_date.replace(tzinfo=None) if offload.created_date else None
        formatted_created_date = created_date.strftime('%d-%m-%Y') if created_date else None  # Format as d-m-Y

        sheet.append([
            formatted_created_date,
            offload.salesman.get_fullname(),  # Call the method to get the full name
            offload.van.van_make,
            offload.van.get_vans_routes(),
            offload.product.product_name,
            offload.quantity,
            offload.stock_type,
        ])

    for coupon in offload_coupons:
        # Remove timezone information and format the datetime
        created_date = coupon.created_date.replace(tzinfo=None) if coupon.created_date else None
        formatted_created_date = created_date.strftime('%d-%m-%Y') if created_date else None  # Format as d-m-Y

        sheet.append([
            formatted_created_date,
            coupon.salesman.get_fullname(),  # Call the method to get the full name
            coupon.van.van_make,
            coupon.van.get_vans_routes(),
            coupon.coupon.book_num,
            coupon.quantity,
            coupon.stock_type,
        ])



    # Prepare the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    start_date_str = start_date.strftime('%Y-%m-%d') if start_date else 'start'
    end_date_str = end_date.strftime('%Y-%m-%d') if end_date else 'end'
    response['Content-Disposition'] = f'attachment; filename=Offloaded_detail_{start_date_str}_to_{end_date_str}.xlsx'

    # Save the workbook to the response
    workbook.save(response)
    return response


def todays_cash_sales(request):
    # Get today's date
    today_date = now().date()
    
    # Filter today's supplies and cash sales
    todays_supply_instances = CustomerSupply.objects.filter(created_date__date=today_date)
    supply_cash_sales_instances = todays_supply_instances.filter(
        amount_recieved__gt=0
    ).exclude(customer__sales_type="CASH COUPON")
    # Calculate totals
    total_supply_qty = sum(sale.get_total_supply_qty() for sale in supply_cash_sales_instances)
    total_grand_total = sum(sale.grand_total for sale in supply_cash_sales_instances)
    total_net_payable = sum(sale.net_payable for sale in supply_cash_sales_instances)
    total_cash_received = sum(sale.amount_recieved for sale in supply_cash_sales_instances)

    # Pass the data to the template
    context = {
        'supply_cash_sales_instances': supply_cash_sales_instances,
        'total_supply_qty': total_supply_qty,
        'total_grand_total': total_grand_total,
        'total_net_payable': total_net_payable,
        'total_cash_received': total_cash_received,
    }
    return render(request, 'master/dashboard/todays_cash_sales.html', context)

def todays_credit_sales(request):
    today = timezone.now().date()
    
    # Filter today's supply instances
    todays_supply_instances = CustomerSupply.objects.filter(created_date__date=today)
    supply_credit_sales_instances = todays_supply_instances.filter(
        amount_recieved__lte=0
    ).exclude(customer__sales_type__in=["FOC", "CASH COUPON"])

    # Calculate totals
    total_supply_qty = sum(sale.get_total_supply_qty() for sale in supply_credit_sales_instances)
    total_grand_total = sum(sale.grand_total for sale in supply_credit_sales_instances)
    total_net_payable = sum(sale.net_payable for sale in supply_credit_sales_instances)
    total_amount_received = sum(sale.amount_recieved for sale in supply_credit_sales_instances)

    context = {
        "supply_credit_sales_instances": supply_credit_sales_instances,
        "total_supply_qty": total_supply_qty,
        "total_grand_total": total_grand_total,
        "total_net_payable": total_net_payable,
        "total_amount_received": total_amount_received,
    }
    return render(request, "master/dashboard/todays_credit_sales.html", context)


def cheque_collections_view(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    route_name = request.GET.get("route_name")
    search_query = request.GET.get("q")

    collections = CollectionPayment.objects.filter(payment_method="cheque")
  
    if start_date:
        collections = collections.filter(created_date__gte=start_date)
    if end_date:
        collections = collections.filter(created_date__lte=end_date)
    if route_name:
        collections = collections.filter(customer__routes__route_name__iexact=route_name)
    if search_query:
        collections = collections.filter(
            Q(customer__customer_name__icontains=search_query) |
            Q(customer__custom_id__icontains=search_query) |
            Q(receipt_number__icontains=search_query)
        )
    
    collection_list = []
    for collection in collections:
        # Initialize invoice number list for each collection
        invoice_numbers = []
        collection_items = collection.collectionitems_set.all()
        
        # Loop through collection items and get invoice numbers
        for item in collection_items:
            invoice_numbers.append(item.invoice.invoice_no)

        if collection.payment_method.lower() == "cheque":
            cheque_details = CollectionCheque.objects.filter(collection_payment=collection)
            collection_list.append({
                "collection_id": collection.id,
                "customer": collection.customer.customer_name,
                "route": collection.customer.routes.route_name if collection.customer.routes else "No Route",
                "custom_id": collection.customer.custom_id,
                "payment_method": "Cheque",
                "details": cheque_details,
                "invoice_numbers":  ", ".join(invoice_numbers), 
                "amount_received": collection.amount_received,
                "created_date": collection.created_date,
                "receipt_number": collection.receipt_number,
            })

    route_list = RouteMaster.objects.all()
    filter_data = {
        "start_date": start_date,
        "end_date": end_date,
        "route_name": route_name,
        "q": search_query,
    }

    return render(request, 'sales_management/cheque_collections.html', {
        'collection_list': collection_list,
        'route_li': route_list,
        'filter_data': filter_data,
    })


def cheque_clearance(request, collection_id):
    try:
        collection = get_object_or_404(CollectionPayment, pk=collection_id)
        customer = collection.customer
        amount_received = collection.amount_received

        with transaction.atomic():
            remaining_amount = Decimal(amount_received)
            invoice_numbers = []

            # Get invoices linked to this cheque
            invoices = Invoice.objects.filter(
                collectionitems__collection_payment=collection
            ).distinct().order_by("created_date")

            for invoice in invoices:
                if invoice.amout_total > invoice.amout_recieved and remaining_amount > 0:
                    invoice_numbers.append(invoice.invoice_no)

                    due_amount = invoice.amout_total - invoice.amout_recieved
                    payment_amount = min(due_amount, remaining_amount)

                    # Record in CollectionItems (only this cheque portion)
                    CollectionItems.objects.create(
                        invoice=invoice,
                        amount=payment_amount,
                        balance=due_amount - payment_amount,  # remaining after this clearance
                        amount_received=payment_amount,
                        collection_payment=collection,
                    )

                    # Update invoice
                    invoice.amout_recieved += payment_amount
                    if invoice.amout_recieved >= invoice.amout_total:
                        invoice.invoice_status = "paid"
                    invoice.save()

                    # Update outstanding
                    outstanding_instance, _ = CustomerOutstandingReport.objects.get_or_create(
                        customer=customer, product_type="amount", defaults={"value": 0}
                    )
                    outstanding_instance.value -= payment_amount
                    outstanding_instance.save()

                    # Decrease remaining cheque amount
                    remaining_amount -= payment_amount

            # Generate / update receipt
            receipt, _ = Receipt.objects.update_or_create(
                instance_id=str(collection.id),
                transaction_type="collection",
                defaults={
                    "amount_received": amount_received,
                    "customer": customer,
                    "invoice_number": ",".join(invoice_numbers),
                    "created_date": datetime.datetime.today().now(),
                }
            )
            collection.receipt_number = receipt.receipt_number
            collection.save()

            # Handle overpayment -> create refund invoice
            if remaining_amount > 0:
                outstanding_instance, _ = CustomerOutstandingReport.objects.get_or_create(
                    customer=customer, product_type="amount", defaults={"value": 0}
                )
                outstanding_instance.value += remaining_amount
                outstanding_instance.save()

                # refund_invoice = Invoice.objects.create(
                #     created_date=datetime.datetime.today(),
                #     net_taxable=-remaining_amount,
                #     vat=0,
                #     discount=0,
                #     amout_total=-remaining_amount,
                #     amout_recieved=0,
                #     customer=customer,
                #     reference_no=f"custom_id{customer.custom_id}"
                # )

                # item = ProdutItemMaster.objects.get(product_name="5 Gallon")
                # InvoiceItems.objects.create(
                #     category=item.category,
                #     product_items=item,
                #     qty=0,
                #     rate=customer.rate,
                #     invoice=refund_invoice,
                #     remarks=f"Refund invoice created from cheque clearance {refund_invoice.reference_no}",
                # )

            # Mark cheque as cleared
            CollectionCheque.objects.filter(collection_payment=collection).update(status="CLEARED")

            redirect_url = reverse("cheque_collections")
            response_data = {
                "status": "true",
                "title": "Cheque Cleared",
                "message": f"Cheque clearance completed successfully for {customer.customer_name}",
                "redirect": "true",
                "redirect_url": redirect_url,
            }
            return HttpResponse(json.dumps(response_data), content_type="application/javascript")

    except Exception as e:
        response_data = {
            "status": "false",
            "title": "Error",
            "message": f"An unexpected error occurred: {str(e)}",
            "redirect": "false",
        }
        return HttpResponse(json.dumps(response_data), content_type="application/javascript")
    
    
# def cheque_clearance(request, collection_id):
#     try:
#         collection = get_object_or_404(CollectionPayment, pk=collection_id)
#         customer = collection.customer
#         amount_received = collection.amount_received
#         receipt_number = collection.receipt_number

#         with transaction.atomic():
#             remaining_amount = Decimal(amount_received)
#             collection_items = collection.collectionitems_set.all()
#             invoice_numbers = []

#             # adjust overpayments
#             for collection_item in collection_items:
#                 invoice = collection_item.invoice
#                 if invoice:
#                     invoice_amount_due = invoice.amout_total - invoice.amout_recieved
#                     if invoice_amount_due < 0:
#                         remaining_amount += abs(invoice_amount_due)

#             # update collection
#             collection.amount_received = amount_received
#             collection.receipt_number = receipt_number
#             collection.save()

#             # apply payments
#             for collection_item in collection_items:
#                 invoice = collection_item.invoice
#                 if invoice.amout_total > invoice.amout_recieved:
#                     invoice_numbers.append(invoice.invoice_no)
#                     due_amount = invoice.amout_total - invoice.amout_recieved
#                     payment_amount = min(remaining_amount, due_amount)

#                     invoice.amout_recieved += payment_amount
#                     invoice.save()

#                     CollectionItems.objects.create(
#                         invoice=invoice,
#                         amount=invoice.amout_total,
#                         balance=invoice.amout_total - invoice.amout_recieved,
#                         amount_received=payment_amount,
#                         collection_payment=collection
#                     )

#                     outstanding_instance, _ = CustomerOutstandingReport.objects.get_or_create(
#                         customer=customer, product_type="amount"
#                     )
#                     outstanding_instance.value -= payment_amount
#                     outstanding_instance.save()

#                     remaining_amount -= payment_amount

#                     if invoice.amout_recieved == invoice.amout_total:
#                         invoice.invoice_status = 'paid'
#                         invoice.save()

#             # create or update receipt
#             Receipt.objects.update_or_create(
#                 instance_id=str(collection.id),
#                 transaction_type="collection",
#                 defaults={
#                     "amount_received": amount_received,
#                     "receipt_number": receipt_number,
#                     "customer": customer,
#                     "invoice_number": ",".join(invoice_numbers)
#                 }
#             )

#             # handle remaining balance
#             if remaining_amount > 0:
#                 outstanding_instance, _ = CustomerOutstandingReport.objects.get_or_create(
#                     customer=customer, product_type="amount"
#                 )
#                 outstanding_instance.value += remaining_amount
#                 outstanding_instance.save()

#                 # invoice = Invoice.objects.update(
#                 #     created_date=collection.created_date or timezone.now(),
#                 #     net_taxable=remaining_amount,
#                 #     vat=0,
#                 #     discount=0,
#                 #     amout_total=remaining_amount,
#                 #     amout_recieved=0,
#                 #     customer=customer,
#                 #     reference_no=f"custom_id{customer.custom_id}"
#                 # )

#                 # item = ProdutItemMaster.objects.get(product_name="5 Gallon")
#                 # InvoiceItems.objects.create(
#                 #     category=item.category,
#                 #     product_items=item,
#                 #     qty=0,
#                 #     rate=customer.rate,
#                 #     invoice=invoice,
#                 #     remarks=f'Invoice generated from collection: {invoice.reference_no}'
#                 # )

#             # mark cheque as cleared
#             CollectionCheque.objects.filter(collection_payment=collection).update(status="CLEARED")

#             # ✅ response in same format as create_customer_outstanding
#             redirect_url = reverse('cheque_collections')  # adjust name if needed
#             response_data = {
#                 "status": "true",
#                 "title": "Cheque Cleared",
#                 "message": f"Cheque clearance completed successfully for {customer.customer_name}",
#                 "redirect": "true",
#                 "redirect_url": redirect_url,
#             }
#             return HttpResponse(json.dumps(response_data), content_type="application/javascript")

#     except Exception as e:
#         response_data = {
#             "status": "false",
#             "title": "Error",
#             "message": f"An unexpected error occurred: {str(e)}",
#             "redirect": "false",
#         }
#         return HttpResponse(json.dumps(response_data), content_type="application/javascript")
    
    

def production_onload_report_view(request):
# Get today's date in dd-mm-yyyy format
    today_str = date.today().strftime('%d-%m-%Y')
    # Get the dates from the request
    start_date_str = request.GET.get("start_date")
    print("start_date_str", start_date_str)
    end_date_str = request.GET.get("end_date")
    print("end_date_str", end_date_str)

    if not start_date_str or not end_date_str:
        # If no dates are provided, default to today's date
        start_date = end_date = date.today()
    else:
        try:
            # Convert input strings to datetime in dd-mm-yyyy format
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = end_date = date.today()
    # Fetch orders within the date range (or only today if no dates were provided)
    orders = Staff_Orders_details.objects.select_related("product_id", "staff_order_id").filter(staff_order_id__order_date__range=[start_date, end_date]).order_by('-created_date')

    issued_data = []
    for order in orders:
        product_item_instance = order.product_id

        # Check if product is "5 Gallon"
        if product_item_instance and product_item_instance.product_name == "5 Gallon":
            van_instance = (
                Van.objects.filter(salesman_id__id=order.staff_order_id.created_by).first()
                if order.staff_order_id else None
            )

            if van_instance:
                van_stock_instance = VanProductStock.objects.filter(
                    created_date=order.staff_order_id.order_date,
                    product=product_item_instance,
                    van=van_instance
                ).first()

                product_stock = ProductStock.objects.filter(
                    product_name=product_item_instance,
                    branch=van_instance.branch_id
                ).first()

                scrap_stock = ScrapStock.objects.filter(product=product_item_instance).first()

                # Fetch additional production damage related to the product
                damage_data = ProductionDamage.objects.filter(
                    product=product_item_instance,
                    created_date__lte=order.staff_order_id.order_date
                ).values('product_from', 'product_to').annotate(
                    total_quantity=Sum('quantity')
                )

                # Variables for service, used, fresh, issued bottles
                service_count = 0
                used_bottle_count = 0
                fresh_bottle_count = 0

                for data in damage_data:
                    if data['product_from'] == 'used' and data['product_to'] == 'service':
                        service_count += data['total_quantity']
                    if data['product_from'] == 'used':
                        used_bottle_count += data['total_quantity']
                    if data['product_from'] == 'fresh':
                        fresh_bottle_count += data['total_quantity']

                issued_bottle_count = order.issued_qty

                issued_data.append({
                    "product_name": product_item_instance.product_name,
                    "van_name": van_instance.van_make,
                    "order_date": order.staff_order_id.order_date.strftime('%d-%m-%Y'),
                    "initial_van_stock": van_stock_instance.stock if van_stock_instance else 0,
                    "updated_van_stock": (van_stock_instance.stock + issued_bottle_count) if van_stock_instance else 0,
                    "initial_product_stock": product_stock.quantity if product_stock else 0,
                    "updated_product_stock": (product_stock.quantity - issued_bottle_count) if product_stock else 0,
                    "scrap_stock": scrap_stock.quantity if scrap_stock else 0,
                    "service_count": service_count,
                    "used_bottle_count": used_bottle_count,
                    "fresh_bottle_count": fresh_bottle_count,
                    "issued_bottle_count": issued_bottle_count,
                })

    context = {
        "issued_data": issued_data,
        "filter_data": {
            'filter_date_from': start_date.strftime('%d-%m-%Y'),
            'filter_date_to': end_date.strftime('%d-%m-%Y'),
        }
    }
    return render(request, "sales_management/production_onload_report.html", context)

def production_onload_print(request):
    # Get today's date in dd-mm-yyyy format
    today_str = date.today().strftime('%d-%m-%Y')
    
    # Get the dates from the request
    start_date_str = request.GET.get("start_date")
    print("start_date_str", start_date_str)
    end_date_str = request.GET.get("end_date")
    print("end_date_str", end_date_str)

    if not start_date_str or not end_date_str:
        # If no dates are provided, default to today's date
        start_date = end_date = date.today()
    else:
        try:
            # Adjust the date format to match the passed 'dd-mm-yyyy' format
            start_date = datetime.datetime.strptime(start_date_str, '%d-%m-%Y').date()
            end_date = datetime.datetime.strptime(end_date_str, '%d-%m-%Y').date()
        except ValueError:
            start_date = end_date = date.today()

    # Fetch orders within the date range (or only today if no dates were provided)
    orders = Staff_Orders_details.objects.select_related("product_id", "staff_order_id").filter(
        staff_order_id__order_date__range=[start_date, end_date]
    ).order_by('-created_date')

    issued_data = []
    for order in orders:
        product_item_instance = order.product_id

        # Check if product is "5 Gallon"
        if product_item_instance and product_item_instance.product_name == "5 Gallon":
            van_instance = (
                Van.objects.filter(salesman_id__id=order.staff_order_id.created_by).first()
                if order.staff_order_id else None
            )

            if van_instance:
                van_stock_instance = VanProductStock.objects.filter(
                    created_date=order.staff_order_id.order_date,
                    product=product_item_instance,
                    van=van_instance
                ).first()

                product_stock = ProductStock.objects.filter(
                    product_name=product_item_instance,
                    branch=van_instance.branch_id
                ).first()

                scrap_stock = ScrapStock.objects.filter(product=product_item_instance).first()

                # Fetch additional production damage related to the product
                damage_data = ProductionDamage.objects.filter(
                    product=product_item_instance,
                    created_date__lte=order.staff_order_id.order_date
                ).values('product_from', 'product_to').annotate(
                    total_quantity=Sum('quantity')
                )

                # Variables for service, used, fresh, issued bottles
                service_count = 0
                used_bottle_count = 0
                fresh_bottle_count = 0

                for data in damage_data:
                    if data['product_from'] == 'used' and data['product_to'] == 'service':
                        service_count += data['total_quantity']
                    if data['product_from'] == 'used':
                        used_bottle_count += data['total_quantity']
                    if data['product_from'] == 'fresh':
                        fresh_bottle_count += data['total_quantity']

                issued_bottle_count = order.issued_qty

                issued_data.append({
                    "product_name": product_item_instance.product_name,
                    "van_name": van_instance.van_make,
                    "order_date": order.staff_order_id.order_date.strftime('%d-%m-%Y'),
                    "initial_van_stock": van_stock_instance.stock if van_stock_instance else 0,
                    "updated_van_stock": (van_stock_instance.stock + issued_bottle_count) if van_stock_instance else 0,
                    "initial_product_stock": product_stock.quantity if product_stock else 0,
                    "updated_product_stock": (product_stock.quantity - issued_bottle_count) if product_stock else 0,
                    "scrap_stock": scrap_stock.quantity if scrap_stock else 0,
                    "service_count": service_count,
                    "used_bottle_count": used_bottle_count,
                    "fresh_bottle_count": fresh_bottle_count,
                    "issued_bottle_count": issued_bottle_count,
                })

    context = {
        "issued_data": issued_data,
        "filter_data": {
            'filter_date_from': start_date.strftime('%d-%m-%Y'),
            'filter_date_to': end_date.strftime('%d-%m-%Y'),
        }
    }
    return render(request, "sales_management/production_onload_print.html", context)



def download_production_onload(request):
    # Get the dates from the request
    start_date_str = request.GET.get("start_date")
    print("start_date_str", start_date_str)
    end_date_str = request.GET.get("end_date")
    print("end_date_str", end_date_str)

    if not start_date_str or not end_date_str:
        # If no dates are provided, default to today's date
        start_date = end_date = date.today()
    else:
        try:
            # Adjust the date format to match the passed 'dd-mm-yyyy' format
            start_date = datetime.datetime.strptime(start_date_str, '%d-%m-%Y').date()
            end_date = datetime.datetime.strptime(end_date_str, '%d-%m-%Y').date()
        except ValueError:
            start_date = end_date = date.today()


    # Fetch orders within the date range (or only today if no dates were provided)
    orders = Staff_Orders_details.objects.select_related("product_id", "staff_order_id").filter(
        staff_order_id__order_date__range=[start_date, end_date]
    ).order_by('-created_date')


    # Create an Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Production Onload Data"

    # Define the header row
    headers = [
        "Sl.No", "Product Name", "Van Name", "Order Date", "Initial Van Stock", 
        "Updated Van Stock", "Initial Product Stock", "Updated Product Stock", 
        "Scrap Stock", "Service Count", "Used Bottle Count", "Fresh Bottle Count", "Issued Bottle Count"
    ]
    sheet.append(headers)

    # Populate the Excel sheet with data
    serial_no = 1
    for order in orders:
        product_item_instance = order.product_id

        if product_item_instance and product_item_instance.product_name == "5 Gallon":
            van_instance = (
                Van.objects.filter(salesman_id__id=order.staff_order_id.created_by).first()
                if order.staff_order_id else None
            )
            
            if van_instance:
                van_stock_instance = VanProductStock.objects.filter(
                    created_date=order.staff_order_id.order_date,
                    product=product_item_instance,
                    van=van_instance
                ).first()
                
                product_stock = ProductStock.objects.filter(
                    product_name=product_item_instance,
                    branch=van_instance.branch_id
                ).first()
                
                scrap_stock = ScrapStock.objects.filter(product=product_item_instance).first()
                
                damage_data = ProductionDamage.objects.filter(
                    product=product_item_instance,
                    created_date__lte=order.staff_order_id.order_date
                ).values('product_from', 'product_to').annotate(
                    total_quantity=Sum('quantity')
                )
                
                service_count = 0
                used_bottle_count = 0
                fresh_bottle_count = 0
                
                for data in damage_data:
                    if data['product_from'] == 'used' and data['product_to'] == 'service':
                        service_count += data['total_quantity']
                    if data['product_from'] == 'used':
                        used_bottle_count += data['total_quantity']
                    if data['product_from'] == 'fresh':
                        fresh_bottle_count += data['total_quantity']

                issued_bottle_count = order.issued_qty
                
                row = [
                    serial_no,
                    product_item_instance.product_name,
                    van_instance.van_make,
                    order.staff_order_id.order_date,
                    van_stock_instance.stock if van_stock_instance else 0,
                    (van_stock_instance.stock + issued_bottle_count) if van_stock_instance else 0,
                    product_stock.quantity if product_stock else 0,
                    (product_stock.quantity - issued_bottle_count) if product_stock else 0,
                    scrap_stock.quantity if scrap_stock else 0,
                    service_count,
                    used_bottle_count,
                    fresh_bottle_count,
                    issued_bottle_count,
                ]
                sheet.append(row)
                serial_no += 1

    # Prepare the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    start_date_str = start_date if start_date else 'start'
    end_date_str = end_date if end_date else 'end'
    response['Content-Disposition'] = f'attachment; filename=Production_Onload_{start_date_str}_to_{end_date_str}.xlsx'

    # Save the workbook to the response
    workbook.save(response)
    return response

def scrap_clearance_report(request):
    product = ProdutItemMaster.objects.filter(product_name="5 Gallon").first()
    scrap_clearance_records = ScrapcleanedStock.objects.filter(product=product)

    # Get today's date
    today = now().date()

    # Check if date filters are applied
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        # Convert string to date objects
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()

            # Filter data for the selected date range
            scrap_clearance_records = scrap_clearance_records.filter(
                created_date__date__range=[start_date, end_date]
            )
        except ValueError:
            pass  # Handle invalid date format gracefully
    else:
        # Get today's records
        todays_records = scrap_clearance_records.filter(created_date__date=today)

        # If today's data is available, use it; otherwise, get all records
        scrap_clearance_records = todays_records if todays_records.exists() else scrap_clearance_records

    context = {
    "scrap_clearance_records": scrap_clearance_records.order_by('-created_date'),
    "product_name": "5 Gallon",
    "filter_data": {
        "filter_date_from": start_date.strftime("%Y-%m-%d") if start_date else "",
        "filter_date_to": end_date.strftime("%Y-%m-%d") if end_date else "",
    },
}


    return render(request, "sales_management/scrap_clearance_report.html", context)

def scrap_clearance_print(request):
    product = ProdutItemMaster.objects.filter(product_name="5 Gallon").first()
    scrap_clearance_records = ScrapcleanedStock.objects.filter(product=product)

    # Get today's date
    today = now().date()

    # Get filter parameters from URL
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            scrap_clearance_records = scrap_clearance_records.filter(
                created_date__date__range=[start_date, end_date]
            )
        except ValueError:
            pass  # Handle invalid date formats

    else:
        # Get today's records
        todays_records = scrap_clearance_records.filter(created_date__date=today)

        # If today's data is available, use it; otherwise, get all records
        scrap_clearance_records = todays_records if todays_records.exists() else scrap_clearance_records

    context = {
        "scrap_clearance_records": scrap_clearance_records.order_by('-created_date'),
        "product_name": "5 Gallon",
        "filter_data": {
            "filter_date_from": start_date or "",
            "filter_date_to": end_date or "",
        },
    }

    return render(request, "sales_management/scrap_clearance_print.html", context)

def scrap_clearance_to_excel(request):
    product = ProdutItemMaster.objects.filter(product_name="5 Gallon").first()
    scrap_clearance_records = ScrapcleanedStock.objects.filter(product=product)

    # Get today's date
    today = now().date()

    # Get filter parameters from URL
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            scrap_clearance_records = scrap_clearance_records.filter(
                created_date__date__range=[start_date, end_date]
            )
        except ValueError:
            pass  # Handle invalid date formats
    else:
        # Get today's records
        todays_records = scrap_clearance_records.filter(created_date__date=today)

        # If today's data is available, use it; otherwise, get all records
        scrap_clearance_records = todays_records if todays_records.exists() else scrap_clearance_records

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Scrap Clearance Data"

    # Define column headers
    headers = ["Sl.No", "Product Name", "Cleared Date", "Cleared Quantity", "Cleared By"]
    sheet.append(headers)

    # Populate Excel rows
    for index, record in enumerate(scrap_clearance_records, start=1):
        sheet.append([
            index,  # Auto-incremented serial number
            record.product.product_name,
            record.created_date.strftime("%Y-%m-%d") if record.created_date else "N/A",
            record.quantity,
            record.created_by
        ])

    # Prepare the response with the correct content type
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Scrap_Clearance_Report.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response

def route_wise_bottle_count(request):
    route_li = RouteMaster.objects.filter()
    context = {'route_li': route_li}
    
    return render(request, "sales_management/route_wise_bottle_count.html", context)

# def custody_custom_list(request):
#     """
#     Fetches all custody records along with their customer details.
#     """
#     custody_records = CustodyCustom.objects.select_related('customer').all().order_by('-created_date')
#     return render(request, 'sales_management/custody_custom_list.html', {'custody_records': custody_records})

def custody_custom_list(request):
    custody_records = CustodyCustom.objects.select_related('customer').all()

    # Get today's date
    today = now().date()

    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get('q', '').strip()  # Get search query and remove extra spaces

    # Filter by date range
    if start_date and end_date:
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            custody_records = custody_records.filter(created_date__date__range=[start_date, end_date])
        except ValueError:
            pass  # Handle invalid date format gracefully
    else:
        todays_records = custody_records.filter(created_date__date=today)
        custody_records = todays_records if todays_records.exists() else custody_records

    # Apply search filter
    if query:
        custody_records = custody_records.filter(
            Q(customer__customer_name__icontains=query) |
            Q(customer__custom_id__icontains=query) |
            Q(agreement_no__icontains=query) |
            Q(reference_no__icontains=query)
        )

    context = {
        "custody_records": custody_records.order_by('-created_date'),
        "filter_data": {
            "filter_date_from": start_date.strftime("%Y-%m-%d") if start_date else "",
            "filter_date_to": end_date.strftime("%Y-%m-%d") if end_date else "",
            "q": query,
        },
    }

    return render(request, 'sales_management/custody_custom_list.html', context)

def custody_custom_detail(request, custody_id):
    custody_record = get_object_or_404(CustodyCustom, custody_custom_id=custody_id)
    custody_items = CustodyCustomItems.objects.filter(custody_custom=custody_record)

    context = {
        "custody_record": custody_record,
        "custody_items": custody_items,
    }
    
    return render(request, 'sales_management/custody_custom_detail.html', context)



def collection(request, customer_id):
    user = request.user
    today = now().date()

    if request.method == "POST":
        payment_date_str = request.POST.get("payment_date")

        if payment_date_str:
            today = datetime.strptime(payment_date_str, "%Y-%m-%d").date()
        else:
            today = now().date()

    customer = get_object_or_404(Customers, pk=customer_id)
    salesmen = CustomUser.objects.filter(is_active=True)


    # ✅ FETCH ALL invoices (status-independent)
    invoices = (
        Invoice.objects
        .filter(customer=customer, is_deleted=False)
        .exclude(amout_total=0)
        .order_by("created_date")   # FIFO
    )

    display_invoices = []
    total_balance = 0

    # ✅ Calculate balance safely
    for inv in invoices:
        balance = inv.amout_total - inv.amout_recieved
        if balance > 0:
            inv.balance_amount = balance
            display_invoices.append(inv)
            total_balance += balance
    total_credit = (
        CustomerCredit.objects
        .filter(customer=customer)
        .aggregate(total=Sum("amount"))["total"]
        or 0
    )
    # Logging (read-only)
    log_activity(
        created_by=user,
        description=f"User {user.username} viewed collection invoices for customer_id={customer_id}."
    )

    total_collection = total_balance - total_credit

    context = {
        "page_title": "Payment Collection",
        "customer": customer,
        "invoices": display_invoices,   # only invoices with balance
        "total_balance": total_balance,
        "total_credit": total_credit, 
        "total_collection": total_collection,
        "today": today.strftime("%Y-%m-%d"),
        "salesmen": salesmen,
    }

    return render(request, "sales_management/collection.html", context)

def parse_payment_date(date_str):
    """
    Convert 'YYYY-MM-DD' string into a timezone-aware datetime.
    If missing/invalid, return now().
    """
    if not date_str:
        return now()
    try:
        # Parse as date
        dt = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        # Combine with midnight and make timezone-aware
        return make_aware(datetime.datetime.combine(dt, datetime.datetime.min.time()))
    except Exception:
        return now()



def collection_payment(request, customer_id):

    if request.method != "POST":
        return JsonResponse({
            "status": "false",
            "title": "Invalid Request",
            "message": "Only POST method is allowed."
        })

    customer = get_object_or_404(Customers, pk=customer_id)
   

    invoices = (
        Invoice.objects
        .filter(customer=customer, is_deleted=False)
        .exclude(amout_total=0)
        .order_by("created_date")   # FIFO
    )

    try:
        with transaction.atomic():

            payment_method = request.POST.get("payment_method")
            amount_received = Decimal(request.POST.get("amount_received", 0))
            payment_date = parse_payment_date(request.POST.get("payment_date"))

            remaining = amount_received
            invoice_numbers = []
            print("payment date:",payment_date)
            # 1️⃣ Create CollectionPayment
            customer = Customers.objects.select_for_update().get(pk=customer_id)
            route = customer.routes

            van_route = Van_Routes.objects.filter(routes=route).select_related("van").first()
            van = van_route.van if van_route else None
            salesman = van.salesman if van else None

            invoices = (
                Invoice.objects
                .filter(customer=customer, is_deleted=False)
                .exclude(amout_total=0)
                .order_by("created_date")   # FIFO
            )

            if not invoices.exists():
                return Response({"message": "No valid invoices found."}, status=400)

            # ================= CUSTOMER CREDIT =================
            customer_credit = (
                CustomerCredit.objects
                .filter(customer=customer)
                .aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
            )

            remaining_cash = amount_received
            remaining_credit = customer_credit

            # ---------------- CREATE COLLECTION ----------------
            collection_payment = CollectionPayment.objects.create(
                payment_method=payment_method.upper(),
                customer=customer,
                salesman=salesman,
                amount_received=amount_received,
                created_date = payment_date
            )

            # # ---------------- CHEQUE ----------------
            # if payment_method.lower() == "cheque":
            #     CollectionCheque.objects.create(
            #         collection_payment=collection_payment,
            #         cheque_amount=amount_received,
            #         cheque_no=cheque_details.get("cheque_no"),
            #         bank_name=cheque_details.get("bank_name"),
            #         cheque_date=cheque_details.get("cheque_date"),
            #     )
            #     return Response({"message": "Cheque payment saved."}, status=201)

            collected_invoice_numbers = []

            # ================= APPLY FIFO =================
            for invoice in invoices:

                invoice_due = invoice.amout_total - invoice.amout_recieved
                if invoice_due <= 0:
                    continue

                used_credit = Decimal("0.00")
                used_cash = Decimal("0.00")

                # ---- Apply CREDIT first ----
                if remaining_credit > 0:
                    used_credit = min(remaining_credit, invoice_due)
                    remaining_credit -= used_credit
                    invoice_due -= used_credit

                # ---- Apply CASH ----
                if invoice_due > 0 and remaining_cash > 0:
                    used_cash = min(invoice_due, remaining_cash)
                    remaining_cash -= used_cash
                    invoice_due -= used_cash

                total_used = used_credit + used_cash
                if total_used == 0:
                    continue

                # Update invoice
                invoice.amout_recieved += total_used
                invoice.invoice_status = (
                    "paid" if invoice.amout_recieved >= invoice.amout_total else "partial"
                )
                invoice.save()

                collected_invoice_numbers.append(invoice.invoice_no)

                # Save collection item
                CollectionItems.objects.create(
                    collection_payment=collection_payment,
                    invoice=invoice,
                    amount=invoice.amout_total,
                    amount_received=total_used,
                    balance=invoice.amout_total - invoice.amout_recieved,
                )

                # Credit usage entry
                if used_credit > 0:
                    CustomerCredit.objects.create(
                        customer=customer,
                        amount=-used_credit,
                        source="invoice_adjustment",
                        remark=f"Used for invoice {invoice.invoice_no}",
                    )

                # Online payment record
                # if payment_method.lower() == "online" and used_cash > 0:
                #     CollectionOnline.objects.create(
                #         collection_payment=collection_payment,
                #         online_amount=used_cash,
                #         transaction_no=online_details.get("transaction_no"),
                #         transaction_date=online_details.get("transaction_date"),
                #         status=online_details.get("status", "PENDING"),
                #     )

                if remaining_cash <= 0 and remaining_credit <= 0:
                    break

            # ---------------- REMAINING CASH → CREDIT ----------------
            if remaining_cash > 0:
                CustomerCredit.objects.create(
                    customer=customer,
                    amount=remaining_cash,
                    source="excess_payment",
                    remark="Extra payment after settlement",
                )

            # 4️⃣ Create Receipt (FINAL)
            receipt = Receipt.objects.create(
                transaction_type="collection",
                instance_id=str(collection_payment.id),
                amount_received=amount_received,
                customer=customer,
                invoice_number=",".join(invoice_numbers),
                receipt_number=generate_receipt_no(str(payment_date.date())),
                created_date=payment_date
            )

            collection_payment.receipt_number = receipt.receipt_number
            collection_payment.save(update_fields=["receipt_number"])

        return JsonResponse({
            "status": "true",
            "title": "Success",
            "message": f"Collection saved successfully. Receipt No: {receipt.receipt_number}",
            "redirect": "true",
            "redirect_url": reverse("customers"),
        })

    except Exception as e:
        return JsonResponse({
            "status": "false",
            "title": "Error",
            "message": str(e)
        })

def route_supply_report_view(request):
    """
    Report: Most Supplied Routes
    """

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # ---------------- Date filter ----------------
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        date_filter = {
            "customer_supply__created_date__date__range": [start_date, end_date]
        }
    else:
        date_filter = {}

    report_data = {}

    # ---------------- SUPPLY (SALES + FOC) ----------------
    supply_qs = CustomerSupplyItems.objects.select_related(
        "customer_supply",
        "customer_supply__customer",
        "customer_supply__customer__routes",
        "product"
    ).filter(**date_filter)

    for item in supply_qs:
        route = item.customer_supply.customer.routes.route_name

        if route not in report_data:
            report_data[route] = {
                "route": route,
                "total_bottles": 0,
                "total_orders": set(),
                "total_amount": 0,
                "custody_issued": 0,
                "custody_pullout": 0,
                "returned": 0,
            }

        # Only count 5 Gallon as bottle supply
        if item.product.product_name == "5 Gallon":
            report_data[route]["total_bottles"] += item.quantity

        # Orders count
        report_data[route]["total_orders"].add(item.customer_supply.id)

        # Amount
        report_data[route]["total_amount"] += float(item.amount or 0)

    # ---------------- CUSTODY ISSUE ----------------
    custody_qs = CustodyCustomItems.objects.select_related(
        "custody_custom",
        "custody_custom__customer",
        "custody_custom__customer__routes",
        "product"
    )

    if start_date and end_date:
        custody_qs = custody_qs.filter(
            custody_custom__created_date__date__range=[start_date, end_date]
        )

    for c in custody_qs:
        route = c.custody_custom.customer.routes.route_name
        if route in report_data:
            report_data[route]["custody_issued"] += c.quantity

    # ---------------- CUSTODY PULLOUT ----------------
    pullout_qs = CustomerReturnItems.objects.select_related(
        "customer_return",
        "customer_return__customer",
        "customer_return__customer__routes",
        "product"
    )

    if start_date and end_date:
        pullout_qs = pullout_qs.filter(
            customer_return__created_date__date__range=[start_date, end_date]
        )

    for p in pullout_qs:
        route = p.customer_return.customer.routes.route_name
        if route in report_data:
            report_data[route]["custody_pullout"] += p.quantity

    # ---------------- CUSTOMER RETURN ----------------
    return_qs = CustomerProductReturn.objects.select_related(
        "customer",
        "customer__routes",
        "product"
    )

    if start_date and end_date:
        return_qs = return_qs.filter(
            created_date__date__range=[start_date, end_date]
        )

    for r in return_qs:
        route = r.customer.routes.route_name
        if route in report_data:
            report_data[route]["returned"] += r.quantity

    # ---------------- FINAL BUILD ----------------
    report = []

    for route, val in report_data.items():
        net_consumption = val["total_bottles"] - val["returned"]

        report.append({
            "route": route,
            "total_bottles": val["total_bottles"],
            "total_orders": len(val["total_orders"]),
            "total_amount": round(val["total_amount"], 2),
            "custody_issued": val["custody_issued"],
            "custody_pullout": val["custody_pullout"],
            "returned": val["returned"],
            "net_consumption": net_consumption
        })

    # Sort by most supplied
    report.sort(key=lambda x: x["total_bottles"], reverse=True)

    return render(request, "sales_management/route_supply_report.html", {
        "report": report,
        "start_date": start_date,
        "end_date": end_date
    })