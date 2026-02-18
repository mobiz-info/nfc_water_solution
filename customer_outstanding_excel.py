
import os
import django
from datetime import datetime
from django.utils import timezone
from django.db.models import Sum
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")  # change if needed
django.setup()
from master.models import RouteMaster
from accounts.models import Customers
from client_management.utils import get_customer_outstanding_amount 
from django.db import transaction
from decimal import Decimal

# -----------------------------------
# DJANGO SETUP
# -----------------------------------




  # adjust import

# -----------------------------------
# INPUT DATE HERE
# -----------------------------------
# INPUT_DATE = "2026-01-02"     # change date
# ROUTE_NAME = "S-13"           # change route name

# DATE = datetime.strptime(INPUT_DATE, "%Y-%m-%d").date()

# # -----------------------------------
# # FETCH ROUTE
# # -----------------------------------
# route = RouteMaster.objects.filter(route_name=ROUTE_NAME).first()

# if not route:
#     print(f"❌ Route '{ROUTE_NAME}' not found")
#     exit()

# # -----------------------------------
# # CREATE EXCEL
# # -----------------------------------
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = f"Outstanding {ROUTE_NAME}"

# # Header
# headers = ["Customer ID", "Customer Name", "Outstanding Amount"]
# ws.append(headers)

# for cell in ws[1]:
#     cell.font = Font(bold=True)

# total_outstanding = 0

# # -----------------------------------
# # FETCH CUSTOMERS BY ROUTE
# # -----------------------------------
# customers = Customers.objects.filter(
#     routes=route,
#     is_deleted=False
# )

# for customer in customers:
#     outstanding = get_customer_outstanding_amount(customer)

#     total_outstanding += outstanding

#     ws.append([
#         customer.custom_id,
#         customer.customer_name,
#         round(outstanding, 2)
#     ])

# # -----------------------------------
# # TOTAL ROW
# # -----------------------------------
# ws.append([])
# ws.append(["", "TOTAL", round(total_outstanding, 2)])

# # -----------------------------------
# # AUTO COLUMN WIDTH
# # -----------------------------------
# for col in ws.columns:
#     max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
#     ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

# # -----------------------------------
# # SAVE FILE
# # -----------------------------------
# filename = f"customer_outstanding_{ROUTE_NAME}_{INPUT_DATE}.xlsx"
# file_path = os.path.join(os.getcwd(), filename)

# wb.save(file_path)

# print(f"\n✅ Excel generated successfully:")
# print(file_path)




from accounts.models import Customers, CustomUser
from invoice_management.models import Invoice
from client_management.models import (
    CustomerOutstanding,
    OutstandingAmount,
    CustomerOutstandingReport,
)
CREATED_DATE = datetime(2025, 12, 20)

# -------------------- CONFIG --------------------
EXCEL_FILE = "s-31 diff.xlsx"   # your excel file
SALESMAN_ID = 815                           # 👈 salesman user ID
CREATED_BY_ID = SALESMAN_ID                # same user
START_SL_NO = 1
END_SL_NO = 43  

def safe_decimal(value):
    try:
        value = str(value).replace(",", "").strip()
        return Decimal(value)
    except Exception:
        return None

          

# -------------------------------------------------
# LOAD EXCEL
# -------------------------------------------------
wb = openpyxl.load_workbook(EXCEL_FILE,data_only=True)
sheet = wb.active

salesman = CustomUser.objects.get(pk=SALESMAN_ID)

print("\n🚀 Import started (Sl No 1–43 only)...\n")

# -------------------------------------------------
# PROCESS EXCEL ROWS
# -------------------------------------------------
for sl_no, row in enumerate(
        sheet.iter_rows(min_row=2, max_row=END_SL_NO + 1, values_only=True),
        start=START_SL_NO
    ):
    """
    Excel columns:
    A -> Sl No
    B -> Customer ID
    C -> Customer Name
    G -> Difference (amount)
    """
    if sl_no > 43:
     break
    
     if not isinstance(row[0], (int, float)):
        continue

    # Difference column text header check
    if row[6] is None:
        continue

    if str(row[6]).strip().lower() in ("difference", "diference"):
        continue

    customer_id = row[1]       # Column B
    customer_name = row[2]     # Column C
    difference = row[6]        # Column G (Difference)

    print("difference----------:",difference)

    if not customer_id or not difference:
        continue

    difference = safe_decimal(difference)

    if difference is None:
        print(f"⚠️ [Sl {sl_no}] Skipped invalid difference: {difference}")
        continue

    # Skip zero difference
    if difference == 0:
        continue

    amount = difference

    try:
        customer = Customers.objects.get(custom_id=customer_id)
    except Customers.DoesNotExist:
        print(f"❌ [Sl {sl_no}] Customer not found: {customer_id} - {customer_name}")
        continue

    try:
        with transaction.atomic():

            # -------------------- CREATE CUSTOMER OUTSTANDING --------------------
            outstanding = CustomerOutstanding.objects.create(
                created_date = CREATED_DATE,
                outstanding_date = CREATED_DATE,
                customer=customer,
                product_type="amount",
                created_by=CREATED_BY_ID,
            )

            # -------------------- OUTSTANDING AMOUNT --------------------
            OutstandingAmount.objects.create(
                customer_outstanding=outstanding,
                amount=amount
            )

            # -------------------- CREATE INVOICE --------------------
            invoice = Invoice.objects.create(
                created_date=CREATED_DATE,
                net_taxable=amount,
                vat=0,
                discount=0,
                amout_total=amount,
                amout_recieved=0,
                customer=customer,
                salesman=salesman,
                reference_no="Excel Outstanding",
                invoice_type="credit_invoice",
                invoice_status="non_paid"
            )

            # Link invoice to outstanding
            outstanding.invoice_no = invoice.invoice_no
            outstanding.save(update_fields=["invoice_no"])

            # -------------------- REBUILD OUTSTANDING REPORT --------------------
            totals = Invoice.objects.filter(
                customer=customer,
                is_deleted=False
            ).aggregate(
                total=Sum("amout_total"),
                received=Sum("amout_recieved")
            )

            outstanding_value = (totals["total"] or 0) - (totals["received"] or 0)

            CustomerOutstandingReport.objects.update_or_create(
                customer=customer,
                product_type="amount",
                defaults={"value": outstanding_value}
            )

            print(
                f"✅ [Sl {sl_no}] Imported → "
                f"{customer.customer_id} | {customer.customer_name} | Amount: {amount}"
            )

    except Exception as e:
        print(f"❌ [Sl {sl_no}] Error: {e}")

print("\n🎉 Import completed successfully!\n")