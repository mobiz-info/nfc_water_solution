import json
import math


from coupon_management.models import CouponStock
from coupon_management.serializers import couponStockSerializers
from master.functions import generate_form_errors
from .models import Van, Van_Routes, Van_License, BottleAllocation
from accounts.models import CustomUser, Customers
from master.models import EmirateMaster, RouteMaster
from customer_care.models import DiffBottlesModel
from client_management.models import *
from sales_management.models import *
from product.models import ProductStock, ScrapProductStock, ScrapStock, Staff_IssueOrders, WashingProductStock, WashingStock

import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from reportlab.pdfgen import canvas
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from datetime import datetime

from django.views import View
from django.urls import reverse
from django.db.models import Max
from django.contrib import messages
from django.utils.timezone import now
from django.db.models import Q,Count,Sum
from django.core.serializers import serialize
from django.db import transaction, IntegrityError
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse

from .forms import  *
from collections import defaultdict
from master.functions import log_activity


def get_van_coupon_bookno(request):
    van_id = request.GET.get("vanId")
    coupon_type = request.GET.get("productName")
    stock_date = request.GET.get("stockDate")
    print("stock_date",stock_date)
    
    if (instances := VanCouponStock.objects.filter(created_date=stock_date,van__pk=van_id,coupon__coupon_type__coupon_type_name=coupon_type,stock__gt=0)).exists():
        instance = instances.values_list('coupon__pk')
        stock_instances = CouponStock.objects.filter(couponbook__pk__in=instance)
        serialized = couponStockSerializers(stock_instances, many=True)
        status_code = 200
        response_data = {
            "status": "true",
            "data": serialized.data,
        }
    else:
        status_code = 404
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "item not found",
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")


# Van
# def van(request):
#     all_van = Van.objects.all()
#     context = {'all_van': all_van}
#     return render(request, 'van_management/van.html', context)
def van(request):
    all_van = Van.objects.filter(van_type="company")
    
    routes_assigned = {}  # Initialize an empty dictionary
    
    # print("All Vans:", all_van)  
    
    for van in all_van:
        van_routes = Van_Routes.objects.filter(van=van)
        route_names = [van_route.routes.route_name for van_route in van_routes]
        routes_assigned[van.van_id] = route_names
        
    # Log the activity for viewing van information
    log_activity(
        created_by=request.user,
        description="Viewed all vans and their assigned routes."
    ) 
    context = {
        'all_van': all_van,
        'routes_assigned': routes_assigned,
    }
    
    # print("Context:", context)  
    
    return render(request, 'van_management/van.html', context)

def create_van(request):
    if request.method == 'POST':
        form = VanForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            data.branch_id = request.user.branch_id
            data.created_by = str(request.user.id)
            data.save()
            # Log successful van creation
            log_activity(
                created_by=request.user,
                description=f"Van created successfully: {data.van_make}"
            )
            messages.success(request, 'Van created successfully!')
            return redirect('van')
        else:
            # Log failed attempt to create van
            log_activity(
                created_by=request.user,
                description="Failed to create van due to invalid form data."
            )
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = VanForm()
    context = {'form': form}
    return render(request, 'van_management/create_van.html', context)

def edit_van(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)
    # Log the initiation of the edit action
    log_activity(
        created_by=request.user,
        description=f"Opened edit van form for {van.van_make} "
    )
    if request.method == 'POST':
        form = EditVanForm(request.POST, instance=van)
        if form.is_valid():
            data = form.save(commit=False)
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.branch_id = request.user.branch_id
            data.save()
            # Log successful edit
            log_activity(
                created_by=request.user,
                description=f"Successfully edited van: {van.van_make}"
            )
            return redirect('van')
        else:
            # Log form submission failure
            log_activity(
                created_by=request.user,
                description=f"Failed to edit van: {van.van_make}  due to invalid form data."
            )
    else:
        form = EditVanForm(instance=van)
    return render(request, 'van_management/edit_van.html', {'form': form, 'van': van})

def view_van(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)
    # Log the view activity
    log_activity(
        created_by=request.user,
        description=f"Viewed details of van: {van.van_make}."
    )
    return render(request, 'van_management/view_van.html', {'van': van})

def delete_van(request, van_id):
    van = Van.objects.get(van_id=van_id)
    if request.method == 'POST':
        # Log the deletion action
        log_activity(
            created_by=request.user,
            description=f"Deleted van: {van.van_make}"
        )
        van.delete()
        return redirect('van')
    return render(request, 'master/confirm_delete.html', {'van': van})



# Van staff assigning
def view_association(request):
    vans_with_associations = Van.objects.exclude(driver=None).exclude(salesman=None)
    # Log the activity
    log_activity(
        created_by=request.user,
        description="Viewed van-driver-salesman associations."
    )
    return render(request, 'van_management/driver_salesman_list.html', {'vans_with_associations': vans_with_associations})

def create_association(request):
    if request.method == 'POST':
        form = VanAssociationForm(request.POST)
        if form.is_valid():
            van_instance = form.cleaned_data['van']
            driver_id = form.cleaned_data['driver']
            salesman_id = form.cleaned_data['salesman']
            van = get_object_or_404(Van, van_id=van_instance.van_id)
            driver_instance = get_object_or_404(CustomUser, id=driver_id)
            salesman_instance = get_object_or_404(CustomUser, id=salesman_id)
            van.driver = driver_instance
            van.salesman = salesman_instance
            van.save()
            # Log activity for association creation
            log_activity(
                created_by=request.user,
                description=(
                    f"Created association: Van {van.van_make} "
                    f"with Driver {driver_instance.username} and "
                    f"Salesman {salesman_instance.username} "
                )
            )
            return redirect('/van_assign')
    else:
        form = VanAssociationForm()
        # Log activity for accessing the form
        log_activity(
            created_by=request.user,
            description="Accessed the van-driver-salesman association creation form."
        )
    return render(request, 'van_management/create_assign.html', {'form': form})


def edit_assign(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)
    if request.method == 'POST':
        form = EditAssignForm(van, request.POST)
        if form.is_valid(): 
            driver_id = form.cleaned_data['driver']
            salesman_id = form.cleaned_data['salesman']
            driver_instance = get_object_or_404(CustomUser, id=driver_id)
            salesman_instance = get_object_or_404(CustomUser, id=salesman_id)
            van.driver = driver_instance
            van.salesman = salesman_instance
            van.save()
            # Log activity for editing the assignment
            log_activity(
                created_by=request.user,
                description=(
                    f"Edited association for Van {van.van_make}  "
                    f"Driver updated to {driver_instance.username} "
                    f"Salesman updated to {salesman_instance.username} "
                )
            )
            return redirect('/van_assign')
    else:
        form = EditAssignForm(van)
        # Log activity for accessing the edit form
        log_activity(
            created_by=request.user,
            description=f"Accessed the edit form for Van {van.van_make} "
        )
    return render(request, 'van_management/edit_assign.html', {'form': form, 'van': van})



def delete_assign(request, van_id):
    van = Van.objects.get(van_id=van_id)
    if request.method == 'POST':
        van.driver = None
        van.salesman = None
        van.save()
        # Log activity for deleting the assignment
        log_activity(
            created_by=request.user,
            description=(
                f"Deleted association for Van {van.van_make}"))
        return redirect('/van_assign')
    return render(request, 'master/confirm_delete_assign.html', {'van': van})

def route_assign(request,van_id):
    form = VanAssignRoutesForm()
    all_van=Van_Routes.objects.filter(van__van_id = van_id)
    van_data = Van.objects.get(van_id = van_id)
    context = {'all_van' : all_van,"form":form,"van_data":van_data}
    if request.method == 'POST':
        form = VanAssignRoutesForm(request.POST)
        context = {'all_van' : all_van,"form":form,"van_data":van_data}
        if form.is_valid():
            data = form.save(commit=False)
            route = data.routes
            route_exists = Van_Routes.objects.filter(van = van_data,routes = route).exists()
            if route_exists:
                # Log the duplicate route assignment attempt
                log_activity(
                    created_by=request.user,
                    description=(
                        f"Attempted to assign a duplicate route '{route.route_name}' "))
                messages.success(request, 'Route is already assigned..')
                context = {'all_van' : all_van,"form":form,"van_data":van_data}
                return render(request, 'van_management/assignroute_tovan.html',context)
            data.van = van_data
            data.created_by = str(request.user)
            data.created_date = datetime.now()
            data.save()
            # Log successful route assignment
            log_activity(
                created_by=request.user,
                description=(
                    f"Successfully assigned route '{route.route_name}' ")
            )
            messages.success(request, 'Van Route Assigned successfully!')
            return redirect('route_assign',van_id)
        else:
            # Log invalid form submission
            log_activity(
                created_by=request.user,
                description=(
                    f"Invalid form submission while assigning a route to "
                    f"Van {van_data.van_make} "
                )
            )
            messages.success(request, 'Invalid form data. Please check the input.')
            return render(request, 'van_management/assignroute_tovan.html',context)
    return render(request, 'van_management/assignroute_tovan.html',context)

def delete_route_assign (request):
    van_route_id = request.POST.get('delete_id')
    van_id = request.POST.get('van_idd')
    van = Van_Routes.objects.get(van_route_id=van_route_id)
    if request.method == 'POST':
        van.delete()
        # Log successful deletion
        log_activity(
            created_by=request.user,
            description=(
                f"Deleted route '{van.routes.route_name}' assigned to "
                f"Van {van.van_make} "
            )
        )
        return redirect('route_assign',van_id)
    return render(request, 'master/confirm_delete.html',{'van': van} )






class Licence_List(View):
    template_name = 'van_management/licence_list.html'
    
    def get(self, request, *args, **kwargs):
        emirates = EmirateMaster.objects.all().order_by('emirate_id')
        licenses = Van_License.objects.all()
        vans = {}
        emirate_names = [emirate.name for emirate in emirates]  # List to store emirate names dynamically
        license_list = []
        # for license in licenses:
        #     van_plate = license.van.plate
        #     emirate_name = license.emirate.name
        #     emirate_id = license.emirate.emirate_id
        #     expiry_date = license.expiry_date
        
        # for license in licenses:
        #     lic= {'license_id':license.van_license_id,'van_plate':license.van.plate,
        #            'emirate_name':license.emirate,"expiry_date":license.expiry_date}
        #     license_list.append(lic)

        #     if van_plate not in vans:
        #         vans[van_plate] = {emirate: None for emirate in emirate_names}
        #     vans[van_plate][emirate_name] = expiry_date
        
        for license in licenses:
            if license.van:
                van_plate = license.van.plate
                emirate_name = license.emirate.name
                emirate_id = license.emirate.emirate_id
                expiry_date = license.expiry_date
            
            
                lic= {'license_id':license.van_license_id,'van_plate':license.van.plate,
                    'emirate_name':license.emirate,"expiry_date":license.expiry_date}
                license_list.append(lic)

                if van_plate not in vans:
                    vans[van_plate] = {emirate: None for emirate in emirate_names}
                vans[van_plate][emirate_name] = expiry_date
        
        return render(request, self.template_name, {'vans': vans, 'emirate_names': emirate_names,'license_list':license_list})
   
    

class Licence_Adding(View):
    template_name = 'van_management/licence_add.html'
    form_class = Licence_Add_Form

    def get(self, request, *args, **kwargs):
        emirate_values = EmirateMaster.objects.all()
        van = Van.objects.all()
        context = {'emirate_values': emirate_values, 'vans': van}
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit = False )
            van_id = request.POST.get('van')
            license_no = request.POST.get('licence_no')
            emirate_id = request.POST.get('emirate')
            expiry_date = request.POST.get('expiry_date')
            van = Van.objects.get(van_id = van_id)
            emirate = EmirateMaster.objects.get(pk = emirate_id)
            data.created_by = str(request.user.id)
            data.license_no = license_no
            data.van = van
            data.emirate = emirate
            data.expiry_date = expiry_date
            data.save()
            messages.success(request, 'Licence Successfully Added.', 'alert-success')
            return redirect('licence_list')
        else:
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, self.template_name, context)
        
class License_Edit(View):
    template_name = 'van_management/licence_edit.html'
    form_class = Licence_Edit_Form

    def get(self, request, pk, *args, **kwargs):
        rec = Van_License.objects.get(van_license_id=pk)
        form = self.form_class(instance=rec)
        context = {'form': form,'rec':rec}
        return render(request, self.template_name, context)

    def post(self, request, pk, *args, **kwargs):
        rec = Van_License.objects.get(van_license_id=pk)
        form = self.form_class(request.POST, request.FILES, instance=rec)
        if form.is_valid():
            data = form.save(commit=False)
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            messages.success(request, 'Licence Successfully Updated', 'alert-success')
            return redirect('licence_list')
        else:
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, self.template_name, context)
        
def licence_edit(request, plate):
    van = get_object_or_404(Van, plate=plate)
    van_id = van.van_id
    van_licence = Van_License.objects.filter(van=van)
    emirate = EmirateMaster.objects.all()
    if request.method == 'POST':
        emirate_id = request.POST.get('emirate')
        expiry_date = request.POST.get('expiry_date')
        existing_van_licence = Van_License.objects.filter(van=van, emirate_id=emirate_id).first()
        if existing_van_licence:
            existing_van_licence.expiry_date = expiry_date
            existing_van_licence.save()
        else:
            Van_License.objects.create(van=van, emirate_id=emirate_id, expiry_date=expiry_date)
        return redirect('licence_list')
    return render(request, 'van_management/licence_edit.html', {'emirate_values' : emirate})

# Need to remove 
def licence_delete(request, plate):
    van = get_object_or_404(Van, plate=plate)
    if request.method == 'POST':

        van = get_object_or_404(Van, plate=plate)
        van_id = van.van_id
        van_licence = Van_License.objects.filter(van_id = van_id)
        van_licence.delete()
        return redirect('licence_list')
    return render(request, 'van_management/licence_delete.html', {'van' : van})


# Trip schedule

def schedule_view(request):
    date_str = request.POST.get('date') if request.method == 'POST' else now().strftime('%Y-%m-%d')
    date = datetime.strptime(date_str, '%Y-%m-%d')
    
    routes = RouteMaster.objects.all()
    route_details = []
    
    for route in routes:
        todays_customers = find_customers(request, date_str, route.route_id)
        if todays_customers:
            customer_count = len(todays_customers)
            bottle_count = sum(customer['no_of_bottles'] if customer['no_of_bottles'] is not None else 0 for customer in todays_customers)
            trips = list(set(customer['trip'] for customer in todays_customers))
            trips.reverse()
            route_details.append({
                'route_name': route.route_name,
                'route_id': route.route_id,
                'no_of_customers': customer_count,
                'no_of_bottles': bottle_count,
                'no_of_trips': len(trips),
                'trips': trips
            })
    
    return render(request, 'van_management/schedule_view.html', {'def_date': date_str, 'details': route_details})

            
def schedule_by_route(request, def_date, route_id, trip):
    route = RouteMaster.objects.get(route_id = route_id)
    todays_customers = find_customers(request, def_date, route_id)
    customers = []
    for customer in todays_customers:
        if customer['trip'] == trip:
            customers.append(customer)
    totale_bottle=0
    no_of_bottles=0
    for customer in customers:
        if customer['no_of_bottles'] :
            no_of_bottles = customer['no_of_bottles']
        totale_bottle+=no_of_bottles
    return render(request, 'van_management/schedule_by_route.html', {
        'def_date':def_date,  
        'route':route, 
        'todays_customers' :customers, 
        'trip':trip, 
        'totale_bottle':totale_bottle })
    

# old code starting for find customers
# def find_customers(request, def_date, route_id):
    # date_str = def_date
    # if date_str:
    #     date = datetime.strptime(date_str, '%Y-%m-%d')
    #     day_of_week = date.strftime('%A')
    #     week_num = (date.day - 1) // 7 + 1
    #     week_number = f'Week{week_num}'
    # route = RouteMaster.objects.get(route_id=route_id)

    # van_route = Van_Routes.objects.filter(routes=route).first()
    # if van_route:
    #     van_capacity = van_route.van.capacity
    # else:
    #     van_capacity = 200
    # todays_customers = []

    # buildings = []
    # for customer in Customers.objects.filter(is_guest=False, routes = route):
    #     if customer.visit_schedule and day_of_week in customer.visit_schedule and week_number in customer.visit_schedule[day_of_week]:
    #         todays_customers.append(customer)
    #         if customer.building_name not in buildings:
    #             buildings.append(customer.building_name)
# new starting code for find customers
# def find_customers(request, def_date, route_id):
#     date_str = def_date
#     if date_str:
#         date = datetime.strptime(date_str, '%Y-%m-%d')
#         day_of_week = date.strftime('%A')
#         week_num = (date.day - 1) // 7 + 1
#         week_number = f'Week{week_num}'
#     route = RouteMaster.objects.get(route_id=route_id)

#     van_route = Van_Routes.objects.filter(routes=route).first()
#     if van_route:
#         van_capacity = van_route.van.capacity
#     else:
#         van_capacity = 200
#     todays_customers = []

#     buildings = []
#     for customer in Customers.objects.filter(is_guest=False, routes=route):
#         if customer.visit_schedule:
#             for day, weeks in customer.visit_schedule.items():
#                 if day in str(day_of_week):
#                     if week_number in str(weeks):
#                         todays_customers.append(customer)
#                         buildings.append(customer.building_name)
                        
#     # Customers on vacation
#     date = datetime.strptime(def_date, '%Y-%m-%d').date()
#     for vacation in Vacation.objects.all():
#         if vacation.start_date <= date <= vacation.end_date:
#             if vacation.customer in todays_customers:
#                 todays_customers.remove(vacation.customer)

#     # Emergency customer
#     special_customers = DiffBottlesModel.objects.filter(delivery_date = date)
#     emergency_customers = []
#     for client in special_customers:
#         if client.customer in todays_customers:
#             emergency_customers.append(client.customer)
#         else:
#             if client.customer.routes == route:
#                 todays_customers.append(client.customer)
#                 emergency_customers.append(client.customer)
#                 if client.customer.building_name not in buildings:
#                     buildings.append(client.customer.building_name)
#     co = 0
#     for cus in todays_customers:
#         if cus.no_of_bottles_required:
#             co+=cus.no_of_bottles_required
#     # print(route,co)
    
#     if not len(buildings) == 0:
#         building_count = {}

#         for building in buildings:
#             for customer in todays_customers:
#                 if customer.building_name == building:
#                     if building in building_count:
#                         building_count[building] += customer.no_of_bottles_required
#                     else:
#                         building_count[building] = customer.no_of_bottles_required

#         building_gps = []

#         for building, bottle_count in building_count.items():
#             c = Customers.objects.filter(is_guest=False, building_name=building, routes=route).first()
#             gps_longitude = c.gps_longitude
#             gps_latitude = c.gps_latitude
#             building_gps.append((building, gps_longitude, gps_latitude, bottle_count))

#         sorted_building_gps = sorted(building_gps, key=lambda x: (x[1] if x[1] is not None else '', x[2] if x[2] is not None else ''))
#         sorted_buildings = [item[0] for item in sorted_building_gps]
#         sorted_building_count = dict(sorted(building_count.items(), key=lambda item: item[1]))

#         trips = {}
#         trip_count = 1
#         current_trip_bottle_count = 0
#         trip_buildings = []
#         bottle_count = 0
#         for building in sorted_buildings:
          
#             for building_data in sorted_building_gps:
#                 if building_data[0]==building:
#                     building = building_data[0]
#                     if building_data[3]:
#                         bottle_count = building_data[3]
                        
#                     if current_trip_bottle_count + bottle_count > van_capacity:
#                         trip_buildings = [building]
#                         trip_count+=1
#                         trips[f"Trip{trip_count}"] = trip_buildings
#                         # print(route,"trip",trip_count)
#                         current_trip_bottle_count = bottle_count
#                     else:
                        
#                         trip_buildings.append(building)
#                         trips[f"Trip{trip_count}"] = trip_buildings
#                         current_trip_bottle_count += bottle_count
            
        

#         # Merge trips if possible to optimize
#         merging_occurred = False
#         for trip_num in range(1, trip_count + 1):
#             for other_trip_num in range(trip_num + 1, trip_count + 1):
#                 trip_key = f"Trip{trip_num}"
#                 other_trip_key = f"Trip{other_trip_num}"
#                 combined_buildings = trips.get(trip_key, []) + trips.get(other_trip_key, [])
#                 total_bottles = sum(sorted_building_count.get(building, 0) for building in combined_buildings)
#                 if total_bottles <= van_capacity:
#                     trips[trip_key] = combined_buildings
#                     del trips[other_trip_key]
#                     merging_occurred = True
#                     break
#             if merging_occurred:  
#                 break
                    
#         # List to store trip-wise customer details
#         trip_customers = []
#         location_name = ""
#         for trip in trips:
#             for building in trips[trip]:
#                 for customer in todays_customers:
#                     if customer.building_name == building:
#                         if customer.location :
#                             location_name = customer.location.location_name
#                         trip_customer = {
#                             "customer_name" : customer.customer_name,
#                             "mobile":customer.mobile_no,
#                             "trip":trip,
#                             "building":customer.building_name,
#                             "route" : customer.routes.route_name,
#                             "no_of_bottles": customer.no_of_bottles_required,
#                             "location" : location_name,
#                             "building": customer.building_name,
#                             "door_house_no": customer.door_house_no,
#                             "floor_no": customer.floor_no,
#                             "gps_longitude": customer.gps_longitude,
#                             "gps_latitude": customer.gps_latitude,
#                             "customer_type": customer.sales_type,
#                             # 'rate': customer.rate,
#                         }
#                         if customer in emergency_customers:
#                             trip_customer['type'] = 'Emergency'
#                             dif = DiffBottlesModel.objects.filter(customer=customer, delivery_date=date).latest('created_date')
#                             trip_customer['no_of_bottles'] = dif.quantity_required
#                         else:
#                             trip_customer['type'] = 'Default'
#                         if customer.sales_type == 'CASH' or  customer.sales_type == 'CREDIT':
#                             trip_customer['rate'] = customer.rate

#                         trip_customers.append(trip_customer)
#         return trip_customers

def find_customers(request, def_date, route_id):
    from datetime import datetime
    date_str = def_date
    if date_str:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        day_of_week = date.strftime('%A')
        week_num = (date.day - 1) // 7 + 1
        week_number = f'Week{week_num}'
    
    route = get_object_or_404(RouteMaster, route_id=route_id)

    van_route = Van_Routes.objects.filter(routes=route).first()
    van_capacity = van_route.van.capacity if van_route else 200
    
    vocation_customer_ids = Vacation.objects.filter(start_date__gte=date,end_date__lte=date).values_list('customer__pk')
    
    todays_customers = []
    buildings = []
    for customer in Customers.objects.filter(is_guest=False, routes=route,is_calling_customer=False,is_deleted=False,is_active=True).exclude(pk__in=vocation_customer_ids):
        if customer.visit_schedule:
            for day, weeks in customer.visit_schedule.items():
                if day in str(day_of_week) and week_number in str(weeks):
                    todays_customers.append(customer)
                    buildings.append(customer.building_name)
   
    # Emergency customers
    special_customers = DiffBottlesModel.objects.filter(delivery_date=date)
    emergency_customers = []
    for client in special_customers:
        if client.customer in todays_customers:
            emergency_customers.append(client.customer)
        else:
            if client.customer.routes == route:
                todays_customers.append(client.customer)
                emergency_customers.append(client.customer)
                if client.customer.building_name not in buildings:
                    buildings.append(client.customer.building_name)
    
    # Calculate total bottle count
    co = sum(cus.no_of_bottles_required or 0 for cus in todays_customers)

    # print(f"Total bottle count: {co}, Van capacity: {van_capacity}")

    if buildings:
        building_count = {}
        for building in buildings:
            for customer in todays_customers:
                if customer.building_name == building:
                    no_of_bottles = customer.no_of_bottles_required or 0  # Use 0 if no_of_bottles_required is None
                    building_count[building] = building_count.get(building, 0) + no_of_bottles

        building_gps = []
        for building, bottle_count in building_count.items():
            c = Customers.objects.filter(is_guest=False, building_name=building, routes=route,is_deleted=False,is_active=True).first()
            building_gps.append((building, c.gps_longitude, c.gps_latitude, bottle_count))

        # Sort buildings by GPS coordinates
        sorted_building_gps = sorted(building_gps, key=lambda x: (x[1] if x[1] is not None else '', x[2] if x[2] is not None else ''))
        sorted_buildings = [item[0] for item in sorted_building_gps]

        # Check if total bottle count exceeds van capacity
        if co <= van_capacity:
            # All buildings can fit into one trip
            trips = {"Trip1": sorted_buildings}
        else:
            # Initialize trips
            trips = {}
            trip_count = 1
            current_trip_bottle_count = 0
            trip_buildings = []

            for building in sorted_buildings:
                bottle_count = building_count[building]
                # print(f"Processing building: {building}, Bottle count: {bottle_count}, Current trip bottle count: {current_trip_bottle_count}")

                if current_trip_bottle_count + bottle_count > van_capacity:
                    # print(f"Creating new trip. Current trip bottle count ({current_trip_bottle_count}) + bottle count ({bottle_count}) exceeds van capacity ({van_capacity}).")
                    trips[f"Trip{trip_count}"] = trip_buildings
                    trip_count += 1
                    trip_buildings = [building]
                    current_trip_bottle_count = bottle_count
                else:
                    trip_buildings.append(building)
                    current_trip_bottle_count += bottle_count

            if trip_buildings:
                trips[f"Trip{trip_count}"] = trip_buildings

            # Merge trips if possible to optimize
            merging_occurred = True
            while merging_occurred:
                merging_occurred = False
                for trip_num in range(1, trip_count):
                    for other_trip_num in range(trip_num + 1, trip_count + 1):
                        trip_key = f"Trip{trip_num}"
                        other_trip_key = f"Trip{other_trip_num}"
                        if trip_key in trips and other_trip_key in trips:
                            combined_buildings = trips[trip_key] + trips[other_trip_key]
                            total_bottles = sum(building_count.get(building, 0) for building in combined_buildings)
                            if total_bottles <= van_capacity:
                                # print(f"Merging trips {trip_key} and {other_trip_key}. Combined bottle count: {total_bottles}")
                                trips[trip_key] = combined_buildings
                                del trips[other_trip_key]
                                trip_count -= 1
                                merging_occurred = True
                                break
                    if merging_occurred:
                        break

        # List to store trip-wise customer details
        trip_customers = []
        for trip, buildings in trips.items():
            for building in buildings:
                for customer in todays_customers:
                    if customer.building_name == building:
                        trip_customer = {
                            "customer_id": customer.customer_id,
                            "custom_id": customer.custom_id,
                            "customer_name": customer.customer_name,
                            "mobile": customer.mobile_no,
                            "trip": trip,
                            "building": customer.building_name,
                            "route": customer.routes.route_name,
                            "no_of_bottles": customer.no_of_bottles_required,
                            "location": customer.location.location_name if customer.location else "",
                            "door_house_no": customer.door_house_no,
                            "floor_no": customer.floor_no,
                            "gps_longitude": customer.gps_longitude,
                            "gps_latitude": customer.gps_latitude,
                            "customer_type": customer.sales_type,
                        }
                        if customer in emergency_customers:
                            trip_customer['type'] = 'Emergency'
                            dif = DiffBottlesModel.objects.filter(customer=customer, delivery_date=date).latest('created_date')
                            trip_customer['no_of_bottles'] = dif.quantity_required
                        else:
                            trip_customer['type'] = 'Default'
                        if customer.sales_type in ['CASH', 'CREDIT']:
                            trip_customer['rate'] = customer.rate

                        trip_customers.append(trip_customer)
        return trip_customers
        
import math

def pdf_download(request, route_id, def_date, trip):
    route = RouteMaster.objects.get(route_id=route_id)
    customers = []
    todays_customers = find_customers(request, def_date, route_id)
    for customer in todays_customers:
        if customer['trip'] == trip:
            customers.append(customer)

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)

    x_route_name, y_route_name = 80, 730
    formatted_date = datetime.strptime(def_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    p.setFont("Helvetica", 14)
    p.drawString(x_route_name, y_route_name, f"Client list for {route.route_name} --> {trip}")
    p.drawString(x_route_name + 380, y_route_name, f"{formatted_date}")
    line_y = y_route_name - 14
    p.line(x_route_name - 12, line_y, x_route_name + 450, line_y)

    x, y = 70, 700
    p.setFont("Helvetica", 9)
    p.drawString(x, y, "Si No")
    p.drawString(x + 30, y, "Client Name")
    p.drawString(x + 135, y, "Location")
    p.drawString(x + 200, y, "No of Bottles")
    p.drawString(x + 267, y, "Outstanding")
    p.drawString(x + 280, y-10, "Bottles")
    p.drawString(x + 330, y, "Outstanding")
    p.drawString(x + 337, y-10, "Coupons")
    p.drawString(x + 400, y, "Type")
    y -= 16
    p.line(x, y, x + 460, y)

    y -= 25
    counter = 1
    multiline_client = False  # Flag to track multiline client name in the row
    for customer in customers:
        if y <= 100:  
            p.showPage()  
            y = 700 
            
            
            # Draw headers on the new page
            p.setFont("Helvetica", 9)
            p.drawString(x, y, "Si No")
            p.drawString(x + 30, y, "Client Name")
            p.drawString(x + 135, y, "Location")
            p.drawString(x + 200, y, "No of Bottles")
            p.drawString(x + 267, y, "Outstanding")
            p.drawString(x + 280, y-10, "Bottles")
            p.drawString(x + 330, y, "Outstanding")
            p.drawString(x + 337, y-10, "Coupons")
            p.drawString(x + 400, y, "Type")
            y -= 16
            p.line(x, y, x + 460, y)
            y -= 50

        # Split the client name from the nearest space if it needs to be split into multiple lines
        client_name = customer['customer_name']
        client_name_parts = []
        if p.stringWidth(client_name, "Helvetica", 9) > 86:
            parts = client_name.split()
            current_line = parts[0]
            for part in parts[1:]:
                if p.stringWidth(current_line + ' ' + part, "Helvetica", 9) <= 86:
                    current_line += ' ' + part
                else:
                    client_name_parts.append(current_line)
                    current_line = part
            client_name_parts.append(current_line)
            multiline_client = True  # Set flag for multiline client name
        else:
            client_name_parts.append(client_name)

        if multiline_client:
            client_name_height = 10 * len(client_name_parts)
            cell_middle_y = y - client_name_height / 2
            p.drawString(x + 5, cell_middle_y + 6, str(counter))
        else:
            y -= 0
            
            p.drawString(x + 5, y, str(counter))
            
        # Draw the client name
        client_name_height = 10 * len(client_name_parts)
        for part in client_name_parts:
            p.drawString(x + 30, y, part)
            y -= 10
            

        # Adjust vertical position for the other cells if multiline client name is printed
        if multiline_client:
            cell_middle_y = y + client_name_height / 2
            cell_middle_y += 10
            p.drawString(x + 135, cell_middle_y, str(customer['location']))
            p.drawString(x + 220, cell_middle_y, str(customer['no_of_bottles']))
            p.drawString(x + 300, cell_middle_y, "")
            p.drawString(x + 350, cell_middle_y, "")
            p.drawString(x + 400, cell_middle_y, "Emergency" if customer['type'] == 'Emergency' else "Default")
        else:
            y += 10
            p.drawString(x + 135, y, str(customer['location']))
            p.drawString(x + 220, y, str(customer['no_of_bottles']))
            p.drawString(x + 300, y, "")
            p.drawString(x + 350, y, "")
            p.drawString(x + 400, y, "Emergency" if customer['type'] == 'Emergency' else "Default")
            y -= 20  # Adjust vertical position for the next entry

        # Draw horizontal line below each row
        p.line(x, y, x + 460, y)
        y -= 25
        counter += 1

    p.save()
    buffer.seek(0)

    filename = f"{route.route_name}_{def_date}_{trip}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


def excel_download(request, route_id, def_date, trip):
    route = RouteMaster.objects.get(route_id=route_id)
    customers = []
    todays_customers = find_customers(request, def_date, route_id)
    
    # Manually generate serial numbers
    serial_number = 1
    for customer in todays_customers:
        if customer['trip'] == trip:
            customer['serial_number'] = serial_number
            customers.append(customer)
            serial_number += 1 

    # Prepare data for DataFrame
    data = {
        'Si No': [customer['serial_number'] for customer in customers],
        'Client Name': [customer['customer_name'] for customer in customers],
        'Mobile': [customer['mobile'] for customer in customers],
        'Locations': [customer['location'] for customer in customers],
        'Building': [customer['building'] for customer in customers],
        'Floor': [customer['floor_no'] for customer in customers],
        'Door': [customer['door_house_no'] for customer in customers],
        'No of Bottles': [customer['no_of_bottles'] for customer in customers],
        'Out- Bottles': [' ' for _ in customers],
        'Out- Coupons': [' ' for _ in customers],
        'Out- Cash': [' ' for _ in customers],
        'Rate': [customer.get('rate', '') for customer in customers],
        'Type': [customer['customer_type'] for customer in customers],
        'Delivery Type': [customer['type'] for customer in customers]
    }
    total_bottle=0
    no_of_bottles=0
    for customer in customers:
        if customer['no_of_bottles'] :
            no_of_bottles = customer['no_of_bottles']
        total_bottle+=no_of_bottles
    # Create DataFrame
    df = pd.DataFrame(data)

    # Convert DataFrame to Excel
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        table_border_format = workbook.add_format({'border': 1})  
        worksheet.conditional_format(4, 0, len(df.index) + 4, len(df.columns) - 1, {'type': 'cell', 'criteria': '>', 'value': 0, 'format': table_border_format})

        # Merge cells and write other information with borders
        merge_format = workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'border': 1})
        worksheet.merge_range('A1:N2', f'Sana Water', merge_format)
        merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
        worksheet.merge_range('A3:D3', f'Route:    {route.route_name}    {trip}', merge_format)
        worksheet.merge_range('E3:I3', f'Date: {def_date}', merge_format)
        worksheet.merge_range('J3:N3', f'Total bottle: {total_bottle}', merge_format)
        merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
        worksheet.merge_range('A4:N4', '', merge_format)

    filename = f"{route.route_name}_{def_date}_{trip}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response
    



#Expence

class ExpenseHeadList(View):
    template_name = 'van_management/expensehead_list.html'
    def get(self, request, *args, **kwargs):
        expence_heads = ExpenseHead.objects.all()
        context = {'expence_heads':expence_heads}
        # Log access to the ExpenseHead list view
        log_activity(
            created_by=request.user,
            description=f"Viewed the list of Expense Heads"
        )
        return render(request, self.template_name, context)



class ExpenseHeadAdd(View):
    template_name = 'van_management/expensehead_add.html'
    form_class = ExpenseHeadForm

    def get(self, request, *args, **kwargs):
        form = self.form_class
        context = {'form': form}
        # Log activity for viewing the add expense head page
        log_activity(
            created_by=request.user,
            description=f"Accessed the 'Add Expense Head' page at {datetime.now()}"
        )
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid:
            expense_head = form.save()
            # Log activity for successfully adding an Expense Head
            log_activity(
                created_by=request.user,
                description=f"Successfully added Expense Head '{expense_head.name}' at {datetime.now()}"
            )
            return redirect('expensehead_list')
        else:
            # Log activity for form validation failure
            log_activity(
                created_by=request.user,
                description=f"Failed to add Expense Head due to validation errors at {datetime.now()}")
            return render(request, self.template_name, {'form':form})

            

class ExpenseHeadEdit(View):
    template_name = 'van_management/expensehead_edit.html'
    form_class = ExpenseHeadForm

    def get(self, request, expensehead_id, *args, **kwargs):
        expensehead = ExpenseHead.objects.get(expensehead_id=expensehead_id)
        form = self.form_class(instance = expensehead)
        context = {'form':form, 'expensehead':expensehead}
        return render(request, self.template_name, context)

    def post(self, request, expensehead_id, *args, **kwargs):
       expensehead = ExpenseHead.objects.get(expensehead_id=expensehead_id)
       form = self.form_class(request.POST, instance=expensehead)
       if form.is_valid():
           form.save()
           log_activity(
                created_by=request.user,
                description=f"Edited ExpenseHead: {expensehead.name}"
            )
           return redirect('expensehead_list')
       else:
           context = {'form':form}
           return render(request, self.template_name, context)
       
class ExpenseHeadDelete(View):
    template_name = 'van_management/expensehead_delete.html'
    def get(self, request, expensehead_id, *args, **kwargs):
        expensehead = ExpenseHead.objects.get(expensehead_id=expensehead_id)
        context = {'expensehead':expensehead}
        return render(request, self.template_name, context)
    def post(self, request, expensehead_id, *args, **kwargs):
        expensehead = ExpenseHead.objects.get(expensehead_id=expensehead_id)
        expensehead.delete()
        log_activity(
            created_by=request.user,
            description=f"Deleted ExpenseHead: {expensehead.name}"
        )
        return redirect('expensehead_list')




class ExpenseList(View):
    template_name = 'van_management/expense_list.html'

    def get(self, request, *args, **kwargs):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        fil_expe = request.GET.get('expense_type')
        
        expenses = Expense.objects.all()  
        
        if from_date and to_date:
            from_date = datetime.strptime(from_date, '%Y-%m-%d')
            to_date = datetime.strptime(to_date, '%Y-%m-%d')
            expenses = expenses.filter(expense_date__range=[from_date, to_date])
        else:
            expenses = expenses.filter(expense_date=datetime.today())
        
        if fil_expe:
            expenses = expenses.filter(expence_type_id=fil_expe)
        expense_types = ExpenseHead.objects.all()
        context = {'expenses': expenses, 'expense_types':expense_types}
        log_activity(
            created_by=request.user,
            description="Viewed the expense list."
        )
        return render(request, self.template_name, context)

class ExpenseAdd(View):
    template_name = 'van_management/expense_add.html'
    form_class = ExpenseAddForm

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {'form': form}
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            form.save()
            log_activity(
                created_by=request.user,
                description=f"Added a new expense."
            )
            return redirect('expense_list')
        else:
            context = {'form': form}
            return render(request, self.template_name, context)

class ExpenseEdit(View):
    template_name = 'van_management/expense_edit.html'
    form_class = ExpenseEditForm

    def get(self, request, expense_id, *args, **kwargs):
        expense = Expense.objects.get(expense_id=expense_id)
        form = self.form_class(instance=expense)
        context = {'form': form, 'expense':expense}
        return render(request, self.template_name, context)

    def post(self, request, expense_id, *args, **kwargs):
        expense = get_object_or_404(Expense, expense_id=expense_id)
        form = self.form_class(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            log_activity(
                created_by=request.user,
                description=f"Edited expense: {expense.remarks}"
            )
            return redirect('expense_list')
        else:
            context = {'form': form, 'expense':expense}
            return render(request, self.template_name, context)

class ExpenseDelete(View):
    template_name = 'van_management/expense_delete.html'

    def get(self, request, expense_id, *args, **kwargs):
        expense = Expense.objects.get(expense_id=expense_id)
        context = {'expense': expense}
        return render(request, self.template_name, context)

    def post(self, request, expense_id, *args, **kwargs):
        expense = Expense.objects.get(expense_id=expense_id)
        expense.delete()
        log_activity(
            created_by=request.user,
            description=f"Deleted expense: {expense.remarks}"
        )
        return redirect('expense_list')
    
    
# ----------VanStock-----------
# class VanStock(View):
    
#     def get(self, request, *args, **kwargs):
#         instances = VanCouponStock.objects.all()
#         van_stock = VanProductStock.objects.all()

#         if request.user.is_authenticated and request.user.user_type == "Salesman":
#             instances = instances.filter(van__salesman_id=request.user.id)
#             van_stock = van_stock.filter(van__salesman_id=request.user.id)
            
#         # Aggregate the total count of each coupon type for each van
#         van_coupon_counts = instances.values('van__van_make', 'coupon__coupon_type__coupon_type_name').annotate(
#             coupon_count=Sum('count')
#         )
        
#         # Calculate opening stock based on total count of coupons
#         for van_coupon in van_coupon_counts:
#             van_coupon['opening_stock'] = van_coupon['coupon_count']
    
#         context = {'van_stock': van_stock, 'van_coupon_counts': van_coupon_counts}
#         return render(request, 'van_management/vanstock_list.html', context)
from datetime import time
class VanStockList(View):
    
    def get(self, request, *args, **kwargs):
        filter_data = {}
        
        date = request.GET.get('date')
        if date:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        else:
            date = datetime.today().date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        
        instances = VanCouponStock.objects.filter(created_date=date)
        van_stock = VanProductStock.objects.filter(created_date=date)
        requested_quantity=CustomerSupply.objects.all()
        issued_orders = Staff_IssueOrders.objects.filter(salesman_id=request.user.id)
        
        offload=Offload.objects.all()
        morning_stock_count=0
        evening_stock_count=0
        salesman_id=request.user.id
        
        if request.user.is_authenticated and request.user.user_type == "Salesman":
            instances = instances.filter(van__salesman_id=request.user.id)
            van_stock = van_stock.filter(van__salesman_id=request.user.id)
            for v in van_stock:
                pro=v.product.id
                morning_stocks = van_stock.filter(product=pro,van__created_date__time__lt=time(12, 0, 0))
                
                morning_stock_count = morning_stocks.filter(van__salesman_id=request.user.id,).count()
            #    o morning_stocks = VanStock.objects.filter(van__salesman_id=request.user.id,created_date__time__lt=time(12, 0, 0))
            #     mrning_stock_count = morning_stocks.filter(van__salesman_id=request.user.id,).count()
                print("morning_stock_count",morning_stock_count)
            evening_stocks = VanStock.objects.filter(van__salesman_id=request.user.id,created_date__time__gte=time(12, 0, 0))
            evening_stock_count = evening_stocks.filter(van__salesman_id=request.user.id).count()

            # requested_quantity=requested_quantity.filter(salesman__salesman_id=request.user.id)
        
        offload=offload.filter(van__salesman_id=salesman_id)
       
        van_coupon_counts = instances.values('van__van_make', 'coupon__coupon_type__coupon_type_name').annotate(
            coupon_count=Sum('count')
        )
        
        # Calculate opening stock based on total count of coupons
        for van_coupon in van_coupon_counts:
            van_coupon['opening_stock'] = van_coupon['coupon_count']
    
        context = {'van_stock': van_stock, 'van_coupon_counts': van_coupon_counts,
                   'issued_orders':issued_orders,'morning_stock_count': morning_stock_count,
            'evening_stock_count': evening_stock_count}
        return render(request, 'van_management/vanstock.html', context)
    
    
class VanProductStockList(View):
    
    def get(self, request, *args, **kwargs):
        filter_data = {}
        
        date = request.GET.get('date')
        van_name = request.GET.get('van_name')
        
        if date:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        else:
            date = datetime.today().date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        
        van_product_stock = VanProductStock.objects.filter(created_date=date)
        
        if van_name:
            van_product_stock = van_product_stock.filter(van__pk=van_name)
            filter_data['filter_van'] = van_name
            
        
        van_instances = Van.objects.all()
    
        context = {
            'van_product_stock': van_product_stock,
            'van_instances': van_instances,
            
            'filter_data': filter_data,
        }
        return render(request, 'van_management/van_product_stock.html', context)
    
class VanProductStockUpdate(View):
    
    def get(self, request, pk, *args, **kwargs):
        instance = VanProductStock.objects.get(pk=pk)
        van_stock_form = VanProductStockForm(instance=instance)
    
        context = {
            'instance': instance,
            'van_stock_form': van_stock_form,
        }
        return render(request, 'van_management/van_product_stock_edit.html', context)
    
    def post(self, request, pk, *args, **kwargs):
        instance = VanProductStock.objects.get(pk=pk)
        van_stock_form = VanProductStockForm(request.POST, instance=instance)
        
        if van_stock_form.is_valid():
            try:
                with transaction.atomic():
                    
                    stock_instance = van_stock_form.save(commit=False)
                    stock_instance.save()
                    
                    log_activity(
                        created_by=request.user,
                        description=f"Van stock Updated from edit button: stock - {stock_instance.stock}, sold - {stock_instance.sold_count} empty {stock_instance.empty_can_count}"
                    )
            
                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "stock updated successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('vanstock_product')
                    }
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(van_stock_form,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    
    
def offload(request):
    instances = Van.objects.all()
    
    context = {
        'instances': instances
        }
    return render(request, 'van_management/offload.html', context)

class View_Item_Details(View):
    template_name = 'van_management/item_details.html'

    @method_decorator(login_required)
    def get(self, request, pk, *args, **kwargs):
        filter_data = {}
        
        date = request.GET.get('date')
        if date:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        else:
            date = datetime.today().date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
            
        product_items = VanProductStock.objects.filter(created_date=date,van__pk=pk)
        coupon_items = ProdutItemMaster.objects.filter(category__category_name__iexact="coupons")
        # VanCouponStock.objects.filter(created_date=date,van__pk=pk,stock__gt=0)
        
        context = {
            'product_items': product_items,
            'coupon_items': coupon_items,
            'van_pk': pk,
            'filter_data': filter_data,
            }
        
        return render(request, self.template_name, context)   
    
class EditProductView(View):
    def post(self, request, pk):
        count = request.POST.get('count')
        stock_type = request.POST.get('stock_type')
        item = VanProductStock.objects.get(pk=pk)
        # print(stock_type)
        # print(count)
        # try:
        #     with transaction.atomic():
        if item.product.product_name == "5 Gallon" and stock_type == "stock":
            # print("stock")
            item.stock -= int(count)
            item.save()
            log_activity(
                created_by=request.user.id,
                description=f"Reduced stock for product '{item.product.product_name}' by {count}."
            )
            
            product_stock = ProductStock.objects.get(branch=item.van.branch_id,product_name=item.product)
            product_stock.quantity += int(count)
            product_stock.save()
            log_activity(
                created_by=request.user.id,
                description=f"Increased product stock for '{item.product.product_name}' by {count}."
            )
            
        elif item.product.product_name == "5 Gallon" and stock_type == "empty_can":
            # print("empty")
            item.empty_can_count -= int(count)
            item.save()
            log_activity(
                created_by=request.user.id,
                description=f"Reduced empty can count for product '{item.product.product_name}' by {count}."
            )
            
            emptycan=EmptyCanStock.objects.create(
                product=item.product,
                quantity=int(count)
            )
            emptycan.save()
            log_activity(
                created_by=request.user.id,
                description=f"Added {count} to empty can stock for product '{item.product.product_name}'."
            )
            
        elif item.product.product_name == "5 Gallon" and stock_type == "return_count":
            # print("return")
            scrap_count = int(request.POST.get('scrap_count'),0)
            washing_count = int(request.POST.get('washing_count'),0)
            
            # print(scrap_count)
            # print(washing_count)
            
            OffloadReturnStocks.objects.create(
                created_by=request.user.id,
                created_date=datetime.today(),
                salesman=item.van.salesman,
                van=item.van,
                product=item.product,
                scrap_count=scrap_count,
                washing_count=washing_count
            )
            log_activity(
                created_by=request.user.id,
                description=f"Created offload return stocks with scrap count {scrap_count} and washing count {washing_count} for product '{item.product.product_name}'."
            )
            
            
            if scrap_count > 0 :
                if not ScrapProductStock.objects.filter(created_date__date=datetime.today().date(),product=item.product).exists():
                    scrap_instance=ScrapProductStock.objects.create(created_by=request.user.id,created_date=datetime.today(),product=item.product)
                else:
                    scrap_instance=ScrapProductStock.objects.get(created_date__date=datetime.today().date(),product=item.product)
                scrap_instance.quantity = scrap_count
                scrap_instance.save()
                
                if ScrapStock.objects.filter(product=scrap_instance.product).exists():
                    scrap_stock,scrap_stock_create = ScrapStock.objects.get_or_create(product=scrap_instance.product)
                    scrap_stock.quantity += scrap_count
                    scrap_stock.save()
            
            if washing_count > 0 :
                if not WashingProductStock.objects.filter(created_date__date=datetime.today().date(),product=item.product).exists():
                    washing_instance=WashingProductStock.objects.create(created_by=request.user.id,created_date=datetime.today(),product=item.product)
                else:
                    washing_instance=WashingProductStock.objects.get(created_date__date=datetime.today().date(),product=item.product)
                washing_instance.quantity = washing_count
                washing_instance.save()
                
                if WashingStock.objects.filter(product=scrap_instance.product).exists():
                    washing_stock,washing_stock_create = WashingStock.objects.get_or_create(product=scrap_instance.product)
                    washing_stock.quantity += washing_count
                    washing_stock.save()
                
            count = scrap_count + washing_count
            # print(count)
            item.return_count -= int(count)
            item.save()
        
        elif item.product.product_name == "5 Gallon" and stock_type == "damage_count":
            # print("return")
            scrap_count = int(count)
            washing_count = 0
            
            OffloadDamageStocks.objects.create(
                created_by=request.user.id,
                created_date=datetime.today(),
                salesman=item.van.salesman,
                van=item.van,
                product=item.product,
                scrap_count=scrap_count,
                washing_count=washing_count
            )
            
            if scrap_count > 0 :
                if not ScrapProductStock.objects.filter(created_date__date=datetime.today().date(),product=item.product).exists():
                    scrap_instance=ScrapProductStock.objects.create(created_by=request.user.id,created_date=datetime.today(),product=item.product)
                else:
                    scrap_instance=ScrapProductStock.objects.get(created_date__date=datetime.today().date(),product=item.product)
                scrap_instance.quantity = scrap_count
                scrap_instance.save()
                
                if ScrapStock.objects.filter(product=scrap_instance.product).exists():
                    scrap_stock,scrap_stock_create = ScrapStock.objects.get_or_create(product=scrap_instance.product)
                    scrap_stock.quantity += scrap_count
                    scrap_stock.save()
            
            count = scrap_count
            # print(count)
            item.damage_count -= int(count)
            item.save()
            
        else : 
            # print("else")
            item.stock -= int(count)
            item.save()
            
            product_stock = ProductStock.objects.get(branch=item.van.branch_id,product_name=item.product)
            product_stock.quantity += int(count)
            product_stock.save()
        
        Offload.objects.create(
            created_by=request.user.id,
            created_date=datetime.today(),
            salesman=item.van.salesman,
            van=item.van,
            product=item.product,
            quantity=int(count),
            stock_type=stock_type
        )
        
        response_data = {
            "status": "true",
            "title": "Successfully Offloaded",
            "message": "Offload successfully.",
            'reload': 'true',
        }
                
        # except IntegrityError as e:
        #     # Handle database integrity error
        #     response_data = {
        #         "status": "false",
        #         "title": "Failed",
        #         "message": str(e),
        #     }

        # except Exception as e:
        #     # Handle other exceptions
        #     response_data = {
        #         "status": "false",
        #         "title": "Failed",
        #         "message": str(e),
        #     }
            
        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
            
        
class EditCouponView(View):
    def post(self, request, van_pk):
        filter_data = {}
        date = request.GET.get('date')
        book_numbers = request.POST.getlist("coupon_book_no")
        
        if date:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        else:
            date = datetime.today().date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        
        for book_number in book_numbers:
            coupon_instance = NewCoupon.objects.get(book_num=book_number)
            
            van_coupon_stock = VanCouponStock.objects.get(van__pk=van_pk,coupon=coupon_instance,created_date=date)
            van_coupon_stock.stock = 0
            van_coupon_stock.save()
            
            log_activity(
                created_by=request.user.id,
                description=f"Updated van coupon stock for book number {book_number} to 0."
            )
            
            product_stock = ProductStock.objects.get(branch=van_coupon_stock.van.branch_id,product_name__product_name=coupon_instance.coupon_type.coupon_type_name)
            product_stock.quantity += 1
            product_stock.save()
            log_activity(
                created_by=request.user.id,
                description=f"Updated product stock for {coupon_instance.coupon_type.coupon_type_name}, increased by 1."
            ) 
            
            coupon = CouponStock.objects.get(couponbook=coupon_instance)
            coupon.coupon_stock = "company"
            coupon.save()
            log_activity(
                created_by=request.user.id,
                description=f"Updated coupon stock for book number {book_number} to 'company'."
            )
            
            OffloadCoupon.objects.create(
                created_by=request.user.id,
                created_date=datetime.now(),
                salesman=van_coupon_stock.van.salesman,
                van=van_coupon_stock.van,
                coupon=van_coupon_stock.coupon,
                quantity = 1,
                stock_type="stock"
            )
            log_activity(
                created_by=request.user.id,
                description=f"Offloaded coupon for book number {book_number}."
            )
            
        response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Coupon Offload successfully.",
                'reload': 'true',
            }
        
    
def salesman_requests(request):
    
    instances = SalesmanRequest.objects.all()
    
    log_activity(
        created_by=request.user,
        description="Accessed Salesman Requests page."
    )
    context = {
        'instances': instances
        }
    return render(request, 'van_management/salesman_requests.html', context)


def BottleAllocationn(request):
    filter_data = {}
    route_details = []
    
    van_routes = Van_Routes.objects.all()
    
    date = request.GET.get('date')
    
    if date:
        date = datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    for route in van_routes:
        # Fetch the 5 Gallon product stock for the given date, or default to 0 if not found
        five_gallon_stock = VanProductStock.objects.filter(
            created_date=date, van=route.van, product__product_name="5 Gallon"
        ).aggregate(stock=Max('stock'))['stock'] or 0
        
        latest_update = CustomerSupply.objects.filter(customer__routes=route.routes).aggregate(latest_update=Max('modified_date'))['latest_update']
        
        route_details.append({
            'route_name': route.routes.route_name,
            'five_gallon_stock': five_gallon_stock,
            'latest_update': latest_update,
            'route_id': route.routes.route_id,
        })
    log_activity(
        created_by=request.user,
        description=f"Viewed Bottle Allocation page ."
    )
    context = {
        'route_details': route_details,
        'filter_data': filter_data,
    }

    return render(request, 'van_management/bottle_allocation.html', context)



def EditBottleAllocation(request, route_id=None):
    if route_id:
        route = get_object_or_404(RouteMaster, route_id=route_id)
        print("route_id",route)
    else:
        route = None

    if request.method == 'POST':
        form = BottleAllocationForm(request.POST)
        if form.is_valid():
            bottle_allocation = form.save(commit=False)
            bottle_allocation.created_by = request.user.username
            bottle_allocation.save()
            
            log_activity(
                created_by=request.user,
                description=f"Edited Bottle Allocation for route ID: {route_id}."
            )
            
            return redirect('bottle_allocation')  # Replace with your list view name
    else:
        form = BottleAllocationForm(initial={'route': route})

    context = {
        'form': form,
    }

    return render(request, 'van_management/edit_bottle_allocation.html', context)



class VanCouponStockList(View):
    def get(self, request, *args, **kwargs):
        filter_data = {}
        
        date = request.GET.get('date')
        if date:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        else:
            date = datetime.today().date()
            filter_data['filter_date'] = date.strftime('%Y-%m-%d')
        
        van_coupon_stock = VanCouponStock.objects.filter(created_date=date)

        log_activity(
            created_by=request.user,
            description=f"Viewed Van Coupon Stock page ."
        )
        
        context = {
            'van_coupon_stock': van_coupon_stock,
            'filter_data': filter_data,
        }
        return render(request, 'van_management/van_coupon_stock.html', context)


# List view
def excess_bottle_count_list(request):
    excess_bottles = ExcessBottleCount.objects.all()
    return render(request, 'van_management/excess_bottle_count_list.html', {'excess_bottles': excess_bottles})

# Create view
def excess_bottle_count_create(request):
    if request.method == 'POST':
        form = ExcessBottleCountForm(request.POST)
        if form.is_valid():
            excess_bottle_count = form.save(commit=False)
            excess_bottle_count.created_by = request.user  # Set the created_by field
            excess_bottle_count.save()

            # Update or create VanProductStock
            van_product_stock, created = VanProductStock.objects.get_or_create(
                van=excess_bottle_count.van,
                created_date=excess_bottle_count.created_date.date(),
                defaults={'excess_bottle': excess_bottle_count.bottle_count}
            )
            if not created:
                van_product_stock.excess_bottle += excess_bottle_count.bottle_count
                van_product_stock.save()
            
            log_activity(
                created_by=request.user,
                description=f"Created excess bottle count entry for van {excess_bottle_count.van} with {excess_bottle_count.bottle_count} bottles."
            )
            
            return redirect('excess_bottle_count_list')
    else:
        form = ExcessBottleCountForm()
    return render(request, 'van_management/excess_bottle_count_create.html', {'form': form})

# Update view
def excess_bottle_count_update(request, pk):
    # Retrieve the existing ExcessBottleCount instance
    excess_bottle_count = get_object_or_404(ExcessBottleCount, pk=pk)
    
    # Store the old bottle count for adjustment
    old_bottle_count = excess_bottle_count.bottle_count
    
    if request.method == 'POST':
        # Initialize the form with POST data and the existing instance
        form = ExcessBottleCountForm(request.POST, instance=excess_bottle_count)
        
        if form.is_valid():
            # Save the form without committing to the database
            excess_bottle_count = form.save(commit=False)
            excess_bottle_count.created_by = request.user.pk  # Update the created_by field
            excess_bottle_count.save()

            # Update the related VanProductStock instance
            van_product_stock = VanProductStock.objects.filter(
                van=excess_bottle_count.van,
                created_date=excess_bottle_count.created_date.date()
            ).first()
            
            if van_product_stock:
                # Adjust the excess_bottle count in VanProductStock
                van_product_stock.excess_bottle += excess_bottle_count.bottle_count - old_bottle_count
                van_product_stock.save()

            log_activity(
                created_by=request.user,
                description=f"Updated excess bottle count for van {excess_bottle_count.van}. Old count: {old_bottle_count}, New count: {excess_bottle_count.bottle_count}."
            )
            # Redirect to the list view after saving
            return redirect('excess_bottle_count_list')
    else:
        # Initialize the form with the existing instance
        form = ExcessBottleCountForm(instance=excess_bottle_count)

    # Render the update template with the form
    return render(request, 'van_management/excess_bottle_count_edit.html', {'form': form})
# Delete view
def excess_bottle_count_delete(request, pk):
    excess_bottle_count = get_object_or_404(ExcessBottleCount, pk=pk)
    if request.method == 'POST':
        # Update VanProductStock
        van_product_stock = VanProductStock.objects.filter(
            van=excess_bottle_count.van,
            created_date=excess_bottle_count.created_date.date()
        ).first()
        if van_product_stock:
            van_product_stock.excess_bottle -= excess_bottle_count.bottle_count
            van_product_stock.save()
        
        log_activity(
            created_by=request.user,
            description=f"Deleted excess bottle count entry for van {excess_bottle_count.van} with {excess_bottle_count.bottle_count} bottles."
        )
        excess_bottle_count.delete()
        return redirect('excess_bottle_count_list')
    return render(request, 'van_management/excess_bottle_count_confirm_delete.html', {'object': excess_bottle_count})


class VanDamageBottleList(View):
    def get(self, request, *args, **kwargs):
        filter_data = {}
        route = request.GET.get('route')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        export = request.GET.get('export')
        print_status = request.GET.get('print')

        if start_date:
            start_date = datetime.strptime(str(start_date), '%Y-%m-%d').date()
        else:
            start_date = datetime.today().date()
            
        if end_date:
            end_date = datetime.strptime(str(end_date), '%Y-%m-%d').date()
        else:
            end_date = datetime.today().date()
        
        route_li = RouteMaster.objects.all()
        filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
        filter_data['end_date'] = end_date.strftime('%Y-%m-%d')
        
        van_damage_stock = VanSaleDamage.objects.filter(
            created_date__gte=start_date,
            created_date__lt=end_date
        )
        
        if route:
            van_route = Van_Routes.objects.get(van__salesman__user_type="Salesman",routes__route_name=route)
            van_damage_stock = van_damage_stock.filter(van=van_route.van)
            filter_data['route_filter'] = route
        
        # Export to Excel if 'export' parameter is present
        if export == 'excel':
            return self.export_to_excel(van_damage_stock)

        log_activity(
            created_by=request.user,
            description=f"Viewed Van Damage stock page."
        )
        
        if not print_status == "print":
            
            context = {
                'van_damage_stock': van_damage_stock.order_by("-created_date"),
                'total_damage_count': van_damage_stock.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0,
                'filter_data': filter_data,
                'route_li': route_li
            }
            return render(request, 'van_management/van_damage_stock.html', context)
        
        else:
            context = {
                'van_damage_stock': van_damage_stock.order_by("-created_date"),
                'total_damage_count': van_damage_stock.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0,
                'filter_data': filter_data,
                'route_li': route_li,
            }
            return render(request, 'van_management/van_damage_stock_print.html', context)

    
    def export_to_excel(self, van_damage_stock):
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Van Damage Stock"

        # Add headers to the worksheet
        headers = ["Date", "Van Name", "Route", "Product", "Quantity"]
        ws.append(headers)

        # Apply styles to headers
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for damage in van_damage_stock:
            ws.append([
                damage.created_date.strftime('%Y-%m-%d') if damage.created_date else "N/A",
                damage.van.salesman.get_fullname() if damage.van else "N/A",
                damage.van.get_van_route() if damage.van and damage.van.get_van_route() else "N/A",
                damage.product.product_name if damage.product else "N/A",
                damage.quantity
            ])

        # Create an HTTP response with the Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="van_damage_stock.xlsx"'
        wb.save(response)
        return response
    
    
class SalesmanCustomerRequestType_List(View):
    template_name = 'van_management/salesman_customer_request_type_list.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        request_li = SalesmanCustomerRequestType.objects.all()
        context = {'request_li': request_li}
        return render(request, self.template_name, context)
    
class SalesmanCustomerRequestType_Create(View):
    template_name = 'van_management/salesman_customerrequest_create.html'
    form_class = SalesmanCustomerRequestTypeForm

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = {'form': self.form_class}
        return render(request, self.template_name, context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = str(request.user.id)
            data.save()
            messages.success(request, 'Customer Request Successfully Added.', 'alert-success')
            return redirect('salesman_customer_request_type_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    print(f"Field: {field}, Error: {error}")
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, self.template_name, context)
        
        
class SalesmanCustomerRequestType_Edit(View):
    template_name = 'van_management/salesman_customer_requesttype_edit.html'
    form_class = SalesmanCustomerRequestType_Edit_Form

    @method_decorator(login_required)
    def get(self, request, pk, *args, **kwargs):
        rec = SalesmanCustomerRequestType.objects.get(id=pk)
        form = self.form_class(instance=rec)
        context = {'form': form,'rec':rec}
        return render(request, self.template_name, context)

    @method_decorator(login_required)
    def post(self, request, pk, *args, **kwargs):
        rec = SalesmanCustomerRequestType.objects.get(id=pk)
        form = self.form_class(request.POST, request.FILES, instance=rec)
        if form.is_valid():
            data = form.save(commit=False)
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            messages.success(request, 'Request Data Successfully Updated', 'alert-success')
            return redirect('salesman_customer_request_type_list')
        else:
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, self.template_name, context)

class SalesmanCustomerRequestType_Delete(View):

    def get(self, request, pk, *args, **kwargs):
        rec = get_object_or_404(SalesmanCustomerRequestType, id=pk)
        rec.delete()
        messages.success(request, 'Request type deleted successfully', 'alert-success')
        return redirect('salesman_customer_request_type_list')

    def post(self, request, pk, *args, **kwargs):
        rec = get_object_or_404(SalesmanCustomerRequestType, id=pk)
        rec.delete()
        messages.success(request, 'Request type deleted successfully', 'alert-success')
        return redirect('salesman_customer_request_type_list')
    
    
def salesman_customer_requests_list(request):
    requests = SalesmanCustomerRequests.objects.all()
    return render(request, 'van_management/requests_list.html', {'requests': requests})

def update_request_status(request, pk):
    request_instance = get_object_or_404(SalesmanCustomerRequests, id=pk)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status:
            request_instance.status = new_status
            request_instance.save()

            SalesmanCustomerRequestStatus.objects.create(
                salesman_customer_request=request_instance,
                status=new_status,
                modified_by=request.user.username
            )
            messages.success(request, "Status updated successfully!")
        else:
            messages.error(request, "Invalid status update.")

    return redirect('salesman_customer_requests_list')

#--------------------------------------auditing-------------------------------------------
from van_management.templatetags.van_template_tags import get_audit_details
def audit_report(request):
    audits = AuditBase.objects.select_related('route', 'marketing_executieve').annotate(
        total_customers=Count('route__customer_route', distinct=True),
        audited_customers=Count('audit_details__customer', distinct=True)
    ).order_by('-start_date')

    return render(request, 'van_management/audit_report.html', {'audits': audits})


def audit_detail(request, audit_id):
    audit = get_object_or_404(AuditBase, id=audit_id)
    audit_details = audit.audit_details.all()

    totals = {
        'previous_outstanding_amount': Decimal("0.00"),
        'audit_outstanding_amount': Decimal("0.00"),
        'amount_variation': Decimal("0.00"),
        'previous_outstanding_coupon': 0,
        'audit_outstanding_coupon': 0,
        'coupon_variation': 0,
        'previous_bottle': 0,
        'audit_bottle': 0,
        'bottle_variation': 0,
        'previous_hot_and_cold': 0,
        'audit_hot_and_cold': 0,
        'hot_and_cold_variation': 0,
        'previous_table': 0,
        'audit_table': 0,
        'table_variation': 0,
    }

    for item in audit_details:
        prev_amount = item.previous_outstanding_amount or 0
        cur_amount = item.outstanding_amount or 0
        prev_coupon = item.previous_outstanding_coupon or 0
        cur_coupon = item.outstanding_coupon or 0
        prev_bottle = item.previous_bottle_outstanding or 0
        cur_bottle = item.bottle_outstanding or 0

        amount_variation = prev_amount - cur_amount
        coupon_variation = prev_coupon - cur_coupon
        bottle_variation = prev_bottle - cur_bottle

        totals['previous_outstanding_amount'] += prev_amount
        totals['audit_outstanding_amount'] += cur_amount
        totals['amount_variation'] += amount_variation

        totals['previous_outstanding_coupon'] += prev_coupon
        totals['audit_outstanding_coupon'] += cur_coupon
        totals['coupon_variation'] += coupon_variation

        totals['previous_bottle'] += prev_bottle
        totals['audit_bottle'] += cur_bottle
        totals['bottle_variation'] += bottle_variation
        
        prev_hotcold = item.previous_hot_and_cold_dispenser or 0
        cur_hotcold = item.hot_and_cold_dispenser or 0
        hotcold_variation = prev_hotcold - cur_hotcold

        prev_table = item.previous_table_dispenser or 0
        cur_table = item.table_dispenser or 0
        table_variation = prev_table - cur_table

        totals['previous_hot_and_cold'] += prev_hotcold
        totals['audit_hot_and_cold'] += cur_hotcold
        totals['hot_and_cold_variation'] += hotcold_variation

        totals['previous_table'] += prev_table
        totals['audit_table'] += cur_table
        totals['table_variation'] += table_variation

    context = {
        'audit': audit,
        'audit_details': audit_details,
        'totals': totals
    }

    return render(request, 'van_management/audit_detail.html', context)



def van_damage_stock_report(request):
    # Get filter parameters
    route_name = request.GET.get('route_name', '')
    product_name = request.GET.get('product_name', '')
    search_query = request.GET.get('q', '')

    # Get all available routes
    routes = RouteMaster.objects.all()
    if route_name:
        routes = routes.filter(route_name=route_name)

    # Collect total damage stock per route and product
    damage_data = []

    for route in routes:
        # Get all vans assigned to this route
        van_routes = Van_Routes.objects.filter(routes=route).select_related('van')
        vans = [van_route.van for van_route in van_routes]

        # Get damaged products for this route, grouped by product
        damaged_products = VanSaleDamage.objects.filter(van__in=vans).values(
            'product__product_name'
        ).annotate(total_quantity=Sum('quantity'))

        # Apply product filter
        if product_name:
            damaged_products = damaged_products.filter(product__product_name__icontains=product_name)

        # Apply search query filter
        if search_query:
            damaged_products = damaged_products.filter(
                Q(product__product_name__icontains=search_query)
            )

        # Append data
        for product in damaged_products:
            damage_data.append({
                'route_id': route.route_id,
                'route': route.route_name,
                'product_name': product['product__product_name'],
                'total_quantity': product['total_quantity'],
            })

    context = {
        'damage_data': damage_data,
        'van_routes': RouteMaster.objects.all(),
        'product_names': ProdutItemMaster.objects.all(),
        'filter_data': {
            'route_name': route_name,
            'product_name': product_name,
            'q': search_query
        },
    }
    return render(request, 'van_management/van_routewise_damage/van_damage_stock_report.html', context)


def van_damage_stock_print(request):
    # Get filter parameters
    route_name = request.GET.get('route_name', '')
    product_name = request.GET.get('product_name', '')
    search_query = request.GET.get('q', '')

    # Get all available routes
    routes = RouteMaster.objects.all()
    if route_name:
        routes = routes.filter(route_name=route_name)

    # Collect total damage stock per route and product
    damage_data = []

    for route in routes:
        # Get all vans assigned to this route
        van_routes = Van_Routes.objects.filter(routes=route).select_related('van')
        vans = [van_route.van for van_route in van_routes]

        # Get damaged products for this route, grouped by product
        damaged_products = VanSaleDamage.objects.filter(van__in=vans).values(
            'product__product_name'
        ).annotate(total_quantity=Sum('quantity'))

        # Apply product filter
        if product_name:
            damaged_products = damaged_products.filter(product__product_name__icontains=product_name)

        # Apply search query filter
        if search_query:
            damaged_products = damaged_products.filter(
                Q(product__product_name__icontains=search_query)
            )

        # Append data
        for product in damaged_products:
            damage_data.append({
                'route_id': route.route_id,
                'route': route.route_name,
                'product_name': product['product__product_name'],
                'total_quantity': product['total_quantity'],
            })

    context = {
        'damage_data': damage_data,
        'van_routes': RouteMaster.objects.all(),
        'product_names': ProdutItemMaster.objects.all(),
        'filter_data': {
            'route_name': route_name,
            'product_name': product_name,
            'q': search_query
        },
    }
    return render(request, 'van_management/van_routewise_damage/van_damage_stock_print.html', context)

def van_damage_stock_excel(request):
    # Get filter parameters
    route_name = request.GET.get('route_name', '')
    product_name = request.GET.get('product_name', '')
    search_query = request.GET.get('q', '')

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Van Damage Stock Report"

    # Add headers
    headers = ["Sl.No", "Route Name", "Product Name", "Total Quantity"]
    sheet.append(headers)

    # Apply bold style to headers
    for col_num, col_name in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=col_name).font = openpyxl.styles.Font(bold=True)

    # Get all available routes
    routes = RouteMaster.objects.all()
    if route_name:
        routes = routes.filter(route_name=route_name)

    # Collect total damage stock per route and product
    damage_data = []
    sl_no = 1  # Auto-incremented index

    for route in routes:
        # Get all vans assigned to this route
        van_routes = Van_Routes.objects.filter(routes=route).select_related('van')
        vans = [van_route.van for van_route in van_routes]

        # Get damaged products for this route, grouped by product
        damaged_products = VanSaleDamage.objects.filter(van__in=vans).values(
            'product__product_name'
        ).annotate(total_quantity=Sum('quantity'))

        # Apply product filter
        if product_name:
            damaged_products = damaged_products.filter(product__product_name__icontains=product_name)

        # Apply search query filter
        if search_query:
            damaged_products = damaged_products.filter(
                Q(product__product_name__icontains=search_query)
            )

        # Append data to Excel
        for product in damaged_products:
            sheet.append([
                sl_no,  # Auto-incremented serial number
                route.route_name,
                product['product__product_name'],
                product['total_quantity'],
            ])
            sl_no += 1  # Increment serial number

    # Prepare the response with the correct content type
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Van_Damage_Stock_Report.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response

def route_damage_detail(request, route_id):
    """
    Display all van sale damage details for a given route with filtering options.
    """
    # Get the selected route
    route = get_object_or_404(RouteMaster, route_id=route_id)

    # Get all vans assigned to this route
    van_routes = Van_Routes.objects.filter(routes=route).select_related("van")
    vans = [van_route.van for van_route in van_routes]  # Get only the van instances

    # Fetch all van sale damage records ONLY for the vans assigned to this route
    damage_records = VanSaleDamage.objects.filter(van__in=vans)

    # Get filter parameters from request
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    product_name = request.GET.get("product_name", "")
    van_name = request.GET.get("van_name", "")
    reason = request.GET.get("reason", "")
    search_query = request.GET.get("q", "")

    # Apply filters
    if start_date:
        damage_records = damage_records.filter(created_date__gte=start_date)
    if end_date:
        damage_records = damage_records.filter(created_date__lte=end_date)
    if product_name:
        damage_records = damage_records.filter(product__product_name__icontains=product_name)
    if van_name:
        damage_records = damage_records.filter(van__van_make__icontains=van_name)
    if reason:
        damage_records = damage_records.filter(reason__reason__icontains=reason)
    if search_query:
        damage_records = damage_records.filter(
            Q(product__product_name__icontains=search_query) |
            Q(reason__reason__icontains=search_query) |
            Q(van__van_make__icontains=search_query)
        )

    # Aggregate total quantities
    damage_records = damage_records.values(
        "id", "created_date", "product__product_name", "reason__reason", "van__van_make"
    ).annotate(total_quantity=Sum("quantity")).order_by("created_date")

    # Calculate grand total quantity
    grand_total_quantity = sum(item["total_quantity"] for item in damage_records)

    # Fetch all van names, product names, and damage reasons for dropdowns (but only for related records)
    available_vans = Van.objects.filter(van_id__in=[van.van_id for van in vans])
    available_products = ProdutItemMaster.objects.filter(id__in=damage_records.values_list("product__id", flat=True))
    available_reasons = ProductionDamageReason.objects.filter(id__in=damage_records.values_list("reason__id", flat=True))

    context = {
        "route": route,
        "damage_records": damage_records,
        "grand_total_quantity": grand_total_quantity,
        "available_vans": available_vans,
        "product_names": available_products,
        "damage_reasons": available_reasons,
        "filter_data": {
            "start_date": start_date,
            "end_date": end_date,
            "product_name": product_name,
            "van_name": van_name,
            "reason": reason,
            "q": search_query,
        },
    }

    return render(request, "van_management/van_routewise_damage/route_damage_detail.html", context)

def route_damage_detail_print(request, route_id):
    """
    Print view for van sale damage details of a given route.
    """
    # Get the selected route
    route = get_object_or_404(RouteMaster, route_id=route_id)

    # Get all vans assigned to this route
    van_routes = Van_Routes.objects.filter(routes=route).select_related("van")
    vans = [van_route.van for van_route in van_routes]  # Extract van instances

    # Fetch damage records only for the vans assigned to this route
    damage_records = VanSaleDamage.objects.filter(van__in=vans)

    # Get filter parameters
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    product_name = request.GET.get("product_name", "")
    van_name = request.GET.get("van_name", "")
    reason = request.GET.get("reason", "")
    search_query = request.GET.get("q", "")

    # Apply filters
    if start_date:
        damage_records = damage_records.filter(created_date__gte=start_date)
    if end_date:
        damage_records = damage_records.filter(created_date__lte=end_date)
    if product_name:
        damage_records = damage_records.filter(product__product_name__icontains=product_name)
    if van_name:
        damage_records = damage_records.filter(van__van_make__icontains=van_name)
    if reason:
        damage_records = damage_records.filter(reason__reason__icontains=reason)
    if search_query:
        damage_records = damage_records.filter(
            Q(product__product_name__icontains=search_query) |
            Q(reason__reason__icontains=search_query) |
            Q(van__van_make__icontains=search_query)
        )

    # Aggregate total quantities
    damage_records = damage_records.values(
        "created_date", "product__product_name", "reason__reason", "van__van_make"
    ).annotate(total_quantity=Sum("quantity")).order_by("created_date")

    # Calculate grand total quantity
    grand_total_quantity = sum(item["total_quantity"] for item in damage_records)

    # Fetch related dropdown data
    available_vans = Van.objects.filter(van_id__in=[van.van_id for van in vans])
    available_products = ProdutItemMaster.objects.filter(id__in=damage_records.values_list("product__id", flat=True))
    available_reasons = ProductionDamageReason.objects.filter(id__in=damage_records.values_list("reason__id", flat=True))

    context = {
        "route": route,
        "damage_records": damage_records,
        "grand_total_quantity": grand_total_quantity,
        "available_vans": available_vans,
        "product_names": available_products,
        "damage_reasons": available_reasons,
        "filter_data": {
            "start_date": start_date,
            "end_date": end_date,
            "product_name": product_name,
            "van_name": van_name,
            "reason": reason,
            "q": search_query,
        },
    }

    return render(request, "van_management/van_routewise_damage/route_damage_detail_print.html", context)


def edit_van_sale_damage(request, pk):
    instance = get_object_or_404(VanSaleDamage, pk=pk)
    van_instance = Van.objects.get(pk=instance.van.pk)

    if request.method == 'POST':
        form = VanSaleDamageForm(request.POST, instance=instance)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Save old values before update
                    old_damage_from = instance.damage_from
                    old_quantity = instance.quantity
                    old_product = instance.product

                    # Get van stock record
                    vanstock = VanProductStock.objects.get(
                        product=old_product,
                        created_date=datetime.now().date(),
                        van=van_instance
                    )

                    # 1️⃣ Revert old stock changes
                    if old_damage_from == "fresh_stock":
                        vanstock.stock += old_quantity
                    elif old_damage_from == "empty_can":
                        vanstock.empty_can_count += old_quantity
                    vanstock.damage_count -= old_quantity

                    # 2️⃣ Apply new data
                    updated_instance = form.save(commit=False)
                    new_damage_from = updated_instance.damage_from
                    new_quantity = updated_instance.quantity

                    if new_damage_from == "fresh_stock":
                        vanstock.stock -= new_quantity
                    elif new_damage_from == "empty_can":
                        vanstock.empty_can_count -= new_quantity
                    vanstock.damage_count += new_quantity

                    vanstock.save()
                    updated_instance.save()

                    # 3️⃣ Log activity
                    log_activity(
                        created_by=request.user.id,
                        description=f"Van Sale Damage updated for {updated_instance.product.product_name} "
                                    f"from {old_quantity} → {new_quantity} ({new_damage_from})."
                    )
                    
                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "Van  Updated Successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('van_damage_stock_report'),
                        "return": True,
                    }


            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(form, formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        form = VanSaleDamageForm(instance=instance)

    return render(request, 'van_management/van_routewise_damage/damage_edit.html', {'form': form, 'instance': instance})


@transaction.atomic
def delete_van_sale_damage(request, pk):
    """
    Deletes a VanSaleDamage record and adjusts van stock accordingly.
    """
    damage_instance = get_object_or_404(VanSaleDamage, pk=pk)
    van = damage_instance.van
    product = damage_instance.product
    quantity = damage_instance.quantity

    try:
        # Get today's van stock (or the stock for the damage date)
        vanstock = VanProductStock.objects.filter(
            product=product,
            created_date=damage_instance.created_date.date(),
            van=van
        ).first()

        if vanstock:
            # Reverse the stock adjustments
            if hasattr(damage_instance, "damage_from") and damage_instance.damage_from == "fresh_stock":
                vanstock.stock += quantity
            elif hasattr(damage_instance, "damage_from") and damage_instance.damage_from == "empty_can":
                vanstock.empty_can_count += quantity

            vanstock.damage_count -= quantity
            if vanstock.damage_count < 0:
                vanstock.damage_count = 0  # prevent negative value
            vanstock.save()

        # Delete the damage record
        damage_instance.delete()

        log_activity(request.user, f"Deleted Van Sale Damage for {product.product_name}, Quantity {quantity}")
        
        response_data = {
            "status": "true",
            "title": "Successfully Deleted",
            "message": "Van damage record deleted successfully and stock adjusted.",
            "reload": "true",
        }
                
        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    except Exception as e:
        response_data = {
            "status": "false",
            "title": "Deletion Failed",
            "message": str(e),
        }
        return HttpResponse(json.dumps(response_data), content_type='application/javascript')


def route_damage_detail_excel(request, route_id):
    """
    Export van sale damage details for a given route as an Excel file with a footer.
    """
    # Get the selected route
    route = get_object_or_404(RouteMaster, route_id=route_id)

    # Get all vans assigned to this route
    van_routes = Van_Routes.objects.filter(routes=route).select_related("van")
    vans = [van_route.van for van_route in van_routes]

    # Fetch all van sale damage records for the assigned vans
    damage_records = VanSaleDamage.objects.filter(van__in=vans)

    # Get filter parameters
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    product_name = request.GET.get("product_name", "")
    van_name = request.GET.get("van_name", "")
    reason = request.GET.get("reason", "")
    search_query = request.GET.get("q", "")

    # Apply filters
    if start_date:
        damage_records = damage_records.filter(created_date__gte=start_date)
    if end_date:
        damage_records = damage_records.filter(created_date__lte=end_date)
    if product_name:
        damage_records = damage_records.filter(product__product_name__icontains=product_name)
    if van_name:
        damage_records = damage_records.filter(van__van_make__icontains=van_name)
    if reason:
        damage_records = damage_records.filter(reason__reason__icontains=reason)
    if search_query:
        damage_records = damage_records.filter(
            Q(product__product_name__icontains=search_query) |
            Q(reason__reason__icontains=search_query) |
            Q(van__van_make__icontains=search_query)
        )

    # Aggregate total quantities
    damage_records = damage_records.values(
        "created_date", "product__product_name", "reason__reason", "van__van_make"
    ).annotate(total_quantity=Sum("quantity")).order_by("created_date")

    # Calculate the grand total
    grand_total_quantity = sum(item["total_quantity"] for item in damage_records)

    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Van Damage Stock Report"

    # Define column headers
    headers = ["Sl. No", "Date", "Van Name", "Product Name", "Reason", "Total Quantity"]
    sheet.append(headers)

    # Style headers (bold)
    for col in range(1, len(headers) + 1):
        sheet.cell(row=1, column=col).font = Font(bold=True)

    # Add data to the sheet with auto-incremented serial numbers
    sl_no = 1
    row_num = 2  # Start writing data from the second row
    for record in damage_records:
        sheet.append([
            sl_no,
            record["created_date"].strftime("%Y-%m-%d"),
            record["van__van_make"],
            record["product__product_name"],
            record["reason__reason"],
            record["total_quantity"],
        ])
        sl_no += 1
        row_num += 1  # Move to the next row

    # Add a footer row (Grand Total)
    footer_row = ["", "", "", "", "Grand Total:", grand_total_quantity]
    sheet.append(footer_row)

    # Style the footer row (bold)
    for col in range(5, 7):  # Make "Grand Total" and total quantity bold
        sheet.cell(row=row_num, column=col).font = Font(bold=True)

    # Prepare the response with the correct content type
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Van_And_Route_Damage_Stock_Report.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response

def freelancevan(request):
    all_van = Van.objects.filter(van_type="freelance")
    
    log_activity(
        created_by=request.user,
        description="Viewed all freelance vans "
    ) 
    context = {
        'all_van': all_van,
       
    }
    
    return render(request, 'van_management/freelancevan.html', context)

# def freelancevan_rate_change(request, pk):
#     van = get_object_or_404(Van, pk=pk)

#     # Ensure only freelance vans can have their rate changed
#     if van.van_type != 'freelance':
#         messages.error(request, "Only freelance vans can have their rate changed.")
#         return redirect('van_list')

#     # Fetch all rate change history
#     rate_change_history = FreelanceVehicleRateChange.objects.filter(van=van).order_by('-created_date')

#     # Get the last rate change entry for old rate reference
#     last_rate_change = rate_change_history.first()
#     old_rate = last_rate_change.new_rate if last_rate_change else 0  

#     if request.method == 'POST':
#         form = FreelanceVehicleRateChangeForm(request.POST)
#         if form.is_valid():
#             rate_change = form.save(commit=False)
#             rate_change.van = van
#             rate_change.old_rate = old_rate 
#             rate_change.created_by = request.user.username
#             rate_change.save()

#             messages.success(request, "Freelance van rate updated successfully.")
            
#             # Reload the same page after successful form submission
#             return redirect(reverse('freelancevan_rate_change', kwargs={'pk': pk}))
#         else:
#             messages.error(request, "Invalid form data. Please check the input.")

#     else:
#         form = FreelanceVehicleRateChangeForm(initial={'old_rate': old_rate})

#     return render(request,'van_management/freelancevan_rate_change.html',{'form': form,'van': van,'rate_change_history': rate_change_history}
#     )
    
    
class FreelanceVanRateHistoryView(View):
    template_name = 'van_management/freelance_van_rate_history.html'

    def get(self, request, pk, *args, **kwargs):
        van_instance = Van.objects.get(pk=pk)
        van_rate_instances = FreelanceVehicleRateChange.objects.filter(van=van_instance).order_by('-created_date')
        other_product_rate_instances = FreelanceVehicleOtherProductChargesChanges.objects.filter(van=van_instance).order_by('-created_date')
        print("other_product_rate_instances",other_product_rate_instances)
        
        last_rate_change = van_rate_instances.first()
        old_rate = last_rate_change.new_rate if last_rate_change else 0  
        # Create the form with the old_rate as initial data
        new_rate_form = FreelanceVehicleRateChangeForm(initial={'old_rate': old_rate})
        
        context = {
            "van_instance": van_instance,
            "van_rate_instances": van_rate_instances,
            "new_rate_form": new_rate_form,
            "other_product_rate_instances": other_product_rate_instances,
            "product_items": ProdutItemMaster.objects.exclude(product_name="5 gallon")
        }
        return render(request, self.template_name, context)
    
    def post(self, request, pk, *args, **kwargs):
        van_instance = Van.objects.get(pk=pk)
        van_rate_instances = FreelanceVehicleRateChange.objects.filter(van=van_instance).order_by('-created_date')
         # Ensure that there is at least one instance in van_rate_instances
        last_rate_change = van_rate_instances.first()
        old_rate = last_rate_change.new_rate if last_rate_change else 0 
    
        form = FreelanceVehicleRateChangeForm(data=request.POST)
        
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = str(request.user.username)
            data.old_rate = old_rate
            data.van = van_instance
            data.save()
            
            van_instance.rate = data.new_rate
            van_instance.save()
            
            # Log activity for rate update
            log_activity(
                created_by=request.user,
                description=f"Updated rate for van: {van_instance.van_make} (ID: {pk}) to {data.new_rate}"
            )
            
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Rate Updated successfully.",
                'redirect': 'true',
                "redirect_url": reverse('freelancevan_rate_change', kwargs={'pk': pk})
            }
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
        else:
            # Log failed update attempt
            log_activity(
                created_by=request.user,
                description=f"Failed to update rate for van: {van_instance.van_make} (ID: {pk}) due to invalid form data"
            )

            messages.error(request, 'Invalid form data. Please check the input.')

class FreelanceVanOtherProductRateChangeView(View):
    def post(self, request, pk, *args, **kwargs):
        van_instance = get_object_or_404(Van, pk=pk)
        
        if not FreelanceVehicleOtherProductCharges.objects.filter(van=van_instance).exists():
            product_items = ProdutItemMaster.objects.exclude(product_name="5 gallon")
            print("product_items",product_items)
            for product_item in product_items:
                current_rate = request.POST.get(str(product_item.pk))
                if not current_rate:
                    continue
                
                try:
                    current_rate = float(current_rate)
                except ValueError:
                    continue
                
                FreelanceVehicleOtherProductChargesChanges.objects.create(
                    created_by=request.user,
                    van=van_instance,
                    product_item=product_item,
                    privious_rate=product_item.rate,
                    current_rate=current_rate
                )
                
                van_charge, created = FreelanceVehicleOtherProductCharges.objects.get_or_create(
                    van=van_instance,
                    product_item=product_item,
                )
                van_charge.current_rate = current_rate
                van_charge.save()
                
                log_activity(
                    created_by=request.user,
                    description=f"Updated rate for van: {van_instance.van_name} (ID: {van_instance.van_id}) to {van_charge.current_rate}"
                )
        else:
            product_item = ProdutItemMaster.objects.get(pk=request.POST.get("other_product_item"))
            current_rate = request.POST.get("other_product_item_value")
            current_rate = float(current_rate)
            
            FreelanceVehicleOtherProductChargesChanges.objects.create(
                created_by=request.user,
                van=van_instance,
                product_item=product_item,
                privious_rate=product_item.rate,
                current_rate=current_rate
            )
            
            van_charge, created = FreelanceVehicleOtherProductCharges.objects.get_or_create(
                van=van_instance,
                product_item=product_item,
            )
            van_charge.current_rate = current_rate
            van_charge.save()
            
            log_activity(
                created_by=request.user,
                description=f"Updated rate for van: {van_instance.van_name} (ID: {van_instance.van_id}) to {van_charge.current_rate}"
            )
        
        response_data = {
            "status": "true",
            "title": "Successfully Created",
            "message": "Rate updated successfully.",
            'redirect': 'true',
            "redirect_url": reverse('freelance_van_rate_history', kwargs={'pk': pk})
        }
        return JsonResponse(response_data)

def freelance_van_bottle_issue(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)
    print("van.branch_id",van.salesman.branch_id)
    if request.method == 'POST':
        form = FreelanceVanProductIssueForm(request.POST)
        
  
        if form.is_valid():
            product = form.cleaned_data.get('product')
            empty_bottles = form.cleaned_data.get('empty_bottles', 0)
            extra_bottles = form.cleaned_data.get('extra_bottles', 0)
            status = form.cleaned_data.get('status')
            total_bottles_issued = extra_bottles + empty_bottles  

            if total_bottles_issued > van.bottle_count :  
                messages.error(request, "Bottle count out of range.")
                return redirect('freelance_van_bottle_issue', van_id=van_id)

            try:
                with transaction.atomic():
                    issue = FreelanceVanProductIssue.objects.create(
                        van=van,
                        product=product,
                        empty_bottles=empty_bottles,
                        extra_bottles=extra_bottles,
                        total_bottles_issued=total_bottles_issued,
                        issued_by=request.user.username,
                        issued_date=now(),
                        status=status,
                    )
                    
                    # van.bottle_count += total_bottles_issued
                    # van.save()

                    product_stock = ProductStock.objects.get(
                        branch=van.salesman.branch_id,
                        product_name=product,
                    )
                    product_stock.quantity -= total_bottles_issued
                    product_stock.save()

                    van_product_stock, _ = VanProductStock.objects.get_or_create(
                        created_date=now().date(),
                        product=product,
                        van=van,
                        defaults={'stock': 0}
                    )
                    van_product_stock.stock += total_bottles_issued
                    van_product_stock.save()
                    
                    if issue.status == "non_paid":
                        if product.product_name == "5 Gallon":
                            product_rate = FreelanceVehicleRateChange.objects.filter(van=van).latest("created_date").new_rate
                        else:
                            product_rate = FreelanceVehicleOtherProductCharges.objects.filter(van=van,product_item=product).latest("created_date").current_rate
                        
                        new_rate = product_rate
                        
                        total_amount = total_bottles_issued * new_rate
                        outstanding = FreelanceVanOutstanding.objects.create(
                            van=van,
                            created_by=request.user.id,
                            created_date=datetime.now(),
                            outstanding_amount=total_amount,
                            )

                    # Log success
                    log_activity(
                        created_by=request.user,
                        description=f"Van Bottle Issue recorded: {van} - {total_bottles_issued} bottles"
                    )

                    messages.success(request, "Van bottle issue recorded successfully!")
                    return redirect('freelancevan')  # Redirect to the van list page

            except IntegrityError as e:
                messages.error(request, f"Integrity error: {str(e)}")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

        else:
            messages.error(request, "Invalid form data. Please check your input.")

    else:
        form = FreelanceVanProductIssueForm()

    context = {
        'form': form,
        'van': van,
        'current_bottle_count': van.bottle_count,
    }
    return render(request, 'van_management/freelance_van_bottle_issue.html', context)

def freelance_van_outstanding_details(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)
    
    outstanding = FreelanceVanOutstanding.objects.filter( van=van)
    
    context = {
        'van': van,
        'outstanding': outstanding,
    }
    
    return render(request, 'van_management/freelance_van_outstanding_details.html', context)

def freelance_van_issue_report(request):
    vans = Van.objects.filter(van_type="freelance")

    report_data = []
    for van in vans:
        total_outstanding = FreelanceVanOutstanding.objects.filter(van=van).aggregate(
            total_outstanding=Sum('outstanding_amount')
        )['total_outstanding'] or 0 

        report_data.append({
            "van": van,
            "total_outstanding": total_outstanding,
        })

    context = {
        "report_data": report_data
    }
    return render(request, "van_management/freelance_van_issue_report.html", context)


def freelance_van_issue_list(request, van_id):
    van = get_object_or_404(Van, van_id=van_id)
    issues = FreelanceVanProductIssue.objects.filter(van=van)

    context = {
        "van": van,
        "issues": issues
    }
    return render(request, "van_management/freelance_van_issue_list.html", context)

