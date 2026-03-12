import random
import uuid
import json
import datetime
import openpyxl

from client_management.serializers import CustomerCreditSerializer
from coupon_management.views import delete_coupon_recharge
from customer_care.models import DiffBottlesModel
from invoice_management.models import Invoice, InvoiceDailyCollection, InvoiceItems
from van_management.models import *
from decimal import Decimal
from collections import defaultdict


from django.views import View
from django.db.models import Q, Sum, Count, DecimalField, F, IntegerField, Value, Case, When, Value, CharField, ExpressionWrapper
from django.urls import reverse
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory, inlineformset_factory
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import localtime
from rest_framework.serializers import Serializer
from rest_framework import status
from rest_framework.response import Response

from .forms import  *
from .models import *
from product.models import *
from accounts.models import *
from master.functions import generate_form_errors, generate_invoice_no, generate_receipt_no
from master.models import RouteMaster, LocationMaster
from van_management.models import Van, Van_Routes
from sales_management.models import CollectionItems, Receipt
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import BasicAuthentication

from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from competitor_analysis.forms import CompetitorAnalysisFilterForm
from django.db.models import Q
import openpyxl

from django.core.files.storage import default_storage
import os
from django.conf import settings
import pandas as pd
from datetime import datetime
from django.utils import timezone
from apiservices.notification import *
from master.functions import log_activity
from client_management.templatetags.client_templatetags import get_outstanding_amount, get_outstanding_bottles, get_outstanding_coupons


def customer_custody_item(request,customer_id):
    customer_instance = Customers.objects.get(customer_id=customer_id)
    CustodyItemsFormset = formset_factory(CustodyCustomItemForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        custody_custom_form = CustodyCustomForm(request.POST)
        custody_items_formset = CustodyItemsFormset(request.POST,prefix='custody_items_formset', form_kwargs={'empty_permitted': False})
        
        if custody_custom_form.is_valid() and custody_items_formset.is_valid():
            try:
                with transaction.atomic():
                    custody_custom_data = custody_custom_form.save(commit=False)
                    custody_custom_data.created_by=request.user.id
                    custody_custom_data.created_date=datetime.today()
                    custody_custom_data.modified_by=request.user.id
                    custody_custom_data.modified_date=datetime.today()
                    
                    custody_custom_data.customer=customer_instance
                    custody_custom_data.save()
                    
                    for form in custody_items_formset:
                        item = form.save(commit=False)
                        item.custody_custom = custody_custom_data
                        item.save()
                        
                        # Update bottle count if necessary
                        # if item.product.product_name == "5 Gallon":
                        #     bottle_count, created = BottleCount.objects.get_or_create(
                        #         van__salesman=request.user,
                        #         created_date__date=custody_custom_data.created_date.date(),
                        #         defaults={'custody_issue': item.quantity}
                        #     )
                        #     if not created:
                        #         bottle_count.custody_issue += item.quantity
                        #         bottle_count.save()

                        # Update or create CustomerCustodyStock
                        if CustomerCustodyStock.objects.filter(customer=custody_custom_data.customer, product=item.product).exists():
                            stock_instance = CustomerCustodyStock.objects.get(customer=custody_custom_data.customer, product=item.product)
                            stock_instance.quantity += item.quantity
                            stock_instance.serialnumber = (stock_instance.serialnumber + ',' + item.serialnumber) if stock_instance.serialnumber else item.serialnumber
                            stock_instance.agreement_no = (stock_instance.agreement_no + ',' + item.custody_custom.agreement_no) if stock_instance.agreement_no else item.custody_custom.agreement_no
                            stock_instance.save()
                        else:
                            CustomerCustodyStock.objects.create(
                                customer=custody_custom_data.customer,
                                agreement_no=custody_custom_data.agreement_no,
                                deposit_type=custody_custom_data.deposit_type,
                                reference_no=custody_custom_data.reference_no,
                                product=item.product,
                                quantity=item.quantity,
                                serialnumber=item.serialnumber,
                                amount=item.amount,
                                can_deposite_chrge=item.can_deposite_chrge,
                                five_gallon_water_charge=item.five_gallon_water_charge,
                                amount_collected=custody_custom_data.amount_collected
                            )
                        
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "custody Item created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('customers')
                    }
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(custody_custom_form, formset=False)
            message += generate_form_errors(custody_items_formset, formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        custody_custom_form = CustodyCustomForm()
        custody_items_formset = CustodyItemsFormset(prefix='custody_items_formset')
        
        context = {
            'custody_custom_form': custody_custom_form,
            'custody_items_formset': custody_items_formset,
            'customer_instance' : customer_instance,
            
            'page_title': 'Create Custody Item',
            'page_name' : 'create Custody Item',
        }
        
        return render(request,'client_management/add_custody_items.html',context)


# def customer_custody_item(request,customer_id):
#     template_name = 'client_management/add_custody_items.html'
#     if request.method == "GET":
#         customer_exists = Customers.objects.filter(is_guest=False, customer_id=customer_id).exists()
#         if customer_exists:
#             customer_data = Customers.objects.get(customer_id=customer_id)
#             products = ProdutItemMaster.objects.all()

#             price_list = []
#             for product in products:
#                 default_rates_exists = Product_Default_Price_Level.objects.filter(product_id=product,customer_type=customer_data.customer_type).exists()
#                 if default_rates_exists:
#                     default_rates = Product_Default_Price_Level.objects.get(product_id=product,customer_type=customer_data.customer_type)
#                     custody_items_exists = CustodyCustomItems.objects.filter(product_id=product,customer=customer_data.customer_id).exists()
#                     if custody_items_exists:
#                         custody_items = CustodyCustomItems.objects.get(product_id=product,customer=customer_data.customer_id)
#                         custody_item_id = custody_items.custody_item_id
#                         product_id = product.pk
#                         product_name = product.product_name
#                         product_rate = custody_items.rate
#                         product_count = custody_items.count
#                     else:
#                         custody_item_id = ''
#                         product_id = product.pk
#                         product_name = product.product_name
#                         product_rate = default_rates.rate
#                         product_count = 0
#                 else:
#                     custody_item_id = ''
#                     product_id = product.pk
#                     product_name = product.product_name
#                     product_rate = 0
#                     product_count = 0

#                 ite = {'custody_item_id':custody_item_id,'product_id': product_id, 'product_name': product_name,'product_rate':product_rate,'product_count':product_count}
#                 price_list.append(ite)
#         context = {'price_list': price_list,'customerid':customer_data.customer_id,'customername':customer_data.customer_name}
#         return render(request, template_name, context)
    
#     if request.method == 'POST':
#         customer_id = request.POST.get('id_customer')
#         product_ids = request.POST.getlist('price_checkbox')
#         rate = request.POST.getlist('rate')
#         count = request.POST.getlist('count')
#         id_custody_items = request.POST.getlist('id_custody_item')
#         if customer_id is not None and product_ids is not None:
#             customer_instance = Customers.objects.get(customer_id=customer_id)
#             for i, item_id in enumerate(product_ids):
#                 product_id, index = item_id.split('+')
#                 index = int(index) - 1

#                 product_instance = ProdutItemMaster.objects.get(pk=product_id)
#                 if id_custody_items[index]=='':
#                     CustodyCustomItems.objects.create(created_by=request.user,
#                                         customer=customer_instance,
#                                         rate=rate[index],
#                                         count=count[index],
#                                         product=product_instance)   
#                 else:
#                     customer_custody_instance = CustodyCustomItems.objects.get(custody_item_id=id_custody_items[index])
#                     customer_custody_instance.rate = rate[index]
#                     customer_custody_instance.count = count[index]
#                     customer_custody_instance.save()

#             messages.success(request, 'Custody Items Successfully Added.', 'alert-success')
#             return redirect('customers')
#         else:
#             messages.success(request, 'Data is not valid.', 'alert-danger')
#             context = {}
#     return render(request, template_name, context)


#ajax
def get_custody_items(request):
    if request.method == "GET":
        customer = request.GET['customer']
        if customer is not None:
            customer_exists = Customers.objects.filter(is_guest=False, customer_id=customer).exists()
            if customer_exists:
                customer_data = Customers.objects.get(customer_id=customer)
                branch_id=request.user.branch_id.branch_id
                branch = BranchMaster.objects.get(branch_id=branch_id)
                products = Product.objects.filter(branch_id=branch)
                price_list = []
                for product in products:
                   default_rates_exists = Product_Default_Price_Level.objects.filter(product_id=product,customer_type=customer_data.customer_type).exists()
                   if default_rates_exists:
                        default_rates = Product_Default_Price_Level.objects.get(product_id=product,customer_type=customer_data.customer_type)
                        custody_items_exists = CustodyCustomItems.objects.filter(product_id=product,customer=customer_data.customer_id).exists()
                        if custody_items_exists:
                            custody_items = CustodyCustomItems.objects.get(product_id=product,customer=customer_data.customer_id)
                            custody_item_id = custody_items.custody_item_id
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = custody_items.rate
                            product_count = custody_items.count
                        else:
                            custody_item_id = ''
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = default_rates.rate
                            product_count = 0
                   else:
                        custody_items_exists = CustodyCustomItems.objects.filter(product_id=product,customer=customer_data.customer_id).exists()
                        if custody_items_exists:
                            custody_items = CustodyCustomItems.objects.get(product_id=product,customer=customer_data.customer_id)
                            custody_item_id = custody_items.custody_item_id
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = custody_items.rate
                            product_count = custody_items.count
                        else:
                            custody_item_id = ''
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = default_rates.rate
                            product_count = 0

                   ite = {'custody_item_id':custody_item_id,'product_id': product_id, 'product_name': product_name,'product_rate':product_rate,'product_count':product_count}
                   price_list.append(ite)
            dat = {'price_list': price_list}   
        return JsonResponse(dat)
    
    # --------------------------------------------------------------------------

# Vaccation
def vacation_route_summary(request):
    today = date.today()

    vacations = (
        Vacation.objects
        .filter(end_date__gte=today)
        .select_related('customer', 'customer__routes')
    )

    route_data = defaultdict(lambda: {
        'vacation_customers': set(),
        'expected_drop': Decimal('0.00'),
        'avg_sales': Decimal('0.00'),
        'route_name': ''
    })

    for v in vacations:
        customer = v.customer
        route = customer.routes

        if not route:
            continue

        # Vacation days
        vacation_days = (v.end_date - v.start_date).days + 1

        # Daily sales
        daily_sales = (
            (customer.no_of_bottles_required or 0) *
            (customer.rate or 0)
        )

        # Expected drop for this vacation
        drop_amount = daily_sales * vacation_days

        route_id = route.route_id

        route_data[route_id]['route_name'] = route.route_name
        route_data[route_id]['vacation_customers'].add(customer.customer_id)
        route_data[route_id]['expected_drop'] += drop_amount
        route_data[route_id]['avg_sales'] += daily_sales
        
        route_summary = []

        for route_id, data in route_data.items():
            route_summary.append({
                'customer__routes__route_id': route_id,
                'customer__routes__route_name': data['route_name'],
                'vacation_customers': len(data['vacation_customers']),
                'expected_drop': round(data['expected_drop'], 2),
                'avg_sales': round(data['avg_sales'], 2)
            })

        route_summary.sort(key=lambda x: x['customer__routes__route_name'])
        
    return render(request, 'client_management/vacation_route_summary.html', {
            'route_summary': route_summary
        })


def vacation_list_by_route(request, route_id):
    today = date.today()

    # Get route info (for page title / breadcrumb)
    route = get_object_or_404(RouteMaster, route_id=route_id)

    # Get ACTIVE vacation customers for this route
    vacation = Vacation.objects.filter(
        customer__routes_id=route_id,
        customer__isnull=False,
        end_date__gte=today
    ).select_related('customer').order_by('start_date')

    context = {
        'vacation': vacation,
        'route': route
    }
    return render(request, 'client_management/customer_vacation_list.html', context)


def vacation_list(request):
    template = 'client_management/vacation_list.html'
    Vacation.objects.filter(end_date__lt= date.today()).delete()
    vacation = Vacation.objects.all()
    context = {'vacation':vacation}
    return render(request, template, context)

class RouteSelection(View):
    def get(self, request):
        template = 'client_management/select_route.html'
        routes = RouteMaster.objects.all()
        return render(request, template, {'routes': routes}) 
    
class Vacation_Add(View):
    def get(self, request):
        template = 'client_management/vacation_add.html'
        form = Vacation_Add_Form
        search_form = CustomerSearchForm()
        selected_route = request.GET.get('route')
        print('root',selected_route)
        customers = Customers.objects.filter(is_guest=False, routes = selected_route)
        search_query = request.GET.get('search_query')
        if search_query:
            customers = customers.filter(
                Q(customer_name__icontains=search_query) |
                Q(mobile_no__icontains=search_query) |
                Q(location__location_name__icontains=search_query) |
                Q(building_name__icontains=search_query) 
            )
        return render(request, template, {'form': form, 'search_form': search_form, 'customers': customers, 'selected_route':selected_route})

    def post(self, request):
        template = 'client_management/vacation_add.html'
        form = Vacation_Add_Form(request.POST)
        if form.is_valid():
            vacation = form.save()
            log_activity(
                created_by=request.user,
                description=f"Added vacation for customer '{vacation.customer.customer_name}' from {vacation.start_date} to {vacation.end_date}."
            )
            return redirect(vacation_route_summary)
        return render(request, template, {'form': form})
    
class Vacation_Edit(View):
    template = 'client_management/vacation_edit.html'
    def get(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        form = Vacation_Edit_Form(instance=vacation)
        return render(request, self.template, {'form': form, 'vacation': vacation})

    def post(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        form = Vacation_Edit_Form(request.POST, instance=vacation)
        if form.is_valid():
            form.save()
            log_activity(
                created_by=request.user,
                description=f"Updated vacation for customer '{vacation.customer.customer_name}' (ID: {vacation.vacation_id})."
            )
            return redirect(vacation_list)
        return render(request, self.template, {'form': form, 'vacation': vacation})

class Vacation_Delete(View):
    template='client_management/vacation_delete.html'
    def get(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        return render(request, self.template, {'vacation':vacation})

    def post(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        vacation.delete()
        log_activity(
            created_by=request.user,
            description=f"Deleted vacation for customer '{vacation.customer.customer_name}' (ID: {vacation.vacation_id})."
        )
        return redirect(vacation_list)
    


class CustomerCustodyList(View):
    template_name = 'client_management/custody_item/customer_custody_list.html'

    def get(self, request, *args, **kwargs):
        # Query all routes from the RouteMaster model
        route_li = RouteMaster.objects.all()

        # Base queryset for customer custody items
        user_li = CustodyCustomItems.objects.select_related('custody_custom__customer', 'product').all().order_by('-custody_custom__created_date')

        # Search query filtering
        query = request.GET.get("q")
        if query:
            user_li = user_li.filter(
                Q(custody_custom__customer__customer_name__icontains=query) |
                Q(custody_custom__customer__mobile_no__icontains=query) |
                Q(custody_custom__customer__building_name__icontains=query)
            )

        # Date filtering
        start_date = request.GET.get('from_date')
        print("start_date",start_date)
        end_date = request.GET.get('end_date')
        print("end_date",end_date)
        if start_date and end_date:
            user_li = user_li.filter(custody_custom__created_date__date__range=[start_date, end_date])

        # Route filtering
        route_filter = request.GET.get('route_name')
        if route_filter:
            user_li = user_li.filter(custody_custom__customer__routes__route_name=route_filter)

        # Aggregating the product counts for each customer
        aggregated_data = (
            user_li.values(
                'custody_custom__created_date',
                'custody_custom__customer__customer_name',
                'custody_custom__customer__mobile_no',
                'custody_custom__customer__building_name',
                'custody_custom__customer__routes__route_name',
                'custody_custom__customer__custom_id'
            )
            .annotate(
                five_gallon_count=Sum('quantity', filter=Q(product__product_name='5 Gallon')),
                dispenser_count=Sum('quantity', filter=Q(product__product_name='Dispenser')),
                water_cooler_count=Sum('quantity', filter=Q(product__product_name='Water Cooler'))
            )
        )

        context = {
            'user_li': aggregated_data,
            'route_li': route_li,
            'filter_data': {
                'start_date': start_date,
                'end_date': end_date,
                'route_name': route_filter,
            }
        }
        return render(request, self.template_name, context)
class AddCustodyItems(View):
    template_name = 'client_management/custody_item/add_custody_items.html'
    form_class = CustodyCustomItemForm

    def get(self, request):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            # Save the instance based on deposit_form value
            if instance.deposit_form:
                instance.amount = instance.deposit_amount
                instance.deposit_form_number = instance.deposit_number
            else:
                instance.amount = None
            instance.save()
            messages.success(request, 'Entry created successfully!')
            return redirect('add_custody_list')
        return render(request, self.template_name, {'form': form})

        
class AddCustodyList(View):
    template_name = 'client_management/custody_item/add_custody_list.html'

    def get(self, request):
        get_addedlist = CustodyCustomItems.objects.all()
        print(get_addedlist,'geddedlist')
        return render(request, self.template_name, {'get_addedlist': get_addedlist })
    
class EditCustodyItem(View):
    template_name = 'client_management/custody_item/add_custody_list.html'

    def get(self, request):
        get_addedlist = CustodyCustomItems.objects.all()
        print(get_addedlist,'geddedlist')
        return render(request, self.template_name, {'get_addedlist': get_addedlist })



class PulloutListView(View):
    template_name = 'client_management/pullout_list.html'

    # def get(self, request):
        # form = CustodyItemFilterForm(request.GET)
    def get(self, request, pk):
        customer = Customers.objects.get(customer_id=pk)
        print('customer',customer)
        custody_items = CustodyCustomItems.objects.filter(customer=customer)
        # custody_pullout_list = Customer_Custody_Items.objects.all()
        print("custody_pullout_list",list(custody_items))
        return render(request, self.template_name, {'custody_items': custody_items,'customer': customer})
    
@transaction.atomic
def change_supply_date(request):
    if request.method == "POST":
        supply_id = request.POST.get("supply_id")
        new_date = request.POST.get("supply_date")

        supply = get_object_or_404(CustomerSupply, pk=supply_id)

        

        # Update supply
        supply.supply_date = new_date
        supply.is_edited = True
        supply.reference_number = "supply date changed"
        # supply.modified_by = request.user.username
        supply.save(update_fields=['supply_date','is_edited','reference_number'])

        # Sync invoice
        Invoice.objects.filter(
            invoice_no=supply.invoice_no,
        ).update(invoice_date=new_date,reference_no="supply date changed")

        # Sync outstanding
        CustomerOutstanding.objects.filter(
            invoice_no=supply.invoice_no
        ).update(outstanding_date=new_date)

        messages.success(request, "Supply date updated successfully.")

    return redirect('customer_supply_list')

@login_required
def customer_supply_list(request):
    """
    Customer Supply List
    :param request:
    :return: CustomerSupplys list view
    """
    filter_data = {}
    
    instances = CustomerSupply.objects.all().order_by("-created_date")
    routes = RouteMaster.objects.all()
    
    if request.GET.get('start_date') and request.GET.get('end_date'):
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
    else:
        start_date_str = datetime.today().date().strftime('%Y-%m-%d')
        end_date_str = datetime.today().date().strftime('%Y-%m-%d')
        
    route_name = request.GET.get('route_name')
    supply_type = request.GET.get("supply_type")
    if supply_type == "Paid":
        instances = instances.filter(subtotal=models.F('amount_recieved'))
        filter_data['supply_type'] = supply_type
    elif supply_type == "Non Paid":
        instances = instances.exclude(subtotal=models.F('amount_recieved'))
        filter_data['supply_type'] = supply_type

        
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        instances = instances.filter(supply_date__date__gte=start_date, supply_date__date__lte=end_date)
        filter_data['start_date'] = start_date
        filter_data['end_date'] = end_date
        
        log_activity(
            created_by=request.user,
            description=f"Filtered customer supplies between {start_date} and {end_date}"
        )
    
    if route_name:
        instances = instances.filter(customer__routes__route_name=route_name)
        filter_data['route_name'] = route_name
        
        log_activity(
            created_by=request.user,
            description=f"Filtered customer supplies by route: {route_name}"
        )
        
    query = request.GET.get("q")
    if query:
        instances = instances.filter(
            Q(customer__custom_id__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__building_name__icontains=query)
        )
        filter_data['q'] = query
        
        log_activity(
            created_by=request.user,
            description=f"Searched customer supplies with query: {query}"
        )
        
    for instance in instances:
        instance.can_edit = (timezone.now().date() - instance.created_date.date()).days <= 3
        
    context = {
        'instances': instances,
        'page_name' : 'Customer Supply List',
        'page_title' : 'Customer Supply List',
        'routes': routes, 
        'is_customer_supply': True,
        'is_need_datetime_picker': True,
        
        'filter_data': filter_data,
    }

    return render(request, 'client_management/customer_supply/list.html', context)

def view_invoice(request, invoice_no):
    invoice = get_object_or_404(Invoice, invoice_no=invoice_no)  
    return render(request, 'client_management/view_invoice.html', {'invoice': invoice})

@login_required
def customer_supply_info(request,pk):
    """
    Customer Supply Info
    :param request:
    :return: CustomerSupplys Info view
    """
    
    supply = CustomerSupply.objects.get(pk=pk)
    instances = CustomerSupplyItems.objects.filter(customer_supply=supply).order_by("-customer_supply__created_date")
    # date_range = ""
    # date_range = request.GET.get('date_range')
    # # print(date_range)

    # if date_range:
    #     start_date_str, end_date_str = date_range.split(' - ')
    #     start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
    #     end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
    #     instances = instances.filter(date__range=[start_date, end_date])
    
    # filter_data = {}
    # query = request.GET.get("q")
    
    # if query:

    #     instances = instances.filter(
    #         Q(customer_supply_no__icontains=query) |
    #         Q(product__customer_supply_id__icontains=query) 
    #     )
    #     title = "Customer Supply List - %s" % query
    #     filter_data['q'] = query
    log_activity(
        created_by=request.user,
        description=f"Viewed customer supply info for supply ID: {supply.pk}"
    )
    context = {
        'supply': supply,
        'instances': instances,
        'page_name' : 'Customer Supply List',
        'page_title' : 'Customer Supply List',
        # 'filter_data' :filter_data,
        # 'date_range': date_range,
        
        'is_customer_supply': True,
        'is_need_datetime_picker': True,
    }

    return render(request, 'client_management/customer_supply/info.html', context)

@login_required
def customer_supply_customers(request):
    filter_data = {}
    
    instances = Customers.objects.filter(is_guest=False, is_deleted=False,is_active=True)
    
    if request.GET.get('route'):
        instances = instances.filter(routes__pk=request.GET.get('route'))
        
    if request.GET.get('building_no'):
        instances = instances.filter(door_house_no=request.GET.get('building_no'))
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(customer_id__icontains=query) |
            Q(mobile_no__icontains=query) |
            Q(whats_app__icontains=query) |
            Q(customer_name__icontains=query) 
        )
        title = "Customer Supply Customers - %s" % query
        filter_data['q'] = query
        
    route_instances = RouteMaster.objects.all()

    context = {
        'instances': instances,
        'route_instances' : route_instances,
        
        'page_title': 'Create customer supply',
        'customer_supply_page': True,
        'is_need_datetime_picker': True
    }
    
    return render(request,'client_management/customer_supply/customer_list.html',context)

# def create_customer_supply(request,pk):
    
#     customer_instance = Customers.objects.get(pk=pk)
#     SupplyItemsFormset = formset_factory(CustomerSupplyItemsForm, extra=2)
#     message = ''
    
#     customer_outstanding_coupon = CustomerOutstanding.objects.none
#     customer_outstanding_empty_can = CustomerOutstanding.objects.none
#     customer_outstanding_amount = CustomerOutstanding.objects.none
    
#     if request.method == 'POST':
#         supply_date_str = request.POST.get("adding_date")
#         balance_empty_bottle = 0
#         supply_date_part = datetime.strptime(supply_date_str, '%Y-%m-%d').date()
#         current_time_part = datetime.now().time()

#         supply_date = datetime.combine(supply_date_part, current_time_part)
        
#         customer_supply_form = CustomerSupplyForm(request.POST)
#         customer_supply_items_formset = SupplyItemsFormset(request.POST, request.FILES,prefix='customer_supply_items_formset',form_kwargs={'empty_permitted': False})
        
#         if customer_supply_form.is_valid() and customer_supply_items_formset.is_valid():
#             # try:
#             #     with transaction.atomic():
#             customer_suply_form_instance = customer_supply_form.save(commit=False)
#             customer_suply_form_instance.customer = customer_instance
#             customer_suply_form_instance.salesman = customer_instance.sales_staff
#             customer_suply_form_instance.created_by = customer_instance.sales_staff.pk
#             customer_suply_form_instance.created_date = supply_date
#             customer_suply_form_instance.save()
            
#             total_fivegallon_qty = 0
#             van = Van.objects.get(salesman=customer_suply_form_instance.customer.sales_staff)
            
#             for form in customer_supply_items_formset:
#                 item_data = form.save(commit=False)
#                 item_data.customer_supply = customer_suply_form_instance  # Associate with the customer supply instance
#                 item_data.save()
            
#                 vanstock = VanProductStock.objects.get(created_date=customer_suply_form_instance.created_date.date(), product=item_data.product, van=van)
#                 if customer_suply_form_instance.created_date.date() == datetime.today().date():
#                     if vanstock.stock >= item_data.quantity:
#                         vanstock.stock -= item_data.quantity
                        
#                         customer_suply_form_instance.van_stock_added = True
#                         customer_suply_form_instance.save()
                        
#                         vanstock.sold_count += item_data.quantity
#                         vanstock.pending_count += item_data.customer_supply.allocate_bottle_to_pending
                        
#                         if item_data.product.product_name == "5 Gallon":
#                             total_fivegallon_qty += Decimal(item_data.quantity)
#                             vanstock.empty_can_count += customer_suply_form_instance.collected_empty_bottle
                            
#                             customer_suply_form_instance.van_emptycan_added = True
#                             customer_suply_form_instance.save()
                            
#                         vanstock.save()
#                     else:
#                         response_data = {
#                             "status": "false",
#                             "title": "Failed",
#                             "message": f"No stock available in {item_data.product.product_name}, only {vanstock.stock} left",
#                         }
#                         return HttpResponse(json.dumps(response_data), content_type='application/javascript')    
                
#             invoice_generated = False
            
#             if customer_suply_form_instance.customer.sales_type != "FOC" :
#                 # empty bottle calculate
#                 if total_fivegallon_qty < Decimal(customer_suply_form_instance.collected_empty_bottle) :
#                     balance_empty_bottle = Decimal(customer_suply_form_instance.collected_empty_bottle) - total_fivegallon_qty
#                     if CustomerOutstandingReport.objects.filter(customer=customer_suply_form_instance.customer,product_type="emptycan").exists():
#                         outstanding_instance = CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="emptycan")
#                         outstanding_instance.value -= Decimal(balance_empty_bottle)
#                         outstanding_instance.save()
                
#                 elif total_fivegallon_qty > Decimal(customer_suply_form_instance.collected_empty_bottle) :
#                     balance_empty_bottle = total_fivegallon_qty - Decimal(customer_suply_form_instance.collected_empty_bottle)
#                     customer_outstanding_empty_can = CustomerOutstanding.objects.create(
#                         customer=customer_suply_form_instance.customer,
#                         product_type="emptycan",
#                         created_by=request.user.id,
#                         created_date=supply_date,
#                     )

#                     outstanding_product = OutstandingProduct.objects.create(
#                         empty_bottle=balance_empty_bottle,
#                         customer_outstanding=customer_outstanding_empty_can,
#                     )
#                     outstanding_instance = {}

#                     try:
#                         outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="emptycan")
#                         outstanding_instance.value += Decimal(outstanding_product.empty_bottle)
#                         outstanding_instance.save()
#                     except:
#                         outstanding_instance = CustomerOutstandingReport.objects.create(
#                             product_type='emptycan',
#                             value=outstanding_product.empty_bottle,
#                             customer=outstanding_product.customer_outstanding.customer
#                         )
                        
#                     customer_suply_form_instance.outstanding_bottle_added = True
#                     customer_suply_form_instance.save()
            
#                 supply_items = CustomerSupplyItems.objects.filter(customer_supply=customer_suply_form_instance) # supply items
                
#                 # Update CustomerSupplyStock
#                 for item_data in supply_items:
#                     customer_supply_stock, _ = CustomerSupplyStock.objects.get_or_create(
#                         customer=customer_suply_form_instance.customer,
#                         product=item_data.product,
#                     )
                    
#                     customer_supply_stock.stock_quantity += item_data.quantity
#                     customer_supply_stock.save()
                    
#                     if Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CASH COUPON" :
#                         # print("cash coupon")
#                         total_coupon_collected = request.POST.get('total_coupon_collected')
                        
#                         if request.POST.get('coupon_method') == "manual" :
#                             collected_coupon_ids = request.POST.getlist('collected_coupon_ids')
#                             # print(collected_coupon_ids)
                            
#                             for c_id in collected_coupon_ids:
#                                 customer_supply_coupon = CustomerSupplyCoupon.objects.create(
#                                     customer_supply=customer_suply_form_instance,
#                                 )
#                                 if CouponLeaflet.objects.filter(pk=c_id).exists():
#                                     leaflet_instance = CouponLeaflet.objects.get(pk=c_id)
#                                     customer_supply_coupon.leaf.add(leaflet_instance)
#                                     leaflet_instance.used=True
#                                     leaflet_instance.save()
                                    
#                                 if FreeLeaflet.objects.filter(pk=c_id).exists():
#                                     leaflet_instance = FreeLeaflet.objects.get(pk=c_id)
#                                     customer_supply_coupon.free_leaf.add(leaflet_instance)
#                                     leaflet_instance.used=True
#                                     leaflet_instance.save()
                                
#                                 if CustomerCouponStock.objects.filter(customer__pk=customer_suply_form_instance.customer.pk,coupon_method="manual",coupon_type_id=leaflet_instance.coupon.coupon_type).exists() :
#                                     customer_stock = CustomerCouponStock.objects.get(customer__pk=customer_suply_form_instance.customer.pk,coupon_method="manual",coupon_type_id=leaflet_instance.coupon.coupon_type)
#                                     customer_stock.count -= 1
#                                     customer_stock.save()
                                    
#                             if total_fivegallon_qty < len(collected_coupon_ids):
#                                 # print("total_fivegallon_qty < len(collected_coupon_ids)", total_fivegallon_qty, "------------------------", len(collected_coupon_ids))
#                                 balance_coupon = Decimal(total_fivegallon_qty) - Decimal(len(collected_coupon_ids))
                                
#                                 customer_outstanding_coupon = CustomerOutstanding.objects.create(
#                                     customer=customer_suply_form_instance.customer,
#                                     product_type="coupons",
#                                     created_by=request.user.id,
#                                 )
                                
#                                 customer_coupon = CustomerCouponStock.objects.filter(customer__pk=customer_suply_form_instance.customer.pk,coupon_method="manual").first()
#                                 outstanding_coupon = OutstandingCoupon.objects.create(
#                                     count=balance_coupon,
#                                     customer_outstanding=customer_outstanding_coupon,
#                                     coupon_type=customer_coupon.coupon_type_id
#                                 )
#                                 outstanding_instance = ""

#                                 try:
#                                     outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="coupons")
#                                     outstanding_instance.value += Decimal(outstanding_coupon.count)
#                                     outstanding_instance.save()
#                                 except:
#                                     outstanding_instance = CustomerOutstandingReport.objects.create(
#                                         product_type='coupons',
#                                         value=outstanding_coupon.count,
#                                         customer=outstanding_coupon.customer_outstanding.customer
#                                     )
                                    
#                                 customer_suply_form_instance.outstanding_bottle_added = True
#                                 customer_suply_form_instance.save()
                            
#                             elif total_fivegallon_qty > len(collected_coupon_ids) :
#                                 balance_coupon = total_fivegallon_qty - len(collected_coupon_ids)
#                                 try :
#                                     outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="coupons")
#                                     outstanding_instance.value += Decimal(balance_coupon)
#                                     outstanding_instance.save()
#                                 except:
#                                     outstanding_instance=CustomerOutstandingReport.objects.create(
#                                         product_type="coupons",
#                                         value=balance_coupon,
#                                         customer=customer_suply_form_instance.customer,
#                                         )
                                    
#                                 customer_suply_form_instance.outstanding_bottle_added = True
#                                 customer_suply_form_instance.save()
                                    
#                         elif request.POST.get('coupon_method') == "digital" :
#                             try : 
#                                 customer_coupon_digital = CustomerSupplyDigitalCoupon.objects.get(
#                                     customer_supply=customer_suply_form_instance,
#                                     )
#                             except:
#                                 customer_coupon_digital = CustomerSupplyDigitalCoupon.objects.create(
#                                     customer_supply=customer_suply_form_instance,
#                                     count = 0,
#                                     )
#                             customer_coupon_digital.count += total_coupon_collected
#                             customer_coupon_digital.save()
                            
#                             customer_stock = CustomerCouponStock.objects.get(customer__pk=customer_suply_form_instance.customer.pk,coupon_method="digital",coupon_type_id__coupon_type_name="Digital")
#                             customer_stock.count -= Decimal(total_coupon_collected)
#                             customer_stock.save()
                            
#                     elif Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CREDIT COUPON" :
#                         pass
#                     elif Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CASH" or Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CREDIT" :
#                         if customer_suply_form_instance.amount_recieved < customer_suply_form_instance.subtotal:
#                             balance_amount = customer_suply_form_instance.subtotal - customer_suply_form_instance.amount_recieved
                            
#                             customer_outstanding_amount = CustomerOutstanding.objects.create(
#                                 product_type="amount",
#                                 created_by=request.user.id,
#                                 customer=customer_suply_form_instance.customer,
#                                 created_date=datetime.today()
#                             )

#                             outstanding_amount = OutstandingAmount.objects.create(
#                                 amount=balance_amount,
#                                 customer_outstanding=customer_outstanding_amount,
#                             )
#                             outstanding_instance = {}

#                             try:
#                                 outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="amount")
#                                 outstanding_instance.value += Decimal(outstanding_amount.amount)
#                                 outstanding_instance.save()
#                             except:
#                                 outstanding_instance = CustomerOutstandingReport.objects.create(
#                                     product_type='amount',
#                                     value=outstanding_amount.amount,
#                                     customer=outstanding_amount.customer_outstanding.customer
#                                 )
                                
#                             customer_suply_form_instance.outstanding_amount_added = True
#                             customer_suply_form_instance.save()
                                
#                         elif customer_suply_form_instance.amount_recieved > customer_suply_form_instance.subtotal:
#                             balance_amount = customer_suply_form_instance.amount_recieved - customer_suply_form_instance.subtotal
                            
#                             customer_outstanding_amount = CustomerOutstanding.objects.create(
#                                 product_type="amount",
#                                 created_by=request.user.id,
#                                 customer=customer_suply_form_instance.customer,
#                             )

#                             outstanding_amount = OutstandingAmount.objects.create(
#                                 amount=balance_amount,
#                                 customer_outstanding=customer_outstanding_amount,
#                             )
                            
#                             outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="amount")
#                             outstanding_instance.value -= Decimal(balance_amount)
#                             outstanding_instance.save()
                            
#                             customer_suply_form_instance.outstanding_amount_added = True
#                             customer_suply_form_instance.save()
                            
#                     # elif Customers.objects.get(pk=customer_suply_form_instance.customer).sales_type == "CREDIT" :
#                         # pass
                
#                 # if customer_suply_form_instance.customer.sales_type == "CASH" or customer_suply_form_instance.customer.sales_type == "CREDIT":
#                 invoice_generated = True
                
#                 if not customer_suply_form_instance.reference_number:
#                     reference_no = f"{customer_suply_form_instance.customer.custom_id}"
#                 else:
#                     reference_no = customer_suply_form_instance.reference_number
#                 # print("invoice_number",invoice_number)    
#                 # Create the invoice
#                 invoice = Invoice.objects.create(
#                     invoice_no=generate_invoice_no(customer_suply_form_instance.created_date.date()),
#                     created_date=customer_suply_form_instance.created_date,
#                     net_taxable=customer_suply_form_instance.net_payable,
#                     vat=customer_suply_form_instance.vat,
#                     discount=customer_suply_form_instance.discount,
#                     amout_total=customer_suply_form_instance.subtotal,
#                     amout_recieved=customer_suply_form_instance.amount_recieved,
#                     customer=customer_suply_form_instance.customer,
#                     reference_no=reference_no
#                 )
                
#                 customer_suply_form_instance.invoice_no = invoice.invoice_no
#                 customer_suply_form_instance.save()
                
#                 # print(customer_outstanding_empty_can)
#                 # if customer_outstanding_empty_can :
#                 #     customer_outstanding_empty_can.invoice_no = invoice.invoice_no
#                 #     customer_outstanding_empty_can.save()
                
#                 # print(customer_outstanding_amount)
#                 # if customer_outstanding_amount :
#                 #     customer_outstanding_amount.invoice_no = invoice.invoice_no
#                 #     customer_outstanding_amount.save()
                    
#                 # if Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CASH COUPON" and customer_outstanding_coupon :
#                 #     customer_outstanding_coupon.invoice_no = invoice.invoice_no
#                 #     customer_outstanding_coupon.save()
                
#                 if customer_suply_form_instance.customer.sales_type == "CREDIT":
#                     invoice.invoice_type = "credit_invoice"
#                     invoice.save()

#                 # Create invoice items
#                 for item_data in supply_items:
#                     item = CustomerSupplyItems.objects.get(pk=item_data.pk)
                    
#                     InvoiceItems.objects.create(
#                         category=item.product.category,
#                         product_items=item.product,
#                         qty=item.quantity,
#                         rate=item.amount,
#                         invoice=invoice,
#                         remarks='invoice genereted from supply items reference no : ' + invoice.reference_no
#                     )
#                     # print("invoice generate")
#                     InvoiceDailyCollection.objects.create(
#                         invoice=invoice,
#                         created_date=supply_date,
#                         customer=invoice.customer,
#                         salesman=request.user,
#                         amount=invoice.amout_recieved,
#                     )

#                 DiffBottlesModel.objects.filter(
#                     delivery_date__date=supply_date,
#                     assign_this_to=customer_suply_form_instance.salesman,
#                     customer=customer_suply_form_instance.customer
#                     ).update(status='supplied')
                
                
#             if customer_instance.sales_staff:
#                 sales_man = customer_instance.sales_staff
#                 try:
#                     salesman_body = f'Customer Supply created successfully and Invoice generated. Fresh cans supplied: {item_data.quantity} ,Empty cans collected:{balance_empty_bottle},Amount collected:{invoice.amout_recieved}'
#                     notification(sales_man.pk, "Dear Customer", salesman_body, "Nationalwaterfcm")
#                     notification(customer_instance.user_id.pk, "New  Supply", "Your  Supply Created Succesfull.", "Nationalwatercustomer")
#                 except CustomUser.DoesNotExist:
#                     messages.error(request, 'Salesman does not exist.', 'alert-danger')
#                 except Send_Notification.DoesNotExist:
#                     messages.error(request, 'No device token found for the salesman.', 'alert-danger')
#                 except Exception as e:
#                     messages.error(request, f'Error sending notification: {e}', 'alert-danger')  
                    
#             log_activity(
#                 created_by=request.user,
#                 description=f'Customer Supply for customer {customer_instance.customer_name} created successfully and Invoice generated. Fresh cans supplied: {item_data.quantity} ,Empty cans collected:{balance_empty_bottle},Amount collected:{invoice.amout_recieved}'
#             )          
#             if invoice_generated:
#                 response_data = {
#                     "status": "true",
#                     "title": "Successfully Created",
#                     "message": "Customer Supply created successfully and Invoice generated.",
#                     'redirect': 'true',
#                     "redirect_url": reverse('customer_supply_list'),
#                     "return": True,
#                 }
#             else:
#                 response_data = {
#                     "status": "true",
#                     "title": "Successfully Created",
#                     "message": "customer supply Created Successfully.",
#                     'redirect': 'true',
#                     "redirect_url": reverse('customer_supply_list'),
#                     "return": True,
#             }
                    
#             # except IntegrityError as e:
#             #     # Handle database integrity error
#             #     response_data = {
#             #         "status": "false",
#             #         "title": "Failed",
#             #         "message": str(e),
#             #     }

#             # except Exception as e:
#             #     # Handle other exceptions
#             #     response_data = {
#             #         "status": "false",
#             #         "title": "Failed",
#             #         "message": str(e),
#             #     }
#         else:
#             message = generate_form_errors(customer_supply_form,formset=False)
#             message += generate_form_errors(customer_supply_items_formset,formset=True)
            
#             response_data = {
#                 "status": "false",
#                 "title": "Failed",
#                 "message": message,
#             }

#         return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
#     else:
#         customer_supply_form = CustomerSupplyForm()
#         customer_supply_items_formset = SupplyItemsFormset(prefix='customer_supply_items_formset',form_kwargs={'empty_permitted': False})
        
#         context = {
#             'customer_instance': customer_instance,
#             'customer_supply_form': customer_supply_form,
#             'customer_supply_items_formset': customer_supply_items_formset,
            
#             'page_title': 'Create customer supply',
#             'customer_supply_page': True,
#             'is_need_datetime_picker': True
#         }
        
#         return render(request,'client_management/customer_supply/create.html',context)


def create_customer_supply(request,pk):
    print("working")
    customer_instance = Customers.objects.get(pk=pk)
    route = customer_instance.routes

    van_route = Van_Routes.objects.filter(routes=route).select_related("van").first()
    van = van_route.van if van_route else None
    salesman = van.salesman if van else None
    # print("customer_instance.sales_staff:",customer_instance.sales_staff)
    van_instance = Van.objects.get(salesman=salesman)
    
    SupplyItemsFormset = formset_factory(CustomerSupplyItemsForm, extra=2)
    message = ''
    
    if request.method == 'POST':
        supply_date_str = request.POST.get("adding_date")
        if not supply_date_str:
            return JsonResponse({
                "status": "false",
                "message": "Supply date is required."
            })

        try:
            supply_date_part = datetime.strptime(supply_date_str, '%Y-%m-%d').date()
            supply_date = datetime.combine(supply_date_part, datetime.now().time())
        except ValueError:
            return JsonResponse({
                "status": "false",
                "message": "Invalid supply date format."
            })

        # Safe to continue here
        balance_empty_bottle = 0
        total_fivegallon_qty = 0

        customer_supply_form = CustomerSupplyForm(request.POST)
        SupplyItemsFormset = formset_factory(CustomerSupplyItemsForm, extra=2)
        customer_supply_items_formset = SupplyItemsFormset(
            request.POST, request.FILES,
            prefix='customer_supply_items_formset',
            form_kwargs={'empty_permitted': False}
        )

        if customer_supply_form.is_valid() and customer_supply_items_formset.is_valid():
            try:
                with transaction.atomic():
                    supply_instance = customer_supply_form.save(commit=False)
                    supply_instance.customer = customer_instance
                    supply_instance.salesman = customer_instance.sales_staff
                    supply_instance.created_by = customer_instance.sales_staff.pk
                    supply_instance.created_date = supply_date
                    supply_instance.supply_date = supply_date
                    supply_instance.save()
                    
                    DiffBottlesModel.objects.filter(
                        delivery_date__date=supply_instance.created_date.date(),
                        assign_this_to=supply_instance.salesman_id,
                        customer=supply_instance.customer_id
                        ).update(status='supplied')

                    print("Salesman:",customer_instance.sales_staff)
                    print("Salesman_id:",customer_instance.sales_staff.pk)
            
                    invoice = Invoice.objects.create(
                        # invoice_no=generate_invoice_no(supply_instance.created_date.date()),
                        created_date=supply_instance.created_date,
                        invoice_date=supply_instance.created_date,
                        net_taxable=supply_instance.net_payable,
                        vat=supply_instance.vat,
                        discount=supply_instance.discount,
                        amout_total=supply_instance.subtotal,
                        amout_recieved=supply_instance.amount_recieved,
                        customer=supply_instance.customer,
                        reference_no=supply_instance.reference_number,
                        salesman=customer_instance.sales_staff
                    )
                    
                    # print("invoice generate")
                    InvoiceDailyCollection.objects.create(
                        invoice=invoice,
                        created_date=supply_instance.created_date,
                        customer=invoice.customer,
                        salesman=request.user,
                        amount=invoice.amout_recieved,
                    )
                    
                    supply_instance.invoice_no = invoice.invoice_no
                    supply_instance.save()

                    if supply_instance.amount_recieved == supply_instance.subtotal:
                        invoice_numbers = []
                        invoice_numbers.append(invoice.invoice_no)
                            
                        receipt = Receipt.objects.create(
                            transaction_type="supply",
                            instance_id=str(supply_instance.id),  
                            amount_received=supply_instance.amount_recieved,
                            customer=supply_instance.customer,
                            invoice_number=",".join(invoice_numbers),
                            created_date=supply_instance.created_date
                        )
                    
                    for item_form in customer_supply_items_formset:
                        item_data = item_form.save(commit=False)
                        item_data.customer_supply = supply_instance
                        item_data.save()

                        customer_supply_stock, _ = CustomerSupplyStock.objects.get_or_create(
                        customer=supply_instance.customer,
                        product=item_data.product,
                            )

                        customer_supply_stock.stock_quantity += item_data.quantity
                        customer_supply_stock.save()
                        
                        InvoiceItems.objects.create(
                            category=item_data.product.category,
                            product_items=item_data.product,
                            qty=item_data.quantity,
                            rate=item_data.amount,
                            invoice=invoice,
                            remarks='invoice genereted from supply items reference no : ' + invoice.reference_no
                        )

                        # Van stock update (as needed)
                        if datetime.today().date() == supply_instance.created_date.date():
                            vanstock = VanProductStock.objects.get(
                                created_date=supply_instance.created_date.date(), product=item_data.product, van=van_instance
                            )
                            if vanstock.stock < item_data.quantity:
                                response_data = {
                                    "status": "false",
                                    "title": "Failed",
                                    "message": f"No stock available in {item_data.product.product_name}, only {vanstock.stock} left",
                                }
                                return HttpResponse(json.dumps(response_data), content_type='application/javascript')

                            vanstock.stock -= (item_data.quantity + supply_instance.allocate_bottle_to_free)
                            if supply_instance.customer.sales_type != "FOC":
                                vanstock.sold_count += item_data.quantity
                            else:
                                vanstock.foc += item_data.quantity
                                supply_instance.van_foc_added = True

                            if item_data.product.product_name == "5 Gallon":
                                total_fivegallon_qty += item_data.quantity
                                
                                vanstock.empty_can_count += supply_instance.collected_empty_bottle
                                supply_instance.van_emptycan_added = True
                                
                            if supply_instance.allocate_bottle_to_free > 0:
                                # vanstock.sold_count += item_data.quantity
                                vanstock.foc += supply_instance.allocate_bottle_to_free
                                supply_instance.van_foc_added = True

                            vanstock.save()

                        supply_instance.van_stock_added = True
                        supply_instance.save()
                    
                    # Save coupon logic
                    if supply_instance.customer.sales_type == "CASH COUPON":
                        collected_coupon_ids = request.POST.getlist('collected_coupon_ids')
                        coupon_method = request.POST.get('coupon_method')
                        total_coupon_collected = request.POST.get('total_coupon_collected')
                        
                        if coupon_method == "manual" and collected_coupon_ids:
                            coupon_entry = CustomerSupplyCoupon.objects.create(customer_supply=supply_instance)

                            for cid in collected_coupon_ids:
                                is_free_leaf = False

                                # Check and add to regular CouponLeaflet
                                if CouponLeaflet.objects.filter(pk=cid).exists():
                                    paid_leaflet = CouponLeaflet.objects.get(pk=cid)
                                    coupon_entry.leaf.add(paid_leaflet)
                                    paid_leaflet.used = True
                                    paid_leaflet.save()
                                    print("regular leaf", paid_leaflet.leaflet_name)

                                    # Reduce from stock
                                    coupon_type = paid_leaflet.coupon.coupon_type
                                    coupon_stock = CustomerCouponStock.objects.filter(
                                        customer=supply_instance.customer,
                                        coupon_method=paid_leaflet.coupon.coupon_method,
                                        coupon_type_id=coupon_type
                                    ).first()
                                    if coupon_stock and coupon_stock.count > 0:
                                        coupon_stock.count -= 1
                                        coupon_stock.save(update_fields=["count"])

                                # Check and add to FreeLeaflet
                                if FreeLeaflet.objects.filter(pk=cid).exists():
                                    free_leaflet = FreeLeaflet.objects.get(pk=cid)
                                    coupon_entry.free_leaf.add(free_leaflet)
                                    free_leaflet.used = True
                                    free_leaflet.save()
                                    is_free_leaf = True
                                    # print("free leaf", free_leaflet.leaflet_name)

                                    # Reduce from stock
                                    coupon_type = free_leaflet.coupon.coupon_type
                                    coupon_stock = CustomerCouponStock.objects.filter(
                                        customer=supply_instance.customer,
                                        coupon_method=free_leaflet.coupon.coupon_method,
                                        coupon_type_id=coupon_type
                                    ).first()
                                    if coupon_stock and coupon_stock.count > 0:
                                        coupon_stock.count -= 1
                                        coupon_stock.save(update_fields=["count"])

                            # Balance check and outstanding logic
                            if total_fivegallon_qty > int(total_coupon_collected):
                                balance = total_fivegallon_qty - int(total_coupon_collected)
                            elif total_fivegallon_qty < int(total_coupon_collected):
                                balance = Decimal(total_coupon_collected) - Decimal(total_fivegallon_qty)
                            else:
                                balance = 0

                            if balance:
                                customer_coupon_type = (
                                    CustomerCouponStock.objects.filter(
                                        customer=supply_instance.customer,
                                        coupon_method="manual"
                                    ).first().coupon_type_id
                                    if CustomerCouponStock.objects.filter(
                                        customer=supply_instance.customer,
                                        coupon_method="manual"
                                    ).exists() else
                                    CouponType.objects.get(coupon_type_name="Digital")
                                )

                                outstanding_obj = CustomerOutstanding.objects.create(
                                    customer=supply_instance.customer,
                                    product_type="coupons",
                                    created_by=customer_instance.sales_staff.pk,
                                    created_date=supply_instance.created_date,
                                    outstanding_date=supply_instance.created_date,
                                    invoice_no=invoice.invoice_no
                                )

                                OutstandingCoupon.objects.create(
                                    count=balance,
                                    customer_outstanding=outstanding_obj,
                                    coupon_type=customer_coupon_type
                                )

                                report, created = CustomerOutstandingReport.objects.get_or_create(
                                    customer=supply_instance.customer,
                                    product_type="coupons",
                                    defaults={"value": balance}
                                )
                                if not created:
                                    report.value += Decimal(balance)
                                    report.save()

                                supply_instance.outstanding_coupon_added = True
                                supply_instance.save()

                        elif coupon_method == "digital":
                            # Digital coupon update
                            digital_coupon_obj, created = CustomerSupplyDigitalCoupon.objects.get_or_create(
                                customer_supply=supply_instance,
                                defaults={"count": 0}
                            )
                            digital_coupon_obj.count += Decimal(total_coupon_collected)
                            digital_coupon_obj.save()

                            customer_stock = CustomerCouponStock.objects.get(
                                customer=supply_instance.customer,
                                coupon_method="digital",
                                coupon_type__coupon_type_name="Digital"
                            )
                            customer_stock.count -= Decimal(total_coupon_collected)
                            customer_stock.save()

                            # Balance logic
                            if total_fivegallon_qty > Decimal(total_coupon_collected):
                                balance = total_fivegallon_qty - Decimal(total_coupon_collected)
                            elif total_fivegallon_qty < Decimal(total_coupon_collected):
                                balance = Decimal(total_coupon_collected) - total_fivegallon_qty
                            else:
                                balance = 0

                            if balance:
                                customer_coupon_type = CouponType.objects.get(coupon_type_name="Digital")

                                outstanding_obj = CustomerOutstanding.objects.create(
                                    customer=supply_instance.customer,
                                    product_type="coupons",
                                    created_by=customer_instance.sales_staff.pk,
                                    created_date=supply_instance.created_date,
                                    outstanding_date=supply_instance.created_date,
                                    invoice_no=invoice.invoice_no
                                )

                                OutstandingCoupon.objects.create(
                                    count=balance,
                                    customer_outstanding=outstanding_obj,
                                    coupon_type=customer_coupon_type
                                )

                                report, created = CustomerOutstandingReport.objects.get_or_create(
                                    customer=supply_instance.customer,
                                    product_type="coupons",
                                    defaults={"value": balance}
                                )
                                if not created:
                                    report.value += Decimal(balance)
                                    report.save()

                                supply_instance.outstanding_coupon_added = True
                                supply_instance.save()

                    # custody amount section 
                    # Assuming `customer` and `customer_supply` are already retrieved earlier in the code:
                    if supply_instance.customer.sales_type in ["CASH", "CREDIT"]:
                        if supply_instance.amount_recieved < supply_instance.subtotal:
                            balance_amount = supply_instance.subtotal - supply_instance.amount_recieved

                            customer_outstanding_amount = CustomerOutstanding.objects.create(
                                product_type="amount",
                                created_by=customer_instance.sales_staff.pk,
                                customer=supply_instance.customer,
                                created_date=supply_instance.created_date,
                                outstanding_date=supply_instance.created_date,
                                invoice_no=invoice.invoice_no
                            )

                            outstanding_amount = OutstandingAmount.objects.create(
                                amount=balance_amount,
                                customer_outstanding=customer_outstanding_amount,
                            )

                            try:
                                outstanding_instance = CustomerOutstandingReport.objects.get(
                                    customer=supply_instance.customer, product_type="amount"
                                )
                                outstanding_instance.value += Decimal(outstanding_amount.amount)
                                outstanding_instance.save()
                            except CustomerOutstandingReport.DoesNotExist:
                                CustomerOutstandingReport.objects.create(
                                    product_type='amount',
                                    value=outstanding_amount.amount,
                                    customer=supply_instance.customer
                                )

                            supply_instance.outstanding_amount_added = True
                            supply_instance.save()

                        elif supply_instance.amount_recieved > supply_instance.subtotal:
                            balance_amount = supply_instance.amount_recieved - supply_instance.subtotal

                            customer_outstanding_amount = CustomerOutstanding.objects.create(
                                product_type="amount",
                                created_by=customer_instance.sales_staff.pk,
                                customer=supply_instance.customer,
                                created_date=supply_instance.created_date,
                                outstanding_date=supply_instance.created_date,
                            )

                            outstanding_amount = OutstandingAmount.objects.create(
                                amount=balance_amount,
                                customer_outstanding=customer_outstanding_amount,
                            )

                            try:
                                outstanding_instance = CustomerOutstandingReport.objects.get(
                                    customer=supply_instance.customer, product_type="amount"
                                )
                                outstanding_instance.value -= Decimal(balance_amount)
                                outstanding_instance.save()
                            except CustomerOutstandingReport.DoesNotExist:
                                CustomerOutstandingReport.objects.create(
                                    product_type='amount',
                                    value=outstanding_amount.amount,
                                    customer=supply_instance.customer
                                )
                            supply_instance.outstanding_amount_added = True
                            supply_instance.save()

                    if customer_instance.sales_staff:
                        sales_man = customer_instance.sales_staff
                        try:
                            salesman_body = f'Customer Supply created successfully and Invoice generated. Fresh cans supplied: {item_data.quantity} ,Empty cans collected:{balance_empty_bottle},Amount collected:{invoice.amout_recieved}'
                            notification(sales_man.pk, "Dear Customer", salesman_body, "Nationalwaterfcm")
                            notification(customer_instance.user_id.pk, "New  Supply", "Your  Supply Created Succesfull.", "Nationalwatercustomer")
                        except CustomUser.DoesNotExist:
                            messages.error(request, 'Salesman does not exist.', 'alert-danger')
                        except Send_Notification.DoesNotExist:
                            messages.error(request, 'No device token found for the salesman.', 'alert-danger')
                        except Exception as e:
                            messages.error(request, f'Error sending notification: {e}', 'alert-danger')  
                            
                    log_activity(
                        created_by=request.user,
                        description=f'Customer Supply for customer {customer_instance.customer_name} created successfully and Invoice generated. Fresh cans supplied: {item_data.quantity} ,Empty cans collected:{balance_empty_bottle},Amount collected:{invoice.amout_recieved}'
                    )          
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Customer Supply created successfully and Invoice generated.",
                        'redirect': 'true',
                        "redirect_url": reverse('customer_supply_list'),
                        "return": True,
                    }
                    
            except IntegrityError as e:
                print("error",e)
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                print("error",e)
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
            print("response==============",response_data)
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        customer_supply_form = CustomerSupplyForm()
        customer_supply_items_formset = SupplyItemsFormset(prefix='customer_supply_items_formset',form_kwargs={'empty_permitted': False})
        
        context = {
            'customer_instance': customer_instance,
            'customer_supply_form': customer_supply_form,
            'customer_supply_items_formset': customer_supply_items_formset,
            
            'page_title': 'Create customer supply',
            'customer_supply_page': True,
            'is_need_datetime_picker': True,
            'customer_rate': customer_instance.rate,
        }
        
        return render(request,'client_management/customer_supply/create.html',context)


@login_required
def edit_customer_supply(request, pk):
    """
    edit operation of customer_supply
    :param request:
    :param pk:
    :return:
    """
    message = ''
    
    customer_instance = CustomerSupply.objects.get(pk=pk).customer
    salesman_instance = CustomerSupply.objects.get(pk=pk).salesman
    created_date = CustomerSupply.objects.get(pk=pk).created_date
    customer_supply_instance = CustomerSupply.objects.get(pk=pk)
    old_invoice_number = CustomerSupply.objects.get(pk=pk).invoice_no
    supply_items_instances = CustomerSupplyItems.objects.filter(customer_supply=customer_supply_instance)
    
    if supply_items_instances.exists():
        extra = 0
    else:
        extra = 1 

    SupplyItemsFormset = inlineformset_factory(
        CustomerSupply,
        CustomerSupplyItems,
        extra=extra,
        form=CustomerSupplyItemsForm,
    )
    
    if request.method == 'POST':
        customer_outstanding_coupon = None
        customer_outstanding_empty_can = None
        customer_outstanding_amount = None
        
        customer_supply_form = EditCustomerSupplyForm(request.POST)
        customer_supply_items_formset = SupplyItemsFormset(
            request.POST, request.FILES,
            instance=customer_supply_instance,
            prefix='customer_supply_items_formset',
            form_kwargs={'empty_permitted': False}
        )
        
        if customer_supply_form.is_valid() and customer_supply_items_formset.is_valid():
            try:
                with transaction.atomic():
                    # #create
                    five_gallon_qty = supply_items_instances.filter(product__product_name="5 Gallon").aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
                    customer_supply_instance = get_object_or_404(CustomerSupply, pk=pk)
                    supply_items_instances = CustomerSupplyItems.objects.filter(customer_supply=customer_supply_instance)
                    five_gallon_qty = supply_items_instances.filter(product__product_name="5 Gallon").aggregate(total_quantity=Sum('quantity', output_field=DecimalField()))['total_quantity'] or 0
                    
                    DiffBottlesModel.objects.filter(
                        delivery_date__date=customer_supply_instance.created_date.date(),
                        assign_this_to=customer_supply_instance.salesman,
                        customer=customer_supply_instance.customer
                    ).update(status='pending')
                    
                    # Handle invoice related deletions
                    # handle_invoice_deletion(customer_supply_instance)
                    
                    # Handle outstanding amount adjustments
                    # handle_outstanding_amounts(customer_supply_instance, five_gallon_qty)
                    
                    # Handle coupon deletions and adjustments
                    handle_coupons(customer_supply_instance, five_gallon_qty)
                    
                    handle_outstanding_coupon(customer_supply_instance, five_gallon_qty)
                    
                    handle_empty_bottle_outstanding(customer_supply_instance, five_gallon_qty)
                    
                    # Update van product stock and empty bottle counts
                    update_van_product_stock(customer_supply_instance, supply_items_instances, five_gallon_qty)
                    
                    # Mark customer supply and items as deleted
                    customer_supply_instance.delete()
                    supply_items_instances.delete()
                    
                    # create new supply
                    customer_suply_form_instance = customer_supply_form.save(commit=False)
                    customer_suply_form_instance.customer = customer_instance
                    customer_suply_form_instance.salesman = salesman_instance
                    customer_suply_form_instance.created_by = customer_instance.sales_staff.pk
                    customer_suply_form_instance.created_date = created_date
                    customer_suply_form_instance.is_edited = True
                    customer_suply_form_instance.save()
                    
                    total_fivegallon_qty = 0
                    van = Van.objects.get(salesman=customer_suply_form_instance.salesman)
                    
                    for form in customer_supply_items_formset:
                        item_data = form.save(commit=False)
                        item_data.customer_supply = customer_suply_form_instance  # Associate with the customer supply instance
                        item_data.save()

                        if created_date.date() == datetime.today().date():
                            vanstock = VanProductStock.objects.get(created_date=created_date.date(), product=item_data.product, van=van)
                            
                            if vanstock.stock >= item_data.quantity:
                                
                                if customer_suply_form_instance.customer.sales_type != "FOC" :
                                    vanstock.sold_count += item_data.quantity
                                
                                if customer_suply_form_instance.customer.sales_type == "FOC" :
                                    vanstock.foc += item_data.quantity
                                    
                                if item_data.product.product_name == "5 Gallon" :
                                    total_fivegallon_qty += Decimal(item_data.quantity)
                                    vanstock.empty_can_count += customer_suply_form_instance.collected_empty_bottle
                                    
                                    if customer_suply_form_instance.customer.customer_type == "WATCHMAN" :
                                        vanstock.foc += customer_suply_form_instance.allocate_bottle_to_free
                                    
                                vanstock.stock -= item_data.quantity
                                vanstock.pending_count += item_data.customer_supply.allocate_bottle_to_pending
                                vanstock.save()
                            else:
                                response_data = {
                                    "status": "false",
                                    "title": "Failed",
                                    "message": f"No stock available in {item_data.product.product_name}, only {vanstock.stock} left",
                                }
                                return HttpResponse(json.dumps(response_data), content_type='application/javascript')    
                        
                    invoice_generated = False
                    
                    if customer_suply_form_instance.customer.sales_type != "FOC" :
                        # empty bottle calculate
                        if total_fivegallon_qty < Decimal(customer_suply_form_instance.collected_empty_bottle) :
                            balance_empty_bottle = Decimal(customer_suply_form_instance.collected_empty_bottle) - total_fivegallon_qty
                            if CustomerOutstandingReport.objects.filter(customer=customer_suply_form_instance.customer,product_type="emptycan").exists():
                                outstanding_instance = CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="emptycan")
                                outstanding_instance.value -= Decimal(balance_empty_bottle)
                                outstanding_instance.save()
                        
                        elif total_fivegallon_qty > Decimal(customer_suply_form_instance.collected_empty_bottle) :
                            balance_empty_bottle = total_fivegallon_qty - Decimal(customer_suply_form_instance.collected_empty_bottle)
                            
                            customer_outstanding_empty_can = CustomerOutstanding.objects.create(
                                customer=customer_suply_form_instance.customer,
                                product_type="emptycan",
                                created_by=customer_instance.sales_staff.pk,
                                created_date=created_date,
                                outstanding_date=created_date,
                            )

                            outstanding_product = OutstandingProduct.objects.create(
                                empty_bottle=balance_empty_bottle,
                                customer_outstanding=customer_outstanding_empty_can,
                            )
                            outstanding_instance = {}

                            try:
                                outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="emptycan")
                                outstanding_instance.value += Decimal(outstanding_product.empty_bottle)
                                outstanding_instance.save()
                            except:
                                outstanding_instance = CustomerOutstandingReport.objects.create(
                                    product_type='emptycan',
                                    value=outstanding_product.empty_bottle,
                                    customer=outstanding_product.customer_outstanding.customer
                                )
                    
                        supply_items = CustomerSupplyItems.objects.filter(customer_supply=customer_suply_form_instance) # supply items
                        
                        # Update CustomerSupplyStock
                        for item_data in supply_items:
                            customer_supply_stock, _ = CustomerSupplyStock.objects.get_or_create(
                                customer=customer_suply_form_instance.customer,
                                product=item_data.product,
                            )
                            
                            customer_supply_stock.stock_quantity += item_data.quantity
                            customer_supply_stock.save()
                            
                            if Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CASH COUPON" :
                                # print("cash coupon")
                                total_coupon_collected = request.data.get('total_coupon_collected')
                                
                                if request.data.get('coupon_method') == "manual" :
                                    collected_coupon_ids = request.data.get('collected_coupon_ids')
                                    
                                    for c_id in collected_coupon_ids:
                                        customer_supply_coupon = CustomerSupplyCoupon.objects.create(
                                            customer_supply=customer_suply_form_instance,
                                        )
                                        leaflet_instance = CouponLeaflet.objects.get(pk=c_id)
                                        customer_supply_coupon.leaf.add(leaflet_instance)
                                        leaflet_instance.used=True
                                        leaflet_instance.save()
                                        
                                        if CustomerCouponStock.objects.filter(customer__pk=customer_suply_form_instance.customer.pk,coupon_method="manual",coupon_type_id=leaflet_instance.coupon.coupon_type).exists() :
                                            customer_stock = CustomerCouponStock.objects.get(customer__pk=customer_suply_form_instance.customer.pk,coupon_method="manual",coupon_type_id=leaflet_instance.coupon.coupon_type)
                                            customer_stock.count -= 1
                                            customer_stock.save()
                                            
                                    if total_fivegallon_qty < len(collected_coupon_ids):
                                        # print("total_fivegallon_qty < len(collected_coupon_ids)", total_fivegallon_qty, "------------------------", len(collected_coupon_ids))
                                        balance_coupon = Decimal(total_fivegallon_qty) - Decimal(len(collected_coupon_ids))
                                        
                                        customer_outstanding_coupon = CustomerOutstanding.objects.create(
                                            customer=customer_suply_form_instance.customer,
                                            product_type="coupons",
                                            created_by=customer_instance.sales_staff.pk,
                                            created_date=created_date,
                                            outstanding_date=created_date,
                                        )
                                        
                                        customer_coupon = CustomerCouponStock.objects.filter(customer__pk=customer_suply_form_instance.customer,coupon_method="manual").first()
                                        outstanding_coupon = OutstandingCoupon.objects.create(
                                            count=balance_coupon,
                                            customer_outstanding=customer_outstanding_coupon,
                                            coupon_type=customer_coupon.coupon_type_id
                                        )
                                        outstanding_instance = ""

                                        try:
                                            outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="coupons")
                                            outstanding_instance.value += Decimal(outstanding_coupon.count)
                                            outstanding_instance.save()
                                        except:
                                            outstanding_instance = CustomerOutstandingReport.objects.create(
                                                product_type='coupons',
                                                value=outstanding_coupon.count,
                                                customer=outstanding_coupon.customer_outstanding.customer
                                            )
                                    
                                    elif total_fivegallon_qty > len(collected_coupon_ids) :
                                        balance_coupon = total_fivegallon_qty - len(collected_coupon_ids)
                                        try :
                                            outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="coupons")
                                            outstanding_instance.value += Decimal(balance_coupon)
                                            outstanding_instance.save()
                                        except:
                                            outstanding_instance=CustomerOutstandingReport.objects.create(
                                                product_type="coupons",
                                                value=balance_coupon,
                                                customer=customer_suply_form_instance.customer,
                                                )
                                            
                                elif request.data.get('coupon_method') == "digital" :
                                    try : 
                                        customer_coupon_digital = CustomerSupplyDigitalCoupon.objects.get(
                                            customer_supply=customer_suply_form_instance,
                                            )
                                    except:
                                        customer_coupon_digital = CustomerSupplyDigitalCoupon.objects.create(
                                            customer_supply=customer_suply_form_instance,
                                            count = 0,
                                            )
                                    customer_coupon_digital.count += total_coupon_collected
                                    customer_coupon_digital.save()
                                    
                                    customer_stock = CustomerCouponStock.objects.get(customer__pk=customer_suply_form_instance.customer.pk,coupon_method="digital",coupon_type_id__coupon_type_name="Digital")
                                    customer_stock.count -= Decimal(total_coupon_collected)
                                    customer_stock.save()
                                    
                            elif Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CREDIT COUPON" :
                                pass
                            elif Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CASH" or Customers.objects.get(pk=customer_suply_form_instance.customer.pk).sales_type == "CREDIT" :
                                
                                if customer_suply_form_instance.amount_recieved < customer_suply_form_instance.subtotal:
                                    print("Working if")
                                    balance_amount = customer_suply_form_instance.subtotal - customer_suply_form_instance.amount_recieved
                                    print("if 1")
                                    update_or_create_outstanding_amount(
                                        customer=customer_suply_form_instance.customer,
                                        old_invoice_no=old_invoice_number,
                                        new_balance_amount=balance_amount,
                                        created_date=created_date,
                                        user=request.user
                                    )
                                    print("if 2")
                                    outstanding_instance = {}

                                    try:
                                        print("if 3")
                                        
                                        outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="amount")
                                        print("if 4")
                                        outstanding_instance.value = balance_amount
                                        print("if 5")
                                        outstanding_instance.save()
                                        print("if 6")
                                    except:
                                        print("if 7")
                                        outstanding_instance = CustomerOutstandingReport.objects.create(
                                            product_type='amount',
                                            value=outstanding_amount.amount,
                                            customer=customer_outstanding_amount.customer_outstanding.customer
                                        )
                                    print("if complete")
                                        
                                elif customer_suply_form_instance.amount_recieved > customer_suply_form_instance.subtotal:
                                    print("Working elif")
                                    balance_amount = customer_suply_form_instance.amount_recieved - customer_suply_form_instance.subtotal
                                    print("elif 1")
                                    credit_amount = -balance_amount
                                    update_or_create_outstanding_amount(
                                        customer=customer_suply_form_instance.customer,
                                        old_invoice_no=old_invoice_number,
                                        new_balance_amount=credit_amount,
                                        created_date=created_date,
                                        user=request.user
                                    )
                                    print("elif 3")
                                    outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_suply_form_instance.customer,product_type="amount")
                                    outstanding_instance.value -= Decimal(balance_amount)
                                    outstanding_instance.save()
                                else :
                                    balance_amount = customer_suply_form_instance.amount_recieved - customer_suply_form_instance.subtotal
                                    update_or_create_outstanding_amount(
                                        customer=customer_suply_form_instance.customer,
                                        old_invoice_no=old_invoice_number,
                                        new_balance_amount=balance_amount,
                                        created_date=created_date,
                                        user=request.user
                                    )
                                    
                            # elif Customers.objects.get(pk=customer_suply_form_instance.customer).sales_type == "CREDIT" :
                                # pass
                        
                        # if customer_suply_form_instance.customer.sales_type == "CASH" or customer_suply_form_instance.customer.sales_type == "CREDIT":
                        invoice_generated = True
                        # print(old_invoice_number)
                        if not old_invoice_number == None :
                            invoice_number = old_invoice_number
                        else:
                            try:
                                invoice_last_no = Invoice.objects.filter(is_deleted=False).latest('created_date')
                                last_invoice_number = invoice_last_no.invoice_no
                                prefix, date_part, number_part = last_invoice_number.split('-')
                                new_number_part = int(number_part) + 1
                                invoice_number = f'{prefix}-{date_part}-{new_number_part:04d}'
                            except Invoice.DoesNotExist:
                                date_part = created_date.strftime('%Y%m%d')
                                random_part = str(random.randint(1000, 9999))
                                invoice_number = f'WTR-{date_part}-{random_part}'
                        
                        if not customer_suply_form_instance.reference_number:
                            reference_no = f"{customer_suply_form_instance.customer.custom_id}"
                        else:
                            reference_no = customer_suply_form_instance.reference_number
                        # print("invoice")
                        # Create the invoice
                        invoice = Invoice.objects.create(
                            invoice_no=invoice_number,
                            created_date=created_date,
                            net_taxable=customer_suply_form_instance.net_payable,
                            vat=customer_suply_form_instance.vat,
                            discount=customer_suply_form_instance.discount,
                            amout_total=customer_suply_form_instance.subtotal,
                            amout_recieved=customer_suply_form_instance.amount_recieved,
                            customer=customer_suply_form_instance.customer,
                            reference_no=reference_no,
                            salesman=request.user
                        )
                        # print(invoice.invoice_no)
                        customer_suply_form_instance.invoice_no = invoice.invoice_no
                        customer_suply_form_instance.save()
                        
                        if customer_outstanding_empty_can:
                            customer_outstanding_empty_can.invoice_no = invoice.invoice_no
                            customer_outstanding_empty_can.save()
                        
                        if customer_outstanding_coupon:
                            customer_outstanding_coupon.invoice_no = invoice.invoice_no
                            customer_outstanding_coupon.save()

                        if customer_outstanding_amount:
                            customer_outstanding_amount.invoice_no = invoice.invoice_no
                            customer_outstanding_amount.save()
                        
                        if customer_suply_form_instance.customer.sales_type == "CREDIT":
                            invoice.invoice_type = "credit_invoice"
                            invoice.save()

                        # Create invoice items
                        for item_data in supply_items:
                            item = CustomerSupplyItems.objects.get(pk=item_data.pk)
                            
                            InvoiceItems.objects.create(
                                category=item.product.category,
                                product_items=item.product,
                                qty=item.quantity,
                                rate=item.amount,
                                invoice=invoice,
                                remarks='invoice genereted from supply items reference no : ' + invoice.reference_no
                            )
                            # print("invoice generate")
                            InvoiceDailyCollection.objects.create(
                                invoice=invoice,
                                created_date=customer_suply_form_instance.created_date,
                                customer=invoice.customer,
                                salesman=request.user,
                                amount=invoice.amout_recieved,
                            )

                        DiffBottlesModel.objects.filter(
                            delivery_date__date=created_date.date(),
                            assign_this_to=customer_suply_form_instance.salesman,
                            customer=customer_suply_form_instance.customer
                            ).update(status='supplied')
                    
                    log_activity(
                        created_by=request.user,
                        description=f'Customer Supply for customer {customer_instance.customer_name} Updated successfully. '
                    )
                    if invoice_generated:
                        response_data = {
                            "status": "true",
                            "title": "Successfully Updated",
                            "message": "Customer Supply updated successfully and Invoice generated.",
                            'redirect': 'true',
                            "redirect_url": reverse('customer_supply_list'),
                            "return": True,
                        }
                    else:
                        response_data = {
                            "status": "true",
                            "title": "Successfully Updated",
                            "message": "customer supply Updated Successfully.",
                            'redirect': 'true',
                            "redirect_url": reverse('customer_supply_list'),
                            "return": True,
                    }
                        
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(customer_supply_form, formset=False)
            message += generate_form_errors(customer_supply_items_formset, formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
                        
    else:
        customer_supply_form = EditCustomerSupplyForm(instance=customer_supply_instance)
        customer_supply_items_formset = SupplyItemsFormset(
            queryset=supply_items_instances,
            prefix='customer_supply_items_formset',
            instance=customer_supply_instance
        )

        context = {
            'customer_supply_form': customer_supply_form,
            'customer_supply_items_formset': customer_supply_items_formset,
            'customer_instance': customer_supply_instance.customer,
            
            'message': message,
            'page_name': 'edit customer supply',
            'customer_supply_page': True,   
            'is_edit': True,        
        }

        return render(request, 'client_management/customer_supply/create.html', context)

    
@login_required
def delete_customer_supply(request, pk):
    """
    customer_supply deletion, it only marks as is_deleted field to true
    :param request:
    :param pk:
    :return:
    """
    try:
        with transaction.atomic():
            customer_supply_instance = get_object_or_404(CustomerSupply, pk=pk)
            supply_items_instances = CustomerSupplyItems.objects.filter(customer_supply=customer_supply_instance)
            five_gallon_qty = supply_items_instances.filter(product__product_name="5 Gallon").aggregate(total_quantity=Sum('quantity', output_field=DecimalField()))['total_quantity'] or 0
            
            qs = CustomerSupply.objects.filter(customer=customer_supply_instance.customer,created_date__date=customer_supply_instance.created_date.date(),invoice_no=customer_supply_instance.invoice_no)

            if qs.count() == 1:
                DiffBottlesModel.objects.filter(
                    delivery_date__date=customer_supply_instance.created_date.date(),
                    assign_this_to=customer_supply_instance.salesman,
                    customer=customer_supply_instance.customer_id
                    ).update(status='pending')
                        
            # Handle invoice related deletions
            handle_invoice_deletion(customer_supply_instance)
                    
                    # Handle outstanding amount adjustments
            handle_outstanding_amounts(customer_supply_instance, five_gallon_qty)
                    
                    # Handle coupon deletions and adjustments
            handle_coupons(customer_supply_instance, five_gallon_qty)
                    
            handle_outstanding_coupon(customer_supply_instance, five_gallon_qty)
                    
            handle_empty_bottle_outstanding(customer_supply_instance, five_gallon_qty)
                    
            # Update van product stock and empty bottle counts
            update_van_product_stock(customer_supply_instance, supply_items_instances, five_gallon_qty)
                    
            CustomerOutstanding.objects.filter(invoice_no=customer_supply_instance.invoice_no).delete()
                    
                    # Mark customer supply and items as deleted
            customer_supply_instance.delete()
            supply_items_instances.delete()
            log_activity(
                created_by=request.user,
                description=f"Customer supply with ID {pk} was successfully deleted."
            )        
            response_data = {
                        "status": "true",
                        "title": "Successfully Deleted",
                        "message": "Customer supply successfully deleted.",
                        "reload": "true",
                    }
                    
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    except Exception as e:
        log_activity(
            created_by=request.user,
            description=f"Failed to delete customer supply with ID {pk}. Error: {str(e)}"
        )
        response_data = {
            "status": "false",
            "title": "Deletion Failed",
            "message": str(e),
        }
        return HttpResponse(json.dumps(response_data), content_type='application/javascript')


def handle_invoice_deletion(customer_supply_instance):
    if (invoice_instances:=Invoice.objects.filter(created_date__date=customer_supply_instance.created_date.date(), customer=customer_supply_instance.customer, invoice_no=customer_supply_instance.invoice_no, amout_total=customer_supply_instance.subtotal,is_deleted=False)).exists():
        invoice_instance = invoice_instances.first()
        InvoiceItems.objects.filter(
            invoice=invoice_instance,
            is_deleted=False
        ).update(is_deleted=True)

        # Soft delete daily collection
        InvoiceDailyCollection.objects.filter(
            invoice=invoice_instance,
            created_date__date=customer_supply_instance.created_date.date(),
            customer=customer_supply_instance.customer,
            salesman=customer_supply_instance.salesman,
            is_deleted=False
        ).update(is_deleted=True)

        # Soft delete invoice
        invoice_instance.is_deleted = True
        invoice_instance.save()

def handle_outstanding_coupon(customer_supply_instance, five_gallon_qty):
    if customer_supply_instance.outstanding_coupon_added :
        if (customet_outstanding_instances:=CustomerOutstanding.objects.filter(invoice_no=customer_supply_instance.invoice_no,product_type="coupons")).exists():
            customet_outstanding_instance = customet_outstanding_instances.first()
            outstanding_coupon_count = OutstandingCoupon.objects.filter(customer_outstanding=customet_outstanding_instance).aggregate(total_count=Sum('count'))['total_count']
            outstanding_report = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer,product_type="coupons")
            outstanding_report.value -= outstanding_coupon_count
            outstanding_report.save()
            
            customet_outstanding_instance.delete()

def update_or_create_outstanding_amount(
    customer,
    old_invoice_no,
    new_balance_amount,
    created_date,
    user
):
    # Try UPDATE first (EDIT case)
    outstanding = CustomerOutstanding.objects.filter(
        customer=customer,
        invoice_no=old_invoice_no,
        product_type="amount"
    ).first()
    print("inside fun")
    if outstanding:
        outstanding_row = OutstandingAmount.objects.filter(
            customer_outstanding=outstanding
        ).first()
        print("inside fun 1")
        if outstanding_row:
            outstanding_row.amount = new_balance_amount
            print("inside fun 2")
            outstanding_row.save(update_fields=["amount"])
            print("inside fun 3")
            return  # ✅ UPDATED — STOP HERE

    # CREATE only if NOT FOUND (NEW supply case)
    if new_balance_amount != 0:
        print("inside fun 4")
        new_outstanding = CustomerOutstanding.objects.create(
            customer=customer,
            invoice_no=old_invoice_no,
            product_type="amount",
            created_by=user.id,
            created_date=created_date,
            outstanding_date=created_date,
        )

        OutstandingAmount.objects.create(
            customer_outstanding=new_outstanding,
            amount=new_balance_amount
        )


def handle_outstanding_amounts(customer_supply_instance, five_gallon_qty):
    if customer_supply_instance.outstanding_amount_added:
        balance_amount = customer_supply_instance.subtotal - customer_supply_instance.amount_recieved
        customer_amount_instances = CustomerOutstanding.objects.filter(
            customer=customer_supply_instance.customer,
            invoice_no=customer_supply_instance.invoice_no,
            product_type="amount"
        ).first()
        
        if customer_supply_instance.amount_recieved < customer_supply_instance.subtotal:
            if CustomerOutstandingReport.objects.filter(customer=customer_supply_instance.customer, product_type="amount").exists():
                customer_outstanding_report_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="amount")
                customer_outstanding_report_instance.value -= Decimal(balance_amount)
                customer_outstanding_report_instance.save()
        elif customer_supply_instance.amount_recieved > customer_supply_instance.subtotal:
            customer_outstanding_report_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="amount")
            customer_outstanding_report_instance.value += Decimal(balance_amount)
            customer_outstanding_report_instance.save()
        
        if customer_amount_instances:
            customer_amount_instances.delete()

def handle_coupons(customer_supply_instance, five_gallon_qty):
    if (manual_coupons := CustomerSupplyCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
        for coupon in manual_coupons:
            for l in coupon.leaf.all():
                if CustomerCouponStock.objects.filter(customer=customer_supply_instance.customer,coupon_method="manual",coupon_type_id=l.coupon.coupon_type).exists() :
                    customer_stock = CustomerCouponStock.objects.get(customer=customer_supply_instance.customer,coupon_method="manual",coupon_type_id=l.coupon.coupon_type)
                    customer_stock.count += 1
                    customer_stock.save()
            coupon.leaf.update(used=False)
        manual_coupons.delete()
        
    elif (digital_coupons := CustomerSupplyDigitalCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
        customer_coupon_digital = CustomerSupplyDigitalCoupon.objects.get(customer_supply=customer_supply_instance)
                                
        customer_stock = CustomerCouponStock.objects.get(customer=customer_supply_instance.customer,coupon_method="digital",coupon_type_id__coupon_type_name="Digital")
        customer_stock.count += Decimal(customer_coupon_digital.count)
        customer_stock.save()
        
        customer_coupon_digital.delete()
    # if (digital_coupons_instances := CustomerSupplyDigitalCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
    #     digital_coupons_instance = digital_coupons_instances.first()
    #     CustomerCouponStock.objects.filter(
    #         coupon_method="digital",
    #         customer=customer_supply_instance.customer,
    #         coupon_type_id__coupon_type_name="Digital"
    #     ).update(count=F('count') + digital_coupons_instance.count)
    
    # elif (manual_coupon_instances := CustomerSupplyCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
    #     manual_coupon_instance = manual_coupon_instances.first()
    #     leaflets_to_update = manual_coupon_instance.leaf.filter(used=True)
    #     updated_count = leaflets_to_update.count()

    #     if updated_count > 0:
    #         first_leaflet = leaflets_to_update.first()

    #         if first_leaflet and CustomerCouponStock.objects.filter(
    #                 customer=customer_supply_instance.customer,
    #                 coupon_method="manual",
    #                 coupon_type_id=first_leaflet.coupon.coupon_type
    #             ).exists():
    #             customer_stock_instance = CustomerCouponStock.objects.get(
    #                 customer=customer_supply_instance.customer,
    #                 coupon_method="manual",
    #                 coupon_type_id=first_leaflet.coupon.coupon_type
    #             )
    #             customer_stock_instance.count += Decimal(updated_count)
    #             customer_stock_instance.save()
                
    #             handle_empty_bottle_outstanding(customer_supply_instance, five_gallon_qty)
                
    #             leaflets_to_update.update(used=False)


def handle_empty_bottle_outstanding(customer_supply_instance, five_gallon_qty):
    if customer_supply_instance.outstanding_bottle_added:
        if five_gallon_qty < Decimal(customer_supply_instance.collected_empty_bottle):
            balance_empty_bottle = Decimal(customer_supply_instance.collected_empty_bottle) - five_gallon_qty
            if CustomerOutstandingReport.objects.filter(customer=customer_supply_instance.customer, product_type="emptycan").exists():
                outstanding_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="emptycan")
                outstanding_instance.value += Decimal(balance_empty_bottle)
                outstanding_instance.save()
        elif five_gallon_qty > Decimal(customer_supply_instance.collected_empty_bottle):
            balance_empty_bottle = five_gallon_qty - Decimal(customer_supply_instance.collected_empty_bottle)
            customer_outstanding_instance = CustomerOutstanding.objects.filter(
                product_type="emptycan",
                created_by=customer_supply_instance.salesman.pk,
                customer=customer_supply_instance.customer,
                created_date=customer_supply_instance.created_date,
            ).first()

            outstanding_product = OutstandingProduct.objects.filter(
                empty_bottle=balance_empty_bottle,
                customer_outstanding=customer_outstanding_instance,
            )

            try:
                outstanding_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="emptycan")
                outstanding_instance.value -= Decimal(outstanding_product.aggregate(total_empty_bottle=Sum('empty_bottle'))['total_empty_bottle'] or 0)
                outstanding_instance.save()
            except:
                pass

            if outstanding_product.exists():
                outstanding_product.delete()
            if customer_outstanding_instance:
                customer_outstanding_instance.delete()


def update_van_product_stock(customer_supply_instance, supply_items_instances, five_gallon_qty):
    for item_data in supply_items_instances:
        # Check if there is a VanProductStock entry for the given date, product, and salesman
        if VanProductStock.objects.filter(created_date=customer_supply_instance.created_date.date(),product=item_data.product,van__salesman=customer_supply_instance.salesman).exists():
            
            # Special handling for "5 Gallon" products
            van_stock = VanProductStock.objects.get(product=item_data.product,created_date=customer_supply_instance.created_date.date(),van__salesman=customer_supply_instance.salesman)
            
            if item_data.product.product_name == "5 Gallon" and customer_supply_instance.van_emptycan_added and van_stock.created_date == datetime.today().date():
                van_stock.empty_can_count -= customer_supply_instance.collected_empty_bottle

            # ✅ Rollback van stock only if the supply was created today
            if van_stock.created_date == datetime.today().date():
                
                if customer_supply_instance.van_stock_added:
                    if customer_supply_instance.customer.sales_type != "FOC" and van_stock.sold_count >= item_data.quantity:
                        van_stock.sold_count -= item_data.quantity
                        van_stock.stock += item_data.quantity
                
                # if customer_supply_instance.van_stock_added :
                #     if customer_supply_instance.customer.sales_type != "FOC" and van_stock.sold_count >= item_data.quantity :
                #         van_stock.sold_count -= item_data.quantity
                #         van_stock.stock += item_data.quantity
                    
                if customer_supply_instance.van_foc_added :
                    if customer_supply_instance.customer.sales_type == "FOC":
                        van_stock.foc -= item_data.quantity
                        van_stock.stock += item_data.quantity
                    else:
                        van_stock.foc -= customer_supply_instance.allocate_bottle_to_free
                        van_stock.stock += customer_supply_instance.allocate_bottle_to_free
                        
            van_stock.save()

#------------------------------REPORT----------------------------------------

def client_report(request):
    instances = Customers.objects.order_by('-created_date')  # Order by latest created date
    return render(request, 'client_management/client_report.html', {'instances': instances})


def clientdownload_pdf(request, customer_id):
    customer = get_object_or_404(Customers, pk=customer_id)
    template_path = 'client_management/client_report_pdf.html'
    context = {'customer': customer}

    # Logic to generate PDF for the specific customer
    pdf_content = f"PDF content for {customer.customer_name}"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{customer.customer_name}_report.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    # Create PDF
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')

    return response


def clientexport_to_csv(request, customer_id):
    customer = get_object_or_404(Customers, pk=customer_id)

    # Create an Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Set the title
    title_cell = ws.cell(row=1, column=1, value="Client Details")
    title_cell.alignment = Alignment(horizontal='center')  # Align the title to the center
    ws.merge_cells('A1:F1')  # Merge cells for the title

    # Define data to be written
    data = [
        ['Customer Name:', customer.customer_name],
        ['Address:', f"{customer.building_name} {customer.door_house_no}"],
        ['Contact:', customer.mobile_no],
        ['Customer Type:', customer.customer_type],
        ['Sales Type:', customer.sales_type],
    ]

    # Write data to the worksheet
    for row in data:
        ws.append(row)

    # Set response headers for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{customer.customer_name}_report.xlsx"'

    # Save the workbook to the HttpResponse
    wb.save(response)

    return response


def custody_items_list_report(request):
    if request.method == 'GET':
        
        instances = CustodyCustom.objects.all()
        # if start_date_str and end_date_str:
        #     start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        #     end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        #     instances = CustodyCustomItems.objects.filter(custody_custom__created_date__range=[start_date, end_date])
        # else:
        
        return render(request, 'client_management/custody_items_list_report.html', {'instances': instances})


def custody_issue(request):
    instances = CustodyCustom.objects.all().order_by("-created_date")

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        instances = instances.filter(created_date__range=[start_date, end_date])
    
        
    customer_product_counts = {}

    for instance in instances:
        customer = instance.customer
        print(customer,"customer")
        if customer not in customer_product_counts:
            customer_product_counts[customer] = {
                '5_gallon_deposit': 0,
                '5_gallon_non_deposit': 0,
                'dispenser_deposit': 0,
                'dispenser_non_deposit': 0,
                'cooler_deposit': 0,
                'cooler_non_deposit': 0,
            }

        custody_items = instance.custodycustomitems_set.all()

        for custody_item in custody_items:
            product = custody_item.product
            if product:
                if product.product_name == '5 Gallon':
                    if custody_item.custody_custom.deposit_type == 'deposit':
                        customer_product_counts[customer]['5_gallon_deposit'] += 1
                    else:
                        customer_product_counts[customer]['5_gallon_non_deposit'] += 1
                elif product.product_name == 'Dispenser':
                    if custody_item.custody_custom.deposit_type == 'deposit':
                        customer_product_counts[customer]['dispenser_deposit'] += 1
                    else:
                        customer_product_counts[customer]['dispenser_non_deposit'] += 1
                elif product.product_name == 'Cooler':
                    if custody_item.custody_custom.deposit_type == 'deposit':
                        customer_product_counts[customer]['cooler_deposit'] += 1
                    else:
                        customer_product_counts[customer]['cooler_non_deposit'] += 1

        # Sum up the total bottles issued
        total_bottles_issued = (
            customer_product_counts[customer]['5_gallon_deposit'] +
            customer_product_counts[customer]['5_gallon_non_deposit'] +
            customer_product_counts[customer]['dispenser_deposit'] +
            customer_product_counts[customer]['dispenser_non_deposit'] +
            customer_product_counts[customer]['cooler_deposit'] +
            customer_product_counts[customer]['cooler_non_deposit']
        )

        # Send notification if sales staff exists
        if customer.sales_staff:
            sales_man = customer.sales_staff
            try:
                # Notification details
                agreement_no = instance.agreement_no  
                deposit_amount = instance.total_amount  
                
                # Salesman notification
                salesman_body = (
                    f'Customer custody issued for {customer}. Bottles issued: {total_bottles_issued}, '
                    f'Quotation deposit amount: {deposit_amount}, Agreement number: {agreement_no}.'
                )
                notification(sales_man.pk, "Custody Issued", salesman_body, "Sanawaterfcm")
                
                # Customer notification
                customer_body = (
                    f'Dear customer, the following custody items have been issued: '
                    f'\nBottles: {total_bottles_issued}, '
                    f'\nQuotation deposit amount: {deposit_amount}, '
                    f'\nAgreement number: {agreement_no}.'
                )
                notification(customer.user_id.pk, "Custody Issued", customer_body, "Sanawatercustomer")

            except CustomUser.DoesNotExist:
                messages.error(request, 'Salesman does not exist.', 'alert-danger')
            except Send_Notification.DoesNotExist:
                messages.error(request, 'No device token found for the salesman.', 'alert-danger')
            except Exception as e:
                messages.error(request, f'Error sending notification: {e}', 'alert-danger')


    for customer, counts in customer_product_counts.items():
        for key, value in counts.items():
            if value == 0:
                customer_product_counts[customer][key] = '--'

    return render(request, 'client_management/custody_issue.html', {'customer_product_counts': customer_product_counts})



def get_customercustody(request, customer_id):
    customer = Customers.objects.get(customer_id=customer_id)  # Use get() if customer_id is unique
    print(customer, "customer")
    
    custody_items = CustodyCustom.objects.filter(customer=customer)
    print("custody_items", custody_items)
    
    custody_items_with_products = []
    
    for custody_item in custody_items:
        custody_custom_items = CustodyCustomItems.objects.filter(custody_custom=custody_item)
        custody_item_data = {
            'custody_custom': custody_item,
            'custody_custom_items': custody_custom_items,
            'products': [item.product for item in custody_custom_items]  
        }
        custody_items_with_products.append(custody_item_data)

    context = {'custody_items_with_products': custody_items_with_products,'customer':customer}
    return render(request, 'client_management/customer_custody_items.html', context)

def custody_report(request):
    instances = CustodyCustom.objects.all().order_by("-created_date")

    start_date_str = request.GET.get('start_date')
    print("start_date_str",    start_date_str)

    end_date_str = request.GET.get('end_date')
    print("end_date_str",end_date_str)
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        instances = instances.filter(created_date__range=[start_date, end_date])

    custody_items_with_products = []
    
    for custody_item in instances:
        custody_custom_items = CustodyCustomItems.objects.filter(custody_custom=custody_item)
        print("custody_custom_items",custody_custom_items)
        custody_item_data = {
            'custody_custom': custody_item,
            'custody_custom_items': custody_custom_items,
            'products': [item.product for item in custody_custom_items]  
        }
        custody_items_with_products.append(custody_item_data)

    context = {'custody_items_with_products': custody_items_with_products,'instances':instances}
    print("context",context)
    return render(request, 'client_management/custody_report.html', context)


class CouponCountList(View):
    template_name = 'client_management/coupon_count_list.html'

    def get(self, request, pk, *args, **kwargs):
        customer = Customers.objects.get(customer_id=pk)
        customers = CustomerCouponStock.objects.filter(customer=customer)

        # Calculate total count
        total_count = customers.aggregate(total_count=Sum('count'))['total_count'] or 0

        context = {
            'customers': customers,
            'pk': pk,  # Pass pk to the template context
            'total_count': total_count,  # Pass total count to the template context
        }

        return render(request, self.template_name, context)







#
# class CouponCountList(View):
#     template_name = 'client_management/coupon_count_list.html'
#
#     def get(self, request, pk, *args, **kwargs):
#         customer = Customers.objects.get(customer_id=pk)
#         customers = CustomerCouponStock.objects.filter(customer=customer)
#
#         context = {
#             'customers': customers,
#             'pk': pk,  # Pass pk to the template context
#         }
#
#         return render(request, self.template_name, context)
#
#     def post(self, request, pk, *args, **kwargs):
#         customer = Customers.objects.get(customer_id=pk)
#         coupon_code = request.POST.get('coupon_code')
#         CustomerCouponStock.objects.create(customer=customer, coupon_code=coupon_code)
#         return redirect('coupon_count_list', pk=pk)



#
# def edit_coupon_count(request, pk):
#     customer_coupon_stock = get_object_or_404(CustomerCouponStock, customer_id=pk)
#     form = CoupenEditForm(instance=customer_coupon_stock)
#
#     if request.method == 'POST':
#         form = CoupenEditForm(request.POST, instance=customer_coupon_stock)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Customer Coupon Stock Updated successfully!')
#             return redirect('customers')
#         else:
#             messages.error(request, 'Invalid form data. Please check the input.')
#
#     return render(request, 'client_management/edit_coupon_count.html', {'form': form})

def new_coupon_count(request,pk):
    if request.method == 'POST':
        form = CoupenEditForm(request.POST)
        if form.is_valid():
            # data = form.save(commit=False)
            # data.customer = Customers.objects.get(pk=pk)
            # data.save()
            coupon_type_id = CouponType.objects.get(pk=form.cleaned_data['coupon_type_id'].pk)
            
            coupon_method = "manual"
            if coupon_type_id.coupon_type_name == "Digital":
                coupon_method = "digital"
            try:
                data = CustomerCouponStock.objects.get(
                    customer=Customers.objects.get(pk=pk),
                    coupon_type_id=coupon_type_id,
                    coupon_method=coupon_method
                )
            except CustomerCouponStock.DoesNotExist:
                data = CustomerCouponStock.objects.create(
                    customer=Customers.objects.get(pk=pk),
                    coupon_type_id=coupon_type_id,
                    coupon_method=coupon_method,
                    count=0
                )

            data.count += Decimal(form.cleaned_data['count'])
            data.save()

            messages.success(request, 'New coupon count added successfully!')
            return redirect('coupon_count_list' ,data.customer_id)
        else:
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = CoupenEditForm()

    return render(request, 'client_management/edit_coupon_count.html', {'form': form})

def delete_count(request, pk):
    customer_coupon_stock = get_object_or_404(CustomerCouponStock, pk=pk)

    if request.method == 'POST':
        customer_pk = customer_coupon_stock.customer.pk
        customer_coupon_stock.delete()
        messages.success(request, 'Coupon count deleted successfully!')
        log_activity(
                created_by=request.user,
                description=f"Coupon count with ID {pk} for customer {customer_pk} was successfully deleted."
            )
        return redirect('coupon_count_list', pk=customer_pk)

    return redirect('coupon_count_list')

# @login_required

from datetime import datetime
from django.db.models import Q, Sum
from django.shortcuts import render

# def customer_outstanding_list(request):
#     """
#     Customer Outstanding List with Excel Export
#     """
#     filter_data = {}
#     q = request.GET.get('q', '')  
#     route_name = request.GET.get('route_name', '')
#     date_str = request.GET.get('date')
#     product_type = request.GET.get('product_type', 'amount')
    
#     route_li = RouteMaster.objects.all()
#     filter_data['product_type'] = product_type

#     if date_str:
#         date = datetime.strptime(date_str, '%Y-%m-%d').date() 
#     else:
#         date = datetime.today().date()
#     filter_data['filter_date'] = date.strftime('%Y-%m-%d')

#     # --- INITIAL FILTERING FOR ROUTE AND DATE ---
#     if route_name:
#         pass # Route filter applied later
#     elif route_li.exists():
#         route_name = route_li.first().route_name
#     filter_data['route_name'] = route_name
    
#     # Base QuerySet for ALL customers on the route
#     all_customers_on_route = Customers.objects.filter(is_guest=False, routes__route_name=route_name)

#     if customer_pk := request.GET.get("customer_pk"):
#         all_customers_on_route = all_customers_on_route.filter(pk=customer_pk)
#         filter_data['customer_pk'] = customer_pk

#     if q:
#         all_customers_on_route = all_customers_on_route.filter(customer_name__icontains=q)
#         filter_data['q'] = q

#     # --- 1. EFFICIENT AGGREGATION (CALCULATES THE TRUE TOTAL: 20694.80) ---

#     # Total Outstanding Debits for the route (113193.20)
#     total_amount_t = OutstandingAmount.objects.filter(
#         customer_outstanding__customer__routes__route_name=route_name,
#         customer_outstanding__created_date__date__lte=date
#     ).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
    
#     # Total Collections Credits for the route (92498.40)
#     total_collection_t = CollectionPayment.objects.filter(
#         created_date__date__lte=date,
#         customer__routes__route_name=route_name
#     ).aggregate(total_amount_received=Sum('amount_received'))['total_amount_received'] or 0

#     # Store the correct net total for final use, rounded to 2 decimals
#     correct_route_net_amount = total_amount_t - total_collection_t
#     total_outstanding_amount = round(correct_route_net_amount, 2)
    
#     # Print the correct value: 20694.80
#     print(f"Correct Net Total: {total_outstanding_amount}") 

#     # --- 2. CUSTOMER LOOP (ITERATE OVER ALL, FILTER NON-ZERO, ACCUMULATE BOTTLES/COUPONS) ---
    
#     instances = []
#     total_outstanding_bottles = 0
#     total_outstanding_coupons = 0
#     # total_outstanding_amount is already correctly set above.

#     for customer in all_customers_on_route:
#         net_amount = 0
#         total_bottles = 0
#         total_coupons = 0

#         # Calculate individual balances regardless of product_type for inclusion check
        
#         # Calculate Amount Balance
#         outstanding_amount = OutstandingAmount.objects.filter(
#             customer_outstanding__customer=customer,
#             customer_outstanding__created_date__date__lte=date
#         ).aggregate(total_amount=Sum('amount'))['total_amount'] or 0

#         collection_amount_customer = CollectionPayment.objects.filter(
#             customer=customer,
#             created_date__date__lte=date
#         ).aggregate(total_amount_received=Sum('amount_received'))['total_amount_received'] or 0

#         net_amount = outstanding_amount - collection_amount_customer
            
#         # Calculate Bottle Balance
#         total_bottles = OutstandingProduct.objects.filter(
#             customer_outstanding__customer=customer,
#             customer_outstanding__created_date__date__lte=date
#         ).aggregate(total_bottles=Sum('empty_bottle'))['total_bottles'] or 0

#         # Calculate Coupon Balance
#         total_coupons = OutstandingCoupon.objects.filter(
#             customer_outstanding__customer=customer,
#             customer_outstanding__created_date__date__lte=date
#         ).aggregate(total_coupons=Sum('count'))['total_coupons'] or 0
        
#         # Only include customers with a non-zero balance for the CURRENTLY SELECTED product_type
#         is_non_zero_balance = False
#         if product_type == "amount" and round(net_amount, 2) != 0:
#              is_non_zero_balance = True
#         elif product_type == "emptycan" and total_bottles != 0:
#              is_non_zero_balance = True
#         elif product_type == "coupons" and total_coupons != 0:
#              is_non_zero_balance = True
             
#         # If the customer has a non-zero balance, add them to the list and accumulate totals
#         if is_non_zero_balance:
#             instances.append(customer)
            
#             # Accumulate bottle and coupon totals
#             total_outstanding_bottles += total_bottles
#             total_outstanding_coupons += total_coupons
            
#     if request.GET.get('export') == 'excel':
#         return export_to_excel(
#             instances,
#             date,
#             total_outstanding_amount, # Uses the correct total (20694.80)
#             total_outstanding_bottles,
#             total_outstanding_coupons,
#             product_type
#         )

#     context = {
#         'instances': instances,
#         'filter_data': filter_data,
#         'route_li': route_li,
#         'date': date,
#         'customer_pk': customer_pk,
#         'route_name': route_name,
#         'page_name': 'Customer Outstanding List',
#         'page_title': 'Customer Outstanding List',
#         'is_customer_outstanding': True,
#         'is_need_datetime_picker': True,
#         'total_outstanding_amount': total_outstanding_amount, 
#         'total_outstanding_bottles': total_outstanding_bottles,
#         'total_outstanding_coupons': total_outstanding_coupons,
#     }

#     return render(request, 'client_management/customer_outstanding/list.html', context)


def customer_outstanding_list(request):
    """
    Customer Outstanding List (FAST VERSION)
    """

    # -------------------- GET FILTERS --------------------
    q = request.GET.get("q", "")
    route_name = request.GET.get("route_name", "")
    date_str = request.GET.get("date")
    product_type = request.GET.get("product_type", "amount")
    salesman_id = request.GET.get("salesman_id", "")

    # -------------------- ROUTES --------------------
    route_li = RouteMaster.objects.all()
    if not route_name and route_li.exists():
        route_name = route_li.first().route_name

    # -------------------- DATE --------------------
    if date_str:
        filter_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        filter_date = datetime.today().date()

    filter_data = {
        "q": q,
        "route_name": route_name,
        "filter_date": filter_date.strftime("%Y-%m-%d"),
        "product_type": product_type,
        "salesman_id": salesman_id,
    }

    # -------------------- SALESMAN LIST --------------------
    salesman_list = CustomUser.objects.filter(user_type="Salesman")

    # -------------------- CUSTOMERS --------------------
    customers_qs = Customers.objects.filter(
        routes__route_name=route_name,
        is_deleted=False
    ).distinct()

    if q:
        customers_qs = customers_qs.filter(customer_name__icontains=q)

    customer_ids = list(customers_qs.values_list("customer_id", flat=True))

    # =====================================================
    # 🔥 AGGREGATIONS (ONLY ONE QUERY PER TABLE)
    # =====================================================

    # -------------------- OUTSTANDING AMOUNT --------------------
    outstanding_qs = OutstandingAmount.objects.filter(
        customer_outstanding__customer_id__in=customer_ids,
        customer_outstanding__outstanding_date__date__lte=filter_date
    )

    if salesman_id:
        outstanding_qs = outstanding_qs.filter(
            customer_outstanding__created_by=salesman_id
        )

    outstanding_map = {
        row["customer_outstanding__customer_id"]: row["total"]
        for row in outstanding_qs.values(
            "customer_outstanding__customer_id"
        ).annotate(total=Sum("amount"))
    }

    # -------------------- COLLECTION --------------------
    collection_qs = CollectionPayment.objects.filter(
        customer_id__in=customer_ids,
        created_date__date__lte=filter_date
    )

    if salesman_id:
        collection_qs = collection_qs.filter(salesman_id=salesman_id)

    collection_map = {
        row["customer_id"]: row["total"]
        for row in collection_qs.values("customer_id").annotate(
            total=Sum("amount_received")
        )
    }

    # -------------------- EMPTY CANS --------------------
    bottle_map = {
        row["customer_outstanding__customer_id"]: row["total"]
        for row in OutstandingProduct.objects.filter(
            customer_outstanding__customer_id__in=customer_ids,
            customer_outstanding__created_date__date__lte=filter_date
        ).values(
            "customer_outstanding__customer_id"
        ).annotate(total=Sum("empty_bottle"))
    }

    # -------------------- COUPONS --------------------
    coupon_map = {
        row["customer_outstanding__customer_id"]: row["total"]
        for row in OutstandingCoupon.objects.filter(
            customer_outstanding__customer_id__in=customer_ids,
            customer_outstanding__created_date__date__lte=filter_date
        ).values(
            "customer_outstanding__customer_id"
        ).annotate(total=Sum("count"))
    }

    # =====================================================
    # 🔥 FINAL LOOP (NO DB HITS)
    # =====================================================

    instances = []
    total_outstanding_amount = Decimal("0.00")
    total_outstanding_bottles = 0
    total_outstanding_coupons = 0

    for customer in customers_qs:
        outstanding_amount = outstanding_map.get(customer.customer_id, Decimal("0.00"))
        collection_amount = collection_map.get(customer.customer_id, Decimal("0.00"))

        net_amount = outstanding_amount - collection_amount

        bottles = bottle_map.get(customer.customer_id, 0)
        coupons = coupon_map.get(customer.customer_id, 0)

        include = False
        if product_type == "amount" and net_amount != 0:
            include = True
        elif product_type == "emptycan" and bottles != 0:
            include = True
        elif product_type == "coupons" and coupons != 0:
            include = True

        if include:
            instances.append({
                "customer": customer,
                "net_amount": net_amount,
                "bottles": bottles,
                "coupons": coupons,
            })
            total_outstanding_amount += net_amount
            total_outstanding_bottles += bottles
            total_outstanding_coupons += coupons

    # -------------------- CONTEXT --------------------
    context = {
        "instances": instances,
        "route_li": route_li,
        "salesman_list": salesman_list,
        "filter_data": filter_data,
        "date": filter_date,
        "route_name": route_name,
        "total_outstanding_amount": total_outstanding_amount,
        "total_outstanding_bottles": total_outstanding_bottles,
        "total_outstanding_coupons": total_outstanding_coupons,
    }

    if request.GET.get('export') == 'excel':
        return export_to_excel(
            instances, 
            filter_date, 
            total_outstanding_amount, 
            total_outstanding_bottles, 
            total_outstanding_coupons, 
            product_type
        )

    return render(
        request,
        "client_management/customer_outstanding/list.html",
        context
    )


def export_to_excel(instances, date, total_amount, total_bottles, total_coupons, product_type):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customer Outstanding List"

    headers = ["Sl No", "Customer ID", "Customer Name", "Route", "Outstanding Amount", "Bottles", "Coupons"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    row_no = 2
    for index, item in enumerate(instances, start=1):
        # Access the dictionary keys correctly
        customer = item['customer'] 
        
        sheet.append([
            index,
            customer.custom_id,
            customer.customer_name,
            customer.routes.route_name if customer.routes else "",
            round(item['net_amount'], 2),
            item['bottles'],
            item['coupons'],
        ])
        row_no += 1

    # Total Row
    sheet[f"D{row_no}"] = "Total"
    sheet[f"E{row_no}"] = round(total_amount, 2)
    sheet[f"F{row_no}"] = total_bottles
    sheet[f"G{row_no}"] = total_coupons

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Customer_Outstanding_{date}.xlsx"'
    workbook.save(response)
    return response




def print_customer_outstanding(request):
    from datetime import datetime
    from decimal import Decimal
    from django.db.models import Sum, F

    filter_data = {}
    q = request.GET.get("q", "")
    route_name = request.GET.get("route_name", "")
    date_str = request.GET.get("date")
    product_type = request.GET.get("product_type", "amount")

    route_li = RouteMaster.objects.all()
    filter_data["product_type"] = product_type

    # ---------------- DATE ----------------
    if date_str:
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        date = datetime.today().date()

    filter_data["filter_date"] = date.strftime("%Y-%m-%d")

    # ---------------- ROUTE ----------------
    if not route_name and route_li.exists():
        route_name = route_li.first().route_name

    filter_data["route_name"] = route_name

    # ---------------- CUSTOMERS ----------------
    customers = Customers.objects.filter(
        is_guest=False,
        routes__route_name=route_name,
        is_deleted=False
    )

    if customer_pk := request.GET.get("customer_pk"):
        customers = customers.filter(pk=customer_pk)
        filter_data["customer_pk"] = customer_pk

    if q:
        customers = customers.filter(customer_name__icontains=q)
        filter_data["q"] = q

    # ---------------- RESULT CONTAINERS ----------------
    instances = []
    total_outstanding_amount = Decimal("0.00")
    total_outstanding_bottles = 0
    total_outstanding_coupons = 0

    # ---------------- MAIN LOOP ----------------
    for customer in customers:

        # -------- AMOUNT (INVOICE BASED – DO NOT USE invoice_status) --------
        invoice_totals = Invoice.objects.filter(
            customer=customer,
            created_date__date__lte=date,
            is_deleted=False,
            invoice_status = "non_paid",
        ).aggregate(
            total_amount=Sum('amout_total'),
            total_received=Sum('amout_recieved')
        )

        total_amount = invoice_totals["total_amount"] or Decimal("0.00")
        total_received = invoice_totals["total_received"] or Decimal("0.00")

        outstanding_amount = total_amount - total_received

        # -------- BOTTLES --------
        bottles = OutstandingProduct.objects.filter(
            customer_outstanding__customer=customer,
            customer_outstanding__created_date__date__lte=date
        ).aggregate(
            total=Sum("empty_bottle")
        )["total"] or 0

        # -------- COUPONS --------
        coupons = OutstandingCoupon.objects.filter(
            customer_outstanding__customer=customer,
            customer_outstanding__created_date__date__lte=date
        ).aggregate(
            total=Sum("count")
        )["total"] or 0

        # -------- PRODUCT TYPE FILTER --------
        include = False
        if product_type == "amount" and outstanding_amount != 0:
            include = True
        elif product_type == "emptycan" and bottles != 0:
            include = True
        elif product_type == "coupons" and coupons != 0:
            include = True

        # -------- ADD CUSTOMER --------
        if include:
            customer.outstanding_amount = outstanding_amount
            customer.outstanding_bottles = bottles
            customer.outstanding_coupons = coupons

            instances.append(customer)
            total_outstanding_amount += outstanding_amount
            total_outstanding_bottles += bottles
            total_outstanding_coupons += coupons

    # ---------------- CONTEXT ----------------

    print("instance:",instances)
    context = {
        "instances": instances,
        "filter_data": filter_data,
        "route_li": route_li,
        "date": date,
        "customer_pk": request.GET.get("customer_pk"),
        "net_total_outstanding": total_outstanding_amount,
        "total_outstanding_bottles": total_outstanding_bottles,
        "total_outstanding_coupons": total_outstanding_coupons,
        "is_customer_outstanding": True,
        "is_need_datetime_picker": True,
        "filter_date": date.strftime("%d-%m-%Y"),
        "page_title": "Print Outstanding Report",
    }

    return render(
        request,
        "client_management/customer_outstanding/print.html",
        context
    )




from io import BytesIO
def excel_customer_outstanding(request):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customer Outstanding Report'

    # Define the headers
    headers = ['Sl No', 'Customer ID', 'Customer', 'Building No', 
               'Room No/Floor No', 'Route', 'Outstanding Amount']
    worksheet.append(headers)

    # Fetching the filter parameters from the request
    date_str = request.GET.get('date', datetime.today().strftime('%Y-%m-%d'))
    date = datetime.strptime(date_str, '%Y-%m-%d').date()  # Convert to date object
    route_name = request.GET.get('route_name', '')
    q = request.GET.get('q', '')

    # Apply filters similar to the view
    outstanding_instances = CustomerOutstanding.objects.filter(created_date__date__lte=date)
    
    if request.GET.get("customer_pk"):
        outstanding_instances = outstanding_instances.filter(customer__pk=request.GET.get("customer_pk"))

    if route_name:
        outstanding_instances = outstanding_instances.filter(customer__routes__route_name=route_name)

    if q:
        outstanding_instances = outstanding_instances.filter(customer__customer_name__icontains=q)
    
    customer_ids = outstanding_instances.values_list('customer__pk', flat=True).distinct()
    instances = Customers.objects.filter(is_guest=False, pk__in=customer_ids)

    # Calculate total outstanding and total collections for net total
    total_outstanding = OutstandingAmount.objects.filter(
        customer_outstanding__customer__in=customer_ids,
        customer_outstanding__created_date__date__lte=date
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_collection = CollectionPayment.objects.filter(
        customer__in=customer_ids,
        created_date__date__lte=date
    ).aggregate(total=Sum('amount_received'))['total'] or 0

    net_total_outstanding = total_outstanding - total_collection

    for idx, customer in enumerate(instances, start=1):
        # Calculate Outstanding Amount
        outstanding_sum = OutstandingAmount.objects.filter(
            customer_outstanding__customer=customer,
            customer_outstanding__created_date__date__lte=date
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Calculate Collection Amount
        collection_sum = CollectionPayment.objects.filter(
            customer=customer,
            created_date__date__lte=date
        ).aggregate(total=Sum('amount_received'))['total'] or 0

        # Compute Net Outstanding Amount
        outstanding_amount = outstanding_sum - collection_sum
        if outstanding_amount !=0:
            # Append each row of data
            worksheet.append([
                idx,
                customer.custom_id,
                customer.customer_name,
                customer.building_name,
                customer.door_house_no,
                customer.routes.route_name,
                outstanding_amount
            ])

    # Append the net total outstanding at the end of the worksheet
    worksheet.append(['', '', '', '', '', 'Net Total Outstanding', net_total_outstanding])

    # Optionally, you can format the header row and the total row
    from openpyxl.styles import Font

    # Bold the header row
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    # Bold the total row
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)  # Move to the beginning of the BytesIO stream

    # Set the response for the Excel file
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customer_outstanding_report.xlsx'
    
    return response


# @login_required
# def customer_outstanding_details(request,customer_pk):
#     """
#     Customer Outstanding details List
#     :param request:
#     :return: Customer Outstanding list view
#     """
#     filter_data = {}
#     query = request.GET.get("q")
#     date = request.GET.get('date')
#     route_filter = request.GET.get('route_name')
#     product_type = request.GET.get('product_type', 'amount')
#     filter_data['product_type'] = product_type
    
#     instances = CustomerOutstanding.objects.filter(customer__pk=customer_pk,product_type=product_type)
#     if date:
#         date = datetime.strptime(date, '%Y-%m-%d').date()
#         filter_data['filter_date'] = date.strftime('%Y-%m-%d')

#     else:
#         date = datetime.today().date()
#         filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
#     # if route_filter:
#     #         instances = instances.filter(customer__routes__route_name=route_filter)
#     # route_li = RouteMaster.objects.all()
    
#     if query:

#         instances = instances.filter(
#             Q(product_type__icontains=query) |
#             Q(invoice_no__icontains=query) 
#         )
#         title = "Outstanding List - %s" % query
#         filter_data['q'] = query
    
#     context = {
#         'instances': instances.order_by("-created_date"),
#         'page_name' : 'Customer Outstanding List',
#         'page_title' : 'Customer Outstanding List',
#         'customer_pk': customer_pk,
        
#         'is_customer_outstanding': True,
#         'is_need_datetime_picker': True,
#         'filter_data': filter_data,
#     }

#     return render(request, 'client_management/customer_outstanding/info_list.html', context)


@login_required
def customer_outstanding_details(request, customer_pk):

    customer = get_object_or_404(Customers, pk=customer_pk)

    query = request.GET.get("q", "")
    date = request.GET.get("date")
    product_type = request.GET.get("product_type", "amount")

    filter_data = {"product_type": product_type}

    rows = []
    total_outstanding = Decimal("0.00")

    # ==================================================
    # 1️⃣ INVOICE OUTSTANDING ROWS
    # ==================================================
    invoices = (
        Invoice.objects
        .filter(customer=customer, is_deleted=False)
        .annotate(balance=F("amout_total") - F("amout_recieved"))
        .filter(balance__gt=0)
    )

    if date:
        date_obj = datetime.strptime(date, "%Y-%m-%d").date()
        invoices = invoices.filter(created_date__date__lte=date_obj)
        filter_data["filter_date"] = date

    for inv in invoices:
        rows.append({
            "created_date": inv.created_date,
            "invoice_no": inv.invoice_no,
            "customer": inv.customer.customer_name,
            "building_no": inv.customer.building_name,
            "route": inv.customer.routes.route_name if inv.customer.routes else "",
            "outstanding_type": "amount",
            "amount": inv.balance,
            "outstanding_pk": None,
        })
        total_outstanding += inv.balance

    # ==================================================
    # 2️⃣ NEGATIVE ADJUSTMENT ROWS
    # ==================================================
    adjustments = OutstandingAmount.objects.filter(
        customer_outstanding__customer=customer,
        amount__lt=0
    ).select_related("customer_outstanding")

    if date:
        adjustments = adjustments.filter(
            customer_outstanding__created_date__date__lte=date_obj
        )

    for adj in adjustments:
        rows.append({
        "created_date": adj.customer_outstanding.outstanding_date,
        "invoice_no": adj.customer_outstanding.invoice_no,
        "customer": adj.customer_outstanding.customer.customer_name,
        "building_no": adj.customer_outstanding.customer.building_name,
        "route": adj.customer_outstanding.customer.routes.route_name if adj.customer_outstanding.customer.routes else "",
        "outstanding_type": "adjustment",
        "amount": adj.amount,
        "outstanding_pk": adj.customer_outstanding.pk,
    })
        total_outstanding += adj.amount

    # ==================================================
    # 3️⃣ SORT BY DATE (LATEST FIRST)
    # ==================================================
    rows.sort(key=lambda x: x["created_date"], reverse=True)

    context = {
        "instances": rows,
        "customer_pk": customer_pk,
        "filter_data": filter_data,
        "total_outstanding": total_outstanding,
    }

    return render(
        request,
        "client_management/customer_outstanding/info_list.html",
        context
    )


@login_required
def customer_outstanding_view(request, customer_pk):
    filter_data = {}
    query = request.GET.get("q")
    date = request.GET.get('date')
    route_filter = request.GET.get('route_name')
    product_type = request.GET.get('product_type', 'amount')
    filter_data['product_type'] = product_type

    customer_details = get_object_or_404(Customers, pk=customer_pk)

    instances = CustomerOutstanding.objects.filter(
        customer__pk=customer_pk,
        product_type=product_type
    )

    if date:
        date = datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')

    if query:
        instances = instances.filter(
            Q(product_type__icontains=query) |
            Q(invoice_no__icontains=query)
        )
        filter_data['q'] = query

    invoice_queryset = OutstandingAmount.objects.filter(
        customer_outstanding__in=instances
    )

    invoices = []
    sl_no = 1
    total_amount = 0
    total_paid = 0
    total_balance = 0

    for invoice_data in invoice_queryset:
        amount = invoice_data.amount or 0
        customer_outstanding = invoice_data.customer_outstanding
        invoice_no = customer_outstanding.invoice_no

        invoice_obj = Invoice.objects.filter(
            invoice_no=invoice_no,
            customer=customer_details,
            is_deleted=False
        ).first()

        if invoice_obj:
            paid = CollectionItems.objects.filter(
                invoice=invoice_obj
            ).aggregate(total=Sum('amount_received'))['total'] or 0
        else:
            paid = 0

        balance = amount - paid

        days_passed = (
            timezone.now().date() - customer_outstanding.created_date.date()
        ).days if customer_outstanding.created_date else None

        invoices.append({
            'sl_no': sl_no,
            'invoice_number': invoice_no,
            'invoice_date': customer_outstanding.created_date,
            'amount': amount,
            'paid': paid,
            'balance': balance,
            'days_passed': days_passed,
        })

        total_amount += amount
        total_paid += paid
        total_balance += balance
        sl_no += 1

    route_name = customer_details.routes.route_name if customer_details.routes else ''

    context = {
        'customer_details': customer_details,
        'invoices': invoices,
        'route_name': route_name,
        'customer_pk': customer_pk,
        'filter_data': filter_data,
        'customer_name': customer_details.customer_name,
        'customer_id': customer_details.custom_id,
        'mobile_no': customer_details.mobile_no,
        'total_amount': total_amount,
        'total_paid': total_paid,
        'total_balance': total_balance,
    }

    return render(
        request,
        'client_management/customer_outstanding/customer_outstanding_view.html',
        context
    )
    
    
@login_required
def customer_outstanding_print(request, customer_pk):
    filter_data = {}
    query = request.GET.get("q")
    date = request.GET.get('date')
    route_filter = request.GET.get('route_name')
    product_type = request.GET.get('product_type', 'amount')
    filter_data['product_type'] = product_type

    customer_details = get_object_or_404(Customers, pk=customer_pk)

    instances = CustomerOutstanding.objects.filter(
        customer__pk=customer_pk,
        product_type=product_type
    )

    if date:
        date = datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')

    if query:
        instances = instances.filter(
            Q(product_type__icontains=query) |
            Q(invoice_no__icontains=query)
        )
        filter_data['q'] = query

    invoice_queryset = OutstandingAmount.objects.filter(
        customer_outstanding__in=instances
    )

    invoices = []
    sl_no = 1
    total_amount = 0
    total_paid = 0
    total_balance = 0

    for invoice_data in invoice_queryset:
        amount = invoice_data.amount or 0
        customer_outstanding = invoice_data.customer_outstanding
        invoice_no = customer_outstanding.invoice_no

        invoice_obj = Invoice.objects.filter(
            invoice_no=invoice_no,
            customer=customer_details,
            is_deleted=False
        ).first()

        if invoice_obj:
            paid = CollectionItems.objects.filter(
                invoice=invoice_obj
            ).aggregate(total=Sum('amount_received'))['total'] or 0
        else:
            paid = 0

        balance = amount - paid

        days_passed = (
            timezone.now().date() - customer_outstanding.created_date.date()
        ).days if customer_outstanding.created_date else None

        invoices.append({
            'sl_no': sl_no,
            'invoice_number': invoice_no,
            'invoice_date': customer_outstanding.created_date,
            'amount': amount,
            'paid': paid,
            'balance': balance,
            'days_passed': days_passed,
        })

        total_amount += amount
        total_paid += paid
        total_balance += balance
        sl_no += 1

    route_name = customer_details.routes.route_name if customer_details.routes else ''

    context = {
        'customer_details': customer_details,
        'invoices': invoices,
        'route_name': route_name,
        'customer_pk': customer_pk,
        'filter_data': filter_data,
        'customer_name': customer_details.customer_name,
        'customer_id': customer_details.custom_id,
        'mobile_no': customer_details.mobile_no,
        'product_type': product_type,
        'date': date.strftime('%Y-%m-%d'),
        'total_amount': total_amount,
        'total_paid': total_paid,
        'total_balance': total_balance,
    }

    return render(
        request,
        'client_management/customer_outstanding/customer_outstanding_view_print.html',
        context
    )
    
    

@login_required
def outstanding_list(request):
    """
    Customer Outstanding  List
    :param request:
    :return: Customer Outstanding list view
    """
    
    filter_data = {}
    instances = CustomerOutstanding.objects.all().order_by('-created_date')
    
    query = request.GET.get("q")
    date = request.GET.get('date')
    route_filter = request.GET.get('route_name')
    sales_type_filter = request.GET.get('sales_type')
    product_type_filter = request.GET.get('product_type')
    
    # Default date filter
    if date:
        date = datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')

    # Apply filters
    if route_filter:
        instances = instances.filter(customer__routes__route_name=route_filter)
        filter_data['route_name'] = route_filter
    route_li = RouteMaster.objects.all()
    
    if sales_type_filter:
        instances = instances.filter(customer__sales_type=sales_type_filter)
        filter_data['sales_type'] = sales_type_filter
    sales_type_li = Customers.objects.values_list('sales_type', flat=True).distinct()
    
    if product_type_filter:
        instances = instances.filter(product_type=product_type_filter)
        filter_data['product_type'] = product_type_filter

    if query:
        instances = instances.filter(
            Q(product_type__icontains=query) |
            Q(invoice_no__icontains=query)
        )
        filter_data['q'] = query

    # Check if filters are applied
    filters_applied = bool(route_filter or sales_type_filter or product_type_filter or query)

    # Calculate total outstanding count only if filters are applied
    total_outstanding_count = sum([item.get_outstanding_count() for item in instances]) if filters_applied else 0
    log_activity(
        created_by=request.user if request.user.is_authenticated else None,
        description=f"Viewed outstanding list with filters: date={date},  Route: {route_filter}, Sales Type: {sales_type_filter}, "
                f"Product Type: {product_type_filter}, Query: {query}"
    )
    context = {
        'instances': instances if filters_applied else [],
        'page_name': 'Customer Outstanding List',
        'page_title': 'Customer Outstanding List',
        'filter_data': filter_data,
        'route_li': route_li,
        'sales_type_li': sales_type_li,
        'total_outstanding_count': total_outstanding_count,
        'product_types': dict(PRODUCT_TYPES),
        'filters_applied': filters_applied,
    }
    
    return render(request, 'client_management/customer_outstanding/outstanding_list.html', context)


def print_outstanding_report(request):
    """
    Print Customer Outstanding Report
    :param request:
    :return: A printable view of Customer Outstanding list
    """
    filter_data = {}
    instances = CustomerOutstanding.objects.all().order_by('-created_date')
    
    query = request.GET.get("q")
    date = request.GET.get('date')
    route_filter = request.GET.get('route_name')
    
    sales_type_filter = request.GET.get('sales_type')
    product_type_filter = request.GET.get('product_type')

    if date:
        date = datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    else:
        date = datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    # Filter by route
    if route_filter:
        instances = instances.filter(customer__routes__route_name=route_filter)
        filter_data['route_name'] = route_filter

    route_li = RouteMaster.objects.all()
    
    # Filter by sales type
    if sales_type_filter:
        instances = instances.filter(customer__sales_type=sales_type_filter)
        filter_data['sales_type'] = sales_type_filter
    sales_type_li = Customers.objects.values_list('sales_type', flat=True).distinct()
    
    # Filter by product type
    if product_type_filter:
        instances = instances.filter(product_type=product_type_filter)
        filter_data['product_type'] = product_type_filter

    # Search functionality
    if query:
        instances = instances.filter(
            Q(product_type__icontains=query) |
            Q(invoice_no__icontains=query)
        )
        filter_data['q'] = query

    # Calculate total outstanding count
    total_outstanding_count = sum([item.get_outstanding_count() for item in instances])
    
     # Log user activity
    log_activity(
        created_by=request.user if request.user.is_authenticated else None,
        description=f"Viewed outstanding list Print with filters: date={date},  Route: {route_filter}, Sales Type: {sales_type_filter}, "
                f"Product Type: {product_type_filter}, Query: {query}"
    )
    
    context = {
        'instances': instances,
        'date': date,
        'route_li': route_li,
        'sales_type_li': sales_type_li,
        'filter_data': filter_data,
        'total_outstanding_count': total_outstanding_count,
        'product_types': dict(PRODUCT_TYPES)
    }

    return render(request, 'client_management/customer_outstanding/print_outstanding_report.html', context)


@login_required
def create_customer_outstanding(request):
    customer_pk = request.GET.get("customer_pk")
    message = ""

    if request.method == "POST":

        # -------------------- FORMS --------------------
        if customer_pk:
            customer_outstanding_form = CustomerOutstandingSingleForm(request.POST)
        else:
            customer_outstanding_form = CustomerOutstandingForm(request.POST)

        amount_form = CustomerOutstandingAmountForm(request.POST)
        bottle_form = CustomerOutstandingBottleForm(request.POST)
        coupon_form = CustomerOutstandingCouponsForm(request.POST)

        product_type = request.POST.get("product_type")
        is_valid = False

        if product_type == "amount":
            is_valid = customer_outstanding_form.is_valid() and amount_form.is_valid()
        elif product_type == "emptycan":
            is_valid = customer_outstanding_form.is_valid() and bottle_form.is_valid()
        elif product_type == "coupons":
            is_valid = customer_outstanding_form.is_valid() and coupon_form.is_valid()

        if not is_valid:
            message = generate_form_errors(customer_outstanding_form, formset=False)
            return JsonResponse({
                "status": "false",
                "title": "Validation Failed",
                "message": message
            })

        try:
            with transaction.atomic():

                # -------------------- SAVE OUTSTANDING MASTER --------------------
                outstanding = customer_outstanding_form.save(commit=False)

                

                if customer_pk:
                    outstanding.customer = Customers.objects.get(pk=customer_pk)
                customer = outstanding.customer

                # 🔥 GET SALESMAN USING ROUTE → VAN → SALESMAN
                route = customer.routes

                van_route = Van_Routes.objects.filter(routes=route).select_related("van").first()
                van = van_route.van if van_route else None
                salesman = van.salesman if van else None
                outstanding.created_by = salesman.id
                outstanding.outstanding_date = outstanding.created_date
                outstanding.save()
                

                # -------------------- HANDLE AMOUNT OUTSTANDING --------------------
                if product_type == "amount":
                    amount_obj = amount_form.save(commit=False)
                    amount_obj.customer_outstanding = outstanding
                    amount_obj.save()

                    amount_value = amount_obj.amount

                    # -------------------- CREATE INVOICE --------------------
                    invoice = Invoice.objects.create(
                        created_date=outstanding.created_date,
                        invoice_date=outstanding.created_date,
                        net_taxable=amount_value,
                        vat=0,
                        discount=0,
                        amout_total=amount_value,
                        amout_recieved=0,
                        customer=customer,
                        reference_no="Outstanding added from backend",
                        salesman=salesman,
                        invoice_type="credit_invoice",
                        invoice_status="non_paid"
                    )

                    outstanding.invoice_no = invoice.invoice_no
                    outstanding.save(update_fields=["invoice_no","outstanding_date"])

                    # # -------------------- CREATE INVOICE ITEM --------------------
                    # item = ProdutItemMaster.objects.get(product_name="5 Gallon")
                    # InvoiceItems.objects.create(
                    #     category=item.category,
                    #     product_items=item,
                    #     qty=0,
                    #     rate=customer.rate or item.rate,
                    #     invoice=invoice,
                    #     remarks="Invoice generated from customer outstanding"
                    # )

                # -------------------- EMPTY CAN --------------------
                elif product_type == "emptycan":
                    bottle_obj = bottle_form.save(commit=False)
                    bottle_obj.customer_outstanding = outstanding
                    bottle_obj.save()

                # -------------------- COUPONS --------------------
                elif product_type == "coupons":
                    coupon_obj = coupon_form.save(commit=False)
                    coupon_obj.customer_outstanding = outstanding
                    coupon_obj.save()

                # -------------------- REBUILD CUSTOMER OUTSTANDING REPORT --------------------
                if product_type == "amount":
                    totals = Invoice.objects.filter(
                        customer=customer,
                        is_deleted=False
                    ).aggregate(
                        total=Sum("amout_total"),
                        received=Sum("amout_recieved")
                    )

                    outstanding_value = (totals["total"] or 0) - (totals["received"] or 0)

                    CustomerOutstandingReport.objects.update_or_create(
                        customer=customer,
                        product_type="amount",
                        defaults={"value": outstanding_value}
                    )

                elif product_type == "emptycan":
                    total_bottles = OutstandingProduct.objects.filter(
                        customer_outstanding__customer=customer
                    ).aggregate(total=Sum("empty_bottle"))["total"] or 0

                    CustomerOutstandingReport.objects.update_or_create(
                        customer=customer,
                        product_type="emptycan",
                        defaults={"value": total_bottles}
                    )

                elif product_type == "coupons":
                    total_coupons = OutstandingCoupon.objects.filter(
                        customer_outstanding__customer=customer
                    ).aggregate(total=Sum("count"))["total"] or 0

                    CustomerOutstandingReport.objects.update_or_create(
                        customer=customer,
                        product_type="coupons",
                        defaults={"value": total_coupons}
                    )

                redirect_url = (
                    reverse("customer_outstanding_list")
                    
                )

                return JsonResponse({
                    "status": "true",
                    "title": "Successfully Created",
                    "message": "Customer outstanding created successfully.",
                    "redirect": "true",
                    "redirect_url": redirect_url,
                })

        except Exception as e:
            return JsonResponse({
                "status": "false",
                "title": "Error",
                "message": str(e)
            })

    # -------------------- GET --------------------
    else:
        if customer_pk:
            customer_outstanding_form = CustomerOutstandingSingleForm()
        else:
            customer_outstanding_form = CustomerOutstandingForm()

        context = {
            "customer_outstanding_form": customer_outstanding_form,
            "customer_outstanding_amount_form": CustomerOutstandingAmountForm(),
            "customer_outstanding_bottles_form": CustomerOutstandingBottleForm(),
            "customer_outstanding_coupon_form": CustomerOutstandingCouponsForm(),
            "customer_pk": customer_pk,
            "url": reverse("create_customer_outstanding"),
            "page_title": "Create Customer Outstanding",
            "customer_outstanding_page": True,
            "is_need_datetime_picker": True,
        }

        return render(
            request,
            "client_management/customer_outstanding/create.html",
            context
        )


def delete_outstanding_adjustment(request, pk):
    """
    Reverse a payment adjustment (negative outstanding)
    """
    print("working")
    try:
        with transaction.atomic():

            # 1️⃣ Get CustomerOutstanding
            print("primaryKey:",pk)
            customer_outstanding = get_object_or_404(CustomerOutstanding, pk=pk)
            

            if customer_outstanding.product_type != "amount":
                return JsonResponse({
                    "status": "false",
                    "message": "Only amount adjustments can be reversed"
                }, status=400)

            # 2️⃣ Get OutstandingAmount (must be negative)
            outstanding = OutstandingAmount.objects.select_for_update().get(
                customer_outstanding=customer_outstanding
            )

            if outstanding.amount >= 0:
                return JsonResponse({
                    "status": "false",
                    "message": "This is not a payment adjustment"
                }, status=400)

            adjust_amount = abs(outstanding.amount)

            # 3️⃣ Restore invoice
            invoice = Invoice.objects.select_for_update().get(
                invoice_no=customer_outstanding.invoice_no,
                customer=customer_outstanding.customer,
                is_deleted=False
            )

            invoice.amout_recieved -= adjust_amount
            if invoice.amout_recieved < 0:
                invoice.amout_recieved = Decimal("0.00")

            if invoice.amout_recieved >= invoice.amout_total:
                invoice.invoice_status = "paid"
            else:
                invoice.invoice_status = "non_paid"

            invoice.save(update_fields=["amout_recieved", "invoice_status"])

            # 4️⃣ Update CustomerOutstandingReport
            report = CustomerOutstandingReport.objects.select_for_update().filter(
                customer=customer_outstanding.customer,
                product_type="amount"
            ).first()

            if report:
                report.value += adjust_amount
                report.save(update_fields=["value"])

            # 5️⃣ Delete adjustment rows
            outstanding.delete()
            customer_outstanding.delete()

            return JsonResponse({
                "status": "true",
                "title": "Adjustment Reversed",
                "message": f"Payment adjustment of {adjust_amount} reversed successfully",
                "reload": "true",
            })

    except Exception as e:
        return JsonResponse({
            "status": "false",
            "title": "Error",
            "message": str(e)
        }, status=500)


@login_required
def delete_outstanding(request, pk):
    """
    outstanding deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    try:
        with transaction.atomic():
            customer_outstanding = CustomerOutstanding.objects.get(pk=pk)
            report = CustomerOutstandingReport.objects.filter(customer=customer_outstanding.customer)
            
            if customer_outstanding.product_type == "amount":
                amount = OutstandingAmount.objects.get(customer_outstanding=customer_outstanding).amount
                report = report.filter(product_type="amount").first()
                report.value -= amount
            if customer_outstanding.product_type == "emptycan":
                emptycan = OutstandingProduct.objects.get(customer_outstanding=customer_outstanding).empty_bottle
                report = report.filter(product_type="emptycan").first()
                report.value -= emptycan
            if customer_outstanding.product_type == "coupons":
                coupons = OutstandingCoupon.objects.filter(customer_outstanding=customer_outstanding).aggregate(total_count=Sum('count'))['total_count'] or 0
                report = report.filter(product_type="coupons").first()
                report.value -= coupons
            
            report.save()
            
            if (invoices:=Invoice.objects.filter(invoice_no=customer_outstanding.invoice_no)).exists():
                for invoice in invoices:
                    if CustomerSupply.objects.filter(invoice_no=invoice.invoice_no).exists():
                        customer_supply_instance = get_object_or_404(CustomerSupply, invoice_no=invoice.invoice_no)
                        supply_items_instances = CustomerSupplyItems.objects.filter(customer_supply=customer_supply_instance)
                        five_gallon_qty = supply_items_instances.filter(product__product_name="5 Gallon").aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
                        
                        DiffBottlesModel.objects.filter(
                            delivery_date__date=customer_supply_instance.created_date.date(),
                            assign_this_to=customer_supply_instance.salesman,
                            customer=customer_supply_instance.customer_id
                            ).update(status='pending')
                    
                        # Handle coupon deletions and adjustments
                        handle_coupons(customer_supply_instance, five_gallon_qty)
                        
                        # Update van product stock and empty bottle counts
                        update_van_product_stock(customer_supply_instance, supply_items_instances, five_gallon_qty)
                        
                        # Mark customer supply and items as deleted
                        customer_supply_instance.delete()
                        supply_items_instances.delete()
                        
                        if CustomerCoupon.objects.filter(invoice_no=invoice.invoice_no).exists():
                            instance = CustomerCoupon.objects.get(invoice_no=invoice.invoice_no)
                            delete_coupon_recharge(instance.invoice_no)
                            
                    invoice.is_deleted=True
                    invoice.save()
                    
                    InvoiceItems.objects.filter(invoice=invoice).update(is_deleted=True)
            
            customer_outstanding.delete()
            
            log_activity(
                created_by=request.user,
                description=f"Outstanding record with ID {pk} was successfully deleted."
            )
            
            status_code = status.HTTP_200_OK
            response_data = {
            "status": "true",
            "title": "Succesfully Deleted",
            "message": "Succesfully Deleted",
            "reload": "true",
            }
            
    except IntegrityError as e:
        # Handle database integrity error
        status_code = status.HTTP_500_INTERNAL_SERVER_ERROR
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }

    except Exception as e:
        # Handle other exceptions
        status_code = status.HTTP_500_INTERNAL_SERVER_ERROR
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }
    return HttpResponse(json.dumps(response_data), status=status_code, content_type='application/javascript')

# customer count

def customer_count(request):
    routes = RouteMaster.objects.all()

    total_cash = 0
    total_credit = 0
    total_coupon = 0
    total_foc = 0
    total_customers = 0

    customer_counts = []

    for route in routes:
        sales_man = ''
        van_route = Van_Routes.objects.filter(routes=route).first()
        if van_route and van_route.van:
            sales_man = van_route.van.salesman

        cash_count = Customers.objects.filter(
            is_guest=False, is_deleted=False, routes=route, sales_type='CASH'
        ).count()

        credit_count = Customers.objects.filter(
            is_guest=False, is_deleted=False, routes=route, sales_type='CREDIT'
        ).count()

        coupon_count = Customers.objects.filter(
            is_guest=False, is_deleted=False, routes=route,
            sales_type__in=['CASH COUPON', 'CREDIT COUPON']
        ).count()

        foc_count = Customers.objects.filter(
            is_guest=False, is_deleted=False, routes=route, sales_type='FOC'
        ).count()

        route_total = cash_count + credit_count + coupon_count + foc_count

        total_cash += cash_count
        total_credit += credit_count
        total_coupon += coupon_count
        total_foc += foc_count
        total_customers += route_total

        if route_total > 0:
            customer_counts.append({
                'route_name': route.route_name,
                'sales_man': sales_man,
                'cash_count': cash_count,
                'credit_count': credit_count,
                'coupon_count': coupon_count,
                'foc_count': foc_count,
                'total_customer': route_total
            })

    # ---- Customers with NO ROUTE ----
    cash_count = Customers.objects.filter(is_guest=False, is_deleted=False, routes=None, sales_type='CASH').count()
    credit_count = Customers.objects.filter(is_guest=False, is_deleted=False, routes=None, sales_type='CREDIT').count()
    coupon_count = Customers.objects.filter(
        is_guest=False, is_deleted=False, routes=None,
        sales_type__in=['CASH COUPON', 'CREDIT COUPON']
    ).count()
    foc_count = Customers.objects.filter(is_guest=False, is_deleted=False, routes=None, sales_type='FOC').count()

    route_total = cash_count + credit_count + coupon_count + foc_count

    total_cash += cash_count
    total_credit += credit_count
    total_coupon += coupon_count
    total_foc += foc_count
    total_customers += route_total

    customer_counts.append({
        'route_name': 'Not Specified',
        'sales_man': '',
        'cash_count': cash_count,
        'credit_count': credit_count,
        'coupon_count': coupon_count,
        'foc_count': foc_count,
        'total_customer': route_total
    })

    context = {
        'customer_counts': customer_counts,
        'total_cash': total_cash,
        'total_credit': total_credit,
        'total_coupon': total_coupon,
        'total_foc': total_foc,
        'total_customers': total_customers,
    }

    return render(request, 'client_management/customer_count.html', context)



def bottle_count(request):
    routes = RouteMaster.objects.all()

    context = {
        'instances': routes,
    }

    return render(request, 'client_management/bottle_count.html', context)
    

def bottle_count_route_wise(request, route_id):
    customers = Customers.objects.filter(is_guest=False, routes__pk=route_id)
    
    context = {
        "instances" : customers
    }

    return render(request, 'client_management/route_details.html', context)
    
@login_required
def customer_orders(request):
    """
    Customer orders List
    :param request:
    :return: Customer orders list view
    """
    filter_data = {}
    instances = CustomerOrders.objects.all()
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(customer__customer_name__icontains=query) |
            Q(customer__customer_id__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__whats_app__icontains=query) |
            Q(customer__email_id__icontains=query)
        )
        title = "Customer Order List - %s" % query
        filter_data['q'] = query
        
    acknowledge_form = CustomerOrdersAcknowledgeForm()
            
    context = {
        'instances': instances,
        'acknowledge_form': acknowledge_form,
        'page_name' : 'Customer Order List',
        'page_title' : 'Customer Order List',
        
        'is_customer_outstanding': True,
        'is_need_datetime_picker': True,
        'filter_data': filter_data,
    }

    return render(request, 'client_management/customer_order_list.html', context)

def customer_order_status_acknowledge(request,pk):
            
    try:
        with transaction.atomic():
            instance = CustomerOrders.objects.get(pk=pk)
            form = CustomerOrdersAcknowledgeForm(request.POST,instance=instance)

            data = form.save(commit=False)
            data.save()
            
            if data.order_status == "approve":
                DiffBottlesModel.objects.create(
                    product_item=data.product,
                    quantity_required=data.quantity,
                    delivery_date=data.delivery_date,
                    assign_this_to=data.customer.sales_staff,
                    mode="paid",
                    amount=data.total_amount,
                    discount_net_total=data.total_net_amount,
                    customer=data.customer,
                    created_by=data.created_by,
                    created_date=datetime.today(),
                )
                
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Acknowledged",
                'reload': 'true',
            }
            
    except IntegrityError as e:
        # Handle database integrity error
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }

    except Exception as e:
        # Handle other exceptions
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


def nonvisitreason_List(request):
    all_nonvisitreason= NonVisitReason.objects.all()
    context = {'all_nonvisitreason': all_nonvisitreason}
    return render(request, 'client_management/NonVisitReason/index_nonvisitReason.html', context)

def create_nonvisitreason(request):
    if request.method == 'POST':
        form = Create_NonVisitReasonForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = str(request.user.id)
            data.save()
            messages.success(request, 'Non Visit Reason created successfully!')
            return redirect('nonvisitreason_List')
        else:
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = Create_NonVisitReasonForm()
    context = {'form': form}
    return render(request, 'client_management/NonVisitReason/create_nonvisitreason.html', context)

def delete_nonvisitreason(request, id):
    delete_nonvisitreason = NonVisitReason.objects.get(id=id)
    if request.method == 'POST':
        delete_nonvisitreason.delete()
        log_activity(
                created_by=request.user,
                description=f"Non-visit reason with ID {id} was successfully deleted."
            )
        return redirect('nonvisitreason_List')
    return render(request, 'client_management/NonVisitReason/delete_nonvisitreason.html', {'delete_nonvisitreason': delete_nonvisitreason})


@transaction.atomic
def populate_models_from_excel(data, user):
    # user = CustomUser.objects.get(username=user.username)
    for index, row in data.iterrows():
        customer_id = row['customer_id']
        customer_name = row['customer_name']
        amount = Decimal(row['amount'])
        str_date = str(row['date'])
        
        if isinstance(str_date, str):
            str_date = str_date.split()[0]  # Take only the date part if it includes time
        date = datetime.strptime(str_date, '%Y-%m-%d')
        
        try:
            customer = Customers.objects.get(custom_id=customer_id)
        except Customers.DoesNotExist:
            print(f"Customer {customer_name} does not exist.")
            continue

        customer_outstanding = CustomerOutstanding.objects.create(
            customer=customer,
            product_type='amount',
            created_by=user.id,
            modified_by=user.id,
            created_date=date,
        )

        outstanding_amount = OutstandingAmount.objects.create(
            customer_outstanding=customer_outstanding,
            amount=amount
        )

        if (instances := CustomerOutstandingReport.objects.filter(customer=customer, product_type='amount')).exists():
            report = instances.first()
        else:
            report = CustomerOutstandingReport.objects.create(customer=customer, product_type='amount')

        report.value += amount
        report.save()
        
        date_part = timezone.now().strftime('%Y%m%d')
        try:
            invoice_last_no = Invoice.objects.filter(is_deleted=False).latest('created_date')
            last_invoice_number = invoice_last_no.invoice_no

            parts = last_invoice_number.split('-')
            if len(parts) == 3 and parts[0] == 'WTR' and parts[1] == date_part:
                prefix, old_date_part, number_part = parts
                new_number_part = int(number_part) + 1
                invoice_number = f'{prefix}-{date_part}-{new_number_part:04d}'
            else:
                random_part = str(random.randint(1000, 9999))
                invoice_number = f'WTR-{date_part}-{random_part}'
        except Invoice.DoesNotExist:
            random_part = str(random.randint(1000, 9999))
            invoice_number = f'WTR-{date_part}-{random_part}'
        
        invoice = Invoice.objects.create(
            invoice_no=invoice_number,
            created_date=outstanding_amount.customer_outstanding.created_date,
            net_taxable=outstanding_amount.amount,
            vat=0,
            discount=0,
            amout_total=outstanding_amount.amount,
            amout_recieved=0,
            customer=outstanding_amount.customer_outstanding.customer,
            reference_no=f"custom_id{outstanding_amount.customer_outstanding.customer.custom_id}",
            salesman=user.id
        )
        customer_outstanding.invoice_no = invoice.invoice_no
        customer_outstanding.save()
        
        if outstanding_amount.customer_outstanding.customer.sales_type == "CREDIT":
            invoice.invoice_type = "credit_invoice"
            invoice.save()

        item = ProdutItemMaster.objects.get(product_name="5 Gallon")
        InvoiceItems.objects.create(
            category=item.category,
            product_items=item,
            qty=0,
            rate=outstanding_amount.customer_outstanding.customer.rate,
            invoice=invoice,
            remarks='invoice generated from backend reference no : ' + invoice.reference_no
        )

        print(f"Processed row {index + 1} for customer {customer_name}")

def upload_outstanding(request):
    if request.method == 'POST':
        form = UploadOutstandingForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            route = form.cleaned_data['route']

            file_name = default_storage.save(excel_file.name, excel_file)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name)

            data = pd.read_excel(file_path)

            populate_models_from_excel(data, request.user)

            default_storage.delete(file_name)

            messages.success(request, "Outstanding uploaded successfully.")
            return redirect(reverse('customer_outstanding_list'))
    else:
        form = UploadOutstandingForm()

    return render(request, 'client_management/customer_outstanding/upload.html', {'form': form})



def edit_customer_outstanding(request, outstanding_pk):
    """
    Edit Customer Outstanding
    """
    customer_pk = request.GET.get("customer_pk")
    print("customer_pk",customer_pk)
    # Get the outstanding instance to edit
    outstanding_instance = get_object_or_404(CustomerOutstanding, id=outstanding_pk)
    
    message = ''
    if request.method == 'POST':
        if not customer_pk:
            customer_outstanding_form = CustomerOutstandingForm(request.POST, instance=outstanding_instance)
        else:
            customer_outstanding_form = CustomerOutstandingSingleForm(request.POST, instance=outstanding_instance)

        customer_outstanding_amount_form = CustomerOutstandingAmountForm(request.POST)
        customer_outstanding_bottles_form = CustomerOutstandingBottleForm(request.POST)
        customer_outstanding_coupon_form = CustomerOutstandingCouponsForm(request.POST)
        
        
        is_form_valid = False
        if request.POST.get('product_type') == "amount":
            if customer_outstanding_form.is_valid() and customer_outstanding_amount_form.is_valid():
                is_form_valid = True
            else:
                message = generate_form_errors(customer_outstanding_form, formset=False)
                message += generate_form_errors(customer_outstanding_amount_form, formset=False)

        if request.POST.get('product_type') == "emptycan":
            if customer_outstanding_form.is_valid() and customer_outstanding_bottles_form.is_valid():
                is_form_valid = True
            else:
                message = generate_form_errors(customer_outstanding_form, formset=False)
                message += generate_form_errors(customer_outstanding_bottles_form, formset=False)

        if request.POST.get('product_type') == "coupons":
            if customer_outstanding_form.is_valid() and customer_outstanding_coupon_form.is_valid():
                is_form_valid = True
            else:
                message = generate_form_errors(customer_outstanding_form, formset=False)
                message += generate_form_errors(customer_outstanding_coupon_form, formset=False)

        if is_form_valid:
            try:
                with transaction.atomic():
                    # Save updated outstanding data
                    outstanding_data = customer_outstanding_form.save(commit=False)
                    outstanding_data.modified_by = request.user.id
                    outstanding_data.modified_date = datetime.today()
                    outstanding_data.save()

                    # Save updated product-specific data
                    if outstanding_data.product_type == "amount":
                        outstanding_amount = customer_outstanding_amount_form.save(commit=False)
                        outstanding_amount.customer_outstanding = outstanding_data
                        outstanding_amount.save()

                        # Update CustomerOutstandingReport
                        report, created = CustomerOutstandingReport.objects.get_or_create(
                            customer=outstanding_data.customer,
                            product_type='amount',
                        )
                        report.value = outstanding_amount.amount
                        report.save()

                    elif outstanding_data.product_type == "emptycan":
                        outstanding_bottle = customer_outstanding_bottles_form.save(commit=False)
                        outstanding_bottle.customer_outstanding = outstanding_data
                        outstanding_bottle.save()

                        # Update report for empty cans
                        report, created = CustomerOutstandingReport.objects.get_or_create(
                            customer=outstanding_data.customer,
                            product_type='emptycan',
                        )
                        report.value = outstanding_bottle.empty_bottle
                        report.save()

                    elif outstanding_data.product_type == "coupons":
                        outstanding_coupon = customer_outstanding_coupon_form.save(commit=False)
                        outstanding_coupon.customer_outstanding = outstanding_data
                        outstanding_coupon.save()

                        # Update report for coupons
                        report, created = CustomerOutstandingReport.objects.get_or_create(
                            customer=outstanding_data.customer,
                            product_type='coupons',
                        )
                        report.value = outstanding_coupon.count
                        report.save()

                    redirect_url = reverse('customer_outstanding_details', args=[outstanding_data.customer.customer_id])
                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "Customer Outstanding updated successfully.",
                        'redirect': 'true',
                        'redirect_url': redirect_url,
                    }

            except IntegrityError as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        # Prepopulate the forms with the existing data
        if not customer_pk:
            customer_outstanding_form = CustomerOutstandingForm(instance=outstanding_instance)
        else:
            customer_outstanding_form = CustomerOutstandingSingleForm(instance=outstanding_instance)

        customer_outstanding_amount_form = CustomerOutstandingAmountForm()
        customer_outstanding_bottles_form = CustomerOutstandingBottleForm()
        customer_outstanding_coupon_form = CustomerOutstandingCouponsForm()
        
        context = {
            'customer_outstanding_form': customer_outstanding_form,
            'customer_outstanding_amount_form': customer_outstanding_amount_form,
            'customer_outstanding_bottles_form': customer_outstanding_bottles_form,
            'customer_outstanding_coupon_form': customer_outstanding_coupon_form,
            'customer_pk': customer_pk,
            "url": reverse('edit_customer_outstanding', args=[outstanding_pk]),
            'page_title': 'Edit customer outstanding',
            'customer_outstanding_page': True,
            'is_need_datetime_picker': True
        }

        return render(request, 'client_management/customer_outstanding/edit.html', context)

def ageing_report_view(request): 
    route_name = request.GET.get('route', None)
    selected_route = None
    routes = RouteMaster.objects.all()

    if route_name:
        selected_route = RouteMaster.objects.filter(route_name=route_name).first()

    totals = {
        'total_less_than_30': 0,
        'total_between_31_and_60': 0,
        'total_between_61_and_90': 0,
        'total_between_91_and_150': 0,
        'total_between_151_and_365': 0,
        'total_more_than_365': 0,
        'total_grand_total': 0,
    }

    if selected_route:
        ageing_data = get_customer_outstanding_aging(selected_route.route_name)
        for item in ageing_data:
            totals['total_less_than_30'] += item['less_than_30']
            totals['total_between_31_and_60'] += item['between_31_and_60']
            totals['total_between_61_and_90'] += item['between_61_and_90']
            totals['total_between_91_and_150'] += item['between_91_and_150']
            totals['total_between_151_and_365'] += item['between_151_and_365']
            totals['total_more_than_365'] += item['more_than_365']
            totals['total_grand_total'] += item['grand_total']

    context = {
        'selected_route': selected_route,
        'routes': routes,
        'totals': totals,
    }

    return render(request, 'client_management/ageing_report.html', context)
   

def print_ageing_report_view(request):
    route_name = request.GET.get('route', None)
    selected_route = None

    routes = RouteMaster.objects.all()

    if route_name:
        try:
            selected_route = RouteMaster.objects.get(route_name=route_name)
        except RouteMaster.DoesNotExist:
            selected_route = None 
    
    context = {
        'selected_route': selected_route,
        'routes': routes,
    }

    return render(request, 'client_management/print_ageing_report.html', context)

from .templatetags.client_templatetags import get_customer_outstanding_aging
  
def ageing_report_excel(request):
    route_name = request.GET.get('route', None)
    selected_route = RouteMaster.objects.get(route_name=route_name) if route_name else None
    aging_report = get_customer_outstanding_aging(selected_route)
    
    # Create DataFrame from the aging report data
    df = pd.DataFrame(aging_report)
    
    # Convert columns to appropriate data types
    for column in df.columns:
        if df[column].dtype == 'object':
            # Attempt to convert to numeric, errors='ignore' ensures non-numeric values stay as they are
            df[column] = pd.to_numeric(df[column], errors='ignore')
    
    # Create the HttpResponse object with the appropriate Excel header
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ageing_report.xlsx"'
    
    # Use Pandas Excel writer
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Ageing Report', index=False)
    
    return response

@login_required
def customer_outstanding_detail(request,customer_id):
    """
    Customer Outstanding details List
    :param request:
    :return: Customer Outstanding list view
    """
    filter_data = {}
    instances = CustomerOutstanding.objects.filter(customer__pk=customer_id,product_type='amount')
    
    query = request.GET.get("q")
    date = request.GET.get('date')
    route_filter = request.GET.get('route_name')
    
    if date:
        date = datetime.strptime(date, '%Y-%m-%d').date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')

    else:
        date = datetime.today().date()
        filter_data['filter_date'] = date.strftime('%Y-%m-%d')
    
    if route_filter:
            instances = instances.filter(customer__routes__route_name=route_filter)
    route_li = RouteMaster.objects.all()
    
    if query:

        instances = instances.filter(
            Q(product_type__icontains=query) |
            Q(invoice_no__icontains=query) 
        )
        title = "Outstanding List - %s" % query
        filter_data['q'] = query
    
    context = {
        'instances': instances.order_by("-created_date"),
        'page_name' : 'Customer Outstanding List',
        'page_title' : 'Customer Outstanding List',
        'customer_id': customer_id,
        'is_customer_outstanding': True,
        'is_need_datetime_picker': True,
        'filter_data': filter_data,
        'route_li':route_li,
    }

    return render(request, 'client_management/customer_outstanding/customer_outstanding_list.html', context)


@login_required
def customer_supply_bottles_list(request, supply_id):
    """
    Displays a list of specific bottles supplied in a customer supply via NFC.
    """
    from bottle_management.models import BottleLedger
    customer_supply = get_object_or_404(CustomerSupply, pk=supply_id)
    
    # Query BottleLedger instead of SupplyItemBottle, since the app's NFC supply API logs it there
    bottles = BottleLedger.objects.filter(
        reference=customer_supply.invoice_no, 
        action__in=["SUPPLY", "FOC"]
    ).select_related('bottle')
    
    context = {
        'customer_supply': customer_supply,
        'bottles': bottles,
        'page_name': 'Supplied Bottles',
        'page_title': f'Bottles - {customer_supply.customer.customer_name}',
    }
    
    return render(request, 'client_management/customer_supply/bottles_list.html', context)

def export_customer_outstanding_to_excel(request):
    instances = CustomerOutstanding.objects.filter(product_type='amount')
    
    if request.GET.get("customer_pk"):
        instances = instances.filter(customer__pk=request.GET.get("customer_pk"))
    if request.GET.get("route_name"):
        instances = instances.filter(customer__routes__route_name=request.GET.get("route_name"))
    if request.GET.get("date"):
        instances = instances.filter(created_date__date__lte=datetime.strptime(request.GET.get("date"), '%Y-%m-%d').date())
    
    # Prepare data for the Excel file
    data = []
    for item in instances:
        # Convert created_date to timezone-naive
        if item.created_date.tzinfo is not None:
            created_date_naive = item.created_date.astimezone(timezone.utc).replace(tzinfo=None)
        else:
            created_date_naive = item.created_date  # Already naive

        data.append([
            created_date_naive,  # Use the naive datetime
            item.invoice_no,
            item.customer.customer_name,
            item.customer.building_name,
            item.customer.routes.route_name,
            item.get_outstanding_count()
        ])
    
    # Create a DataFrame
    df = pd.DataFrame(data, columns=['Created Date', 'Invoice No', 'Customer', 'Building No', 'Route', 'Count'])
    
    # Create the HttpResponse object with the appropriate Excel headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customer_outstanding_report.xlsx"'
    
    # Use pandas to write to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Outstanding Report')
    
    return response

def print_customer_outstandings(request, customer_id):
    instances = CustomerOutstanding.objects.filter(customer__pk=customer_id, product_type='amount')
    context = {
        'instances': instances,
        'page_title': 'Customer Outstanding Print',
    }
    return render(request, 'client_management/customer_outstanding/customer_outstanding_print.html', context)

def customer_transaction_list(request):
    """
    Customer Transaction List
    :param request:
    :return: Customer Transaction list view
    """
    filter_data = {}
    q = request.GET.get('q', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Initialize totals
    total_amount = total_discount = total_net_payable = total_vat = total_qty =0
    total_grand_total = total_amount_recieved = 0
    customer_pk = request.GET.get("customer_pk")
    
    customer = None
    route_name = ""

    if customer_pk:
        try:
            customer = Customers.objects.get(pk=customer_pk)
            route_name = customer.routes.route_name if customer.routes else ""
        except Customers.DoesNotExist:
            pass
    # Get CustomerSupply items excluding certain sales types
    sales = CustomerSupplyItems.objects.filter(
        customer_supply__customer__pk=customer_pk
    ).exclude(customer_supply__customer__sales_type__in=["CASH COUPON", "CREDIT COUPON"]).order_by('-customer_supply__supply_date')

    # Get CustomerCoupon items
    coupons = CustomerCouponItems.objects.filter(
        customer_coupon__customer__pk=customer_pk
    )

    # Handle date filtering
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            filter_data['filter_start_date'] = start_date.strftime('%Y-%m-%d')
            filter_data['filter_end_date'] = end_date.strftime('%Y-%m-%d')

            # Apply date filtering to sales and coupons
            sales = sales.filter(customer_supply__supply_date__date__range=(start_date, end_date))
            coupons = coupons.filter(customer_coupon__created_date__date__range=(start_date, end_date))
        except ValueError:
            start_date = None  # Invalid date format, reset
            end_date = None

    sales_report_data = []

    # Process CustomerSupply data
    for sale in sales:
        sales_report_data.append({
            'date': sale.customer_supply.supply_date,
            'ref_invoice_no': sale.customer_supply.reference_number,
            'invoice_number': sale.customer_supply.invoice_no,
            'product_name': sale.product.product_name,
            'sales_type': sale.customer_supply.customer.sales_type,
            'qty': sale.quantity,
            'amount': sale.customer_supply.grand_total,
            'discount': sale.customer_supply.discount,
            'net_taxable': sale.customer_supply.subtotal,
            'vat_amount': sale.customer_supply.vat,
            'grand_total': sale.customer_supply.grand_total,
            'amount_collected': sale.customer_supply.amount_recieved,
        })

        # Update totals
        total_qty += sale.quantity
        total_amount += sale.customer_supply.grand_total
        total_discount += sale.customer_supply.discount
        total_net_payable += sale.customer_supply.net_payable
        total_vat += sale.customer_supply.vat
        total_grand_total += sale.customer_supply.grand_total
        total_amount_recieved += sale.customer_supply.amount_recieved

    # Process CustomerCoupon data
    for coupon in coupons:
        vat_rate = Tax.objects.get(name="VAT").percentage if Tax.objects.filter(name="VAT").exists() else 0
        sales_report_data.append({
            'date': coupon.customer_coupon.created_date,
            'ref_invoice_no': coupon.customer_coupon.reference_number,
            'invoice_number': coupon.customer_coupon.invoice_no,
            'product_name': coupon.coupon.book_num,
            'sales_type': coupon.customer_coupon.customer.sales_type,
            'qty': 1,
            'amount': coupon.customer_coupon.grand_total,
            'discount': coupon.customer_coupon.discount,
            'net_taxable': coupon.customer_coupon.net_amount,
            'vat_amount': vat_rate,
            'grand_total': coupon.customer_coupon.grand_total,
            'amount_collected': coupon.customer_coupon.amount_recieved,
        })

        # Update totals
        total_amount += coupon.customer_coupon.grand_total
        total_discount += coupon.customer_coupon.discount
        total_net_payable += coupon.customer_coupon.net_amount
        total_vat += vat_rate
        total_grand_total += coupon.customer_coupon.grand_total
        total_amount_recieved += coupon.customer_coupon.amount_recieved

    # Fetch Customer Outstanding instances
    customer_outstanding_instances = CustomerOutstanding.objects.filter(customer__pk=customer_pk)

    # Apply date filter if provided
    if start_date and end_date:
        customer_outstanding_instances = customer_outstanding_instances.filter(
            outstanding_date__date__range=(start_date, end_date)
        )

    customer_outstanding_instances = customer_outstanding_instances.order_by('-outstanding_date')

    # Prepare totals and other necessary data
    total_outstanding_amount = 0
    total_outstanding_coupons = 0
    total_outstanding_emptycan = 0

    # Loop through each outstanding instance to calculate totals
    for outstanding in customer_outstanding_instances:
        # Calculate amounts for each instance
        outstanding_amount = outstanding.outstandingamount_set.aggregate(total=Sum('amount'))['total'] or 0
        outstanding_coupons = outstanding.outstandingcoupon_set.aggregate(total=Sum('count'))['total'] or 0
        outstanding_emptycan = outstanding.outstandingproduct_set.aggregate(total=Sum('empty_bottle'))['total'] or 0

        # Update totals
        total_outstanding_amount += outstanding_amount
        total_outstanding_coupons += outstanding_coupons
        total_outstanding_emptycan += outstanding_emptycan

        # Add calculated values to outstanding instance for template rendering
        outstanding.total_amount = outstanding_amount
        outstanding.total_coupons = outstanding_coupons
        outstanding.total_emptycan = outstanding_emptycan
        
    # Filter CollectionPayment instances for the customer
    collection_payment_instance = CollectionPayment.objects.filter(
        customer__pk=customer_pk
    )

    # Apply date filter if provided
    if start_date and end_date:
        collection_payment_instance = collection_payment_instance.filter(
            created_date__date__range=(start_date, end_date)
        )

    collection_payment_instance = collection_payment_instance.order_by('-created_date')

    # Calculate total amounts
    total_amount_received = collection_payment_instance.aggregate(Sum('amount_received'))['amount_received__sum'] or 0
    total_discounts = sum(payment.total_discounts() for payment in collection_payment_instance)
    total_net_taxable = sum(payment.total_net_taxeble() for payment in collection_payment_instance)
    total_collection_vat = sum(payment.total_vat() for payment in collection_payment_instance)
    total_collected_amount = sum(payment.collected_amount() for payment in collection_payment_instance)

    # Filter redeemed_coupon_instances for the customer and sales_type="CASH COUPON"
    redeemed_coupon_instances = CustomerSupply.objects.filter(
        customer__pk=customer_pk,
        customer__sales_type="CASH COUPON"
    )

    # Apply date filter if provided
    if start_date and end_date:
        redeemed_coupon_instances = redeemed_coupon_instances.filter(
            created_date__date__range=(start_date, end_date)
        )

    redeemed_coupon_instances = redeemed_coupon_instances.order_by('-created_date')

    # Calculate totals for manual and digital coupons
    total_manual_coupons = 0
    total_digital_coupons = 0

    for coupon in redeemed_coupon_instances:
        total_coupons = coupon.total_coupon_recieved()  # Assuming this method returns a dict
        total_manual_coupons += total_coupons.get('manual_coupon', 0)
        total_digital_coupons += total_coupons.get('digital_coupon', 0)
        
    log_activity(
        created_by=request.user if request.user.is_authenticated else None,
        description="Viewed the customer transaction list with filters applied."
    )

    context = {
        'sales_report_data': sales_report_data,
        'total_qty':total_qty,
        'total_amount': total_amount,
        'total_discount': total_discount,
        'total_net_payable': total_net_payable,
        'total_vat': total_vat,
        'total_grand_total': total_grand_total,
        'total_amount_recieved': total_amount_recieved,
        'filter_start_date': filter_data.get('filter_start_date', ''),
        'filter_end_date': filter_data.get('filter_end_date', ''),
        'total_outstanding_amount': total_outstanding_amount,
        'total_outstanding_coupons': total_outstanding_coupons,
        'total_outstanding_emptycan': total_outstanding_emptycan,
        'customer_outstanding_instances': customer_outstanding_instances,
        'collection_payment_instance': collection_payment_instance,
        'total_amount_received': total_amount_received,
        'total_discounts': total_discounts,
        'total_net_taxable': total_net_taxable,
        'total_collection_vat': total_collection_vat,
        'total_collected_amount': total_collected_amount,
        'redeemed_coupon_instances': redeemed_coupon_instances,
        'total_manual_coupons': total_manual_coupons,
        'total_digital_coupons': total_digital_coupons,
        'customer_pk': customer_pk,
        'customer_id': customer.custom_id if customer else '',
        'customer_name': customer.customer_name if customer else '',
        'route_name': route_name,
    }

    return render(request, 'client_management/customer_transaction/customer_transaction_list.html', context)


def customer_transaction_print(request):
    """
    Customer Transaction Print
    :param request:
    :return: Customer Transaction Print view
    """
    filter_data = {}
    q = request.GET.get('q', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Initialize totals
    total_amount = total_discount = total_net_payable = total_vat = 0
    total_grand_total = total_amount_recieved = 0
    customer_pk = request.GET.get("customer_pk")
    
    customer = None
    route_name = ""

    if customer_pk:
        try:
            customer = Customers.objects.get(pk=customer_pk)
            route_name = customer.routes.route_name if customer.routes else ""
        except Customers.DoesNotExist:
            pass
        
    # Get CustomerSupply items excluding certain sales types
    sales = CustomerSupplyItems.objects.filter(
        customer_supply__customer__pk=customer_pk
    ).exclude(customer_supply__customer__sales_type__in=["CASH COUPON", "CREDIT COUPON"]).order_by('-customer_supply__created_date')

    # Get CustomerCoupon items
    coupons = CustomerCouponItems.objects.filter(
        customer_coupon__customer__pk=customer_pk
    )

    # Handle date filtering
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            filter_data['filter_start_date'] = start_date.strftime('%Y-%m-%d')
            filter_data['filter_end_date'] = end_date.strftime('%Y-%m-%d')

            # Apply date filtering to sales and coupons
            sales = sales.filter(customer_supply__created_date__date__range=(start_date, end_date))
            coupons = coupons.filter(customer_coupon__created_date__date__range=(start_date, end_date))
        except ValueError:
            start_date = None  # Invalid date format, reset
            end_date = None

    sales_report_data = []

    # Process CustomerSupply data
    for sale in sales:
        sales_report_data.append({
            'date': sale.customer_supply.created_date.date(),
            'ref_invoice_no': sale.customer_supply.reference_number,
            'invoice_number': sale.customer_supply.invoice_no,
            'product_name': sale.product.product_name,
            'sales_type': sale.customer_supply.customer.sales_type,
            'qty': sale.quantity,
            'amount': sale.customer_supply.grand_total,
            'discount': sale.customer_supply.discount,
            'net_taxable': sale.customer_supply.subtotal,
            'vat_amount': sale.customer_supply.vat,
            'grand_total': sale.customer_supply.grand_total,
            'amount_collected': sale.customer_supply.amount_recieved,
        })

        # Update totals
        total_amount += sale.customer_supply.grand_total
        total_discount += sale.customer_supply.discount
        total_net_payable += sale.customer_supply.net_payable
        total_vat += sale.customer_supply.vat
        total_grand_total += sale.customer_supply.grand_total
        total_amount_recieved += sale.customer_supply.amount_recieved

    # Process CustomerCoupon data
    for coupon in coupons:
        vat_rate = Tax.objects.get(name="VAT").percentage if Tax.objects.filter(name="VAT").exists() else 0
        sales_report_data.append({
            'date': coupon.customer_coupon.created_date.date(),
            'ref_invoice_no': coupon.customer_coupon.reference_number,
            'invoice_number': coupon.customer_coupon.invoice_no,
            'product_name': coupon.coupon.book_num,
            'sales_type': coupon.customer_coupon.customer.sales_type,
            'qty': 1,
            'amount': coupon.customer_coupon.grand_total,
            'discount': coupon.customer_coupon.discount,
            'net_taxable': coupon.customer_coupon.net_amount,
            'vat_amount': vat_rate,
            'grand_total': coupon.customer_coupon.grand_total,
            'amount_collected': coupon.customer_coupon.amount_recieved,
        })

        # Update totals
        total_amount += coupon.customer_coupon.grand_total
        total_discount += coupon.customer_coupon.discount
        total_net_payable += coupon.customer_coupon.net_amount
        total_vat += vat_rate
        total_grand_total += coupon.customer_coupon.grand_total
        total_amount_recieved += coupon.customer_coupon.amount_recieved

    # Fetch Customer Outstanding instances
    customer_outstanding_instances = CustomerOutstanding.objects.filter(customer__pk=customer_pk)

    # Apply date filter if provided
    if start_date and end_date:
        customer_outstanding_instances = customer_outstanding_instances.filter(
            created_date__date__range=(start_date, end_date)
        )

    customer_outstanding_instances = customer_outstanding_instances.order_by('-created_date')

    # Prepare totals and other necessary data
    total_outstanding_amount = 0
    total_outstanding_coupons = 0
    total_outstanding_emptycan = 0

    # Loop through each outstanding instance to calculate totals
    for outstanding in customer_outstanding_instances:
        # Calculate amounts for each instance
        outstanding_amount = outstanding.outstandingamount_set.aggregate(total=Sum('amount'))['total'] or 0
        outstanding_coupons = outstanding.outstandingcoupon_set.aggregate(total=Sum('count'))['total'] or 0
        outstanding_emptycan = outstanding.outstandingproduct_set.aggregate(total=Sum('empty_bottle'))['total'] or 0

        # Update totals
        total_outstanding_amount += outstanding_amount
        total_outstanding_coupons += outstanding_coupons
        total_outstanding_emptycan += outstanding_emptycan

        # Add calculated values to outstanding instance for template rendering
        outstanding.total_amount = outstanding_amount
        outstanding.total_coupons = outstanding_coupons
        outstanding.total_emptycan = outstanding_emptycan
        
    # Filter CollectionPayment instances for the customer
    collection_payment_instance = CollectionPayment.objects.filter(
        customer__pk=customer_pk
    )

    # Apply date filter if provided
    if start_date and end_date:
        collection_payment_instance = collection_payment_instance.filter(
            created_date__date__range=(start_date, end_date)
        )

    collection_payment_instance = collection_payment_instance.order_by('-created_date')

    # Calculate total amounts
    total_amount_received = collection_payment_instance.aggregate(Sum('amount_received'))['amount_received__sum'] or 0
    total_discounts = sum(payment.total_discounts() for payment in collection_payment_instance)
    total_net_taxable = sum(payment.total_net_taxeble() for payment in collection_payment_instance)
    total_collection_vat = sum(payment.total_vat() for payment in collection_payment_instance)
    total_collected_amount = sum(payment.collected_amount() for payment in collection_payment_instance)

    # Filter redeemed_coupon_instances for the customer and sales_type="CASH COUPON"
    redeemed_coupon_instances = CustomerSupply.objects.filter(
        customer__pk=customer_pk,
        customer__sales_type="CASH COUPON"
    )

    # Apply date filter if provided
    if start_date and end_date:
        redeemed_coupon_instances = redeemed_coupon_instances.filter(
            created_date__date__range=(start_date, end_date)
        )

    redeemed_coupon_instances = redeemed_coupon_instances.order_by('-created_date')

    # Calculate totals for manual and digital coupons
    total_manual_coupons = 0
    total_digital_coupons = 0

    for coupon in redeemed_coupon_instances:
        total_coupons = coupon.total_coupon_recieved()  # Assuming this method returns a dict
        total_manual_coupons += total_coupons.get('manual_coupon', 0)
        total_digital_coupons += total_coupons.get('digital_coupon', 0)

    context = {
        'sales_report_data': sales_report_data,
        'total_amount': total_amount,
        'total_discount': total_discount,
        'total_net_payable': total_net_payable,
        'total_vat': total_vat,
        'total_grand_total': total_grand_total,
        'total_amount_recieved': total_amount_recieved,
        'filter_start_date': filter_data.get('filter_start_date', ''),
        'filter_end_date': filter_data.get('filter_end_date', ''),
        'total_outstanding_amount': total_outstanding_amount,
        'total_outstanding_coupons': total_outstanding_coupons,
        'total_outstanding_emptycan': total_outstanding_emptycan,
        'customer_outstanding_instances': customer_outstanding_instances,
        'collection_payment_instance': collection_payment_instance,
        'total_amount_received': total_amount_received,
        'total_discounts': total_discounts,
        'total_net_taxable': total_net_taxable,
        'total_collection_vat': total_collection_vat,
        'total_collected_amount': total_collected_amount,
        'redeemed_coupon_instances': redeemed_coupon_instances,
        'total_manual_coupons': total_manual_coupons,
        'total_digital_coupons': total_digital_coupons,
        'customer_pk': customer_pk,
        'customer_id': customer.custom_id if customer else '',
        'customer_name': customer.customer_name if customer else '',
        'route_name': route_name,
    }
    log_activity(
        created_by=request.user if request.user.is_authenticated else None,
        description="Viewed the customer transaction Print with filters applied."
    )
    return render(request, 'client_management/customer_transaction/customer_transaction_print.html', context)


@login_required
def eligible_customers_conditions(request):
    instances = EligibleCustomerConditions.objects.all()
    
    context = {
        'instances': instances,   
        'page_title': 'Eligible Customer Conditions',
    }
    
    return render(request,'client_management/eligible_customers/conditions_list.html',context)

def create_eligible_customers_condition(request):
    
    message = ''
    if request.method == 'POST':
        form = EligibleCustomerConditionsForm(request.POST)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    instance = form.save(commit=False)
                    instance.save()
                    
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Condition created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('eligible_customers_conditions')
                    }
                    
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(form,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        form = EligibleCustomerConditionsForm()
        
        context = {
            'form': form,
        
            'page_title': 'Create Eligible Customer Conditions',
        }
        
        return render(request,'client_management/eligible_customers/condition_create.html',context)


@login_required
def edit_eligible_customers_condition(request, pk):
    """
    Edit operation for eligible customers condition.
    """
    # Fetch the model instance
    instance = get_object_or_404(EligibleCustomerConditions, pk=pk)

    message = ''

    if request.method == 'POST':
        form = EligibleCustomerConditionsForm(request.POST, instance=instance)

        if form.is_valid():
            # Save the updated data
            data = form.save(commit=False)
            data.save()

            response_data = {
                "status": "true",
                "title": "Successfully Updated",
                "message": "Invoice Updated Successfully.",
                "redirect": "true",
                "redirect_url": reverse('eligible_customers_conditions'),
                "return": True,
            }
        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        form = EligibleCustomerConditionsForm(instance=instance)

        context = {
            'form': form,
            'message': message,
            'page_name': 'Edit Eligible Condition',
            'url': reverse('edit_eligible_customers_condition', args=[instance.pk]),
        }

        return render(request, 'client_management/eligible_customers/condition_create.html', context)

    
def delete_eligible_customers_condition(request, pk):
    """
    deletion
    :param request:
    :param pk:
    :return:
    """
    try:
        with transaction.atomic():
            instance = get_object_or_404(EligibleCustomerConditions, pk=pk)
            instance.delete()           
            log_activity(
                created_by=request.user,
                description=f"Eligible customer condition with ID {pk} was successfully deleted."
            )    
            response_data = {
                "status": "true",
                "title": "Successfully Deleted",
                "message": "eligible condition successfully deleted.",
                "redirect": "true",
                "redirect_url": reverse('eligible_customers_conditions'),
            }
            
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    except EligibleCustomerConditions.DoesNotExist:
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "Condition not found.",
        }
        return HttpResponse(json.dumps(response_data), status=status.HTTP_404_NOT_FOUND, content_type='application/javascript')

    except Exception as e:
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }
        return HttpResponse(json.dumps(response_data), status=status.HTTP_500_INTERNAL_SERVER_ERROR, content_type='application/javascript')




from datetime import datetime
from django.utils.timezone import now
import calendar

def eligible_customers(request):
    filter_data = {}
    from_date = request.GET.get('from_date', datetime.today().strftime('%Y-%m-%d'))
    to_date = request.GET.get('to_date', datetime.today().strftime('%Y-%m-%d'))
    month_filter = request.GET.get('month_filter', '')
    route_name = request.GET.get('route_name')
    custody_filter = request.GET.get('custody')
    status_filter = request.GET.get('status')
    query = request.GET.get('q', '')
    
    today = now().date()
    if month_filter and month_filter != 'custom':
        # Calculate first and last day of the selected month
        year = today.year
        first_day = datetime(year, int(month_filter), 1).date()
        last_day = datetime(year, int(month_filter), calendar.monthrange(year, int(month_filter))[1]).date()
        from_date = first_day.strftime('%Y-%m-%d')  
        to_date = last_day.strftime('%Y-%m-%d')

    custody_items = CustomerCustodyStock.objects.exclude(product__product_name="5 Gallon")
    
    if query:

        custody_items = custody_items.filter(
            Q(customer__custom_id__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__mobile_no__icontains=query)
        )
        title = "Eligible Customers - %s" % query
        filter_data['q'] = query
    
    if route_name:
        custody_items = custody_items.filter(customer__routes__route_name=route_name)
    if custody_filter:
        custody_items = custody_items.filter(product__product_name=custody_filter)
    # Generate eligibility status dynamically
    custody_data = []
    for custody_item in custody_items:
        eligibility = custody_item.get_eligibility_status(from_date, to_date)
        custody_data.append({
            "customer_code": custody_item.customer.custom_id,
            "customer_name": custody_item.customer.customer_name,
            "customer_route": custody_item.customer.routes.route_name if custody_item.customer.routes else "No Route",
            # "customer_route": custody_item.customer.routes.route_name ,
            "customer_mob_no": custody_item.customer.mobile_no ,
            "customer_product_name": custody_item.product.product_name if custody_item.product else "No Product",
            # "customer_product_name": custody_item.product.product_name,
            "eligibility_status": eligibility.get("status"),
            "supply_count": eligibility.get("supply_count"),
            "condition_count": eligibility.get("condition_count"),
            "eligible_count": eligibility.get("eligible_count"),
        })
        
    if status_filter:
        custody_data = [item for item in custody_data if item['eligibility_status'] == status_filter]

    routes = RouteMaster.objects.all()
    item = ProdutItemMaster.objects.exclude(product_name="5 Gallon").exclude(category__category_name="Coupons")
    
    context = {
        'instances': custody_data,
        'routes': routes,
        'from_date': from_date,
        'to_date': to_date,
        'route_name': route_name,
        'custody_filter': custody_filter,
        'status_filter': status_filter,
        'item': item,
        'month_filter': month_filter,
        'filter_data': filter_data,
    }

    return render(request, 'client_management/customer_supply/eligible_customers.html', context)


#-----------------------Audit------------------------------------

class MarketingExecutiveRoutesView(View):
    def get(self, request):
        executives = None
        selected_executive = None

        if request.user.user_type == "marketing_executive":
            selected_executive = request.user.id  
        else:
            executives = CustomUser.objects.filter(user_type="marketing_executive")
            selected_executive = request.GET.get("executive")

        routes = None
        if selected_executive:
            assigned_routes = Van_Routes.objects.filter(van__salesman=selected_executive).values_list("routes", flat=True)
            routes = RouteMaster.objects.filter(route_id__in=assigned_routes)

        return render(
            request,
            "client_management/audit/select_executive_routes.html",
            {
                "executives": executives,
                "routes": routes,
                "selected_executive": selected_executive,
                "is_marketing_executive": request.user.user_type == "marketing_executive", 
            },
        )

    def post(self, request):
        selected_executive = request.POST.get("executive") or request.user.id  
        selected_route = request.POST.get("route")

        if selected_executive and selected_route:
            request.session["selected_executive"] = selected_executive
            return redirect("audit_customer_list", route_id=selected_route)

        return redirect("select_executive")
    
    
class CustomerListView(View):
    def get(self, request, route_id):
        route = get_object_or_404(RouteMaster, route_id=route_id)

        customers = Customers.objects.filter(is_guest=False, routes=route)

        query = request.GET.get("q", "")
        if query:
            customers = customers.filter(
                Q(customer_name__icontains=query) | 
                Q(custom_id__icontains=query) |
                Q(mobile_no__icontains=query) |
                Q(building_name__icontains=query)
            )
        audit = AuditBase.objects.filter(route=route, end_date__isnull=True).first()
        audit_in_progress = bool(audit)  

        return render(
            request,
            "client_management/audit/customer_list.html",
            {
                "route": route,
                "customers": customers,
                "audit_in_progress": audit_in_progress,  
                "filter_data": {"q": query},
            },
        )
        
    def post(self, request, route_id):
        route = get_object_or_404(RouteMaster, route_id=route_id)

        selected_executive_id = request.POST.get("executive") or request.session.get("selected_executive")
       

        marketing_executive = None
        if selected_executive_id:
            marketing_executive = CustomUser.objects.filter(id=selected_executive_id, user_type="marketing_executive").first()

        audit = AuditBase.objects.filter(route=route, end_date__isnull=True).first()

        if audit:
            audit.end_date = timezone.now()
            audit.save()
        else:
            audit = AuditBase.objects.create(
                route=route, 
                start_date=timezone.now(),
                marketing_executieve=marketing_executive  
            )

        return redirect("audit_customer_list", route_id=route_id)

class AuditDetailsView(View):
    def get(self, request, customer_id):
        customer = get_object_or_404(Customers, customer_id=customer_id)

        audit_base, created = AuditBase.objects.get_or_create(
            route=customer.routes,  
            end_date__isnull=True,  
            defaults={"start_date": timezone.now()},
        )

        audit_details, created = AuditDetails.objects.get_or_create(
            audit_base=audit_base, 
            customer=customer,
        )

        form = AuditDetailsForm(instance=audit_details)

        return render(
            request,
            "client_management/audit/audit_details.html",
            {
                "customer": customer,
                "form": form,
            },
        )

    def post(self, request, customer_id):
        customer = get_object_or_404(Customers, customer_id=customer_id)

        audit_base = AuditBase.objects.filter(
            route=customer.routes, 
            end_date__isnull=True   
        ).first()

        if not audit_base:
            messages.error(request, "No active audit found for this route. Start an audit first.")
            return redirect("audit_customer_list", route_id=customer.routes.route_id)

        audit_details, created = AuditDetails.objects.get_or_create(
            audit_base=audit_base,  
            customer=customer
        )

        form = AuditDetailsForm(request.POST, instance=audit_details)
        if form.is_valid():
            audit_obj = form.save()
            # # ------- Calculate customer's current outstanding amount ----------
            # today = timezone.now().date()
            # print("today",today)
            # # 1. Get amount outstanding
            # outstanding_amount = OutstandingAmount.objects.filter(
            #     customer_outstanding__customer=customer,
            #     customer_outstanding__created_date__date__lte=today
            # ).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
            # print("outstanding_amount",outstanding_amount)
            # # 2. Get collection amount
            # collection_amount = CollectionPayment.objects.filter(
            #     customer=customer,
            #     created_date__date__lte=today
            # ).aggregate(total_amount_received=Sum('amount_received'))['total_amount_received'] or 0

            # current_outstanding_amount = outstanding_amount - collection_amount
            # print("current_outstanding_amount",current_outstanding_amount)
            # print("audit_obj.outstanding_amount ",audit_obj.outstanding_amount )
            # # ------- Compare & Update ----------
            # if audit_obj.outstanding_amount > current_outstanding_amount:
            #     # You can choose how to update: create new entry or update existing
            #     # Example: create a new OutstandingAmount entry
            #     CustomerOutstanding.objects.create(
            #         customer=customer,
            #         created_date=timezone.now(),
            #         product_type='amount',  # or however you tag this
            #     )
            #     ab=OutstandingAmount.objects.create(
            #         customer_outstanding=customer.customer_outstanding_set.latest('created_date'),  # assuming FK relation
            #         amount=audit_obj.outstanding_amount 
            #     )
            #     print("ab",ab)
            return redirect("audit_customer_list", route_id=customer.routes.route_id)

        return render(
            request,
            "client_management/audit/audit_details.html",
            {
                "customer": customer,
                "form": form,
            },
        )


def customer_custody_stock(request):
    custody_stocks = CustomerCustodyStock.objects.all()
    context = {"custody_stocks": custody_stocks}
    return render(request, "client_management/customer_custody_stock/customer_custody_stock.html", context)

def custody_item_return_view(request, stock_id):
    custody_stock_instance = get_object_or_404(CustomerCustodyStock, pk=stock_id)

    if request.method == "POST":
        form = CustodyItemReturnForm(request.POST)

        if form.is_valid():
            try:
                customer = custody_stock_instance.customer
                agreement_no = custody_stock_instance.agreement_no
                deposit_type = custody_stock_instance.deposit_type
                product = custody_stock_instance.product
                reference_no = custody_stock_instance.reference_no
                print("reference_no:", reference_no)

                quantity = form.cleaned_data['quantity']
                print("quantity", quantity)
                total_amount = form.cleaned_data['amount']
                print("total_amount", total_amount)
                serialnumber = form.cleaned_data['serialnumber']
                print("serialnumber", serialnumber)

                custody_return_instance = CustomerReturn.objects.create(
                    customer=customer,
                    agreement_no=agreement_no,
                    deposit_type=deposit_type,
                    reference_no=reference_no
                )

                item_instance = form.save(commit=False)
                item_instance.customer_return = custody_return_instance
                item_instance.product = product
                item_instance.save()

                try:
                    stock_instance, created = CustomerReturnStock.objects.get_or_create(
                        customer=customer, product=product,
                        defaults={
                            'agreement_no': agreement_no,
                            'deposit_type': deposit_type,
                            'reference_no': reference_no,
                            'quantity': quantity,
                            'serialnumber': serialnumber,
                            'amount': total_amount
                        }
                    )

                    if not created:
                        stock_instance.agreement_no += ', ' + agreement_no
                        stock_instance.serialnumber += ', ' + serialnumber
                        stock_instance.quantity += quantity
                        stock_instance.amount += total_amount
                        stock_instance.save()

                except CustomerReturnStock.DoesNotExist:
                    CustomerReturnStock.objects.create(
                        agreement_no=agreement_no,
                        deposit_type=deposit_type,
                        reference_no=reference_no,
                        product=product,
                        quantity=quantity,
                        serialnumber=serialnumber,
                        amount=total_amount
                    )

                custody_stock_instance.amount_collected -= total_amount
                custody_stock_instance.quantity -= quantity
                custody_stock_instance.save()

                vanstock = VanProductStock.objects.get(
                    created_date=datetime.today().date(),
                    product=product,
                    van__salesman=request.user
                )
                vanstock.return_count += quantity
                vanstock.save()

                if item_instance.product.product_name == "5 Gallon":
                    bottle_count_queryset = BottleCount.objects.filter(
                        van__salesman=request.user,
                        created_date__date=custody_return_instance.created_date.date()
                    )
                    if bottle_count_queryset.exists():
                        bottle_count = bottle_count_queryset.first()
                        bottle_count.custody_issue += item_instance.quantity
                        bottle_count.save()

                # Fix: Redirect instead of returning JSON
                return redirect("customer_custody_stock")

            except Exception as e:
                return HttpResponse(json.dumps({
                    "status": "false",
                    "title": "Error",
                    "message": str(e)
                }), content_type="application/javascript")

        else:
            messages.error(request, "Invalid form data. Please check the input.")

    else:
        form = CustodyItemReturnForm()

    context = {
        "form": form,
        "custody_stock_instance": custody_stock_instance
    }
    return render(request, "client_management/customer_custody_stock/custody_item_return.html", context)


def customer_outstanding_report(request):
    filter_data = {}
    outstanding_customer_ids = []
    total_opening_amount = 0
    total_collection_upto = 0
    
    route = request.GET.get('route')
    closing_date = request.GET.get('closing_date')
    to_date = request.GET.get('to_date')
    
    filter_closing_date = filter_to_date = filter_opening_date = datetime.today().date()
    
    if closing_date :
        filter_closing_date = datetime.strptime(closing_date, '%Y-%m-%d').date()
        filter_opening_date = datetime.strptime(closing_date, '%Y-%m-%d').date() + timedelta(days=1)
        
    if to_date :
        filter_to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        
    filter_data['filter_route'] = route
    filter_data['filter_closing_date'] = filter_closing_date.strftime('%Y-%m-%d')
    filter_data['filter_to_date'] = filter_to_date.strftime('%Y-%m-%d')
    filter_data['filter_opening_date'] = filter_opening_date.strftime('%Y-%m-%d')
    filter_data['filter_opening_date_display'] = filter_opening_date.strftime('%d-%m-%Y')
    filter_data['filter_to_date_display'] = filter_to_date.strftime('%d-%m-%Y')

    
    if route:
        filter_data['filter_route_name'] = RouteMaster.objects.get(pk=route).route_name
        outstanding_customer_ids = CustomerOutstanding.objects.filter(customer__routes__pk=route,created_date__date__lt=closing_date,product_type='amount').values_list("customer__pk", flat=True).distinct()
        
        total_opening_amount = OutstandingAmount.objects.filter(customer_outstanding__customer__pk__in=outstanding_customer_ids, customer_outstanding__created_date__date__lte=closing_date).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        total_opening_collection = CollectionPayment.objects.filter(customer__pk__in=outstanding_customer_ids, created_date__date__lte=closing_date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0
        total_opening_amount = total_opening_amount - total_opening_collection
        
        total_collection_upto = CollectionPayment.objects.filter(customer__pk__in=outstanding_customer_ids, created_date__date__gt=closing_date, created_date__date__lte=to_date).aggregate(total_amount=Sum('amount_received'))['total_amount'] or 0

    customer_instances = Customers.objects.filter(is_guest=False, pk__in=outstanding_customer_ids)
    
    context = {
        "filter_data": filter_data,
        "customer_instances": customer_instances,
        "routes": RouteMaster.objects.all(),
        "total_opening_amount": total_opening_amount,
        "total_collection_upto": total_collection_upto,
        "total_balance_amount": total_opening_amount - total_collection_upto,
        
        "page_title": "Sana - Customer Outstanding Report"
        }
    
    if request.GET.get('print'):
        return render(request, "client_management/customer_outstanding/customer_outstanding_report_print.html", context)
    else:
        return render(request, "client_management/customer_outstanding/customer_outstanding_report.html", context)
    
    
def salesman_outstanding_list(request):

    filter_data = {}
    date = request.GET.get("date")

    if date:
        date = datetime.strptime(date, "%Y-%m-%d").date()
    else:
        date = datetime.today().date()

    filter_data["filter_date"] = date.strftime("%Y-%m-%d")

    salesmen = CustomUser.objects.filter(
        user_type__in=["Salesman", "Driver cum Salesman"],
        is_active=True
    )

    results = []
    grand_total = Decimal("0.00")

    for salesman in salesmen:

        invoices = Invoice.objects.filter(
            salesman=salesman,
            created_date__date__lte=date,
            is_deleted=False,
            invoice_status="non_paid"
        )

        total_amount = invoices.aggregate(
            total=Sum("amout_total")
        )["total"] or Decimal("0.00")

        total_received = invoices.aggregate(
            received=Sum("amout_recieved")
        )["received"] or Decimal("0.00")

        outstanding = total_amount - total_received

        if outstanding > 0:
            results.append({
                "salesman": salesman,
                "outstanding": outstanding
            })
            grand_total += outstanding

    context = {
        "instances": results,
        "date": date,
        "grand_total": grand_total,
        "page_title": "Salesman Outstanding Report",
    }

    return render(
        request,
        "client_management/customer_outstanding/route_outstanding_list.html",
        context
    )

def salesman_invoice_list(request):
    salesman_id = request.GET.get("salesman")
    date = request.GET.get("date")

    if not salesman_id:
        return render(request, "404.html")

    salesman = get_object_or_404(CustomUser, id=salesman_id)

    if date:
        date = datetime.strptime(date, "%Y-%m-%d").date()
    else:
        date = datetime.today().date()

    invoices = Invoice.objects.filter(
        salesman=salesman,
        created_date__date__lte=date,
        is_deleted=False,
        invoice_status="non_paid"
    ).order_by("-created_date")
    invoice_data = []

    for invoice in invoices:
        outstanding = (
            invoice.amout_total or Decimal("0.00")
        ) - (
            invoice.amout_recieved or Decimal("0.00")
        )

        invoice_data.append({
            "invoice": invoice,
            "outstanding": outstanding
        })
    total_amount = invoices.aggregate(
        total=Sum("amout_total")
    )["total"] or Decimal("0.00")

    total_received = invoices.aggregate(
        received=Sum("amout_recieved")
    )["received"] or Decimal("0.00")

    outstanding = total_amount - total_received

    context = {
        "salesman": salesman,
        "invoices": invoice_data,
        "date": date,
        "total_amount": total_amount,
        "total_received": total_received,
        "outstanding": outstanding,
        "page_title": "Salesman Invoice List",
    }

    return render(
        request,
        "client_management/customer_outstanding/salesman_invoice_list.html",
        context
    )
    
class CustomerCreditAPI(APIView):
    authentication_classes = [BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        route_id = request.GET.get("route_id")
        customer_id = request.query_params.get("customer_id")

        # 1. Handle Route ID logic
        if not route_id:
            van_route = Van_Routes.objects.filter(van__salesman=request.user).first()
            if van_route:
                route_id = van_route.routes.pk
            else:
                return Response({"status": False, "message": "No route found for this user"}, status=400)

        # 2. Get IDs of customers with credit
        credit_customer_ids = (
            CustomerCredit.objects
            .values("customer")
            .annotate(total=Sum("amount"))
            .filter(total__gt=0)
            .values_list("customer", flat=True)
        )

        if not credit_customer_ids:
            return Response({"status": True, "data": []}, status=200)

        # 3. Define the base customers queryset FIRST
        customers = Customers.objects.filter(
            routes__pk=route_id,
            is_deleted=False,
            pk__in=credit_customer_ids # Use pk__in to match the credit list
        )

        # 4. Now safely apply the customer_id filter if it exists
        if customer_id:
            customers = customers.filter(pk=customer_id)

        serializer = CustomerCreditSerializer(customers, many=True)

        return Response({
            "status": True,
            "data": serializer.data
        }, status=200)


