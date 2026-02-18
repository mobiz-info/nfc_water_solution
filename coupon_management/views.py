from email.utils import parsedate
import re
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from coupon_management.serializers import couponStockSerializers
from lxml.etree import HTML
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate

from client_management.models import *
from competitor_analysis.forms import CompetitorAnalysisFilterForm
from master.functions import generate_form_errors
from product.models import Staff_Orders_details
from van_management.models import Van_Routes
from . models import *
from .forms import  *
from accounts.models import CustomUser, Customers
from invoice_management.models import Invoice
from sales_management.models import *
from master.models import EmirateMaster, BranchMaster, RouteMaster
import json
from django.core.serializers import serialize
from django.views import View
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from master.functions import log_activity
from django.db.models.functions import Coalesce
from django.utils.dateparse import parse_date
from django.db.models import DecimalField, Value

# Create your views here.
# def increment_alphabetic_part(alphabetic_part):
#     """Increment the alphabetic part (handles multi-character strings)."""
#     if not alphabetic_part:
#         return "A"
    
#     alphabetic_list = list(alphabetic_part)
#     i = len(alphabetic_list) - 1

#     # Traverse from the last character backward
#     while i >= 0:
#         if alphabetic_list[i] == "Z":
#             alphabetic_list[i] = "A"  # Reset to 'A'
#             i -= 1
#         else:
#             alphabetic_list[i] = chr(ord(alphabetic_list[i]) + 1)  # Increment
#             break
#     else:
#         # If all characters are 'Z', add a new 'A' at the start
#         alphabetic_list.insert(0, "A")

#     return "".join(alphabetic_list)

import re
from django.http import JsonResponse
from .models import CouponType, NewCoupon, CouponLeaflet, FreeLeaflet


def get_next_coupon_bookno(request):
    coupon_type = request.GET.get("coupon_type")
    next_coupon_bookno = ""
    next_leaf_no = ""
    end_leaf_no = ""
    next_free_leaf_no = ""
    end_free_leaf_no = ""

    # Fetch free leaflets for the coupon type
    coupon_type_freeleaf_count = CouponType.objects.get(pk=coupon_type).free_leaflets

    # Retrieve the last coupon of the selected type
    last_coupon = NewCoupon.objects.filter(coupon_type__pk=coupon_type)
    if last_coupon.exists():
        last_coupon = last_coupon.latest("created_date")
        last_coupon_bookno = last_coupon.book_num

        # Match and separate alphabetic and numeric parts of the book number
        match = re.match(r"([a-zA-Z]*)(\d+)", last_coupon_bookno)
        if match:
            alphabetic_part, numeric_part = match.groups()
            next_numeric_part = str(int(numeric_part) + 1).zfill(len(numeric_part))
            next_coupon_bookno = f"{alphabetic_part}{next_numeric_part}"
        else:
            # Fallback for purely numeric book numbers
            next_coupon_bookno = str(int(last_coupon_bookno) + 1)

        # Handle leaflet numbers
        if (leaflet := CouponLeaflet.objects.filter(coupon=last_coupon)).exists():
            last_leaf_number = leaflet.latest("created_date").leaflet_name

            if last_leaf_number:
                match = re.match(r"([a-zA-Z]*)(\d+)", last_leaf_number)
                if match:
                    leaf_alphabetic_part, leaf_name_part = match.groups()
                    next_leaf_number = str(int(leaf_name_part) + 1).zfill(len(leaf_name_part))
                    end_leaf_number = str(int(next_leaf_number) + int(last_coupon.valuable_leaflets) - 1).zfill(len(leaf_name_part))
                    next_leaf_no = f"{leaf_alphabetic_part}{next_leaf_number}"
                    end_leaf_no = f"{leaf_alphabetic_part}{end_leaf_number}"
                else:
                    try:
                        next_leaf_number = str(int(last_leaf_number) + 1).zfill(len(last_leaf_number))
                        end_leaf_number = str(int(next_leaf_number) + int(last_coupon.valuable_leaflets) - 1).zfill(len(last_leaf_number))
                        next_leaf_no = next_leaf_number
                        end_leaf_no = end_leaf_number
                    except ValueError:
                        next_leaf_no = "1"
                        end_leaf_no = str(int(next_leaf_no) + int(last_coupon.valuable_leaflets) - 1)
            else:
                next_leaf_no = "1"
                end_leaf_no = str(int(next_leaf_no) + int(last_coupon.valuable_leaflets) - 1)

        # Handle free leaflet numbers
        if (free_leaflet := FreeLeaflet.objects.filter(coupon=last_coupon)).exists():
            last_free_leaf_number = free_leaflet.latest("created_date").leaflet_name

            if last_free_leaf_number:
                match = re.match(r"([a-zA-Z]*)(\d+)", last_free_leaf_number)
                if match:
                    free_leaf_alphabetic_part, free_leaf_name_part = match.groups()
                    next_free_leaf_number = str(int(free_leaf_name_part) + 1).zfill(len(free_leaf_name_part))
                    end_free_leaf_number = str(int(next_free_leaf_number) + int(coupon_type_freeleaf_count) - 1).zfill(len(free_leaf_name_part))
                    next_free_leaf_no = f"{free_leaf_alphabetic_part}{next_free_leaf_number}"
                    end_free_leaf_no = f"{free_leaf_alphabetic_part}{end_free_leaf_number}"
                else:
                    try:
                        next_free_leaf_number = str(int(last_free_leaf_number) + 1).zfill(len(last_free_leaf_number))
                        end_free_leaf_number = str(int(next_free_leaf_number) + int(coupon_type_freeleaf_count) - 1).zfill(len(last_free_leaf_number))
                        next_free_leaf_no = next_free_leaf_number
                        end_free_leaf_no = end_free_leaf_number
                    except ValueError:
                        next_free_leaf_no = "1"
                        end_free_leaf_no = str(int(next_free_leaf_no) + int(coupon_type_freeleaf_count) - 1)
            else:
                next_free_leaf_no = "1"
                end_free_leaf_no = str(int(next_free_leaf_no) + int(coupon_type_freeleaf_count) - 1)

    data = {
        'next_coupon_bookno': next_coupon_bookno,
        "next_leaf_no": next_leaf_no,
        "end_leaf_no": end_leaf_no,
        "next_free_leaf_no": next_free_leaf_no,
        "end_free_leaf_no": end_free_leaf_no,
        "coupon_type_freeleaf_count": coupon_type_freeleaf_count
    }
    return JsonResponse(data, safe=False)

def get_leaf_used_status_change(request):
    leaf_id = request.GET.get("leaf_id")
    customer_id = request.GET.get("customer_id")
    
    if (valueable_coupon:=CouponLeaflet.objects.filter(pk=leaf_id)).exists():
        stock = CustomerCouponStock.objects.get(customer__pk=customer_id,coupon_type_id=valueable_coupon.first().coupon.coupon_type)
        
        if not valueable_coupon.first().used :
            valueable_coupon.update(used=True)
            stock.count -= 1
            
            status_code = 200
            response_data = {
                "status": "true",
                "message": "leaf mark as used",
            }
        else:
            valueable_coupon.update(used=False)
            stock.count += 1
            
            status_code = 200
            response_data = {
                "status": "false",
                "message": "leaf mark as not used",
            }
        
        stock.save()
        
    elif (free_coupon:=FreeLeaflet.objects.filter(pk=leaf_id)).exists():
        stock = CustomerCouponStock.objects.get(customer__pk=customer_id,coupon_type_id=free_coupon.first().coupon.coupon_type)
        
        if not free_coupon.first().used :
            free_coupon.update(used=True)
            stock.count -= 1
            
            status_code = 200
            response_data = {
                "status": "true",
                "message": "leaf mark as used",
            }
        else:
            free_coupon.update(used=False)
            stock.count += 1
            
            status_code = 200
            response_data = {
                "status": "false",
                "message": "leaf mark as not used",
            }
        
        stock.save()
        
    else:
        status_code = 404
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "item not found",
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")

def get_coupon_bookno(request):
    request_id = request.GET.get("request_id")
    
    if (instances:=Staff_Orders_details.objects.filter(pk=request_id)).exists():
        instance = instances.first()
        stock_instances = CouponStock.objects.filter(couponbook__coupon_type__coupon_type_name=instance.product_id.product_name,coupon_stock="company")
        serialized = couponStockSerializers(stock_instances, many=True)
        
        status_code = 200
        response_data = {
            "status": "true",
            "data": serialized.data,
        }
    else:
        status_code = 404
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "item not found",
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")

def couponType(request):
    all_couponType = CouponType.objects.all()
    context = {'all_couponType': all_couponType}
    return render(request, 'coupon_management/index_couponType.html', context)

def create_couponType(request):
    if request.method == 'POST':
        form = CreateCouponTypeForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = str(request.user.id)
            data.save()
            messages.success(request, 'Coupon Type created successfully!')
            return redirect('couponType')
        else:
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = CreateCouponTypeForm()
    context = {'form': form}
    return render(request, 'coupon_management/create_couponType.html', context)


def view_couponType(request, coupon_type_id):
    view_couponType = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)
    return render(request, 'coupon_management/view_couponType.html', {'view_couponType': view_couponType})

def edit_CouponType(request, coupon_type_id):
    edit_coupon = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)
    if request.method == 'POST':
        form = EditCouponTypeForm(request.POST, instance=edit_coupon)
        if form.is_valid():
            data = form.save(commit=False)
            # print(data,"data")
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            return redirect('couponType')
    else:
        form = EditCouponTypeForm(instance=edit_coupon)
    return render(request, 'coupon_management/edit_couponType.html', {'form': form, 'edit_coupon': edit_coupon})

def delete_couponType(request, coupon_type_id):
    deleteCouponType = CouponType.objects.get(coupon_type_id=coupon_type_id)
    if request.method == 'POST':
        deleteCouponType.delete()
        log_activity(request.user.id, f"Coupon Type deleted: {deleteCouponType.coupon_type_name}")
        return redirect('couponType')
    return render(request, 'coupon_management/delete_couponType.html', {'deleteCouponType': deleteCouponType})

#------------------------New Coupon
def new_coupon(request):
    filter_data = {}
    
    query = request.GET.get("q")
    status_type = "company"
    
    if request.GET.get('status_type'):
        status_type = request.GET.get('status_type')
    
    filter_data['status_type'] = status_type
    
    coupon_ids = CouponStock.objects.filter(coupon_stock=status_type).values_list("couponbook__pk")
    instances = NewCoupon.objects.filter(pk__in=coupon_ids).order_by("-created_date")
         
    if query:

        instances = instances.filter(
            Q(book_num__icontains=query) |
            Q(coupon_type__coupon_type_name__icontains=query)
        )
        title = "Coupon List - %s" % query
        filter_data['q'] = query
    
    context = {
        'instances': instances,
        'filter_data': filter_data
        }
    
    return render(request, 'coupon_management/index_Newcoupon.html', context)

def create_Newcoupon(request):
    if request.method == 'POST':
        form = CreateNewCouponForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            
            coupon_type_id = request.POST.get('coupon_type')
            book_num = request.POST.get('book_num')
            valuable_leafs = request.POST.get('valuable_leafs')
            free_leafs = request.POST.get('free_leafs')
            
            selected_coupon_type = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)

            data.coupon_type = selected_coupon_type
            data.book_num = book_num
            data.no_of_leaflets = selected_coupon_type.no_of_leaflets
            data.valuable_leaflets = selected_coupon_type.valuable_leaflets
            data.free_leaflets = selected_coupon_type.free_leaflets
            data.created_by = str(request.user.id)
            
            branch_id = request.user.branch_id.branch_id
            branch = BranchMaster.objects.get(branch_id=branch_id)
            data.branch_id = branch           
            data.save()
            log_activity(request.user, f"New coupon created: {data.coupon_id}")
            
            for v in valuable_leafs.split(', '):
                CouponLeaflet.objects.create(
                    coupon=data,
                    leaflet_number=data.valuable_leaflets,
                    leaflet_name=v,
                    created_by=request.user.id,
                    created_date=datetime.now(),
                )
            log_activity(request.user, f"Valuable leaflets created for coupon: {data.coupon_id}")
            if int(data.free_leaflets) > 0:
                for f in free_leafs.split(', '):
                    FreeLeaflet.objects.create(
                        coupon=data,
                        leaflet_number=data.free_leaflets,
                        leaflet_name=f,
                        created_by=request.user.id,
                        created_date=datetime.now(),
                    )
                    
                log_activity(request.user.id, f"Free leaflets created for coupon: {data.coupon_id}")
            # Create CouponStock instance
            CouponStock.objects.create(
                couponbook=data, 
                coupon_stock='company', 
                created_by=str(request.user.id)
                )
            log_activity(request.user.id, f"Coupon stock created for coupon: {data.coupon_id}")
            product_instance=ProdutItemMaster.objects.get(product_name=data.coupon_type.coupon_type_name)
            if (stock_intances:=ProductStock.objects.filter(product_name=product_instance,branch=branch)).exists():
                stock_intance = stock_intances.first()
                stock_intance.quantity += 1
                stock_intance.save()
            else:
                ProductStock.objects.create(
                    product_name=product_instance,
                    branch=branch,
                    quantity=1
                )
            log_activity(request.user.id, f"Product stock updated for coupon: {data.coupon_id}")
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Coupon Generation successfully.",
                'redirect': 'true',
                "redirect_url": reverse('new_coupon')
            }
            return JsonResponse(response_data, status=200)
        else:
            message = generate_form_errors(form, formset=False)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }
            return JsonResponse(response_data, status=200)
    else:
        form = CreateNewCouponForm()
    
    context = {'form': form}
    return render(request, 'coupon_management/create_Newcoupon.html', context)

def generate_leaflets(request, coupon_id):
    coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    leaflets = []
    no_of_leaflets = int(coupon.coupon_type.no_of_leaflets)
    for leaflet_num in range(1, no_of_leaflets + 1):
        leaflet = CouponLeaflet(coupon=coupon, leaflet_number=str(leaflet_num))
        leaflets.append(leaflet)
        leaflet.save()
        

    context = {'coupon': coupon, 'leaflets': leaflets}
    return render(request, 'coupon_management/create_Newcoupon.html', context)

def get_leaflet_serial_numbers(request):
    if request.method == 'GET':
        coupon_type_id = request.GET.get('coupon_type')

        # Fetch leaflets based on the provided coupon type
        try:
            leaflets = CouponLeaflet.objects.filter(coupon__coupon_type_id=coupon_type_id)
            leaflet_data = [{'leaflet_number': leaflet.leaflet_number, 'is_used': leaflet.used} for leaflet in leaflets]
            return JsonResponse(leaflet_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    

def save_coupon_data(request):
    if request.method == 'POST':
        form = CreateNewCouponForm(request.POST)
        if form.is_valid():
            try:
                # Save the coupon data to the database
                new_coupon = form.save()
                # You can also save leaflet data here if necessary
                return JsonResponse({'message': 'Coupon data saved successfully'})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
        else:
            return JsonResponse({'error': 'Invalid form data'}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

def view_Newcoupon(request, coupon_id):
    view_coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    return render(request, 'coupon_management/view_Newcoupon.html', {'view_coupon': view_coupon})

def edit_NewCoupon(request, coupon_id):
    edit_coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    if request.method == 'POST':
        form = EditNewCouponForm(request.POST, instance=edit_coupon)
        if form.is_valid():
            data = form.save(commit=False)
            # print(data,"data")
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            return redirect('new_coupon')
    else:
        form = EditNewCouponForm(instance=edit_coupon)
    return render(request, 'coupon_management/edit_Newcoupon.html', {'form': form, 'edit_coupon': edit_coupon})

def delete_Newcoupon(request, coupon_id):
    deleteCoupon = NewCoupon.objects.get(coupon_id=coupon_id)
    if request.method == 'POST':
        deleteCoupon.delete()
        log_activity(
            created_by=request.user,
            description=f"deleting coupon with ID {coupon_id}"
        )
        return redirect('new_coupon')
    return render(request, 'coupon_management/delete_Newcoupon.html', {'deleteCoupon': deleteCoupon})


# def customer_stock(request):
#
#     couponstock=CustomerCouponStock.objects.select_related('customer').all()
#     context = {
#
#         'couponstock':couponstock
#     }
#
#     return render(request, 'coupon_management/customer_stock.html', context)


# def customer_stock(request):
#     coupenstock = CustomerCouponStock.objects.select_related('customer').all()
#     return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock})



#
# @login_required
# def customer_stock(request):
#     # Get the user's user_type
#     user_type = request.user.user_type
#
#     # Filter customers based on user_type
#     if user_type == 'Salesman':
#         coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
#         route_li = RouteMaster.objects.filter(branch_id=request.user.branch_id)
#     else:
#         coupenstock = CustomerCouponStock.objects.select_related('customer').all()
#         route_li = RouteMaster.objects.all()
#
#     return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock, 'route_li': route_li})



@login_required
def customer_stock(request):
    # Get the user's user_type
    user_type = request.user.user_type

    # Get all routes for the dropdown
    route_li = RouteMaster.objects.all()

    # Filter customers based on user_type and selected route
    selected_route = request.GET.get('route_name')
    if user_type == 'Salesman':
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user, customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
    else:
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').all()


    created_date = request.GET.get('created_date')
    if created_date:
        coupenstock = coupenstock.filter(
            coupon_type_id__created_date=created_date
        )

    return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock, 'route_li': route_li})

@login_required
def customer_stock_coupon_details(request,customer):
    
    customer_instance = Customers.objects.get(pk=customer)
    
    customer_manual_coupons = CustomerCouponItems.objects.filter(customer_coupon__customer=customer_instance)
    
    context = {
        'customer_instance': customer_instance,
        'customer_manual_coupons': customer_manual_coupons,
    }

    return render(request, 'coupon_management/available_coupon_details.html',context
                  )



from openpyxl.styles import Font, Alignment

@login_required
def generate_excel(request):
    # Get the user's user_type
    user_type = request.user.user_type

    # Get all routes for the dropdown
    route_li = RouteMaster.objects.all()

    # Filter customers based on user_type and selected route
    selected_route = request.GET.get('route_name')
    if user_type == 'Salesman':
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user, customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
    else:
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customer_stock.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customer Stock'

    # Formatting styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='left', vertical='center')

    # Write headers with formatting
    headers = [
        'Sl.No', 'Customer Name / Mobile No', 'Building Name / House No',
        'Digital Coupon Count', 'Manual Coupon Count', 'Total Count'
    ]
    for col_num, header_title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data with formatting
    for row_num, stock in enumerate(coupenstock, start=2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1).alignment = data_alignment
        worksheet.cell(row=row_num, column=2, value=stock.customer.customer_name + ', ' + stock.customer.mobile_no).alignment = data_alignment
        worksheet.cell(row=row_num, column=3, value=stock.customer.building_name + ', ' + stock.customer.door_house_no).alignment = data_alignment
        if stock.coupon_method == 'digital':
            worksheet.cell(row=row_num, column=4, value=stock.count).alignment = data_alignment
        else:
            worksheet.cell(row=row_num, column=5, value=stock.count).alignment = data_alignment

    # Calculate and write total counts
    total_digital_count = sum(stock.count for stock in coupenstock if stock.coupon_method == 'digital')
    total_manual_count = sum(stock.count for stock in coupenstock if stock.coupon_method == 'manual')
    total_row_num = len(coupenstock) + 2
    worksheet.cell(row=total_row_num, column=4, value=total_digital_count).alignment = data_alignment
    worksheet.cell(row=total_row_num, column=5, value=total_manual_count).alignment = data_alignment
    worksheet.cell(row=total_row_num, column=6, value=total_digital_count + total_manual_count).alignment = data_alignment

    # Autofit column width
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response

from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


def customer_stock_pdf(request):
    # Retrieve customer stock data
    coupenstock = CustomerCouponStock.objects.select_related('customer').all()

    # Check if coupenstock is not empty
    if coupenstock:
        # Create the HTTP response with PDF content type and attachment filename
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="customer_stock.pdf"'

        # Create a PDF document
        pdf_buffer = BytesIO()
        pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        data = []

        # Add headers
        headers = ['Sl.No', 'Customer Name / Mobile No', 'Building Name / House No', 'Digital', 'Manual', 'Total Count']
        data.append(headers)

        # Add data to the PDF document
        for index, stock in enumerate(coupenstock):
            customer_name_mobile = f"{stock.customer.customer_name}, {stock.customer.mobile_no}"
            building_house = f"{stock.customer.building_name}, {stock.customer.door_house_no}"
            digital = stock.count if stock.coupon_method == 'digital' else ''
            manual = stock.count if stock.coupon_method == 'manual' else ''
            total_count = stock.count
            data.append([index + 1, customer_name_mobile, building_house, digital, manual, total_count])

        table = Table(data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])
        table.setStyle(style)

        # Add the table to the PDF document
        elements = [table]
        pdf.build(elements)

        # Get the value of the BytesIO buffer and write it to the response
        pdf_value = pdf_buffer.getvalue()
        pdf_buffer.close()
        response.write(pdf_value)

        return response
    else:
        # Return an empty HTTP response with a message indicating no data available
        return HttpResponse('No customer stock data available.')
    

def redeemed_history(request):
    filter_data = {}
    query = request.GET.get("q", "").strip()  # Get the search query

    start_date = date.today()
    end_date = date.today()
    route_name =request.GET.get('route_name')
    
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    filter_data["start_date"] = start_date.strftime('%Y-%m-%d')
    filter_data["end_date"] = end_date.strftime('%Y-%m-%d')
    
    instances = CustomerSupply.objects.filter(created_date__date__gte=start_date,created_date__date__lte=end_date,customer__sales_type="CASH COUPON").order_by("-created_date")
    
    if route_name:
        instances = instances.filter(customer__routes__route_name=route_name)
        filter_data['route_name'] = route_name
    # Apply search filter to instances if a query exists
    if query:
        instances = instances.filter(
            Q(customer__custom_id__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__location__location_name__icontains=query) |
            Q(customer__building_name__icontains=query)
        )
        filter_data['q'] = query
        
    # Calculate totals for manual and digital coupons manually
    total_manual_coupons = 0
    total_digital_coupons = 0
    for instance in instances:
        total_coupons = instance.total_coupon_recieved()  # Using the method from the model
        total_manual_coupons += total_coupons.get('manual_coupon', 0)
        total_digital_coupons += total_coupons.get('digital_coupon', 0)
        
    # Get all route names for the dropdown
    route_li = RouteMaster.objects.all()
    context = {
        'instances': instances,
        "filter_data": filter_data,
        "route_li": route_li,
        "total_manual_coupons": total_manual_coupons,
        "total_digital_coupons": total_digital_coupons,
    }

    return render(request, 'coupon_management/redeemed_history.html', context)

def redeemed_coupon_details(request,supply_pk):
    supply_instance = CustomerSupply.objects.get(pk=supply_pk)
    coupon_instances = CustomerSupplyCoupon.objects.filter(customer_supply__pk=supply_pk)

    context = {
        'supply_instance': supply_instance,
        'coupon_instances': coupon_instances,
    }

    return render(request, 'coupon_management/redeemed_coupon_datails.html', context)

def print_redeemed_history(request):
    filter_data = {}
    query = request.GET.get("q", "").strip()  # Get the search query

    start_date = date.today()
    end_date = date.today()
    route_name =request.GET.get('route_name')
    
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    filter_data["start_date"] = start_date.strftime('%Y-%m-%d')
    filter_data["end_date"] = end_date.strftime('%Y-%m-%d')
    
    instances = CustomerSupply.objects.filter(created_date__date__gte=start_date,created_date__date__lte=end_date,customer__sales_type="CASH COUPON").order_by("-created_date")
    
    if route_name:
        instances = instances.filter(customer__routes__route_name=route_name)
        filter_data['route_name'] = route_name
    # Apply search filter to instances if a query exists
    if query:
        instances = instances.filter(
            Q(customer__custom_id__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__location__location_name__icontains=query) |
            Q(customer__building_name__icontains=query)
        )
        filter_data['q'] = query
        
    # Calculate totals for manual and digital coupons manually
    total_manual_coupons = 0
    total_digital_coupons = 0
    for instance in instances:
        total_coupons = instance.total_coupon_recieved()  # Using the method from the model
        total_manual_coupons += total_coupons.get('manual_coupon', 0)
        total_digital_coupons += total_coupons.get('digital_coupon', 0)
        
    # Get all route names for the dropdown
    route_li = RouteMaster.objects.all()
    context = {
        'instances': instances,
        "filter_data": filter_data,
        "route_li": route_li,
        "total_manual_coupons": total_manual_coupons,
        "total_digital_coupons": total_digital_coupons,
    }

    return render(request, 'coupon_management/print_redeemed_history.html', context)

def coupon_recharge_list(request):
    filter_data = {}
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")
    payment_type = request.GET.get('payment_type', '')
    coupon_method = request.GET.get('coupon_method', '')

    coupon_customer = CustomerCoupon.objects.all().order_by('-created_date')
    
    if start_date and end_date:
        filter_data['start_date'] = start_date
        filter_data['end_date'] = start_date
        # If both dates are provided, filter between them
        coupon_customer = coupon_customer.filter(created_date__range=(start_date, end_date))
    if query:
        coupon_customer = coupon_customer.filter(
            Q(customer__custom_id__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__location__location_name__icontains=query) |
            Q(customer__building_name__icontains=query) |
            Q(customercouponitems__coupon__book_num__icontains=query)
        )
        filter_data['q'] = query
    # Apply payment_type and coupon_method filters
    if payment_type:
        coupon_customer = coupon_customer.filter(payment_type=payment_type)
        filter_data['payment_type'] = payment_type
    if coupon_method:
        coupon_customer = coupon_customer.filter(coupon_method=coupon_method)
        filter_data['coupon_method'] = coupon_method

    log_activity(
            created_by=request.user,
            description=f"Viewed coupon recharge list."
        )
    context={
        'coupon_customer':coupon_customer,
        'filter_data': filter_data,
    }

    return render(request,'coupon_management/coupon_recharge_list.html', context)


def edit_coupon_recharge(request, pk):
    coupon_recharge = get_object_or_404(CustomerCoupon, id=pk)
    invoice = get_object_or_404(Invoice, invoice_no=coupon_recharge.invoice_no)
    receipt = get_object_or_404(Receipt, invoice_number=coupon_recharge.invoice_no)
    
    if request.method == "POST":
        form = CustomerCouponForm(request.POST, instance=coupon_recharge)
        if form.is_valid():
            updated_coupon = form.save()

            invoice.net_taxable = updated_coupon.grand_total - updated_coupon.discount
            invoice.amout_total = updated_coupon.grand_total
            invoice.amout_recieved = updated_coupon.amount_recieved
            invoice.save()
            
            receipt.amount_received = updated_coupon.amount_recieved
            receipt.save()
            
            log_activity(
                        created_by=request.user,
                        description=f"Edited coupon recharge with ID {pk}, updated related invoice and receipt."
                    )

            messages.success(request, "Coupon recharge and related invoice updated successfully.")
            return redirect("coupon_recharge")
    else:
        form = CustomerCouponForm(instance=coupon_recharge)

    context = {
        "form": form,
        "coupon_recharge": coupon_recharge,
    }
    return render(request, "coupon_management/edit_coupon_recharge.html", context)

def delete_coupon_recharge(request, pk):
    """
    Rollback and delete a coupon recharge with all linked transactions.
    """
    try:
        with transaction.atomic():
            coupon_recharge = get_object_or_404(CustomerCoupon, id=pk)

            # Rollback Outstanding
            try:
                if coupon_recharge.balance > 0:
                    outstanding_report = CustomerOutstandingReport.objects.get(
                        customer=coupon_recharge.customer,
                        product_type="amount"
                    )
                    outstanding_report.value -= Decimal(coupon_recharge.balance)
                    outstanding_report.save()

                    # delete linked outstanding and amounts
                    CustomerOutstanding.objects.filter(
                        customer=coupon_recharge.customer,
                        product_type="amount"
                    ).delete()
            except CustomerOutstandingReport.DoesNotExist:
                pass

            # Rollback Coupon Items & Stocks
            coupon_items = CustomerCouponItems.objects.filter(customer_coupon=coupon_recharge)
            for item in coupon_items:
                coupon = item.coupon

                # rollback CustomerCouponStock
                try:
                    stock = CustomerCouponStock.objects.get(
                        coupon_method=coupon.coupon_method,
                        customer_id=coupon_recharge.customer.pk,
                        coupon_type_id=coupon.coupon_type_id
                    )
                    stock.count -= Decimal(coupon.no_of_leaflets)
                    stock.save()
                except CustomerCouponStock.DoesNotExist:
                    pass

                # rollback VanCouponStock
                try:
                    van_stock = VanCouponStock.objects.get(
                        created_date=coupon_recharge.created_date.date(),
                        coupon=coupon
                    )
                    van_stock.stock += 1
                    van_stock.sold_count -= 1
                    van_stock.save()
                except VanCouponStock.DoesNotExist:
                    pass

                # reset coupon stock status
                CouponStock.objects.filter(couponbook=coupon).update(coupon_stock="van")

            coupon_items.delete()

            # Rollback Invoice
            try:
                invoice = get_object_or_404(Invoice, invoice_no=coupon_recharge.invoice_no)
                InvoiceItems.objects.filter(invoice=invoice).delete()
                InvoiceDailyCollection.objects.filter(invoice=invoice).delete()
                invoice.delete()
            except Invoice.DoesNotExist:
                pass

            # Rollback Receipt
            try:
                receipt = get_object_or_404(Receipt, invoice_number=coupon_recharge.invoice_no)
                receipt.delete()
            except Receipt.DoesNotExist:
                pass

            # Rollback Cheque Payment
            ChequeCouponPayment.objects.filter(reference_number=coupon_recharge.reference_number).delete()

            # Finally delete CustomerCoupon
            reference_no = coupon_recharge.reference_number
            coupon_recharge.delete()

            log_activity(
                created_by=request.user,
                description=f"Rolled back coupon recharge with ID {pk}, invoice no {reference_no}, and related transactions."
            )

            messages.success(request, "Coupon recharge and all related transactions deleted successfully.")
            return redirect("coupon_recharge")

    except Exception as e:
        messages.error(request, f"Error during rollback: {str(e)}")
        return redirect("coupon_recharge")
    

def route_wise_coupon_report(request):

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = date.today().strftime('%Y-%m-%d')

    # If not provided, set default = today
    if not start_date:
        start_date = today
    if not end_date:
        end_date = today

    filters_issue = {}
    

    filters_sale = {}

    if start_date:
        filters_sale['customer_coupon__created_date__date__gte'] = start_date

    if end_date:
        filters_sale['customer_coupon__created_date__date__lte'] = end_date


    # -------------------------------
    # VAN → ROUTE MAP
    # -------------------------------
    van_route_map = {}
    for vr in Van_Routes.objects.select_related('van', 'routes'):
        van_route_map[vr.van_id] = {
            'route_id': vr.routes.route_id,
            'route_name': vr.routes.route_name
        }

    # -------------------------------
    # ISSUED BOOKS
    # -------------------------------
    issued_qs = (
        Staff_IssueOrders.objects
        .filter(coupon_book__isnull=False, van__isnull=False, **filters_issue)
        .values('van', 'coupon_book')
        .distinct()
    )

    route_issued = {}

    for row in issued_qs:
        van_id = row['van']
        coupon_id = row['coupon_book']

        if van_id not in van_route_map:
            continue

        route_id = van_route_map[van_id]['route_id']
        route_name = van_route_map[van_id]['route_name']

        route_issued.setdefault(route_id, {
            'route_name': route_name,
            'issued_books': set()
        })

        route_issued[route_id]['issued_books'].add(coupon_id)

    for r in route_issued:
        route_issued[r]['issued_count'] = len(route_issued[r]['issued_books'])

    # -------------------------------
    # SOLD BOOKS
    # -------------------------------
    sold_qs = (
        CustomerCouponItems.objects
        .filter(coupon__isnull=False, **filters_sale)
        .values('coupon')
        .distinct()
    )

    # coupon → van map
    issue_map = (
        Staff_IssueOrders.objects
        .filter(coupon_book__isnull=False, van__isnull=False)
        .values('coupon_book', 'van')
        .distinct()
    )

    coupon_van_map = {row['coupon_book']: row['van'] for row in issue_map}

    # -------------------------------
    # SALE AMOUNT
    # -------------------------------
    # -------------------------------
# SALE AMOUNT
# -------------------------------
    sales_amount_qs = (
     CustomerCouponItems.objects
    .filter(**filters_sale)
    .values('coupon')
    .annotate(total_sale=Sum('customer_coupon__net_amount'))
    )

    coupon_sale_amount_map = {}
    for row in sales_amount_qs:
        cid = row['coupon']          # ✅ correct key
        if cid:
            coupon_sale_amount_map[cid] = row['total_sale']



    route_sold = {}

    for row in sold_qs:
        coupon_id = row['coupon']

        if coupon_id not in coupon_van_map:
            continue

        van_id = coupon_van_map[coupon_id]

        if van_id not in van_route_map:
            continue

        route_id = van_route_map[van_id]['route_id']
        route_name = van_route_map[van_id]['route_name']

        route_sold.setdefault(route_id, {
            'route_name': route_name,
            'sold_count': 0,
            'sale_amount': 0
        })

        route_sold[route_id]['sold_count'] += 1
        route_sold[route_id]['sale_amount'] += coupon_sale_amount_map.get(coupon_id, 0)

    # -------------------------------
    # FINAL MERGE
    # -------------------------------
    final_data = []

    all_routes = set(route_issued.keys()) | set(route_sold.keys())

    for route_id in all_routes:
        issued = route_issued.get(route_id, {}).get('issued_count', 0)
        sold = route_sold.get(route_id, {}).get('sold_count', 0)
        sale_amount = route_sold.get(route_id, {}).get('sale_amount', 0)

        route_name = (
            route_issued.get(route_id, {}).get('route_name')
            or route_sold.get(route_id, {}).get('route_name')
        )

        final_data.append({
            'route_id': route_id,
            'route_name': route_name,
            'books_issued': issued,
            'books_sold': sold,
            'sale_amount': sale_amount,
            'balance_books': issued - sold
        })

    context = {
        'routes': final_data,
        'filter_data': {
            'start_date': start_date,
            'end_date': end_date
        },
        'page_title': 'Route Wise Coupon Book Report'
    }
    

    return render(request, 'coupon_management/route_wise_coupon_report.html', context)


def route_balance_books(request, route_id):

    vans = Van_Routes.objects.filter(routes_id=route_id).values_list('van_id', flat=True)

    issued = set(
        Staff_IssueOrders.objects
        .filter(van_id__in=vans, coupon_book__isnull=False)
        .values_list('coupon_book', flat=True)
    )

    sold = set(
        CustomerCouponItems.objects
        .filter(coupon__in=issued)
        .values_list('coupon', flat=True)
    )

    balance_ids = issued - sold

    balance_books = NewCoupon.objects.filter(coupon_id__in=balance_ids)

    context = {
        'route': RouteMaster.objects.get(route_id=route_id),
        'balance_books': balance_books,
        'page_title': 'Balance Coupon Books'
    }

    return render(request, 'coupon_management/route_balance_books.html', context)
